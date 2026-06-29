"""
BridgeLine Partners - Daily MIS Generator
Reads BridgeLine Accounts.xlsx -> outputs DD-Mon-YYYY BridgeLine MIS Package.zip
containing: Consolidated MIS PDF, per-cluster MIS PDFs, per-case Disbursement Memo PDFs.

Standing rules:
  - EXCLUDED_CLUSTERS = ['Mandya'] - never appears in cluster MIS or open-cases tables
  - all_cases_full (incl. Mandya) used ONLY for consolidated dashboard totals
  - CONFIRMED_CLOSED  = []         - reset daily
  - Never recalculate Charges / GST / Total - use sheet values
  - Balance column mixed sign convention: abs() applied universally
  - Debit Note: extract UTR (HDFC[A-Z0-9]+) from col V — do NOT show raw SMS text
  - Beat April: PERMANENTLY REMOVED (04-May-2026)
  - Available for Disbursement = Total Invested - ABS(sum of all open outstanding)
  - Pending Cases = count where ABS(balance) >= 1 across ALL clusters
"""

import sys, os, io, zipfile, datetime, math, tempfile, difflib
from collections import defaultdict
from urllib.parse import quote
import openpyxl
from fpdf import FPDF

# --- CONSTANTS ---------------------------------------------------------------
EXCEL_PATH        = sys.argv[1] if len(sys.argv) > 1 else '/tmp/BridgeLine_Accounts.xlsx'
LOGO_PATH         = os.path.join(tempfile.gettempdir(), 'bl_logo.png')
SIGNATURE_PATH    = os.path.join(tempfile.gettempdir(), 'prem_signature.png')
QR_PATH           = os.path.join(tempfile.gettempdir(), 'bl_qr.png')
OUTPUT_DIR        = sys.argv[2] if len(sys.argv) > 2 else '/tmp'
CONFIRMED_CLOSED  = []
EXCLUDED_CLUSTERS = ['Mandya']
TODAY             = datetime.date.today()
ROI_TARGET_PCT    = 4.0
COLLECTION_CARD_THRESHOLD = 100000   # Collection Card generated if disbursed amount OR outstanding balance is below this

# --- DESIGN TOKENS -----------------------------------------------------------
C_NAVY      = (10,  24,  40)
C_PRI_NAVY  = (44,  62, 107)
C_GOLD_RULE = (184, 150,  46)
C_GOLD_LBL  = (212, 175,  90)
C_CREAM     = (240, 235, 225)
C_LT_GOLD   = (249, 243, 227)
C_LT_SLATE  = (237, 240, 246)
C_RULE      = (200, 196, 188)
C_TEXT_DARK = (26,  26,  46)
C_TEXT_MED  = (74,  74, 106)
C_RED       = (192,  57,  43)
C_HDR_TXT   = (200, 216, 240)
C_FOOTER_BG = (247, 244, 239)
C_MEMO_CREAM= (249, 246, 240)
C_WHITE     = (255, 255, 255)
C_GREEN     = (39,  174,  96)
C_ORANGE    = (230, 126,  34)

HDR_H  = 48
FOOT_H = 8
L_MAR  = 10
R_MAR  = 10

CO_NAME  = "BridgeLine Partners"
CO_ADDR1 = "My Office Space, BM Habitat Mall, 2nd Floor"
CO_ADDR2 = "Jayalakshmi Puram, Mysore 570012"
CO_PH    = "+91 99862 88166  |  +91 98451 22023"
CO_EMAIL = "principal@bridgelinepartners.in"
CO_GSTIN = "29ABGFB6346P1ZR"
PAYEE_NAME = "PREM NARAYAN S A"
UPI_ID     = "9845122023.1@hdfc"

def build_upi_link(amount, note, case_id=''):
    """Tappable upi://pay deep link, prefilled with payee + amount + a short note.
    Opens the customer's UPI app directly when tapped on a phone (no scanning needed).

    Includes mc=0000 (personal/non-merchant code, matching the bank's own QR -- see
    bl_qr.png's decoded payload) and a unique tr (transaction reference) -- both
    recommended by NPCI's spec once 'am' is present.

    pn/tn ARE percent-encoded (corrected 25-Jun-2026, 2nd pass). First fix attempt
    matched the bank's QR by using raw, unencoded spaces in pn/tn -- still failed on
    a real-device retest. The QR can get away with raw spaces because a scanning app
    reads the payload as a plain string with its own parser; a *tapped* link instead
    goes through the OS's standard URI resolver before reaching the UPI app, and an
    unencoded space is not a valid URI character there -- every NPCI/PSP integration
    guide (Razorpay, EximPe, GitHub upi-deeplink-builder, Android Uri.Builder docs)
    builds pn/tn through a URL-encoding call for exactly this reason. Back to quote()."""
    try:
        amt = f'{float(amount):.2f}'
    except (TypeError, ValueError):
        amt = ''
    tr = str(case_id).replace('-', '').replace(' ', '')[:35] or 'BLP'
    pn = quote(PAYEE_NAME, safe='')
    # ver/mode/purpose/qrMedium match the bank's own static QR exactly (decoded
    # 25-Jun-2026: "upi://pay?ver=01&mode=01&purpose=00&mc=0000&qrMedium=02&
    # pa=...&pn=..."). Without these, real-device testing showed UPI apps accept
    # the link and reach PIN entry, but the transaction is declined afterward
    # with a generic error -- HDFC's backend appears to require these fields.
    params = f'ver=01&mode=01&purpose=00&mc=0000&qrMedium=02&pa={UPI_ID}&pn={pn}&tr={tr}&cu=INR'
    if amt:
        params += f'&am={amt}'
    if note:
        params += f'&tn={quote(note, safe="")}'
    return f'upi://pay?{params}'

PAY_PAGE_BASE_URL = 'https://bridgeline-pay.vercel.app'

def build_pay_page_url(case_id):
    """URL for the hosted payment page (bridgeline-pay), used as the *tap*
    target for the QR area and PAY NOW button instead of a raw upi://pay link.

    A bare upi://pay deep link tapped inside a PDF viewer is unreliable: PDF
    readers don't implement Android's intent://-with-package-name convention
    (confirmed via device testing -- it's Chrome-specific syntax), and iOS has
    no mechanism at all to pick which UPI app handles a upi:// link (confirmed
    via NPCI's own spec docs -- whichever app's intent filter wins the race
    gets it, which on a real test device was WhatsApp, which mishandled it).
    The hosted page solves this: Android gets real Pay-with-[App] buttons via
    intent://, iOS gets a best-effort upi:// button plus an always-visible
    QR/copy-ID fallback. The static QR *image* embedded in the PDF (bl_qr.png)
    is untouched -- camera-scanning a UPI QR has always worked reliably; only
    the tappable link needed to change.
    """
    return f'{PAY_PAGE_BASE_URL}/i/{case_id}'

# --- HELPERS -----------------------------------------------------------------
def inr(n):
    try: n = float(n)
    except: return str(n)
    if n < 0: return '-' + inr(-n)
    n_int = int(round(n))
    s = str(n_int)
    if len(s) <= 3: return s
    last3 = s[-3:]
    rest = s[:-3]
    parts = []
    while len(rest) > 2:
        parts.append(rest[-2:])
        rest = rest[:-2]
    if rest: parts.append(rest)
    parts.reverse()
    return ','.join(parts) + ',' + last3

def inr_dec(n, d=2):
    try: n = float(n)
    except: return str(n)
    dec_part = f"{abs(n) % 1:.{d}f}"[1:]
    sign = '-' if n < 0 else ''
    return sign + inr(int(abs(n))) + dec_part

def std_num(n):
    try: return f"{int(round(float(n))):,}"
    except: return str(n)

def fmt_date(d):
    if isinstance(d, datetime.datetime): d = d.date()
    if isinstance(d, datetime.date): return d.strftime('%d-%b-%Y')
    return str(d) if d else '-'

def parse_date(raw):
    """Robustly parse date from datetime, date, or string (DD-MM-YYYY / YYYY-MM-DD / DD/MM/YYYY)."""
    if isinstance(raw, datetime.datetime): return raw.date()
    if isinstance(raw, datetime.date): return raw
    if isinstance(raw, str) and raw.strip():
        for fmt in ('%d-%m-%Y', '%Y-%m-%d', '%d/%m/%Y', '%d-%b-%Y', '%m/%d/%Y'):
            try: return datetime.datetime.strptime(raw.strip(), fmt).date()
            except: pass
    return raw  # return as-is if unparseable

def extract_utr(raw):
    """Extract UTR reference (HDFC[A-Z0-9]+) from HDFC SMS alert text in col V."""
    import re
    if not raw or str(raw).strip() in ('', 'None', '-'):
        return '-'
    m = re.search(r'HDFC[A-Z0-9]+', str(raw))
    return m.group(0) if m else str(raw).strip()[:40]

def charge_rate_label(tat):
    try: t = float(str(tat).split()[0])
    except: return '0.50%'
    if t <= 2:   return '0.50%'
    elif t == 3: return '1.00%'
    elif t == 4: return '1.50%'
    else:        return f'{0.5 * t:.2f}%'

def status_color(status):
    if status == 'Closed':     return C_GREEN
    if status == 'Critical':   return C_RED
    if status == 'Follow Up!': return C_ORANGE
    return C_PRI_NAVY


# --- M COLL LOADER -----------------------------------------------------------
def load_mcoll(wb):
    """
    Read M Coll sheet (cols A-D: DisbID | CollDate | CollAmt | Notes).
    Returns {disb_id: {total_collected, latest_date, notes_list}}.
    Cases not in M Coll are handled by Accounts sheet cols L/M/O as before.
    """
    if 'M Coll' not in wb.sheetnames:
        return {}
    ws = wb['M Coll']
    raw_map = defaultdict(lambda: {'amounts': [], 'dates': [], 'notes': []})
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]: continue
        disb_id  = str(row[0]).strip()
        cdate    = parse_date(row[1])
        camt     = float(row[2]) if row[2] is not None else 0.0
        note     = str(row[3]).strip() if len(row) > 3 and row[3] else ''
        raw_map[disb_id]['amounts'].append(camt)
        raw_map[disb_id]['dates'].append(cdate)
        raw_map[disb_id]['notes'].append(note)
    result = {}
    for disb_id, data in raw_map.items():
        total_collected = sum(data['amounts'])
        if total_collected == 0:
            continue   # skip rows with no valid collection amounts
        valid_dates = [d for d in data['dates'] if isinstance(d, datetime.date)]
        result[disb_id] = {
            'total_collected': total_collected,
            'latest_date':     max(valid_dates) if valid_dates else None,
            'notes':           data['notes'],
        }
    return result

# --- CONTACT LOADING ---------------------------------------------------------
def load_contacts(wb):
    """
    Parse Contact sheet -> (cluster_mgrs, branch_contacts).
    cluster_mgrs    = { 'Mysore': {'name': '...', 'phone': '...'}, ... }
    branch_contacts = { ('Mysore', 'KUVEMPU NAGAR'): {'name': '...', 'phone': '...'}, ... }
    """
    import re as _re
    if 'Contact' not in wb.sheetnames:
        return {}, {}
    ws = wb['Contact']
    CLUSTER_ALIASES = {
        'MYSORE': 'Mysore', 'MANAGLORE': 'Mangalore', 'MANGALORE': 'Mangalore',
        'HASSAN': 'Hassan', 'BELLARY': 'Bellary', 'HUBLI': 'Hubli',
        'MANGALORE SAL': 'Mangalore Sal',
    }
    def _fmt_phone(p):
        s = str(int(p)) if isinstance(p, float) else str(p)
        s = _re.sub(r'\D', '', s)
        return (s[:5] + ' ' + s[5:]) if len(s) == 10 else s

    def _cell(row, i):
        return row[i] if i < len(row) else None

    cluster_mgrs    = {}
    branch_contacts = {}
    rows = list(ws.iter_rows(min_row=1, values_only=True))
    header = [str(c).strip().upper() if c else '' for c in (rows[0] if rows else [])]
    new_format = len(header) >= 2 and header[0] == 'CLUSTER' and header[1] == 'NAME'

    if new_format:
        # New layout: Cluster | Name | Designation | Branch | Phone | Email
        for row in rows[1:]:
            cluster_raw = str(_cell(row, 0)).strip() if _cell(row, 0) else ''
            name   = str(_cell(row, 1)).strip() if _cell(row, 1) else ''
            desig  = str(_cell(row, 2)).strip() if _cell(row, 2) else ''
            branch = str(_cell(row, 3)).strip() if _cell(row, 3) else ''
            phone  = _cell(row, 4)
            if not cluster_raw or not name or not phone: continue
            cluster = CLUSTER_ALIASES.get(cluster_raw.upper(), cluster_raw)
            ph = _fmt_phone(phone)
            if 'CLUSTER MANAGER' in desig.upper():
                cluster_mgrs[cluster] = {'name': name.title(), 'phone': ph}
            elif branch:
                branch_contacts[(cluster, branch.upper())] = {'name': name.title(), 'phone': ph}
    else:
        # Old layout: Name | Designation | Branch | Phone, with cluster name as its own header row
        current_cluster = None
        for row in rows:
            name   = str(_cell(row, 0)).strip() if _cell(row, 0) else ''
            desig  = str(_cell(row, 1)).strip() if _cell(row, 1) else ''
            branch = str(_cell(row, 2)).strip() if _cell(row, 2) else ''
            phone  = _cell(row, 3)
            if not name or name == 'None': continue
            name_up = name.upper()
            if name_up in CLUSTER_ALIASES and not phone:
                current_cluster = CLUSTER_ALIASES[name_up]; continue
            if name_up in CLUSTER_ALIASES and desig.upper() == 'DESIGNATION':
                current_cluster = CLUSTER_ALIASES[name_up]; continue
            if not current_cluster or not phone: continue
            ph = _fmt_phone(phone)
            if 'CLUSTER MANAGER' in desig.upper():
                cluster_mgrs[current_cluster] = {'name': name.title(), 'phone': ph}
            elif branch:
                branch_contacts[(current_cluster, branch.upper())] = {'name': name.title(), 'phone': ph}
    return cluster_mgrs, branch_contacts

def find_branch_contact(cluster, branch, branch_contacts, cluster_mgrs):
    """Match branch name; fall back to cluster manager (returns info, is_fallback).

    Match order: (1) exact, (2) word-overlap, (3) fuzzy spelling match
    (handles Contact-sheet typos like 'Valancia'/'Valencia', 'Shimogga'/'Shimoga'
    that would otherwise silently fall back to the cluster manager), (4) cluster
    manager fallback.
    """
    branch_up = branch.upper()
    key = (cluster, branch_up)
    if key in branch_contacts:
        return branch_contacts[key], False
    branch_words = set(branch_up.split())
    for (cl, br), info in branch_contacts.items():
        if cl != cluster: continue
        if branch_words & set(br.split()):
            return info, False
    # Fuzzy fallback: catch near-identical spellings within the same cluster
    # before giving up to the cluster manager.
    best_ratio, best_info = 0.0, None
    for (cl, br), info in branch_contacts.items():
        if cl != cluster: continue
        ratio = difflib.SequenceMatcher(None, branch_up, br).ratio()
        if ratio > best_ratio:
            best_ratio, best_info = ratio, info
    if best_info is not None and best_ratio >= 0.8:
        return best_info, False
    mgr = cluster_mgrs.get(cluster)
    if mgr: return mgr, True
    return {'name': '-', 'phone': '-'}, True

# --- DATA LOADING ------------------------------------------------------------
def load_data(path):
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb['Accounts']
    rows = []
    for raw in ws.iter_rows(min_row=3, values_only=True):
        # Read 22 cols — col A(0) through col V(21) to capture debit note
        row = list(raw[:22])
        row += [None] * (22 - len(row))
        if not row[0] or str(row[0]).strip() == '': continue
        rows.append(row)
    # Read DashBoard for Total Invested only — all other metrics computed from raw data
    db_raw = {}
    if 'DashBoard' in wb.sheetnames:
        ws_db = wb['DashBoard']
        for raw in ws_db.iter_rows(min_row=1, max_row=11, values_only=True):
            label = raw[0] if len(raw) > 0 else None
            val   = raw[1] if len(raw) > 1 else None
            if label: db_raw[str(label).strip()] = val
    mcoll    = load_mcoll(wb)
    contacts = load_contacts(wb)
    wb.close()
    return rows, db_raw, mcoll, contacts

def parse_cases(rows, mcoll=None):
    all_cases = []
    for row in rows:
        disb_id    = str(row[0]).strip()
        disb_date  = parse_date(row[1])
        cust_name  = str(row[2]).strip() if row[2] else ''
        chq_no     = str(row[3]).strip() if row[3] else '-'
        company    = str(row[4]).strip() if row[4] else ''
        cluster    = str(row[5]).strip() if row[5] else ''
        branch     = str(row[6]).strip() if row[6] else ''
        amount     = float(row[7])  if row[7]  is not None else 0
        charges    = float(row[8])  if row[8]  is not None else 0
        gst        = float(row[9])  if row[9]  is not None else 0
        total      = float(row[10]) if row[10] is not None else 0
        coll_date  = parse_date(row[11])
        coll_amt   = float(row[12]) if row[12] not in (None, '') else 0
        discount   = float(row[13]) if row[13] not in (None, '') else 0  # col N — negative in sheet
        balance_raw = float(row[14]) if row[14] not in (None, '') else 0  # col O — mixed sign convention
        tat_raw    = row[15]
        status     = str(row[18]).strip() if row[18] else ''
        debit_note = extract_utr(row[21])  # col V — len guaranteed 22 by load_data

        # Balance: M Coll cases use total(col K) − total_collected (exact, can be negative).
        # Non-M Coll cases use abs(col O) per existing mixed-sign convention.
        # EXCEPTION (fixed 22-Jun-2026): Accounts sheet Status is authoritative — if a case
        # is marked 'Closed' there, trust its own balance/collection columns even if it also
        # has a (possibly stale/incomplete) M Coll row. Prevents stale M Coll sub-ledger
        # entries from resurrecting already-closed cases as phantom outstanding.
        if mcoll and disb_id in mcoll and balance_raw != 0 and status != 'Closed':
            mc          = mcoll[disb_id]
            # M Coll's own running total can be incomplete (e.g. early instalments
            # entered straight into Accounts before per-instalment M Coll logging
            # existed for that case) — never let it understate what Accounts'
            # own cumulative "Collected Amount" column already shows.
            coll_amt    = max(mc['total_collected'], coll_amt)
            if mc['latest_date']:
                coll_date = mc['latest_date']
            balance_display = total - coll_amt   # exact; can be negative if over-collected
        else:
            balance_display = abs(balance_raw)
        days_out = (TODAY - disb_date).days if isinstance(disb_date, datetime.date) else 0
        try: tat_num = float(str(tat_raw).split()[0])
        except: tat_num = days_out

        all_cases.append({
            'id': disb_id, 'date': disb_date, 'customer': cust_name,
            'chq': chq_no, 'company': company, 'cluster': cluster,
            'branch': branch, 'amount': amount, 'charges': charges,
            'gst': gst, 'total': total, 'coll_date': coll_date,
            'coll_amt': coll_amt, 'discount': discount,
            'balance': balance_display, 'balance_raw': balance_raw,
            'tat': tat_num, 'tat_raw': tat_raw, 'days_out': days_out,
            'status': status, 'debit_note': debit_note,
        })

    # all_cases_full: every cluster including Mandya — used ONLY for consolidated dashboard totals
    all_cases_full = list(all_cases)
    all_cases = [c for c in all_cases if c['cluster'] not in EXCLUDED_CLUSTERS]
    open_cases = [c for c in all_cases
                  if c['balance'] >= 1.0
                  and c['status'] != 'Closed'
                  and c['customer'] not in CONFIRMED_CLOSED]
    open_cases.sort(key=lambda c: (c['date'] if isinstance(c['date'], datetime.date) else datetime.date.min, c['cluster']))
    return open_cases, all_cases, all_cases_full

# --- CORRECT METRICS COMPUTATION ---------------------------------------------
def compute_dashboard_metrics(all_cases_full, db_raw):
    """
    Compute correct dashboard metrics directly from raw case data (ALL clusters incl. Mandya).
    DashBoard sheet formulas are incorrect — do NOT rely on them.
    The only value trusted from DashBoard is Total Invested (hardcoded there).
    """
    total_invested = 8500000  # 82.5L base + 2.5L Harsha (07-May-2026) = 85L

    total_disbursed = sum(c['amount']   for c in all_cases_full)
    total_charges   = sum(c['charges']  for c in all_cases_full)
    total_gst       = sum(c['gst']      for c in all_cases_full)
    total_discount  = sum(c['discount'] for c in all_cases_full)  # negative in sheet
    net_charges     = total_charges + total_discount

    # Open = ABS(balance) >= 1 AND not marked Closed in Accounts sheet (all clusters incl. Mandya).
    # Status='Closed' is authoritative (fixed 22-Jun-2026) — small rounding residuals on an
    # already-closed case must not resurrect it as "pending."
    open_all          = [c for c in all_cases_full if c['balance'] >= 1.0 and c['status'] != 'Closed']
    total_outstanding = sum(c['balance'] for c in open_all)

    # Available for Disbursement = Capital - Outstanding (can be negative = overdeployed)
    available = total_invested - total_outstanding

    pending_cases = len(open_all)
    overdue_cases = len([c for c in all_cases_full if c['status'] == 'Critical'])

    # MTD charges (current month, all clusters)
    mtd_cases   = [c for c in all_cases_full
                   if isinstance(c['date'], datetime.date)
                   and c['date'].month == TODAY.month
                   and c['date'].year  == TODAY.year]
    mtd_charges = sum(c['charges'] for c in mtd_cases)
    mtd_roi     = (mtd_charges / total_invested) if total_invested > 0 else 0

    # All-time ROI
    all_time_roi = (net_charges / total_invested) if total_invested > 0 else 0

    return {
        'Total Invested':             total_invested,
        'Total Disbursed':            total_disbursed,
        'Available for Disbursement': available,
        'Total Charges Earned':       total_charges,
        'Total GST':                  total_gst,
        'Net Charges Earned':         net_charges,
        'Total Pending Cases':        pending_cases,
        'Total Overdue Cases':        overdue_cases,
        'MTD Charges':                mtd_charges,
        'MTD ROI':                    mtd_roi,
        'ROI':                        all_time_roi,
    }

def _dashboard_label_map(metrics):
    return [
        ('Total Invested',             metrics['Total Invested']),
        ('Total Disbursed',            metrics['Total Disbursed']),
        ('Available for Disbursement', metrics['Available for Disbursement']),
        ('Total Charges Earned',       metrics['Total Charges Earned']),
        ('Total GST',                  metrics['Total GST']),
        ('Net Charges Earned',         metrics['Net Charges Earned']),
        ('Total Pending Cases',        metrics['Total Pending Cases']),
        ('Total Overdue Cases',        metrics['Total Overdue Cases']),
        ('MTD Charges',                metrics['MTD Charges']),
        ('MTD ROI',                    metrics['MTD ROI']),
        ('ROI',                        metrics['ROI']),
    ]


def write_claude_dashboard(excel_path, metrics, mcoll, out_dir):
    """
    Write computed metrics as literal values to a Claude_Dashboard sheet.
    Saves updated workbook to BridgeLine MIS workspace folder.
    """
    wb = openpyxl.load_workbook(excel_path)
    # Col O is now a live formula in the sheet — no write-back needed
    if 'Claude_Dashboard' in wb.sheetnames:
        del wb['Claude_Dashboard']
    ws = wb.create_sheet('Claude_Dashboard')
    ws.cell(row=1, column=1, value='Metric')
    ws.cell(row=1, column=2, value='Value')
    for i, (lbl, val) in enumerate(_dashboard_label_map(metrics), start=2):
        ws.cell(row=i, column=1, value=lbl)
        v = round(float(val), 6) if isinstance(val, float) else val
        ws.cell(row=i, column=2, value=v)
    out_path = os.path.join(out_dir, 'BridgeLine Accounts.xlsx')
    wb.save(out_path)
    wb.close()
    print(f'Claude_Dashboard written -> {out_path}')
    return out_path


def write_claude_dashboard_to_sheet(sh, metrics):
    """
    gspread equivalent of write_claude_dashboard() for the live-Google-Sheet path:
    recreates the Claude_Dashboard worksheet directly on the live spreadsheet
    instead of writing a separate local Excel file.
    """
    try:
        sh.del_worksheet(sh.worksheet('Claude_Dashboard'))
    except Exception:
        pass
    ws = sh.add_worksheet(title='Claude_Dashboard', rows=20, cols=2)
    rows = [['Metric', 'Value']]
    for lbl, val in _dashboard_label_map(metrics):
        v = round(float(val), 6) if isinstance(val, float) else val
        rows.append([lbl, v])
    ws.update('A1', rows)
    print('Claude_Dashboard written to live sheet')

# --- PDF BASE CLASS ----------------------------------------------------------
class BLPdf(FPDF):
    def __init__(self, subtitle='', report_date=None):
        super().__init__('P', 'mm', 'A4')
        self.set_auto_page_break(False)
        self.set_margins(0, 0, 0)
        self.alias_nb_pages('{NB}')
        self._subtitle    = subtitle
        self._report_date = report_date or fmt_date(TODAY)
        self._page_count  = 0

    def set_page_count(self, n): self._page_count = n

    def draw_header(self):
        pw = self.w
        self.set_fill_color(*C_NAVY)
        self.rect(0, 0, pw, HDR_H, 'F')
        self.set_fill_color(*C_GOLD_RULE)
        self.rect(0, HDR_H - 0.8, pw, 0.8, 'F')

        logo_h = HDR_H
        logo_w = round(logo_h * (1414 / 2000), 1)
        lx = L_MAR - 2
        try:
            if os.path.getsize(LOGO_PATH) > 0:
                self.image(LOGO_PATH, x=lx, y=0, w=logo_w, h=logo_h)
        except Exception:
            pass  # Logo missing/corrupt — skip gracefully

        tx = lx + logo_w + 3
        self.set_xy(tx, 22)
        self.set_font('Helvetica', '', 8)
        self.set_text_color(*C_HDR_TXT)
        self.cell(80, 4, self._subtitle, border=0, ln=False)
        self.set_xy(tx, 30)
        self.set_font('Helvetica', 'B', 7)
        self.set_text_color(*C_GOLD_LBL)
        self.cell(27, 4, 'REPORTING DATE', border=0, ln=False)
        self.set_xy(tx + 29, 30)
        self.set_font('Helvetica', 'B', 7)
        self.set_text_color(*C_WHITE)
        self.cell(55, 4, self._report_date, border=0, ln=False)

        col_w   = 85
        x_right = pw - col_w - 6
        lbl_w   = 20
        val_w   = col_w - lbl_w - 2
        fields = [
            ('ADDRESS', CO_ADDR1, 8),
            ('',        CO_ADDR2, 13),
            ('CONTACT', CO_PH,    21),
            ('EMAIL',   CO_EMAIL, 29),
            ('GSTIN',   CO_GSTIN, 37),
        ]
        for lbl, val, y in fields:
            if lbl:
                self.set_xy(x_right, y)
                self.set_font('Helvetica', 'B', 6.5)
                self.set_text_color(*C_GOLD_LBL)
                self.cell(lbl_w, 4, lbl, align='R', border=0, ln=False)
            self.set_xy(x_right + lbl_w + 1, y)
            self.set_font('Helvetica', '', 7)
            self.set_text_color(*C_WHITE)
            self.cell(val_w, 4, val, align='R', border=0, ln=False)

        self.set_text_color(*C_TEXT_DARK)
        self.set_y(HDR_H + 5)

    def draw_footer(self):
        pw, ph = self.w, self.h
        y = ph - FOOT_H
        self.set_fill_color(*C_FOOTER_BG)
        self.rect(0, y, pw, FOOT_H, 'F')
        self.set_fill_color(*C_GOLD_RULE)
        self.rect(0, y, pw, 0.4, 'F')
        self.set_xy(8, y + 2)
        self.set_font('Helvetica', 'B', 7)
        self.set_text_color(*C_PRI_NAVY)
        self.cell(60, 4, CO_NAME, border=0, ln=False)
        self.set_xy((pw - 40) / 2, y + 2)
        self.set_font('Helvetica', '', 7)
        self.set_text_color(*C_TEXT_MED)
        self.cell(40, 4, 'Page ' + str(self.page_no()) + ' of {NB}', align='C', border=0, ln=False)
        self.set_xy(pw - 72, y + 2)
        self.set_font('Helvetica', '', 7)
        self.set_text_color(*C_TEXT_MED)
        self.cell(64, 4, f'Confidential | Generated: {fmt_date(TODAY)}', align='R', border=0, ln=False)
        self.set_text_color(*C_TEXT_DARK)

    def header(self): self.draw_header()
    def footer(self): self.draw_footer()
    def content_top(self): return HDR_H + 5
    def content_bottom(self): return self.h - FOOT_H - 3
    def inner_w(self): return self.w - L_MAR - R_MAR

    def section_title(self, title):
        x, w = L_MAR, self.inner_w()
        y = self.get_y()
        self.set_fill_color(*C_LT_SLATE)
        self.rect(x, y, w, 7, 'F')
        self.set_fill_color(*C_GOLD_RULE)
        self.rect(x, y, 1.5, 7, 'F')
        self.set_xy(x + 3.5, y + 1.5)
        self.set_font('Helvetica', 'B', 8.5)
        self.set_text_color(*C_PRI_NAVY)
        self.cell(w - 5, 4, title.upper(), border=0, ln=False)
        self.set_text_color(*C_TEXT_DARK)
        self.set_y(y + 8)

# --- KPI STRIP ---------------------------------------------------------------
def draw_kpi_strip(pdf, open_cases, metrics=None, all_cases=None):
    total_principal   = sum(c['amount']  for c in open_cases)
    total_outstanding = sum(c['balance'] for c in open_cases)
    boxes = [
        ('OPEN CASES',        str(len(open_cases))),
        ('TOTAL PRINCIPAL',   'Rs ' + inr(total_principal)),
        ('TOTAL OUTSTANDING', 'Rs ' + inr(total_outstanding)),
    ]
    if metrics is not None:
        boxes.append(('NET CHARGES (MTD)', 'Rs ' + inr(metrics.get('MTD Charges', 0))))
    elif all_cases is not None:
        _mtd = [c for c in all_cases
                if isinstance(c['date'], datetime.date)
                and c['date'].month == TODAY.month
                and c['date'].year  == TODAY.year]
        boxes.append(('NET CHARGES (MTD)', 'Rs ' + inr(sum(c['charges'] for c in _mtd))))

    n     = len(boxes)
    gap   = 3
    box_w = (pdf.inner_w() - (n - 1) * gap) / n
    box_h = 22
    y     = pdf.content_top()

    for i, (lbl, val) in enumerate(boxes):
        x = L_MAR + i * (box_w + gap)
        pdf.set_fill_color(*C_LT_SLATE)
        pdf.rect(x, y, box_w, box_h, 'F')
        pdf.set_fill_color(*C_GOLD_RULE)
        pdf.rect(x, y, box_w, 1, 'F')
        pdf.set_xy(x, y + 3)
        pdf.set_font('Helvetica', 'B', 6.5)
        pdf.set_text_color(*C_TEXT_MED)
        pdf.cell(box_w, 4, lbl, align='C', border=0, ln=False)
        val_fsize = 13
        pdf.set_font('Helvetica', 'B', val_fsize)
        while pdf.get_string_width(val) > box_w - 4 and val_fsize > 7:
            val_fsize -= 0.5
            pdf.set_font('Helvetica', 'B', val_fsize)
        pdf.set_xy(x, y + 10)
        pdf.set_text_color(*C_PRI_NAVY)
        pdf.cell(box_w, 6, val, align='C', border=0, ln=False)

    pdf.set_text_color(*C_TEXT_DARK)
    pdf.set_y(y + box_h + 5)

# --- OPEN CASES TABLE --------------------------------------------------------
OPEN_COLS = [
    ('#',           6,  'C'),
    ('DATE',       18,  'C'),
    ('CUSTOMER',   30,  'L'),
    ('CLUSTER',    18,  'C'),
    ('BRANCH',     20,  'C'),
    ('PRINCIPAL',  22,  'R'),
    ('CHARGES*',   15,  'R'),
    ('GST',        11,  'R'),
    ('OUTSTANDING',22,  'R'),
    ('DAYS',        8,  'C'),
    ('RATE',       11,  'C'),
    ('STATUS',     15,  'C'),
]

def draw_open_cases_table(pdf, open_cases, cluster_filter=None):
    cases = [c for c in open_cases if cluster_filter is None or c['cluster'] == cluster_filter]
    if not cases: return

    lbl = 'Open Cases' + (f' - {cluster_filter} Cluster' if cluster_filter else ' - All Clusters')
    pdf.section_title(lbl)
    x0      = L_MAR
    rh      = 6.5
    total_w = sum(c[1] for c in OPEN_COLS)

    pdf.set_fill_color(*C_PRI_NAVY)
    hdr_y = pdf.get_y()
    pdf.rect(x0, hdr_y, total_w, rh, 'F')
    xc = x0
    for col, w, align in OPEN_COLS:
        pdf.set_font('Helvetica', 'B', 6.5)
        pdf.set_text_color(*C_WHITE)
        pdf.set_xy(xc, hdr_y + 1)
        pdf.cell(w, rh - 2, col, align=align, border=0, ln=False)
        xc += w
    pdf.set_y(hdr_y + rh)

    for idx, c in enumerate(cases):
        y = pdf.get_y()
        if idx % 2 == 0: pdf.set_fill_color(*C_LT_GOLD)
        else:            pdf.set_fill_color(*C_WHITE)
        pdf.rect(x0, y, total_w, rh, 'F')

        row_vals = [
            str(idx + 1), fmt_date(c['date']), c['customer'], c['cluster'], c['branch'],
            'Rs ' + inr(c['amount']), 'Rs ' + inr(c['charges']), 'Rs ' + inr(c['gst']),
            'Rs ' + inr(c['balance']), str(c['days_out']), charge_rate_label(c['tat']), c['status'],
        ]
        xc = x0
        for (col, w, align), val in zip(OPEN_COLS, row_vals):
            if col == 'STATUS':
                pdf.set_font('Helvetica', 'B', 6.5)
                pdf.set_text_color(*status_color(c['status']))
            elif col == 'OUTSTANDING':
                pdf.set_font('Helvetica', 'B', 6.5)
                pdf.set_text_color(*C_TEXT_DARK)
            elif col == 'DAYS' and c['days_out'] > 2:
                pdf.set_font('Helvetica', 'B', 6.5)
                pdf.set_text_color(*C_RED)
            else:
                pdf.set_font('Helvetica', '', 6.5)
                pdf.set_text_color(*C_TEXT_DARK)
            pdf.set_xy(xc, y + 1)
            pdf.cell(w, rh - 2, val, align=align, border=0, ln=False)
            xc += w
        pdf.set_y(y + rh)

    pdf.set_fill_color(*C_GOLD_RULE)
    pdf.rect(x0, pdf.get_y(), total_w, 0.4, 'F')
    pdf.set_xy(x0, pdf.get_y() + 1)
    pdf.set_font('Helvetica', 'I', 6)
    pdf.set_text_color(*C_TEXT_MED)
    pdf.cell(total_w, 4, '* Charges fixed at disbursement per BLP/CIR/001/2026-27. Not recalculated daily.', border=0, ln=False)
    pdf.set_text_color(*C_TEXT_DARK)
    pdf.set_y(pdf.get_y() + 5)

# --- CLUSTER SUMMARY TABLE ---------------------------------------------------
def draw_cluster_summary(pdf, open_cases):
    pdf.section_title('Cluster Summary')
    x0  = L_MAR
    rh  = 7
    cols = [
        ('CLUSTER',     38, 'L'), ('CASES',       18, 'C'), ('PRINCIPAL',   35, 'R'),
        ('CHARGES',     28, 'R'), ('GST',         18, 'R'), ('OUTSTANDING', 37, 'R'),
    ]
    total_w = sum(c[1] for c in cols)

    pdf.set_fill_color(*C_PRI_NAVY)
    hdr_y = pdf.get_y()
    pdf.rect(x0, hdr_y, total_w, rh, 'F')
    xc = x0
    for col, w, align in cols:
        pdf.set_font('Helvetica', 'B', 7)
        pdf.set_text_color(*C_WHITE)
        pdf.set_xy(xc, hdr_y + 1.5)
        pdf.cell(w, rh - 3, col, align=align, border=0, ln=False)
        xc += w
    pdf.set_y(hdr_y + rh)

    by_cluster = defaultdict(list)
    for c in open_cases: by_cluster[c['cluster']].append(c)
    tot = [0, 0, 0, 0, 0]

    for idx, (cl, cases) in enumerate(sorted(by_cluster.items())):
        y = pdf.get_y()
        if idx % 2 == 0: pdf.set_fill_color(*C_LT_SLATE)
        else:            pdf.set_fill_color(*C_WHITE)
        pdf.rect(x0, y, total_w, rh, 'F')

        n   = len(cases)
        amt = sum(c['amount']  for c in cases)
        chg = sum(c['charges'] for c in cases)
        gst = sum(c['gst']     for c in cases)
        bal = sum(c['balance'] for c in cases)
        tot[0] += n; tot[1] += amt; tot[2] += chg; tot[3] += gst; tot[4] += bal
        vals = [cl, str(n), 'Rs ' + inr(amt), 'Rs ' + inr(chg), 'Rs ' + inr(gst), 'Rs ' + inr(bal)]
        xc = x0
        for (col, w, align), val in zip(cols, vals):
            is_bal = (col == 'OUTSTANDING')
            pdf.set_font('Helvetica', 'B' if is_bal else '', 7)
            pdf.set_text_color(*C_PRI_NAVY if is_bal else C_TEXT_DARK)
            pdf.set_xy(xc, y + 1.5)
            pdf.cell(w, rh - 3, val, align=align, border=0, ln=False)
            xc += w
        pdf.set_y(y + rh)

    y = pdf.get_y()
    pdf.set_fill_color(*C_PRI_NAVY)
    pdf.rect(x0, y, total_w, rh, 'F')
    t_vals = ['TOTAL', str(tot[0]), 'Rs ' + inr(tot[1]), 'Rs ' + inr(tot[2]),
              'Rs ' + inr(tot[3]), 'Rs ' + inr(tot[4])]
    xc = x0
    for (col, w, align), t_val in zip(cols, t_vals):
        pdf.set_font('Helvetica', 'B', 7)
        pdf.set_text_color(*C_WHITE)
        pdf.set_xy(xc, y + 1.5)
        pdf.cell(w, rh - 3, t_val, align=align, border=0, ln=False)
        xc += w
    pdf.set_y(y + rh + 4)
    pdf.set_text_color(*C_TEXT_DARK)

# --- CLUSTER ANALYTICS -------------------------------------------------------
def draw_cluster_analytics(pdf, all_cases, metrics):
    import calendar as _cal
    _fy_start = datetime.date(TODAY.year if TODAY.month >= 4 else TODAY.year - 1, 4, 1)
    _fy_label = f'FY {_fy_start.year}-{str(_fy_start.year + 1)[2:]} YTD'
    pdf.section_title(f'Cluster Analytics - TAT-Adjusted Performance ({_fy_label})')
    by_cluster = defaultdict(list)
    for c in all_cases:
        if isinstance(c['date'], datetime.date) and c['date'] >= _fy_start:
            by_cluster[c['cluster']].append(c)

    rows = []
    for cl, cases in sorted(by_cluster.items()):
        if cl in EXCLUDED_CLUSTERS: continue
        closed  = [c for c in cases if c['status'] == 'Closed']
        tats    = [c['tat'] for c in closed if c['tat'] >= 0]
        avg_tat = sum(tats) / len(tats) if tats else None
        vol     = sum(c['amount']  for c in cases)
        chg     = sum(c['charges'] for c in cases)
        roi_pct = (chg / metrics['Total Invested'] * 100) if metrics.get('Total Invested') else 0
        rows.append({'cl': cl, 'cases': len(cases), 'closed': len(closed),
                     'vol': vol, 'chg': chg, 'avg_tat': avg_tat, 'roi': roi_pct})
    if not rows: return
    rows.sort(key=lambda r: r['roi'], reverse=True)

    x0, rh = L_MAR, 7
    cols = [
        ('RANK', 10, 'C'), ('CLUSTER', 30, 'L'), ('CASES', 14, 'C'),
        ('CLOSED', 14, 'C'), ('VOLUME', 32, 'R'), ('CHARGES', 28, 'R'),
        ('AVG TAT', 18, 'C'), ('ROI %', 18, 'C'),
    ]
    total_w = sum(c[1] for c in cols)

    pdf.set_fill_color(*C_PRI_NAVY)
    hdr_y = pdf.get_y()
    pdf.rect(x0, hdr_y, total_w, rh, 'F')
    xc = x0
    for col, w, align in cols:
        pdf.set_font('Helvetica', 'B', 7)
        pdf.set_text_color(*C_WHITE)
        pdf.set_xy(xc, hdr_y + 1.5)
        pdf.cell(w, rh - 3, col, align=align, border=0, ln=False)
        xc += w
    pdf.set_y(hdr_y + rh)

    for rank, r in enumerate(rows, 1):
        y = pdf.get_y()
        if rank % 2 == 0: pdf.set_fill_color(*C_LT_GOLD)
        else:             pdf.set_fill_color(*C_WHITE)
        pdf.rect(x0, y, total_w, rh, 'F')
        tat_str  = f"{r['avg_tat']:.1f}" if r['avg_tat'] is not None else '-'
        tat_fail = r['avg_tat'] is not None and r['avg_tat'] > 1.5
        vals = [f"#{rank}", r['cl'], str(r['cases']), str(r['closed']),
                'Rs ' + inr(r['vol']), 'Rs ' + inr(r['chg']), tat_str, f"{r['roi']:.3f}%"]
        xc = x0
        for (col, w, align), val in zip(cols, vals):
            if col == 'AVG TAT' and tat_fail:
                pdf.set_font('Helvetica', 'B', 7); pdf.set_text_color(*C_RED)
            elif col == 'ROI %':
                pdf.set_font('Helvetica', 'B', 7); pdf.set_text_color(*C_PRI_NAVY)
            else:
                pdf.set_font('Helvetica', '', 7); pdf.set_text_color(*C_TEXT_DARK)
            pdf.set_xy(xc, y + 1.5)
            pdf.cell(w, rh - 3, val, align=align, border=0, ln=False)
            xc += w
        pdf.set_y(y + rh)

    pdf.set_fill_color(*C_GOLD_RULE)
    pdf.rect(x0, pdf.get_y(), total_w, 0.4, 'F')
    pdf.set_text_color(*C_TEXT_DARK)
    pdf.set_y(pdf.get_y() + 4)

# --- ROI GOALPOST ------------------------------------------------------------
def draw_roi_goalpost(pdf, metrics):
    import calendar as _cal
    _days_in_month = _cal.monthrange(TODAY.year, TODAY.month)[1]
    _month_lbl = TODAY.strftime('%B %Y')
    pdf.section_title(f'Monthly ROI Goalpost - {_month_lbl}  (Target: 4.00%)')

    capital        = metrics['Total Invested']
    earned         = metrics['MTD Charges']
    target         = capital * ROI_TARGET_PCT / 100
    remaining      = max(0, target - earned)
    pct            = min(100, (earned / target * 100) if target else 0)
    days_remaining = _days_in_month - TODAY.day
    daily_needed   = remaining / days_remaining if days_remaining > 0 else 0
    daily_disburse = daily_needed / 0.005 if daily_needed > 0 else 0

    x0, pw = L_MAR, pdf.inner_w()
    y = pdf.get_y()

    bh = 9
    pdf.set_fill_color(*C_LT_SLATE)
    pdf.rect(x0, y, pw, bh, 'F')
    fw = pw * pct / 100
    pdf.set_fill_color(*C_PRI_NAVY)
    if fw > 0: pdf.rect(x0, y, fw, bh, 'F')
    pdf.set_xy(x0 + 2, y + 2.5)
    pdf.set_font('Helvetica', 'B', 7.5)
    pdf.set_text_color(*C_WHITE)
    pdf.cell(fw - 2, 4, f'{pct:.1f}% of target achieved', border=0, ln=False)
    pdf.set_xy(x0 + pw - 60, y + 2.5)
    pdf.set_font('Helvetica', '', 7)
    pdf.set_text_color(*C_TEXT_MED)
    pdf.cell(58, 4, f'Target: Rs {inr(target)}', align='R', border=0, ln=False)
    pdf.set_y(y + bh + 3)

    stat_boxes = [
        ('CAPITAL DEPLOYED',     'Rs ' + inr(capital)),
        ('CHARGES EARNED',       'Rs ' + inr(earned)),
        ('REMAINING',            'Rs ' + inr(remaining)),
        (f'DAYS LEFT IN {TODAY.strftime("%B").upper()}', str(days_remaining)),
        ('DAILY CHARGE NEEDED',  'Rs ' + inr(daily_needed)),
        ('DAILY DISBURSAL NEED', 'Rs ' + inr(daily_disburse)),
    ]
    n = len(stat_boxes)
    bw = (pw - (n - 1) * 2) / n
    box_h = 17
    y2 = pdf.get_y()
    for i, (lbl, val) in enumerate(stat_boxes):
        bx = x0 + i * (bw + 2)
        pdf.set_fill_color(*C_LT_GOLD)
        pdf.rect(bx, y2, bw, box_h, 'F')
        pdf.set_fill_color(*C_GOLD_RULE)
        pdf.rect(bx, y2, bw, 0.6, 'F')
        pdf.set_xy(bx, y2 + 2)
        pdf.set_font('Helvetica', 'B', 5.5)
        pdf.set_text_color(*C_TEXT_MED)
        pdf.cell(bw, 4, lbl, align='C', border=0, ln=False)
        fsize = 9
        pdf.set_font('Helvetica', 'B', fsize)
        while pdf.get_string_width(val) > bw - 2 and fsize > 6:
            fsize -= 0.5
            pdf.set_font('Helvetica', 'B', fsize)
        pdf.set_xy(bx, y2 + 8)
        pdf.set_text_color(*C_PRI_NAVY)
        pdf.cell(bw, 5, val, align='C', border=0, ln=False)
    pdf.set_text_color(*C_TEXT_DARK)
    pdf.set_y(y2 + box_h + 4)

# --- PORTFOLIO DASHBOARD (corrected metrics) ----------------------------------
def draw_portfolio_dashboard(pdf, metrics):
    """
    Shows 7 key portfolio metrics computed directly from raw data (all clusters incl. Mandya).
    Available for Disbursement may be negative when overdeployed.
    """
    pdf.section_title('Portfolio Dashboard')

    avail = metrics['Available for Disbursement']
    avail_str = ('Rs ' + inr(avail)) if avail >= 0 else ('-Rs ' + inr(abs(avail)) + ' (OVERDEPLOYED)')

    fields = [
        ('Total Invested',             'Rs ' + inr(metrics['Total Invested'])),
        ('Total Disbursed (All Time)', 'Rs ' + inr(metrics['Total Disbursed'])),
        ('Available for Disbursement', avail_str),
        ('Net Charges Earned',         'Rs ' + inr(metrics['Net Charges Earned'])),
        ('Total Pending Cases',        str(int(metrics['Total Pending Cases']))),
        ('Total Overdue Cases',        str(int(metrics['Total Overdue Cases']))),
        ('All-Time ROI',               f"{metrics['ROI'] * 100:.4f}%"),
    ]

    n   = 4
    pw  = pdf.inner_w()
    bw  = (pw - (n - 1) * 3) / n
    bh  = 19
    y   = pdf.get_y()

    for i, (lbl, val) in enumerate(fields):
        row = i // n; col = i % n
        bx  = L_MAR + col * (bw + 3)
        by  = y + row * (bh + 3)
        is_overdeployed = (lbl == 'Available for Disbursement' and avail < 0)
        bg_color = (255, 235, 235) if is_overdeployed else C_LT_SLATE
        pdf.set_fill_color(*bg_color)
        pdf.rect(bx, by, bw, bh, 'F')
        accent = C_RED if is_overdeployed else C_GOLD_RULE
        pdf.set_fill_color(*accent)
        pdf.rect(bx, by, 1.2, bh, 'F')
        pdf.set_xy(bx + 3, by + 2.5)
        pdf.set_font('Helvetica', 'B', 6)
        pdf.set_text_color(*C_TEXT_MED)
        pdf.cell(bw - 4, 4, lbl.upper(), border=0, ln=False)
        fsize = 10
        pdf.set_font('Helvetica', 'B', fsize)
        while pdf.get_string_width(val) > bw - 6 and fsize > 6:
            fsize -= 0.5
            pdf.set_font('Helvetica', 'B', fsize)
        pdf.set_xy(bx + 3, by + 9)
        pdf.set_text_color(*C_RED if is_overdeployed else C_PRI_NAVY)
        pdf.cell(bw - 4, 6, val, border=0, ln=False)

    rows_used = math.ceil(len(fields) / n)
    pdf.set_text_color(*C_TEXT_DARK)
    pdf.set_y(y + rows_used * (bh + 3) + 2)

    pdf.set_xy(L_MAR, pdf.get_y())
    pdf.set_font('Helvetica', 'I', 6)
    pdf.set_text_color(*C_TEXT_MED)
    pdf.cell(pdf.inner_w(), 4,
             'Metrics computed from raw Accounts data (all clusters incl. Mandya). '
             'Available = Total Invested - ABS(all open outstanding). Overdeployed when negative.',
             border=0, ln=False)
    pdf.set_text_color(*C_TEXT_DARK)
    pdf.set_y(pdf.get_y() + 6)

# --- MONTHLY DASHBOARD -------------------------------------------------------
def draw_monthly_dashboard(pdf, all_cases_full, metrics):
    """Month-wise breakdown of all key metrics - ALL clusters including Mandya."""

    months_seen = set()
    for c in all_cases_full:
        if isinstance(c['date'], datetime.date):
            months_seen.add((c['date'].year, c['date'].month))
    if not months_seen:
        return

    sorted_months = sorted(months_seen)

    month_rows = []
    for (yr, mo) in sorted_months:
        month_cases  = [c for c in all_cases_full
                        if isinstance(c['date'], datetime.date)
                        and c['date'].year == yr and c['date'].month == mo]
        total_cases  = len(month_cases)
        volume       = sum(c['amount']   for c in month_cases)
        charges      = sum(c['charges']  for c in month_cases)
        gst          = sum(c['gst']      for c in month_cases)
        total_billed = sum(c['total']    for c in month_cases)
        collected    = sum(c['coll_amt'] for c in month_cases)
        # Status='Closed' is authoritative (fixed 22-Jun-2026) — a closed case never counts
        # as outstanding/open, even if a small rounding residual remains in its balance.
        outstanding  = sum(c['balance']  for c in month_cases if c['balance'] >= 1.0 and c['status'] != 'Closed')
        closed_cnt   = len([c for c in month_cases if c['balance'] < 1.0 or c['status'] == 'Closed'])
        open_cnt     = total_cases - closed_cnt
        closed_lst   = [c for c in month_cases if (c['balance'] < 1.0 or c['status'] == 'Closed') and c['tat'] >= 0]
        tats         = [c['tat'] for c in closed_lst]
        avg_tat      = sum(tats) / len(tats) if tats else None
        roi_pct      = (charges / metrics['Total Invested'] * 100) if metrics.get('Total Invested') else 0
        month_lbl    = datetime.date(yr, mo, 1).strftime('%b %Y')
        month_rows.append({
            'label': month_lbl, 'cases': total_cases, 'volume': volume,
            'charges': charges, 'gst': gst, 'billed': total_billed,
            'collected': collected, 'outstanding': outstanding,
            'closed': closed_cnt, 'open': open_cnt, 'avg_tat': avg_tat, 'roi': roi_pct,
        })

    pdf.section_title('Portfolio Dashboard - Month-Wise (All Clusters incl. Mandya)')
    x0 = L_MAR; pw = pdf.inner_w(); rh = 7

    cols = [
        ('MONTH',       22, 'L'), ('CASES',       12, 'C'), ('VOLUME',      28, 'R'),
        ('CHARGES',     24, 'R'), ('GST',         18, 'R'), ('TOTAL BILLED',24, 'R'),
        ('COLLECTED',   24, 'R'), ('OUTSTANDING', 24, 'R'), ('CLOSED',      13, 'C'),
        ('OPEN',        10, 'C'), ('AVG TAT',     15, 'C'), ('ROI %',       14, 'C'),
    ]
    scale   = pw / sum(c[1] for c in cols)
    cols    = [(l, w * scale, a) for l, w, a in cols]
    total_w = sum(c[1] for c in cols)

    pdf.set_fill_color(*C_PRI_NAVY)
    hdr_y = pdf.get_y()
    pdf.rect(x0, hdr_y, total_w, rh, 'F')
    xc = x0
    for col, w, align in cols:
        pdf.set_font('Helvetica', 'B', 6.5); pdf.set_text_color(*C_WHITE)
        pdf.set_xy(xc, hdr_y + 1.5); pdf.cell(w, rh - 3, col, align=align, border=0, ln=False)
        xc += w
    pdf.set_y(hdr_y + rh)

    tot = defaultdict(float)
    for idx, r in enumerate(month_rows):
        is_cur = (r['label'] == TODAY.strftime('%b %Y'))
        y = pdf.get_y()
        if is_cur:         pdf.set_fill_color(*C_LT_GOLD)
        elif idx % 2 == 0: pdf.set_fill_color(*C_LT_SLATE)
        else:              pdf.set_fill_color(*C_WHITE)
        pdf.rect(x0, y, total_w, rh, 'F')
        if is_cur:
            pdf.set_fill_color(*C_GOLD_RULE); pdf.rect(x0, y, 1.2, rh, 'F')

        tat_str  = f"{r['avg_tat']:.1f}" if r['avg_tat'] is not None else '-'
        tat_fail = r['avg_tat'] is not None and r['avg_tat'] > 1.5
        row_vals = [
            r['label'], str(r['cases']), 'Rs ' + inr(r['volume']), 'Rs ' + inr(r['charges']),
            'Rs ' + inr(r['gst']), 'Rs ' + inr(r['billed']), 'Rs ' + inr(r['collected']),
            'Rs ' + inr(r['outstanding']), str(r['closed']), str(r['open']),
            tat_str, f"{r['roi']:.3f}%",
        ]
        xc = x0
        for (col, w, align), val in zip(cols, row_vals):
            if col == 'MONTH' and is_cur:
                pdf.set_font('Helvetica', 'B', 6.5); pdf.set_text_color(*C_PRI_NAVY)
            elif col == 'AVG TAT' and tat_fail:
                pdf.set_font('Helvetica', 'B', 6.5); pdf.set_text_color(*C_RED)
            elif col == 'ROI %':
                pdf.set_font('Helvetica', 'B', 6.5); pdf.set_text_color(*C_PRI_NAVY)
            elif col == 'OUTSTANDING':
                pdf.set_font('Helvetica', 'B', 6.5)
                pdf.set_text_color(*C_RED if r['outstanding'] > 0 else C_GREEN)
            elif col == 'OPEN' and r['open'] > 0:
                pdf.set_font('Helvetica', 'B', 6.5); pdf.set_text_color(*C_ORANGE)
            else:
                pdf.set_font('Helvetica', '', 6.5); pdf.set_text_color(*C_TEXT_DARK)
            pdf.set_xy(xc, y + 1.5); pdf.cell(w, rh - 3, val, align=align, border=0, ln=False)
            xc += w
        pdf.set_y(y + rh)
        for k in ['cases', 'volume', 'charges', 'gst', 'billed', 'collected', 'outstanding', 'closed', 'open']:
            tot[k] += r[k]

    y = pdf.get_y()
    pdf.set_fill_color(*C_PRI_NAVY); pdf.rect(x0, y, total_w, rh, 'F')
    overall_roi = (tot['charges'] / tot['volume'] * 100) if tot['volume'] else 0
    tot_vals = [
        'TOTAL', str(int(tot['cases'])), 'Rs ' + inr(tot['volume']), 'Rs ' + inr(tot['charges']),
        'Rs ' + inr(tot['gst']), 'Rs ' + inr(tot['billed']), 'Rs ' + inr(tot['collected']),
        'Rs ' + inr(tot['outstanding']), str(int(tot['closed'])), str(int(tot['open'])),
        '-', f"{overall_roi:.3f}%",
    ]
    xc = x0
    for (col, w, align), val in zip(cols, tot_vals):
        pdf.set_font('Helvetica', 'B', 6.5); pdf.set_text_color(*C_WHITE)
        pdf.set_xy(xc, y + 1.5); pdf.cell(w, rh - 3, val, align=align, border=0, ln=False)
        xc += w
    pdf.set_y(y + rh + 2)
    pdf.set_xy(x0, pdf.get_y()); pdf.set_font('Helvetica', 'I', 6); pdf.set_text_color(*C_TEXT_MED)
    pdf.cell(total_w, 4,
             'Highlighted = current month  |  TAT red = avg > 1.5d  |  Outstanding red = pending  |  Includes all clusters incl. Mandya',
             border=0, ln=False)
    pdf.set_text_color(*C_TEXT_DARK); pdf.set_y(pdf.get_y() + 5)


# --- MTD / YTD REPORT -------------------------------------------------------
def draw_mtd_ytd_report(pdf, all_cases_full, metrics):
    """
    MTD vs YTD Performance Report — month-wise MTD figures alongside running
    YTD cumulative totals.  Financial year runs April–March.
    """
    fy_start = datetime.date(TODAY.year if TODAY.month >= 4 else TODAY.year - 1, 4, 1)
    fy_label = f'FY {fy_start.year}-{str(fy_start.year + 1)[2:]}'
    cur_month_lbl = TODAY.strftime('%b %Y')

    months_seen = set()
    for c in all_cases_full:
        if isinstance(c['date'], datetime.date) and c['date'] >= fy_start:
            months_seen.add((c['date'].year, c['date'].month))
    if not months_seen:
        return

    sorted_months = sorted(months_seen)

    ytd_cases_cum   = 0
    ytd_volume_cum  = 0.0
    ytd_charges_cum = 0.0
    ytd_gst_cum     = 0.0
    ytd_closed_cum  = 0

    rows = []
    for (yr, mo) in sorted_months:
        month_cases = [c for c in all_cases_full
                       if isinstance(c['date'], datetime.date)
                       and c['date'].year == yr and c['date'].month == mo]
        m_cases   = len(month_cases)
        m_volume  = sum(c['amount']  for c in month_cases)
        m_charges = sum(c['charges'] for c in month_cases)
        m_gst     = sum(c['gst']     for c in month_cases)
        m_closed  = len([c for c in month_cases if c['balance'] < 1.0 or c['status'] == 'Closed'])
        m_open    = m_cases - m_closed
        m_roi     = (m_charges / metrics['Total Invested'] * 100) if metrics.get('Total Invested') else 0

        ytd_cases_cum   += m_cases
        ytd_volume_cum  += m_volume
        ytd_charges_cum += m_charges
        ytd_gst_cum     += m_gst
        ytd_closed_cum  += m_closed
        ytd_roi = (ytd_charges_cum / metrics['Total Invested'] * 100) if metrics.get('Total Invested') else 0

        month_lbl = datetime.date(yr, mo, 1).strftime('%b %Y')
        rows.append({
            'label': month_lbl,
            'm_cases': m_cases, 'm_volume': m_volume, 'm_charges': m_charges,
            'm_gst': m_gst, 'm_closed': m_closed, 'm_open': m_open, 'm_roi': m_roi,
            'ytd_cases': ytd_cases_cum, 'ytd_volume': ytd_volume_cum,
            'ytd_charges': ytd_charges_cum, 'ytd_gst': ytd_gst_cum,
            'ytd_closed': ytd_closed_cum, 'ytd_roi': ytd_roi,
        })

    pdf.section_title(f'MTD / YTD Performance - {fy_label} (All Clusters incl. Mandya)')
    x0 = L_MAR; pw = pdf.inner_w(); rh = 7

    cols = [
        ('MONTH',       22, 'L'),
        # MTD columns
        ('CASES',       11, 'C'), ('VOLUME',      26, 'R'),
        ('CHARGES',     22, 'R'), ('GST',         15, 'R'),
        ('CLOSED',      12, 'C'), ('ROI %',       13, 'C'),
        # YTD columns
        ('YTD CASES',   13, 'C'), ('YTD VOLUME',  26, 'R'),
        ('YTD CHARGES', 22, 'R'), ('YTD GST',     15, 'R'),
        ('YTD CLOSED',  13, 'C'), ('YTD ROI %',   13, 'C'),
    ]
    scale   = pw / sum(c[1] for c in cols)
    cols    = [(l, w * scale, a) for l, w, a in cols]
    total_w = sum(c[1] for c in cols)

    month_w = cols[0][1]
    mtd_w   = sum(c[1] for c in cols[1:7])
    ytd_w   = sum(c[1] for c in cols[7:])
    ytd_x   = x0 + month_w + mtd_w

    # Group header
    grp_rh  = 5
    hdr_y   = pdf.get_y()
    pdf.set_fill_color(*C_NAVY)
    pdf.rect(x0, hdr_y, total_w, grp_rh, 'F')
    pdf.set_xy(x0 + month_w, hdr_y + 0.5)
    pdf.set_font('Helvetica', 'B', 6.5); pdf.set_text_color(*C_GOLD_LBL)
    pdf.cell(mtd_w, grp_rh - 1, 'MONTH TO DATE  (MTD)', align='C', border=0, ln=False)
    pdf.set_fill_color(*C_GOLD_RULE)
    pdf.rect(ytd_x - 0.4, hdr_y, 0.8, grp_rh, 'F')
    pdf.set_xy(ytd_x, hdr_y + 0.5)
    pdf.set_font('Helvetica', 'B', 6.5); pdf.set_text_color(*C_GOLD_LBL)
    pdf.cell(ytd_w, grp_rh - 1, f'YEAR TO DATE  (YTD) - {fy_label}', align='C', border=0, ln=False)
    pdf.set_y(hdr_y + grp_rh)

    # Column headers
    hdr_y2 = pdf.get_y()
    pdf.set_fill_color(*C_PRI_NAVY)
    pdf.rect(x0, hdr_y2, total_w, rh, 'F')
    xc = x0
    for col, w, align in cols:
        pdf.set_font('Helvetica', 'B', 5.5); pdf.set_text_color(*C_WHITE)
        pdf.set_xy(xc, hdr_y2 + 1.5); pdf.cell(w, rh - 3, col, align=align, border=0, ln=False)
        xc += w
    pdf.set_fill_color(*C_GOLD_RULE)
    pdf.rect(ytd_x - 0.4, hdr_y2, 0.8, rh, 'F')
    pdf.set_y(hdr_y2 + rh)

    # Data rows
    for idx, r in enumerate(rows):
        is_cur = (r['label'] == cur_month_lbl)
        y = pdf.get_y()
        if is_cur:
            pdf.set_fill_color(*C_LT_GOLD)
        elif idx % 2 == 0:
            pdf.set_fill_color(*C_LT_SLATE)
        else:
            pdf.set_fill_color(*C_WHITE)
        pdf.rect(x0, y, total_w, rh, 'F')
        if is_cur:
            pdf.set_fill_color(*C_GOLD_RULE); pdf.rect(x0, y, 1.2, rh, 'F')
        # divider between MTD and YTD
        pdf.set_fill_color(*C_RULE)
        pdf.rect(ytd_x - 0.2, y, 0.4, rh, 'F')

        row_vals = [
            r['label'],
            str(r['m_cases']), 'Rs ' + inr(r['m_volume']), 'Rs ' + inr(r['m_charges']),
            'Rs ' + inr(r['m_gst']), str(r['m_closed']), f"{r['m_roi']:.3f}%",
            str(r['ytd_cases']), 'Rs ' + inr(r['ytd_volume']), 'Rs ' + inr(r['ytd_charges']),
            'Rs ' + inr(r['ytd_gst']), str(r['ytd_closed']), f"{r['ytd_roi']:.3f}%",
        ]
        xc = x0
        for i, ((col, w, align), val) in enumerate(zip(cols, row_vals)):
            if col == 'MONTH' and is_cur:
                pdf.set_font('Helvetica', 'B', 6); pdf.set_text_color(*C_PRI_NAVY)
            elif col in ('ROI %', 'YTD ROI %'):
                pdf.set_font('Helvetica', 'B', 6); pdf.set_text_color(*C_PRI_NAVY)
            elif i >= 7:  # YTD columns
                pdf.set_font('Helvetica', '', 6); pdf.set_text_color(*C_TEXT_MED)
            else:
                pdf.set_font('Helvetica', '', 6); pdf.set_text_color(*C_TEXT_DARK)
            pdf.set_xy(xc, y + 1.5); pdf.cell(w, rh - 3, val, align=align, border=0, ln=False)
            xc += w
        pdf.set_y(y + rh)

    # YTD total row
    if rows:
        last = rows[-1]
        y = pdf.get_y()
        pdf.set_fill_color(*C_PRI_NAVY); pdf.rect(x0, y, total_w, rh, 'F')
        tot_vals = [
            f'YTD TOTAL ({fy_label})',
            '-', '-', '-', '-', '-', '-',
            str(last['ytd_cases']), 'Rs ' + inr(last['ytd_volume']),
            'Rs ' + inr(last['ytd_charges']), 'Rs ' + inr(last['ytd_gst']),
            str(last['ytd_closed']), f"{last['ytd_roi']:.3f}%",
        ]
        xc = x0
        for (col, w, align), val in zip(cols, tot_vals):
            pdf.set_font('Helvetica', 'B', 6); pdf.set_text_color(*C_WHITE)
            pdf.set_xy(xc, y + 1.5); pdf.cell(w, rh - 3, val, align=align, border=0, ln=False)
            xc += w
        pdf.set_y(y + rh + 2)

    pdf.set_xy(x0, pdf.get_y())
    pdf.set_font('Helvetica', 'I', 6); pdf.set_text_color(*C_TEXT_MED)
    pdf.cell(total_w, 4,
             f'MTD = per-month figures  |  YTD = cumul. from {fy_start.strftime("%d-%b-%Y")} ({fy_label})  |  Highlighted = current month  |  All clusters incl. Mandya',
             border=0, ln=False)
    pdf.set_text_color(*C_TEXT_DARK); pdf.set_y(pdf.get_y() + 5)

# --- MANDATORY GUIDELINES ----------------------------------------------------
def draw_guidelines(pdf):
    pdf.section_title('Mandatory Guidelines')
    lines = [
        '1. All refunds must originate from the SAME account as disbursement (BLP/CIR/002/2026-27). Violation is taxable under S.68 IT Act.',
        '2. Service charges are computed at disbursement and stored in sheet. Do NOT recalculate on collection day (BLP/CIR/001/2026-27).',
        '3. Close cases within TAT target. TAT directly impacts charge rate - longer TAT means higher charges per BLP/CIR/001/2026-27.',
        '4. GST @ 18% applies on all charges from 01-Apr-2026 onwards.',
        '5. Disbursements must be submitted before EOD for same-day processing priority.',
    ]
    x0, pw = L_MAR, pdf.inner_w()
    y = pdf.get_y()
    line_h  = 5.5
    total_h = len(lines) * line_h + 6
    pdf.set_fill_color(*C_MEMO_CREAM)
    pdf.set_draw_color(*C_GOLD_RULE)
    pdf.rect(x0, y, pw, total_h, 'FD')
    pdf.set_fill_color(*C_GOLD_RULE)
    pdf.rect(x0, y, 1.5, total_h, 'F')
    for i, line in enumerate(lines):
        pdf.set_xy(x0 + 4, y + 3 + i * line_h)
        pdf.set_font('Helvetica', '', 7)
        pdf.set_text_color(*C_TEXT_DARK)
        pdf.cell(pw - 6, 4, line, border=0, ln=False)
    pdf.set_y(y + total_h + 4)

# --- DISBURSEMENT MEMO -------------------------------------------------------
# Bold redesign (21-Jun-2026): stat-bar header (status/days/rate + headline
# TOTAL PAYABLE), every figure bold/oversized, transparent-bg logo. Only this
# function changed — BLPdf header/footer untouched. See memory:
# project_memo_redesign_status.md for the design history.
def generate_memo(case):
    pdf = BLPdf(subtitle='DISBURSEMENT MEMO', report_date=fmt_date(TODAY))
    pdf.add_page()

    pw  = pdf.w
    x0  = L_MAR
    pwi = pdf.inner_w()
    y   = pdf.content_top() + 2

    # --- Title bar (Ref + Date together) ---
    pdf.set_fill_color(*C_PRI_NAVY)
    pdf.rect(x0, y, pwi, 9, 'F')
    pdf.set_xy(x0, y + 1.5)
    pdf.set_font('Helvetica', 'B', 12)
    pdf.set_text_color(*C_WHITE)
    pdf.cell(pwi * 0.5, 6, 'DISBURSEMENT MEMO', align='C', border=0, ln=False)
    pdf.set_font('Helvetica', 'B', 8.5)
    pdf.set_text_color(*C_GOLD_LBL)
    pdf.cell(pwi * 0.5, 6, f'Ref: {case["id"]}   |   Date: {fmt_date(case["date"])}', align='R', border=0, ln=False)
    y += 11

    # --- STAT BAR: 3 accent cells (status/days/rate) + 1 navy headline cell (total) ---
    sc = status_color(case['status'])
    sb_h = 15
    cells = [
        ('STATUS', case['status']),
        ('DAYS OUTSTANDING', str(case['days_out'])),
        ('CHARGE RATE', charge_rate_label(case['tat'])),
    ]
    cw = pwi * 0.22
    total_w = pwi - 3 * cw
    for i, (lbl, val) in enumerate(cells):
        cx = x0 + i * cw
        pdf.set_fill_color(*sc)
        pdf.rect(cx, y, cw - 1, sb_h, 'F')
        pdf.set_xy(cx + 3, y + 2)
        pdf.set_font('Helvetica', 'B', 6); pdf.set_text_color(*C_WHITE)
        pdf.cell(cw - 6, 3.5, lbl, border=0, ln=False)
        pdf.set_xy(cx + 3, y + 6.5)
        pdf.set_font('Helvetica', 'B', 12.5); pdf.set_text_color(*C_WHITE)
        pdf.cell(cw - 6, 6, val, border=0, ln=False)
    tx = x0 + 3 * cw
    pdf.set_fill_color(*C_NAVY)
    pdf.rect(tx, y, total_w, sb_h, 'F')
    pdf.set_fill_color(*C_GOLD_RULE); pdf.rect(tx, y, total_w, 0.8, 'F')
    pdf.set_xy(tx + 3, y + 2)
    pdf.set_font('Helvetica', 'B', 6); pdf.set_text_color(*C_GOLD_LBL)
    pdf.cell(total_w - 6, 3.5, 'TOTAL PAYABLE (INR)', border=0, ln=False)
    pdf.set_xy(tx + 3, y + 6.5)
    pdf.set_font('Helvetica', 'B', 14); pdf.set_text_color(*C_WHITE)
    pdf.cell(total_w - 6, 6.5, f'Rs {inr_dec(case["total"])}', border=0, ln=False)
    y += sb_h + 4

    # --- DISBURSED TO panel ---
    ph = 34
    pdf.set_fill_color(*C_MEMO_CREAM)
    pdf.set_draw_color(*C_GOLD_RULE)
    pdf.rect(x0, y, pwi, ph, 'FD')
    pdf.set_fill_color(*C_GOLD_RULE)
    pdf.rect(x0, y, 1.5, ph, 'F')
    pdf.set_xy(x0 + 4, y + 2)
    pdf.set_font('Helvetica', 'B', 7)
    pdf.set_text_color(*C_TEXT_MED)
    pdf.cell(30, 4, 'DISBURSED TO', border=0, ln=False)

    left_f  = [('Customer', case['customer']), ('Company', case['company']), ('Cluster', case['cluster'])]
    right_f = [('Branch', case['branch']), ('Chq / Ref', case['chq']), ('Debit Note', case.get('debit_note', '-'))]
    NUMERIC_FIELDS = {'Chq / Ref', 'Debit Note'}

    for i, (lbl, val) in enumerate(left_f):
        pdf.set_xy(x0 + 4, y + 8 + i * 5.5)
        pdf.set_font('Helvetica', 'B', 7); pdf.set_text_color(*C_TEXT_MED)
        pdf.cell(22, 4, lbl + ':', border=0, ln=False)
        pdf.set_font('Helvetica', 'B', 8); pdf.set_text_color(*C_TEXT_DARK)
        pdf.cell(60, 4, val, border=0, ln=False)
    for i, (lbl, val) in enumerate(right_f):
        pdf.set_xy(x0 + 90, y + 8 + i * 5.5)
        pdf.set_font('Helvetica', 'B', 7); pdf.set_text_color(*C_TEXT_MED)
        pdf.cell(22, 4, lbl + ':', border=0, ln=False)
        if lbl in NUMERIC_FIELDS:
            pdf.set_font('Helvetica', 'B', 8.5); pdf.set_text_color(*C_PRI_NAVY)
        else:
            pdf.set_font('Helvetica', 'B', 8); pdf.set_text_color(*C_TEXT_DARK)
        pdf.cell(75, 4, val, border=0, ln=False)
    y += ph + 4

    # --- Charge breakdown (left) + Collection details (right) ---
    lw = pwi * 0.56
    rw = pwi * 0.40
    gap = pwi - lw - rw
    rh  = 8.2

    charge_rows = [
        ('Amount Disbursed', f'Rs {inr_dec(case["amount"])}'),
        ('Service Charges',  f'Rs {inr_dec(case["charges"])}'),
        ('GST (18%)',         f'Rs {inr_dec(case["gst"])}'),
        ('TOTAL PAYABLE',    f'Rs {inr_dec(case["total"])}'),
    ]
    pdf.set_fill_color(*C_PRI_NAVY)
    pdf.rect(x0, y, lw, rh, 'F')
    pdf.set_xy(x0 + 2, y + 1.8)
    pdf.set_font('Helvetica', 'B', 7); pdf.set_text_color(*C_WHITE)
    pdf.cell(lw * 0.6, rh - 3, 'DESCRIPTION', align='L', border=0, ln=False)
    pdf.cell(lw * 0.4 - 2, rh - 3, 'AMOUNT', align='R', border=0, ln=False)
    for i, (desc, amt) in enumerate(charge_rows):
        ry, is_tot = y + rh + i * rh, (desc == 'TOTAL PAYABLE')
        if is_tot: pdf.set_fill_color(*C_LT_GOLD)
        elif i % 2: pdf.set_fill_color(*C_WHITE)
        else:       pdf.set_fill_color(*C_LT_SLATE)
        pdf.rect(x0, ry, lw, rh, 'F')
        if is_tot:
            pdf.set_fill_color(*C_GOLD_RULE); pdf.rect(x0, ry, lw, 0.7, 'F')
        # description label
        pdf.set_xy(x0 + 2, ry + 2)
        pdf.set_font('Helvetica', 'B', 8 if is_tot else 7.5)
        pdf.set_text_color(*C_PRI_NAVY if is_tot else C_TEXT_DARK)
        pdf.cell(lw * 0.6 - 2, rh - 3, desc, align='L', border=0, ln=False)
        # amount — bolded/oversized
        pdf.set_xy(x0, ry + 1.2)
        pdf.set_font('Helvetica', 'B', 13 if is_tot else 10.5)
        pdf.set_text_color(*C_PRI_NAVY)
        pdf.cell(lw - 2, rh - 1.5, amt, align='R', border=0, ln=False)

    rx    = x0 + lw + gap
    cdate = fmt_date(case['coll_date']) if case['coll_date'] else '-'
    camt  = f'Rs {inr_dec(case["coll_amt"])}' if case['coll_amt'] else '-'
    bstr  = f'Rs {inr_dec(case["balance"])}'
    bcol  = C_RED if case['balance'] >= 1 else C_GREEN

    coll_rows = [
        ('Collected Date',   cdate,                 C_TEXT_DARK, False),
        ('Collected Amount', camt,                  C_PRI_NAVY,  False),
        ('Balance O/S',      bstr,                  bcol,        True),
        ('Days Outstanding', str(case['days_out']), C_RED if case['days_out'] > 2 else C_TEXT_DARK, False),
    ]
    pdf.set_fill_color(*C_PRI_NAVY)
    pdf.rect(rx, y, rw, rh, 'F')
    pdf.set_xy(rx, y + 1.8)
    pdf.set_font('Helvetica', 'B', 7); pdf.set_text_color(*C_WHITE)
    pdf.cell(rw, rh - 3, 'COLLECTION DETAILS', align='C', border=0, ln=False)
    for i, (desc, val, vc, is_bal) in enumerate(coll_rows):
        ry2 = y + rh + i * rh
        if is_bal: pdf.set_fill_color(*C_LT_GOLD)
        elif i % 2: pdf.set_fill_color(*C_WHITE)
        else:       pdf.set_fill_color(*C_LT_SLATE)
        pdf.rect(rx, ry2, rw, rh, 'F')
        if is_bal:
            pdf.set_fill_color(*C_GOLD_RULE); pdf.rect(rx, ry2, rw, 0.7, 'F')
        # label
        pdf.set_xy(rx + 2, ry2 + 2)
        pdf.set_font('Helvetica', 'B', 7); pdf.set_text_color(*C_TEXT_MED)
        pdf.cell(rw * 0.55, rh - 3, desc, align='L', border=0, ln=False)
        # value — bolded/oversized
        pdf.set_xy(rx + 2, ry2 + 1.2)
        pdf.set_font('Helvetica', 'B', 12.5 if is_bal else 10)
        pdf.set_text_color(*vc)
        pdf.cell(rw - 3, rh - 1.5, val, align='R', border=0, ln=False)

    y += rh + len(charge_rows) * rh + 5

    # --- TAT charge policy note ---
    note_h = 18
    pdf.set_fill_color(*C_MEMO_CREAM)
    pdf.set_draw_color(*C_GOLD_RULE)
    pdf.rect(x0, y, pwi, note_h, 'FD')
    pdf.set_fill_color(*C_GOLD_RULE); pdf.rect(x0, y, 1.5, note_h, 'F')
    pdf.set_xy(x0 + 4, y + 2)
    pdf.set_font('Helvetica', 'B', 7); pdf.set_text_color(*C_PRI_NAVY)
    pdf.cell(0, 4, 'TAT Charge Policy (BLP/CIR/001/2026-27)', border=0, ln=False)
    pdf.set_xy(x0 + 4, y + 7)
    pdf.set_font('Helvetica', 'B', 7); pdf.set_text_color(*C_TEXT_DARK)
    pdf.cell(0, 4, '<=2d: 0.50%  |  3d: 1.00%  |  4d: 1.50%  |  5d+: 0.50% x days  |  GST 18% on all charges', border=0, ln=False)
    pdf.set_xy(x0 + 4, y + 12)
    pdf.set_font('Helvetica', 'BI', 6.5); pdf.set_text_color(*C_TEXT_MED)
    pdf.cell(0, 4, 'Charges fixed at disbursement. Refunds must be from same disbursing account (BLP/CIR/002/2026-27).', border=0, ln=False)
    y += note_h + 8

    # --- Signatures (drops in actual signature image once SIGNATURE_PATH exists) ---
    sig_w = (pwi - 10) / 2
    for i, (label, name) in enumerate([('AUTHORISED BY', 'Prem / Harsha'), ('RECEIVED BY', case['customer'])]):
        sx = x0 + i * (sig_w + 10)
        if i == 0 and os.path.exists(SIGNATURE_PATH) and os.path.getsize(SIGNATURE_PATH) > 0:
            sig_img_h = 12
            pdf.image(SIGNATURE_PATH, x=sx + sig_w / 2 - 18, y=y + 18 - sig_img_h - 1, w=36, h=sig_img_h)
        pdf.set_draw_color(*C_RULE)
        pdf.line(sx, y + 18, sx + sig_w, y + 18)
        pdf.set_xy(sx, y + 20)
        pdf.set_font('Helvetica', 'B', 7); pdf.set_text_color(*C_TEXT_MED)
        pdf.cell(sig_w, 4, label, align='C', border=0, ln=False)
        pdf.set_xy(sx, y + 25)
        pdf.set_font('Helvetica', 'B', 7.5); pdf.set_text_color(*C_TEXT_DARK)
        pdf.cell(sig_w, 4, name, align='C', border=0, ln=False)

    pdf.set_text_color(*C_TEXT_DARK)

    # --- SCAN OR TAP TO PAY (UPI) panel — only for cases under the Collection Card threshold ---
    if case['amount'] < COLLECTION_CARD_THRESHOLD or case['balance'] < COLLECTION_CARD_THRESHOLD:
        y += 33
        panel_w = 55
        px0 = x0 + (pwi - panel_w) / 2
        pdf.set_fill_color(*C_NAVY)
        pdf.rect(px0, y, panel_w, 6.5, 'F')
        pdf.set_xy(px0, y + 1)
        pdf.set_font('Helvetica', 'B', 7); pdf.set_text_color(*C_GOLD_LBL)
        pdf.cell(panel_w, 4.5, 'SCAN OR TAP TO PAY (UPI)', align='C', border=0, ln=False)
        body_y = y + 6.5
        body_h = 40
        pdf.set_fill_color(*C_WHITE)
        pdf.set_draw_color(*C_GOLD_RULE)
        pdf.rect(px0, body_y, panel_w, body_h, 'FD')
        qr_size = min(body_h - 16, panel_w - 8)
        qx = px0 + 4
        qy = body_y + 3
        pay_page_url = build_pay_page_url(case['id'])
        if os.path.exists(QR_PATH) and os.path.getsize(QR_PATH) > 0:
            pdf.image(QR_PATH, x=qx, y=qy, w=qr_size, h=qr_size)
        else:
            pdf.set_draw_color(*C_RULE)
            pdf.rect(qx, qy, qr_size, qr_size, 'D')
            pdf.set_xy(qx, qy + qr_size / 2 - 3)
            pdf.set_font('Helvetica', 'BI', 7); pdf.set_text_color(*C_RULE)
            pdf.cell(qr_size, 6, 'QR PENDING', align='C', border=0, ln=False)
        pdf.link(qx, qy, qr_size, qr_size, pay_page_url)
        # Pay Now button — fills the slack to the right of the QR within the same box, same tap target
        btn_x = qx + qr_size + 3
        btn_w = px0 + panel_w - 2 - btn_x
        btn_h = qr_size
        if btn_w > 14:
            pdf.set_fill_color(*C_NAVY)
            pdf.rect(btn_x, qy, btn_w, btn_h, 'F')
            pdf.set_xy(btn_x, qy + btn_h * 0.30)
            pdf.set_font('Helvetica', 'B', 7); pdf.set_text_color(*C_WHITE)
            pdf.cell(btn_w, 5, 'PAY NOW', align='C', border=0, ln=False)
            pdf.set_xy(btn_x, qy + btn_h * 0.30 + 5)
            pdf.set_font('Helvetica', 'I', 5.5); pdf.set_text_color(*C_GOLD_LBL)
            pdf.cell(btn_w, 4, 'Tap to Pay', align='C', border=0, ln=False)
            pdf.link(btn_x, qy, btn_w, btn_h, pay_page_url)
        pdf.set_xy(px0, qy + qr_size + 2)
        pdf.set_font('Helvetica', 'B', 7); pdf.set_text_color(*C_TEXT_DARK)
        pdf.cell(panel_w, 4, PAYEE_NAME, align='C', border=0, ln=False)
        pdf.set_xy(px0, qy + qr_size + 6)
        pdf.set_font('Helvetica', 'B', 6.5); pdf.set_text_color(*C_PRI_NAVY)
        pdf.cell(panel_w, 4, f'UPI: {UPI_ID}', align='C', border=0, ln=False)
        pdf.set_text_color(*C_TEXT_DARK)

    pdf.set_page_count(pdf.page_no())
    return pdf.output()

# --- COLLECTION CARD (per case, disbursed amount OR outstanding balance < COLLECTION_CARD_THRESHOLD) ---
def generate_collection_card(case, branch_contacts, cluster_mgrs):
    pdf = BLPdf(subtitle='COLLECTION CARD', report_date=fmt_date(TODAY))
    pdf.add_page()

    x0  = L_MAR
    pwi = pdf.inner_w()
    y   = pdf.content_top() + 2

    # --- Title bar (Ref + Date together) ---
    pdf.set_fill_color(*C_PRI_NAVY)
    pdf.rect(x0, y, pwi, 9, 'F')
    pdf.set_xy(x0, y + 1.5)
    pdf.set_font('Helvetica', 'B', 12)
    pdf.set_text_color(*C_WHITE)
    pdf.cell(pwi * 0.5, 6, 'COLLECTION CARD', align='C', border=0, ln=False)
    pdf.set_font('Helvetica', 'B', 8.5)
    pdf.set_text_color(*C_GOLD_LBL)
    pdf.cell(pwi * 0.5, 6, f'Ref: {case["id"]}   |   Date: {fmt_date(TODAY)}', align='R', border=0, ln=False)
    y += 11

    # --- STAT BAR: 3 accent cells (status/days/branch) + 1 navy headline cell (balance outstanding) ---
    sc = status_color(case['status'])
    sb_h = 15
    cells = [
        ('STATUS', case['status']),
        ('DAYS OUTSTANDING', str(case['days_out'])),
        ('BRANCH', case['branch'][:16]),
    ]
    cw = pwi * 0.22
    total_w = pwi - 3 * cw
    for i, (lbl, val) in enumerate(cells):
        cx = x0 + i * cw
        pdf.set_fill_color(*sc)
        pdf.rect(cx, y, cw - 1, sb_h, 'F')
        pdf.set_xy(cx + 3, y + 2)
        pdf.set_font('Helvetica', 'B', 6); pdf.set_text_color(*C_WHITE)
        pdf.cell(cw - 6, 3.5, lbl, border=0, ln=False)
        pdf.set_xy(cx + 3, y + 6.5)
        pdf.set_font('Helvetica', 'B', 12.5); pdf.set_text_color(*C_WHITE)
        pdf.cell(cw - 6, 6, val, border=0, ln=False)
    tx = x0 + 3 * cw
    pdf.set_fill_color(*C_NAVY)
    pdf.rect(tx, y, total_w, sb_h, 'F')
    pdf.set_fill_color(*C_RED); pdf.rect(tx, y, total_w, 0.8, 'F')
    pdf.set_xy(tx + 3, y + 2)
    pdf.set_font('Helvetica', 'B', 6); pdf.set_text_color(*C_GOLD_LBL)
    pdf.cell(total_w - 6, 3.5, 'BALANCE OUTSTANDING (INR)', border=0, ln=False)
    pdf.set_xy(tx + 3, y + 6.5)
    pdf.set_font('Helvetica', 'B', 14); pdf.set_text_color(*C_WHITE)
    pdf.cell(total_w - 6, 6.5, f'Rs {inr_dec(case["balance"])}', border=0, ln=False)
    y += sb_h + 4

    # --- CASE DETAILS panel ---
    ph = 34
    pdf.set_fill_color(*C_MEMO_CREAM)
    pdf.set_draw_color(*C_GOLD_RULE)
    pdf.rect(x0, y, pwi, ph, 'FD')
    pdf.set_fill_color(*C_GOLD_RULE)
    pdf.rect(x0, y, 1.5, ph, 'F')
    pdf.set_xy(x0 + 4, y + 2)
    pdf.set_font('Helvetica', 'B', 7)
    pdf.set_text_color(*C_TEXT_MED)
    pdf.cell(30, 4, 'CASE DETAILS', border=0, ln=False)

    left_f  = [('Customer', case['customer']), ('Cluster', case['cluster']), ('Branch', case['branch'])]
    right_f = [('Disbursed On', fmt_date(case['date'])), ('Amount Disbursed', f'Rs {inr_dec(case["amount"])}'), ('Total Payable', f'Rs {inr_dec(case["total"])}')]
    NUMERIC_FIELDS = {'Amount Disbursed', 'Total Payable'}

    for i, (lbl, val) in enumerate(left_f):
        pdf.set_xy(x0 + 4, y + 8 + i * 5.5)
        pdf.set_font('Helvetica', 'B', 7); pdf.set_text_color(*C_TEXT_MED)
        pdf.cell(22, 4, lbl + ':', border=0, ln=False)
        pdf.set_font('Helvetica', 'B', 8); pdf.set_text_color(*C_TEXT_DARK)
        pdf.cell(60, 4, val, border=0, ln=False)
    for i, (lbl, val) in enumerate(right_f):
        pdf.set_xy(x0 + 90, y + 8 + i * 5.5)
        pdf.set_font('Helvetica', 'B', 7); pdf.set_text_color(*C_TEXT_MED)
        pdf.cell(30, 4, lbl + ':', border=0, ln=False)
        if lbl in NUMERIC_FIELDS:
            pdf.set_font('Helvetica', 'B', 8.5); pdf.set_text_color(*C_PRI_NAVY)
        else:
            pdf.set_font('Helvetica', 'B', 8); pdf.set_text_color(*C_TEXT_DARK)
        pdf.cell(67, 4, val, border=0, ln=False)
    y += ph + 4

    # --- AMOUNT SUMMARY (left) + SCAN TO PAY QR (right) ---
    lw  = pwi * 0.56
    rw  = pwi * 0.40
    gap = pwi - lw - rw
    rh  = 9.6

    camt = f'Rs {inr_dec(case["coll_amt"])}' if case['coll_amt'] else '-'
    amt_rows = [
        ('Amount Disbursed',    f'Rs {inr_dec(case["amount"])}',  C_PRI_NAVY, False),
        ('Charges + GST',       f'Rs {inr_dec(case["charges"] + case["gst"])}', C_PRI_NAVY, False),
        ('Total Payable',       f'Rs {inr_dec(case["total"])}',   C_PRI_NAVY, False),
        ('Collected Amount',    camt,                              C_PRI_NAVY, False),
        ('BALANCE OUTSTANDING', f'Rs {inr_dec(case["balance"])}',  C_RED,      True),
    ]
    pdf.set_fill_color(*C_PRI_NAVY)
    pdf.rect(x0, y, lw, rh, 'F')
    pdf.set_xy(x0 + 2, y + 1.8)
    pdf.set_font('Helvetica', 'B', 7); pdf.set_text_color(*C_WHITE)
    pdf.cell(lw * 0.6, rh - 3, 'AMOUNT SUMMARY', align='L', border=0, ln=False)
    pdf.cell(lw * 0.4 - 2, rh - 3, 'AMOUNT', align='R', border=0, ln=False)
    for i, (desc, amt, col, is_hi) in enumerate(amt_rows):
        ry = y + rh + i * rh
        if is_hi:   pdf.set_fill_color(*C_LT_GOLD)
        elif i % 2: pdf.set_fill_color(*C_WHITE)
        else:       pdf.set_fill_color(*C_LT_SLATE)
        pdf.rect(x0, ry, lw, rh, 'F')
        if is_hi:
            pdf.set_fill_color(*C_GOLD_RULE); pdf.rect(x0, ry, lw, 0.7, 'F')
        pdf.set_xy(x0 + 2, ry + 2)
        pdf.set_font('Helvetica', 'B', 8 if is_hi else 7.5)
        pdf.set_text_color(*C_PRI_NAVY if is_hi else C_TEXT_DARK)
        pdf.cell(lw * 0.6 - 2, rh - 3, desc, align='L', border=0, ln=False)
        pdf.set_xy(x0, ry + 1.2)
        pdf.set_font('Helvetica', 'B', 14 if is_hi else 11)
        pdf.set_text_color(*col)
        pdf.cell(lw - 2, rh - 1.5, amt, align='R', border=0, ln=False)

    qr_h = rh + len(amt_rows) * rh   # match amount-summary block height exactly
    rx = x0 + lw + gap
    pdf.set_fill_color(*C_NAVY)
    pdf.rect(rx, y, rw, rh, 'F')
    pdf.set_xy(rx, y + 1.8)
    pdf.set_font('Helvetica', 'B', 7); pdf.set_text_color(*C_WHITE)
    pdf.cell(rw, rh - 3, 'SCAN OR TAP TO PAY (UPI)', align='C', border=0, ln=False)
    body_y = y + rh
    body_h = qr_h - rh
    pdf.set_fill_color(*C_WHITE)
    pdf.set_draw_color(*C_GOLD_RULE)
    pdf.rect(rx, body_y, rw, body_h, 'FD')
    qr_size = min(body_h - 16, rw - 8)
    qx = rx + 4
    qy = body_y + 3
    pay_page_url = build_pay_page_url(case['id'])
    if os.path.exists(QR_PATH) and os.path.getsize(QR_PATH) > 0:
        pdf.image(QR_PATH, x=qx, y=qy, w=qr_size, h=qr_size)
    else:
        pdf.set_draw_color(*C_RULE)
        pdf.rect(qx, qy, qr_size, qr_size, 'D')
        pdf.set_xy(qx, qy + qr_size / 2 - 3)
        pdf.set_font('Helvetica', 'BI', 7); pdf.set_text_color(*C_RULE)
        pdf.cell(qr_size, 6, 'QR PENDING', align='C', border=0, ln=False)
    pdf.link(qx, qy, qr_size, qr_size, pay_page_url)
    # Pay Now button — fills the slack to the right of the QR within the same box, same tap target
    btn_x = qx + qr_size + 3
    btn_w = rx + rw - 2 - btn_x
    btn_h = qr_size
    if btn_w > 14:
        pdf.set_fill_color(*C_NAVY)
        pdf.rect(btn_x, qy, btn_w, btn_h, 'F')
        pdf.set_xy(btn_x, qy + btn_h * 0.32)
        pdf.set_font('Helvetica', 'B', 7.5); pdf.set_text_color(*C_WHITE)
        pdf.cell(btn_w, 5, 'PAY NOW', align='C', border=0, ln=False)
        pdf.set_xy(btn_x, qy + btn_h * 0.32 + 5)
        pdf.set_font('Helvetica', 'I', 6); pdf.set_text_color(*C_GOLD_LBL)
        pdf.cell(btn_w, 4, 'Tap to Pay', align='C', border=0, ln=False)
        pdf.link(btn_x, qy, btn_w, btn_h, pay_page_url)
    pdf.set_xy(rx, qy + qr_size + 2)
    pdf.set_font('Helvetica', 'B', 7.5); pdf.set_text_color(*C_TEXT_DARK)
    pdf.cell(rw, 4, PAYEE_NAME, align='C', border=0, ln=False)
    pdf.set_xy(rx, qy + qr_size + 6.5)
    pdf.set_font('Helvetica', 'B', 7); pdf.set_text_color(*C_PRI_NAVY)
    pdf.cell(rw, 4, f'UPI: {UPI_ID}', align='C', border=0, ln=False)

    pdf.set_text_color(*C_TEXT_DARK)
    y += qr_h + 6

    # --- Branch contact note ---
    contact, fallback = find_branch_contact(case['cluster'], case['branch'], branch_contacts, cluster_mgrs)
    note_h = 16
    pdf.set_fill_color(*C_MEMO_CREAM)
    pdf.set_draw_color(*C_GOLD_RULE)
    pdf.rect(x0, y, pwi, note_h, 'FD')
    pdf.set_fill_color(*C_GOLD_RULE); pdf.rect(x0, y, 1.5, note_h, 'F')
    pdf.set_xy(x0 + 4, y + 2)
    pdf.set_font('Helvetica', 'B', 7); pdf.set_text_color(*C_PRI_NAVY)
    pdf.cell(0, 4, 'BRANCH CONTACT FOR FOLLOW-UP', border=0, ln=False)
    contact_line = f'{contact["name"]}{"  (cluster mgr fallback)" if fallback else ""}   |   {contact["phone"]}'
    pdf.set_xy(x0 + 4, y + 7.5)
    pdf.set_font('Helvetica', 'B', 9.5); pdf.set_text_color(*C_TEXT_DARK)
    pdf.cell(0, 4, contact_line, border=0, ln=False)
    y += note_h + 6

    # --- Signature ---
    sig_w = 70
    sx = x0 + (pwi - sig_w) / 2
    if os.path.exists(SIGNATURE_PATH) and os.path.getsize(SIGNATURE_PATH) > 0:
        sig_img_h = 12
        pdf.image(SIGNATURE_PATH, x=sx + sig_w / 2 - 18, y=y + 14 - sig_img_h, w=36, h=sig_img_h)
    pdf.set_draw_color(*C_RULE)
    pdf.line(sx, y + 14, sx + sig_w, y + 14)
    pdf.set_xy(sx, y + 16)
    pdf.set_font('Helvetica', 'B', 7); pdf.set_text_color(*C_TEXT_MED)
    pdf.cell(sig_w, 4, 'AUTHORISED BY', align='C', border=0, ln=False)
    pdf.set_xy(sx, y + 21)
    pdf.set_font('Helvetica', 'B', 7.5); pdf.set_text_color(*C_TEXT_DARK)
    pdf.cell(sig_w, 4, 'Prem / Harsha', align='C', border=0, ln=False)
    pdf.set_text_color(*C_TEXT_DARK)
    y += 27

    # --- Footer note ---
    pdf.set_xy(x0, y)
    pdf.set_font('Helvetica', 'I', 6.5); pdf.set_text_color(*C_TEXT_MED)
    pdf.multi_cell(pwi, 3.6,
        'Scan the QR code above to settle the outstanding balance instantly via UPI. '
        'For collection use only - please retain a copy for branch records.',
        border=0, align='L')
    pdf.set_text_color(*C_TEXT_DARK)

    pdf.set_page_count(pdf.page_no())
    return pdf.output()

# --- CALLING FOLLOW-UP PDF ---------------------------------------------------
class FollowUpPDF(BLPdf):
    """
    Landscape A4 — extends BLPdf so draw_header/draw_footer/footer/header are
    IDENTICAL to disbursement memos. self.w = 297 (landscape) so all pw-based
    calculations in draw_header adapt automatically.
    """
    def __init__(self):
        # Bypass BLPdf.__init__ and call FPDF directly to get landscape orientation
        FPDF.__init__(self, 'L', 'mm', 'A4')
        self.set_auto_page_break(False)
        self.set_margins(0, 0, 0)
        self.alias_nb_pages('{NB}')
        self._subtitle    = 'CALLING FOLLOW-UP SHEET'
        self._report_date = fmt_date(TODAY)
        self._page_count  = 0
    # draw_header, draw_footer, header, footer, content_top, content_bottom,
    # inner_w, section_title — all inherited unchanged from BLPdf


def generate_calling_followup_pdf(open_cases, cluster_mgrs, branch_contacts):
    """
    Calling follow-up sheet on BridgeLine letterhead (landscape A4).
    Groups open cases by cluster, sorted by days outstanding descending.
    Returns raw PDF bytes for inclusion in ZIP.
    """
    from collections import defaultdict as _dd

    # Landscape inner width = 277mm
    FOLLOW_COLS = [
        (26,  '#',            'C'),
        (28,  'BLP REF',      'L'),
        (42,  'CUSTOMER',     'L'),
        (28,  'BRANCH',       'L'),
        (30,  'OUTSTANDING',  'R'),
        (16,  'DAYS',         'C'),
        (35,  'CONTACT',      'L'),
        (32,  'PHONE',        'L'),
        (40,  'NOTES',        'L'),
    ]
    # total cols = 277mm

    by_cluster = _dd(list)
    for c in open_cases:
        by_cluster[c['cluster']].append(c)
    for cl in by_cluster:
        by_cluster[cl].sort(key=lambda x: x['days_out'], reverse=True)

    if not by_cluster:
        return None

    pdf = FollowUpPDF()
    pdf.add_page()
    rh = 6.5

    def _cluster_bar(cl, mgr, n_cases, total_due):
        x0 = L_MAR
        pw = pdf.inner_w()
        y  = pdf.get_y()
        pdf.set_fill_color(*C_PRI_NAVY)
        pdf.rect(x0, y, pw, 9, 'F')
        pdf.set_fill_color(*C_GOLD_RULE)
        pdf.rect(x0, y, 1.5, 9, 'F')
        pdf.set_xy(x0 + 4, y + 2)
        pdf.set_font('Helvetica', 'B', 9.5)
        pdf.set_text_color(*C_HDR_TXT)
        pdf.cell(55, 5, cl.upper(), border=0, ln=False)
        pdf.set_font('Helvetica', '', 8)
        pdf.set_text_color(*C_GOLD_LBL)
        pdf.cell(50, 5, f'Cluster Mgr: {mgr["name"]}', border=0, ln=False)
        pdf.set_font('Helvetica', 'B', 8)
        pdf.set_text_color(*C_WHITE)
        pdf.cell(42, 5, mgr.get('phone', '-'), border=0, ln=False)
        summary = f'{n_cases} case{"s" if n_cases != 1 else ""}  |  Rs {inr(total_due)} outstanding'
        pdf.set_font('Helvetica', '', 7.5)
        pdf.set_text_color(*C_GOLD_LBL)
        remain = pw - 4 - 55 - 50 - 42
        pdf.set_xy(x0 + pw - remain, y + 2)
        pdf.cell(remain, 5, summary, align='R', border=0, ln=False)
        pdf.set_text_color(*C_TEXT_DARK)
        pdf.set_y(y + 11)

    def _table_header():
        x0 = L_MAR
        y  = pdf.get_y()
        total_w = sum(c[0] for c in FOLLOW_COLS)
        pdf.set_fill_color(*C_NAVY)
        pdf.rect(x0, y, total_w, rh, 'F')
        xc = x0
        for w, lbl, align in FOLLOW_COLS:
            pdf.set_font('Helvetica', 'B', 6.5)
            pdf.set_text_color(*C_WHITE)
            pdf.set_xy(xc, y + 1)
            pdf.cell(w, rh - 2, lbl, align=align, border=0, ln=False)
            xc += w
        pdf.set_y(y + rh)

    any_fallback = False

    for cluster in sorted(by_cluster.keys()):
        cl_cases  = by_cluster[cluster]
        mgr       = cluster_mgrs.get(cluster, {'name': '-', 'phone': '-'})
        total_due = sum(c['balance'] for c in cl_cases)

        if pdf.get_y() > pdf.content_bottom() - 30:
            pdf.add_page()

        _cluster_bar(cluster, mgr, len(cl_cases), total_due)
        _table_header()

        for idx, c in enumerate(cl_cases):
            contact, fallback = find_branch_contact(
                cluster, c['branch'], branch_contacts, cluster_mgrs
            )
            if fallback:
                any_fallback = True

            if pdf.get_y() > pdf.content_bottom() - rh:
                pdf.add_page()
                _table_header()

            x0      = L_MAR
            y       = pdf.get_y()
            total_w = sum(col[0] for col in FOLLOW_COLS)
            d       = c['days_out']

            # Row stripe
            if idx % 2 == 0: pdf.set_fill_color(*C_LT_GOLD)
            else:             pdf.set_fill_color(*C_WHITE)
            pdf.rect(x0, y, total_w, rh, 'F')

            # Days badge colour
            if d >= 5:    badge_fg = C_RED
            elif d >= 3:  badge_fg = (180, 120, 20)
            elif d == 0:  badge_fg = (24,  95, 165)
            else:         badge_fg = C_GREEN

            contact_disp = (contact['name'] + ' *' if fallback else contact['name'])[:20]
            row_data = [
                (str(idx + 1),  'C', False, C_TEXT_MED),
                ('-'.join(c['id'].split('-')[1:]),  'L', False, C_TEXT_MED),
                (c['customer'][:24] if len(c['customer']) > 24 else c['customer'],
                                'L', True,  C_TEXT_DARK),
                (c['branch'][:16] if len(c['branch']) > 16 else c['branch'],
                                'L', False, C_TEXT_MED),
                ('Rs ' + inr(c['balance']),
                                'R', True,  C_RED if c['balance'] >= 500000 else C_TEXT_DARK),
                (str(d),        'C', True,  badge_fg),
                (contact_disp,  'L', False, C_TEXT_MED),
                (contact['phone'], 'L', True, C_PRI_NAVY),
                ('',            'L', False, C_RULE),
            ]

            xc = x0
            for i, ((w, _, _align), (val, align, bold, fg)) in enumerate(
                    zip(FOLLOW_COLS, row_data)):
                pdf.set_xy(xc, y + 1)
                pdf.set_font('Helvetica', 'B' if bold else '', 6.5)
                pdf.set_text_color(*fg)
                if FOLLOW_COLS[i][1] == 'NOTES':
                    pdf.set_text_color(*C_RULE)
                    val = '_____________________________'
                pdf.cell(w, rh - 2, val, align=align, border=0, ln=False)
                xc += w
            pdf.set_y(y + rh)

        # Gold divider after each cluster
        x0 = L_MAR
        pdf.set_fill_color(*C_GOLD_RULE)
        pdf.rect(x0, pdf.get_y(), sum(c[0] for c in FOLLOW_COLS), 0.4, 'F')
        pdf.set_y(pdf.get_y() + 5)

    # --- EXTREME FOLLOW-UP: outstanding under Rs 1,000, all clusters combined ---
    EXTREME_THRESHOLD = 1000
    extreme_cases = [c for c in open_cases if c['balance'] < EXTREME_THRESHOLD]
    extreme_cases.sort(key=lambda x: x['balance'])

    if extreme_cases:
        EXTREME_COLS = [
            (22,  '#',            'C'),
            (28,  'BLP REF',      'L'),
            (40,  'CUSTOMER',     'L'),
            (24,  'CLUSTER',      'L'),
            (24,  'BRANCH',       'L'),
            (28,  'OUTSTANDING',  'R'),
            (16,  'DAYS',         'C'),
            (35,  'CONTACT',      'L'),
            (32,  'PHONE',        'L'),
            (28,  'NOTES',        'L'),
        ]
        # total cols = 277mm

        pdf.add_page()

        def _extreme_bar(n_cases, total_due):
            x0 = L_MAR
            pw = pdf.inner_w()
            y  = pdf.get_y()
            pdf.set_fill_color(*C_PRI_NAVY)
            pdf.rect(x0, y, pw, 9, 'F')
            pdf.set_fill_color(*C_RED)
            pdf.rect(x0, y, 1.5, 9, 'F')
            pdf.set_xy(x0 + 4, y + 2)
            pdf.set_font('Helvetica', 'B', 9.5)
            pdf.set_text_color(*C_HDR_TXT)
            pdf.cell(100, 5,
                     f'EXTREME FOLLOW-UP - DUES UNDER Rs {inr(EXTREME_THRESHOLD)}',
                     border=0, ln=False)
            summary = f'{n_cases} case{"s" if n_cases != 1 else ""}  |  Rs {inr(total_due)} outstanding'
            pdf.set_font('Helvetica', 'B', 8)
            pdf.set_text_color(*C_WHITE)
            remain = pw - 4 - 100
            pdf.set_xy(x0 + pw - remain, y + 2)
            pdf.cell(remain, 5, summary, align='R', border=0, ln=False)
            pdf.set_text_color(*C_TEXT_DARK)
            pdf.set_y(y + 11)

        def _extreme_table_header():
            x0 = L_MAR
            y  = pdf.get_y()
            total_w = sum(c[0] for c in EXTREME_COLS)
            pdf.set_fill_color(*C_NAVY)
            pdf.rect(x0, y, total_w, rh, 'F')
            xc = x0
            for w, lbl, align in EXTREME_COLS:
                pdf.set_font('Helvetica', 'B', 6.5)
                pdf.set_text_color(*C_WHITE)
                pdf.set_xy(xc, y + 1)
                pdf.cell(w, rh - 2, lbl, align=align, border=0, ln=False)
                xc += w
            pdf.set_y(y + rh)

        _extreme_bar(len(extreme_cases), sum(c['balance'] for c in extreme_cases))
        _extreme_table_header()

        for idx, c in enumerate(extreme_cases):
            contact, fallback = find_branch_contact(
                c['cluster'], c['branch'], branch_contacts, cluster_mgrs
            )
            if fallback:
                any_fallback = True

            if pdf.get_y() > pdf.content_bottom() - rh:
                pdf.add_page()
                _extreme_table_header()

            x0      = L_MAR
            y       = pdf.get_y()
            total_w = sum(col[0] for col in EXTREME_COLS)
            d       = c['days_out']

            # Row stripe
            if idx % 2 == 0: pdf.set_fill_color(*C_LT_GOLD)
            else:             pdf.set_fill_color(*C_WHITE)
            pdf.rect(x0, y, total_w, rh, 'F')

            # Days badge colour
            if d >= 5:    badge_fg = C_RED
            elif d >= 3:  badge_fg = (180, 120, 20)
            elif d == 0:  badge_fg = (24,  95, 165)
            else:         badge_fg = C_GREEN

            contact_disp = (contact['name'] + ' *' if fallback else contact['name'])[:18]
            row_data = [
                (str(idx + 1),  'C', False, C_TEXT_MED),
                ('-'.join(c['id'].split('-')[1:]),  'L', False, C_TEXT_MED),
                (c['customer'][:22] if len(c['customer']) > 22 else c['customer'],
                                'L', True,  C_TEXT_DARK),
                (c['cluster'][:14] if len(c['cluster']) > 14 else c['cluster'],
                                'L', False, C_TEXT_MED),
                (c['branch'][:14] if len(c['branch']) > 14 else c['branch'],
                                'L', False, C_TEXT_MED),
                ('Rs ' + inr(c['balance']),
                                'R', True,  C_RED),
                (str(d),        'C', True,  badge_fg),
                (contact_disp,  'L', False, C_TEXT_MED),
                (contact['phone'], 'L', True, C_PRI_NAVY),
                ('',            'L', False, C_RULE),
            ]

            xc = x0
            for i, ((w, _, _align), (val, align, bold, fg)) in enumerate(
                    zip(EXTREME_COLS, row_data)):
                pdf.set_xy(xc, y + 1)
                pdf.set_font('Helvetica', 'B' if bold else '', 6.5)
                pdf.set_text_color(*fg)
                if EXTREME_COLS[i][1] == 'NOTES':
                    pdf.set_text_color(*C_RULE)
                    val = '_____________________'
                pdf.cell(w, rh - 2, val, align=align, border=0, ln=False)
                xc += w
            pdf.set_y(y + rh)

        # Red divider — marks this list as the priority/extreme section
        x0 = L_MAR
        pdf.set_fill_color(*C_RED)
        pdf.rect(x0, pdf.get_y(), sum(c[0] for c in EXTREME_COLS), 0.4, 'F')
        pdf.set_y(pdf.get_y() + 5)

        pdf.set_xy(L_MAR, pdf.get_y())
        pdf.set_font('Helvetica', 'I', 6)
        pdf.set_text_color(*C_RED)
        pdf.cell(pdf.inner_w(), 4,
                 f'Outstanding under Rs {inr(EXTREME_THRESHOLD)} - close these out immediately. '
                 f'Sorted by outstanding amount, ascending.',
                 border=0, ln=False)
        pdf.set_text_color(*C_TEXT_DARK)
        pdf.set_y(pdf.get_y() + 6)

    # Footnote
    pdf.set_xy(L_MAR, pdf.get_y())
    pdf.set_font('Helvetica', 'I', 6)
    pdf.set_text_color(*C_TEXT_MED)
    note = 'Days = calendar days since disbursement as of reporting date.'
    if any_fallback:
        note += '  * Branch not in Contact sheet - cluster manager used as fallback.'
    pdf.cell(pdf.inner_w(), 4, note, border=0, ln=False)
    pdf.set_text_color(*C_TEXT_DARK)

    pdf.set_page_count(pdf.page_no())
    return pdf.output()

# --- CONSOLIDATED MIS PDF ----------------------------------------------------
def generate_consolidated_mis(open_cases, all_cases, all_cases_full, metrics):
    pdf = BLPdf(subtitle='Consolidated MIS Report - All Clusters', report_date=fmt_date(TODAY))
    pdf.add_page()

    draw_kpi_strip(pdf, open_cases, metrics=metrics)
    draw_open_cases_table(pdf, open_cases)

    if pdf.get_y() > pdf.content_bottom() - 30: pdf.add_page()
    draw_cluster_summary(pdf, open_cases)

    if pdf.get_y() > pdf.content_bottom() - 50: pdf.add_page()
    draw_cluster_analytics(pdf, all_cases, metrics)

    if pdf.get_y() > pdf.content_bottom() - 55: pdf.add_page()
    draw_roi_goalpost(pdf, metrics)

    # Portfolio Dashboard — corrected metrics, all clusters incl. Mandya
    pdf.add_page()
    draw_portfolio_dashboard(pdf, metrics)

    # Monthly breakdown — all clusters incl. Mandya
    if pdf.get_y() > pdf.content_bottom() - 60: pdf.add_page()
    draw_monthly_dashboard(pdf, all_cases_full, metrics)

    # MTD / YTD Report — month-wise MTD + running YTD cumulative
    if pdf.get_y() > pdf.content_bottom() - 60: pdf.add_page()
    draw_mtd_ytd_report(pdf, all_cases_full, metrics)

    pdf.set_page_count(pdf.page_no())
    return pdf.output()

# --- CLUSTER MIS PDF ---------------------------------------------------------
def generate_cluster_mis(cluster, open_cases, all_cases, metrics):
    pdf = BLPdf(subtitle=f'{cluster} Cluster MIS Report', report_date=fmt_date(TODAY))
    pdf.add_page()

    cluster_open = [c for c in open_cases if c['cluster'] == cluster]
    draw_kpi_strip(pdf, cluster_open)
    draw_open_cases_table(pdf, open_cases, cluster_filter=cluster)

    if pdf.get_y() > pdf.content_bottom() - 50: pdf.add_page()
    draw_guidelines(pdf)

    pdf.set_page_count(pdf.page_no())
    return pdf.output()

# --- MAIN --------------------------------------------------------------------
def main():
    excel_path = sys.argv[1] if len(sys.argv) > 1 else EXCEL_PATH
    global OUTPUT_DIR
    if len(sys.argv) > 2:
        OUTPUT_DIR = sys.argv[2]

    # Derive BridgeLine MIS folder from OUTPUT_DIR sibling path
    # OUTPUT_DIR = .../MIS Reports  ->  BL_MIS_DIR = .../BridgeLine MIS (on same mount)
    bl_mis_dir = os.path.join(os.path.dirname(OUTPUT_DIR), 'BridgeLine MIS')

    print(f'Reading: {excel_path}')
    print(f'Output:  {OUTPUT_DIR}')

    # Logo lives in /tmp across the session (FUSE mounts deadlock on reads — never read from them)
    if os.path.exists(LOGO_PATH) and os.path.getsize(LOGO_PATH) > 0:
        print(f'Logo ready at {LOGO_PATH}')
    else:
        print(f'WARNING: Logo not in /tmp — header renders without logo (run setup_logo.py to cache it)')

    rows, db_raw, mcoll, (cluster_mgrs, branch_contacts) = load_data(excel_path)
    open_cases, all_cases, all_cases_full = parse_cases(rows, mcoll)

    print(f'Open cases: {len(open_cases)}')
    for c in open_cases:
        print(f'  {c["id"]} | {c["customer"]} | {c["cluster"]} | Bal: Rs {inr(c["balance"])} | {c["status"]}')

    # Compute corrected metrics from raw data
    metrics = compute_dashboard_metrics(all_cases_full, db_raw)
    print(f'\n--- Corrected Dashboard Metrics ---')
    print(f'  Total Invested:          Rs {inr(metrics["Total Invested"])}')
    print(f'  Total Disbursed:         Rs {inr(metrics["Total Disbursed"])}')
    print(f'  Available for Disbursal: Rs {inr(metrics["Available for Disbursement"])}')
    print(f'  Net Charges Earned:      Rs {inr(metrics["Net Charges Earned"])}')
    print(f'  Total Pending Cases:     {metrics["Total Pending Cases"]}')
    print(f'  MTD Charges:             Rs {inr(metrics["MTD Charges"])}')
    print(f'  All-Time ROI:            {metrics["ROI"]*100:.4f}%')

    # Write Claude_Dashboard sheet to Excel in BridgeLine MIS folder
    try:
        write_claude_dashboard(excel_path, metrics, mcoll, bl_mis_dir)
    except Exception as e:
        print(f'WARNING: Could not write Claude_Dashboard: {e}')

    active_clusters = sorted(set(c['cluster'] for c in open_cases))
    print(f'\nActive clusters: {active_clusters}')

    zip_bytes = build_zip(open_cases, all_cases, all_cases_full, metrics,
                           cluster_mgrs, branch_contacts, active_clusters)

    date_human = TODAY.strftime('%d-%b-%Y')
    zip_name   = f'{date_human} BridgeLine MIS Package.zip'
    zip_path   = os.path.join(OUTPUT_DIR, zip_name)
    with open(zip_path, 'wb') as f:
        f.write(zip_bytes)
    print(f'\nDone -> {zip_path}  ({len(zip_bytes)//1024} KB)')
    return zip_path


def build_zip(open_cases, all_cases, all_cases_full, metrics,
               cluster_mgrs, branch_contacts, active_clusters):
    """Builds the full BridgeLine MIS Package zip (in memory) and returns its bytes.

    Shared by the CLI entrypoint (main()) and any other caller (e.g. a web route
    that sources its data live from Google Sheets instead of a local Excel file)
    so both paths produce byte-identical PDFs from the same generation code.
    """
    date_label = TODAY.strftime('%d%m%y')
    date_human = TODAY.strftime('%d-%b-%Y')

    buf = io.BytesIO()
    with zipfile.ZipFile(buf, 'w', zipfile.ZIP_DEFLATED) as zf:
        print('\nGenerating Consolidated MIS...')
        zf.writestr(f'{date_label} Consolidated BridgeLine MIS.pdf',
                    generate_consolidated_mis(open_cases, all_cases, all_cases_full, metrics))

        # --- COLLECTION CARDS folder: Calling Follow-Up + one Collection Card per
        # open case disbursed below COLLECTION_CARD_THRESHOLD, all together for the
        # field/calling team each day ---
        cc_folder = f'{date_label} Collection Cards'

        print('Generating Calling Follow-Up...')
        followup_bytes = generate_calling_followup_pdf(open_cases, cluster_mgrs, branch_contacts)
        if followup_bytes:
            zf.writestr(f'{cc_folder}/{date_human} BridgeLine Calling Follow-Up.pdf', followup_bytes)

        card_cases = [c for c in open_cases
                      if c['amount'] < COLLECTION_CARD_THRESHOLD
                      or c['balance'] < COLLECTION_CARD_THRESHOLD]
        print(f'Generating {len(card_cases)} Collection Cards (disbursed < Rs {inr(COLLECTION_CARD_THRESHOLD)} OR outstanding < Rs {inr(COLLECTION_CARD_THRESHOLD)})...')
        for c in card_cases:
            safe = c['customer'].replace('/', '-').replace('\\', '-')
            card_name = f'{date_label} {c["cluster"]} {safe} {c["id"]} Collection Card.pdf'
            print(f'  Card: {card_name}')
            zf.writestr(f'{cc_folder}/{card_name}', generate_collection_card(c, branch_contacts, cluster_mgrs))

        for cluster in active_clusters:
            folder = f'{date_label} {cluster}'
            print(f'Generating {cluster} cluster MIS...')
            zf.writestr(f'{folder}/{date_label} {cluster} BridgeLine MIS.pdf',
                        generate_cluster_mis(cluster, open_cases, all_cases, metrics))
            for c in [x for x in open_cases if x['cluster'] == cluster]:
                safe = c['customer'].replace('/', '-').replace('\\', '-')
                memo_name = f'{date_label} {cluster} {safe} {c["id"]} Disbursement Memo.pdf'
                print(f'  Memo: {memo_name}')
                zf.writestr(f'{folder}/{memo_name}', generate_memo(c))

    return buf.getvalue()


if __name__ == '__main__':
    main()
