#!/usr/bin/env python3
"""
BridgeLine Accounts Widget — Google Sheets edition.
Hosted on Vercel via api/index.py (WSGI entrypoint into the `app` object below).
"""

import re
import csv
import json
import io
import threading
import time
import os
import urllib.request
from dataclasses import dataclass
from typing import Optional
from datetime import datetime, date, timedelta, timezone
from flask import Flask, request, jsonify, render_template_string, Response, send_from_directory

# ── Google Sheets setup ───────────────────────────────────────────────────────
import gspread
from google.oauth2.credentials import Credentials
from google.auth.transport.requests import Request
from googleapiclient.discovery import build as _build_google_client
from googleapiclient.http import MediaIoBaseUpload

SPREADSHEET_ID = "1LKhDNyOd1u48UFgQafbz3oP4Ehgf1hJBt59F9A-8H7U"
SHEET_NAME     = "Accounts"
SCOPES         = ["https://www.googleapis.com/auth/spreadsheets", "https://www.googleapis.com/auth/drive"]
IST            = timezone(timedelta(hours=5, minutes=30))

# ── Ops alerting ───────────────────────────────────────────────────────────────
# Basic push-notification hook for the highest-stakes routes (disbursement/
# repayment recording, bank-file export) so a real failure surfaces
# immediately instead of only being noticed if someone happens to check.
# ntfy.sh needs no account/API key -- Prem subscribes by opening this exact
# URL (or the ntfy app) on his phone: https://ntfy.sh/bridgeline-ops-25eb13e871
NTFY_TOPIC = "bridgeline-ops-25eb13e871"

def notify_ops(context, exc):
    """Best-effort push notification -- must NEVER raise or block the real
    request, since a notification failure is not a reason to also fail the
    actual disbursement/repayment/export the officer is waiting on.

    Skips ValueError on purpose: every raise ValueError(...) in this codebase
    is a deliberate, expected, user-facing rejection (missing UTR, case not
    found, stale row, bad account number) -- not a real failure. Alerting on
    those would page Prem every time someone forgets a UTR, which trains
    everyone to ignore the channel within a day. Anything else (a real
    exception -- gspread/API errors, bugs, timeouts) still alerts."""
    if isinstance(exc, ValueError):
        return
    try:
        msg = f"{context}: {exc}"[:500]
        req = urllib.request.Request(
            f"https://ntfy.sh/{NTFY_TOPIC}",
            data=msg.encode("utf-8"),
            headers={"Title": f"BridgeLine widget error: {context}",
                     "Priority": "urgent", "Tags": "rotating_light"},
            method="POST",
        )
        urllib.request.urlopen(req, timeout=5)
    except Exception:
        pass

import requests

def trigger_ledger_rebuild():
    """Marks caches dirty only — does NOT call the ledger webhook itself.
    The actual webhook call (rebuild_ledger_now()) reliably takes 10-15s on
    the Apps Script side (it reformats the whole sheet's borders/column
    widths every time), so it must never run inline inside a save request —
    it used to, and that's why every disbursement/repayment save visibly
    took 10-15+ seconds. Save routes call this (cheap, in-process only);
    the frontend separately fires a non-blocking POST /rebuild-ledger right
    after showing save success, so the ledger still catches up within a
    few seconds without the user waiting on it."""
    global _accounts_cache, _config_cache
    _recent_activity_cache.clear()
    _accounts_cache = None
    _config_cache = None

def rebuild_ledger_now():
    """The actual synchronous webhook call, factored out of
    trigger_ledger_rebuild() so it can be invoked from its own independent
    request (POST /rebuild-ledger) instead of inline during a save. Configure
    ledger_webhook_url / ledger_webhook_token in the Config sheet tab; if
    either is blank, this is a silent no-op. A failure here must never break
    anything else, so all errors are swallowed."""
    cfg = load_config()
    url = (cfg.get("ledger_webhook_url") or "").strip()
    token = (cfg.get("ledger_webhook_token") or "").strip()
    if not url or not token:
        return
    try:
        # Apps Script /exec URLs always 302-redirect to the real content
        # URL, and that redirect drops a POST body (gets converted to GET
        # by both curl and requests, per standard 301/302 behavior). Using
        # GET with the token in the query string avoids that entirely.
        requests.get(url, params={"token": token}, timeout=15)
    except Exception as e:
        print(f"[ledger webhook] call failed: {e}")

_gspread_client_cache = None  # (timestamp, client)
_GSPREAD_CLIENT_TTL = 45 * 60  # seconds — Google access tokens last ~60 min
_drive_client_cache = None  # (timestamp, client)

def get_gspread_client():
    """Auth via a long-lived Google OAuth refresh token stored as Vercel env
    vars — no local token.pickle / browser consent flow, since this runs as
    a stateless serverless function with no persistent disk and no display.

    Cached on a warm serverless instance: without this, a single save could
    trigger 2-5 completely independent OAuth token-refresh round trips
    (one per helper function that needed a client), each an extra network
    hop before any actual Sheets read/write even happened."""
    global _gspread_client_cache
    if _gspread_client_cache and (time.time() - _gspread_client_cache[0]) < _GSPREAD_CLIENT_TTL:
        return _gspread_client_cache[1]
    creds = Credentials(
        None,
        refresh_token=os.environ["GOOGLE_REFRESH_TOKEN"],
        client_id=os.environ["GOOGLE_CLIENT_ID"],
        client_secret=os.environ["GOOGLE_CLIENT_SECRET"],
        token_uri="https://oauth2.googleapis.com/token",
        scopes=SCOPES,
    )
    creds.refresh(Request())
    # BackOffHTTPClient retries 429 quota errors with exponential backoff —
    # without it a burst of reads (multi-step edits, recon rebuilds) can die
    # mid-operation and leave Accounts/M Coll half-updated.
    client = gspread.authorize(creds, http_client=gspread.BackOffHTTPClient)
    _gspread_client_cache = (time.time(), client)
    return client

def get_sheet():
    gc = get_gspread_client()
    sh = gc.open_by_key(SPREADSHEET_ID)
    return sh.worksheet(SHEET_NAME)

def get_drive_client():
    """Same cached-creds pattern as get_gspread_client() -- SCOPES already
    includes Drive, this just builds a second client (Sheets vs Drive are
    different Google APIs) from the same refresh token."""
    global _drive_client_cache
    if _drive_client_cache and (time.time() - _drive_client_cache[0]) < _GSPREAD_CLIENT_TTL:
        return _drive_client_cache[1]
    creds = Credentials(
        None,
        refresh_token=os.environ["GOOGLE_REFRESH_TOKEN"],
        client_id=os.environ["GOOGLE_CLIENT_ID"],
        client_secret=os.environ["GOOGLE_CLIENT_SECRET"],
        token_uri="https://oauth2.googleapis.com/token",
        scopes=SCOPES,
    )
    creds.refresh(Request())
    client = _build_google_client("drive", "v3", credentials=creds)
    _drive_client_cache = (time.time(), client)
    return client

def _drive_folder_id_from_url(kyc_folder_url):
    """The Accounts/Requests 'KYC Folder' cell stores a full
    https://drive.google.com/drive/folders/<id> link (written by the field
    app's upload-kyc-docs.py) -- this repo only ever needs the id portion to
    upload into or rename an already-existing folder; it never creates the
    per-request folder itself, that's the field app's job."""
    if not kyc_folder_url:
        return ''
    return kyc_folder_url.rstrip('/').rsplit('/', 1)[-1]

def _upsert_drive_file(drive, folder_id, filename, content_bytes, mime_type):
    """Idempotent by filename -- mirrors upload-kyc-docs.py's upsert_file()
    exactly (a re-generated invoice, e.g. from a corrected amount, replaces
    the earlier file rather than stacking duplicates). supportsAllDrives is
    required throughout -- the KYC Documents folder lives inside the
    BridgeLine Partners Shared Drive, and Drive's API silently can't see
    Shared Drive content without it (confirmed live while building the
    upload side of this)."""
    q = f"'{folder_id}' in parents and name='{filename}' and trashed=false"
    existing = drive.files().list(
        q=q, fields="files(id)", supportsAllDrives=True, includeItemsFromAllDrives=True,
    ).execute().get("files", [])
    media = MediaIoBaseUpload(io.BytesIO(content_bytes), mimetype=mime_type, resumable=False)
    if existing:
        return drive.files().update(fileId=existing[0]["id"], media_body=media, supportsAllDrives=True).execute()
    return drive.files().create(
        body={"name": filename, "parents": [folder_id]}, media_body=media, fields="id",
        supportsAllDrives=True,
    ).execute()

def _rename_drive_folder(drive, folder_id, new_name):
    drive.files().update(fileId=folder_id, body={"name": new_name}, supportsAllDrives=True).execute()

_accounts_bank_col_ensured = False

def _ensure_accounts_bank_account_column(ws):
    """'Bank Account' was added to the schema after the tab already existed
    with exactly 24 columns — the grid must grow before col 25 is
    addressable, same pattern as get_requests_sheet()'s 'Bank' auto-heal.
    The Accounts header row isn't necessarily row 1 (read_accounts_from_gsheet
    finds it dynamically by searching for 'Disbursement ID'), so this must
    use the same lookup rather than assuming row 1. Checked once per warm
    instance, not once per call."""
    global _accounts_bank_col_ensured
    if _accounts_bank_col_ensured:
        return
    all_vals = ws.get_all_values()
    header_idx = next((i for i, r in enumerate(all_vals) if r and 'Disbursement ID' in r), 1)
    header = all_vals[header_idx] if header_idx < len(all_vals) else []
    if 'Bank Account' not in header:
        if ws.col_count < COL['bank_account']:
            ws.resize(cols=COL['bank_account'])
        ws.update_cell(header_idx + 1, COL['bank_account'], 'Bank Account')
    _accounts_bank_col_ensured = True

_accounts_charge_plan_col_ensured = False

def _ensure_accounts_charge_plan_column(ws):
    """Same auto-heal pattern as _ensure_accounts_bank_account_column() —
    'Charge Plan' marks which BIB HDB Karnataka disbursements use the new
    day-based schedule (stamped 'NEW') vs the flat legacy formula (left
    blank — covers every non-BIB-HDB-Karnataka row AND pre-cutover BIB HDB
    Karnataka rows alike, needing no backfill)."""
    global _accounts_charge_plan_col_ensured
    if _accounts_charge_plan_col_ensured:
        return
    all_vals = ws.get_all_values()
    header_idx = next((i for i, r in enumerate(all_vals) if r and 'Disbursement ID' in r), 1)
    header = all_vals[header_idx] if header_idx < len(all_vals) else []
    if 'Charge Plan' not in header:
        if ws.col_count < COL['charge_plan']:
            ws.resize(cols=COL['charge_plan'])
        ws.update_cell(header_idx + 1, COL['charge_plan'], 'Charge Plan')
    _accounts_charge_plan_col_ensured = True

_accounts_kyc_folder_col_ensured = False

def _ensure_accounts_kyc_folder_column(ws):
    """Same auto-heal pattern as _ensure_accounts_bank_account_column() —
    'KYC Folder' is stamped once at save_disbursement() time (looked up from
    the Requests row, never mutated by update_disbursement()/save_repayment()
    afterward), so this only ever needs checking from that one call site."""
    global _accounts_kyc_folder_col_ensured
    if _accounts_kyc_folder_col_ensured:
        return
    all_vals = ws.get_all_values()
    header_idx = next((i for i, r in enumerate(all_vals) if r and 'Disbursement ID' in r), 1)
    header = all_vals[header_idx] if header_idx < len(all_vals) else []
    if 'KYC Folder' not in header:
        if ws.col_count < COL['kyc_folder']:
            ws.resize(cols=COL['kyc_folder'])
        ws.update_cell(header_idx + 1, COL['kyc_folder'], 'KYC Folder')
    _accounts_kyc_folder_col_ensured = True

_accounts_request_id_col_ensured = False

def _ensure_accounts_request_id_column(ws):
    """Same auto-heal pattern as _ensure_accounts_kyc_folder_column() —
    'Request ID' is stamped once at save_disbursement() time (the value is
    already in the posted data, no lookup needed), never mutated afterward."""
    global _accounts_request_id_col_ensured
    if _accounts_request_id_col_ensured:
        return
    all_vals = ws.get_all_values()
    header_idx = next((i for i, r in enumerate(all_vals) if r and 'Disbursement ID' in r), 1)
    header = all_vals[header_idx] if header_idx < len(all_vals) else []
    if 'Request ID' not in header:
        if ws.col_count < COL['request_id']:
            ws.resize(cols=COL['request_id'])
        ws.update_cell(header_idx + 1, COL['request_id'], 'Request ID')
    _accounts_request_id_col_ensured = True

REQUESTS_SHEET_NAME = "Requests"
REQUESTS_HEADERS = [
    "Request ID", "Submitted At", "Customer Name", "Cluster", "Branch",
    "Amount", "Account No", "IFSC", "Phone", "SO Name", "Gold Weight",
    "Status", "Disb ID", "Notes", "Bank", "Debit Account", "KYC Folder",
    "Company"
]

def get_requests_sheet():
    gc = get_gspread_client()
    sh = gc.open_by_key(SPREADSHEET_ID)
    try:
        ws = sh.worksheet(REQUESTS_SHEET_NAME)
    except Exception:
        ws = sh.add_worksheet(title=REQUESTS_SHEET_NAME, rows=1000, cols=len(REQUESTS_HEADERS))
        ws.append_row(REQUESTS_HEADERS)
        return sh, ws
    # Both "Bank" and "Debit Account" were appended to the schema after the
    # tab already existed with fewer columns — the grid must grow before a
    # new column is addressable. Loop over every header so any future
    # schema addition follows the same auto-heal automatically.
    header = ws.row_values(1)
    missing = [h for h in REQUESTS_HEADERS if h not in header]
    if missing:
        if ws.col_count < len(REQUESTS_HEADERS):
            ws.resize(cols=len(REQUESTS_HEADERS))
        for h in missing:
            ws.update_cell(1, REQUESTS_HEADERS.index(h) + 1, h)
    return sh, ws

# ── Bank bulk-upload templates ────────────────────────────────────────────────
# Declarative layouts for the bulk NEFT/RTGS files each portal accepts. Field
# names resolve against the per-request dict built in /requests/export-bulk;
# ('const', X) emits the literal X. Portals validate their own template files
# literally (IDFC especially), so adjust headers here after diffing against
# the portal's current downloaded template — never hardcode layouts elsewhere.
BANK_TEMPLATES = {
    'hdfc': {
        'label': 'HDFC NetBanking Plus',
        'filetype': 'xlsx',
        'filename': 'HDFC_Bulk_{date}.xlsx',
        'columns': [
            ('Beneficiary Name', 'customer'),
            ('Beneficiary Account Number', 'account_no'),
            ('IFSC Code', 'ifsc'),
            ('Amount', 'amount'),
            ('Payment Mode', 'mode'),
            ('Debit Account Number', 'debit_account'),
            ('Value Date', 'value_date'),
            ('Narration', 'narration'),
            ('Beneficiary Mobile', 'phone'),
        ],
    },
    # Column order/names match IDFC's own downloaded bulk-pay template
    # EXACTLY (BLKPAY_YYYYMMDD-idfc.xlsx) — the portal validates this
    # literally, so never reorder/rename without diffing against a fresh
    # download from IDFC first. 'Transaction Type', 'ifsc'/blank-IFSC
    # handling, and the debit-account/email defaults are IFT-rule-specific
    # (see _resolve_idfc_row()), not generic across templates.
    'idfc': {
        'label': 'IDFC FIRST Bank',
        'filetype': 'xlsx',
        'filename': None,  # computed per-export via _next_idfc_batch_filename()
        # Built via _build_idfc_bulk_xlsx() (see below), which starts from
        # IDFC's actual template file byte-for-byte rather than rebuilding
        # a workbook from scratch — cell formats come from the template's
        # own sample row, not a manually-maintained set here.
        'columns': [
            ('Beneficiary Name', 'customer'),
            ('Beneficiary Account Number', 'account_no'),
            ('IFSC', 'ifsc'),
            ('Transaction Type', 'txn_type'),
            ('Debit Account Number', 'debit_account'),
            ('Transaction Date', 'value_date'),
            ('Amount', 'amount'),
            ('Currency', ('const', 'INR')),
            ('Beneficiary Email ID', ('const', 'principal@bridgelinepartners.in')),
            ('Remarks', 'narration'),
            ('Request ID', 'request_id'),
            # Company (HDB/ICICI) isn't captured on the field-app request —
            # it's chosen later when the disbursement is actually entered
            # into Accounts, which happens AFTER this export. Left blank;
            # fill it in manually in the exported file if needed before
            # upload.
            ('Company', ('const', '')),
            ('Cluster', 'cluster'),
            ('Branch', 'branch'),
            ('Customer Phone No', 'phone'),
        ],
    },
    'razorpayx': {
        'label': 'RazorpayX',
        'filetype': 'csv',
        'filename': 'RazorpayX_Bulk_{date}.csv',
        'columns': [
            ('RazorpayX Account Number', 'debit_account'),
            ('Payout Amount', 'amount'),
            ('Payout Currency', ('const', 'INR')),
            ('Payout Mode', 'mode'),
            ('Payout Purpose', ('const', 'payout')),
            ('Fund Account Type', ('const', 'bank_account')),
            ('Contact Name', 'customer'),
            ('Account IFSC', 'ifsc'),
            ('Account Number', 'account_no'),
            ('Contact Phone', 'phone'),
            ('Payout Narration', 'narration'),
            ('Payout Reference Id', 'request_id'),
        ],
    },
}

RTGS_MIN_AMOUNT = 200000

# BridgeLine Partners' own IDFC FIRST account, used as the fixed debit
# account on every IDFC bulk-pay export (Mysuru Branch; IFSC IDFB0080571 /
# SWIFT IDFBINBBMUM aren't needed in the bulk file itself, only for record).
IDFC_DEBIT_ACCOUNT = '52202388167'

# ── Config ────────────────────────────────────────────────────────────────────
COMPANIES = ["HDB", "ICICI", "ESAF"]
CLUSTERS  = ["Bellary", "BIB HDB Karnataka", "Hassan", "Hubli", "Mandya", "Mangalore", "Mysore", "Other"]

# Day-based, GST-inclusive charge schedule for BIB HDB Karnataka only (every
# other cluster stays on the flat 0.5%+18%-on-top formula). Only disbursements
# on/after this date use the new schedule — already-open BIB HDB Karnataka
# cases keep their original flat terms exactly as already recorded. Once a
# disbursement is stamped with its methodology (COL['charge_plan']), this
# constant is never consulted again for that row — see calc_charges().
BIB_HDB_TAT_CUTOVER_DATE = date(2026, 7, 23)
BRANCHES  = sorted([
    "Adyar", "Beejadi", "Bellary", "Chitradurga", "Chitrapady", "Davangere",
    "Gulbarga", "Hassan", "Hospet", "JP Nagar", "Kedinje", "Kollegala", "Kuvempu Nagar",
    "Mandya", "Mysore", "Other", "Puttur", "Saligrama", "Santhekatte",
    "Shankarpura", "Shikaripura", "Shimoga", "Thokoot", "Tumkur", "Udupi",
    "Vadarasse", "Valencia", "Vijay Nagar"
])

# ── Daily MIS Package (generate_mis.py, bundled in this repo) ─────────────────
# generate_mis.py and its image assets are bundled directly into this repo
# (no Drive-mount dynamic import — there's no Drive mount on a serverless host).

import generate_mis as mis

ASSETS_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "assets")

# generate_mis.py expects these cached at its own LOGO_PATH/SIGNATURE_PATH/QR_PATH
# (tempfile.gettempdir(), which is writable per-invocation on Vercel too).
_MIS_ASSET_SOURCES = {
    "bl_logo.png":      "LOGO_PATH",
    "bl_signature.png": "SIGNATURE_PATH",
    "bl_qr.png":        "QR_PATH",
}

def ensure_mis_assets_cached():
    import shutil
    for src_name, attr in _MIS_ASSET_SOURCES.items():
        src_path = os.path.join(ASSETS_DIR, src_name)
        dest_path = getattr(mis, attr)
        if os.path.exists(src_path):
            shutil.copyfile(src_path, dest_path)

class _SheetShim:
    """Wraps gspread's get_all_values() rows so generate_mis.load_mcoll() /
    load_contacts() (written against openpyxl's ws.iter_rows API) work unchanged
    against live Sheets data instead of an Excel workbook."""
    def __init__(self, rows):
        self._rows = rows

    def iter_rows(self, min_row=1, max_row=None, values_only=True):
        start = min_row - 1
        end = max_row if max_row is not None else len(self._rows)
        for r in self._rows[start:end]:
            yield tuple(r)

class _WorkbookShim:
    def __init__(self, sheets):
        self._sheets = sheets

    @property
    def sheetnames(self):
        return list(self._sheets.keys())

    def __getitem__(self, name):
        return self._sheets[name]

# Numeric columns (0-indexed) in the Accounts sheet that parse_cases() reads
# with float(): amount, charges, gst, total, coll_amt, discount, balance.
# openpyxl(data_only=True) returns the computed number; gspread's
# get_all_values() returns the *displayed* string (e.g. "2,61,000.00"), so
# these need comma-stripping before generate_mis's float() calls see them.
_MIS_NUMERIC_COLS = {7, 8, 9, 10, 12, 13, 14}

def _clean_numeric_cell(v):
    if isinstance(v, str) and v.strip():
        cleaned = v.replace(',', '').strip()
        try:
            float(cleaned)
            return cleaned
        except ValueError:
            return v
    return v

def load_data_from_sheet(sh):
    """Live-Sheets equivalent of generate_mis.load_data(): same return shape
    (rows, db_raw, mcoll, (cluster_mgrs, branch_contacts, area_mgrs, territory_to_area), capital_log),
    sourced from the same spreadsheet the widget already reads/writes,
    instead of a manually downloaded Excel snapshot."""
    acc_vals = sh.worksheet(SHEET_NAME).get_all_values()
    # Find header row dynamically — it's the row containing "Disbursement ID"
    # (same approach as read_accounts_from_gsheet()). A row inserted/removed
    # above the header otherwise makes it get read as a data row and crashes
    # float() conversion on header text like 'Amount'.
    header_idx = next((i for i, r in enumerate(acc_vals) if r and 'Disbursement ID' in r), 1)
    raw_rows = list(acc_vals[header_idx + 1:])

    # Monthly archive tabs (cases relocated out of Accounts but still need
    # follow-up until Closed). Same column order as Accounts, but unlike the
    # comment here used to claim, not all of them are header-row-free (e.g.
    # 'Jun 26' has a title row + column-header row) — filter to actual case
    # rows by disb-id prefix, same as read_accounts_from_gsheet()'s archive
    # loop already does, instead of assuming a fixed/no-header shape.
    for archive_name in get_archive_tab_names():
        try:
            raw_rows += [r for r in sh.worksheet(archive_name).get_all_values()
                         if r and str(r[0]).strip().startswith('BLP-')]
        except gspread.exceptions.WorksheetNotFound:
            continue

    rows = []
    for raw in raw_rows:
        row = list(raw[:26])
        row += [None] * (26 - len(row))
        row = [None if c == '' else c for c in row]  # match openpyxl's blank-cell None
        for i in _MIS_NUMERIC_COLS:
            row[i] = _clean_numeric_cell(row[i])
        if not row[0] or str(row[0]).strip() == '':
            continue
        rows.append(row)

    sheets = {}
    try:
        mcoll_vals = sh.worksheet('M Coll').get_all_values()
        mcoll_vals = [
            [_clean_numeric_cell(c) if i == 2 else c for i, c in enumerate(r)]
            for r in mcoll_vals
        ]
        sheets['M Coll'] = _SheetShim(mcoll_vals)
    except gspread.exceptions.WorksheetNotFound:
        pass
    try:
        sheets['Contact'] = _SheetShim(sh.worksheet('Contact').get_all_values())
    except gspread.exceptions.WorksheetNotFound:
        pass
    try:
        sheets['Capital Log'] = _SheetShim(sh.worksheet('Capital Log').get_all_values())
    except gspread.exceptions.WorksheetNotFound:
        pass
    wb_shim = _WorkbookShim(sheets)

    mcoll = mis.load_mcoll(wb_shim)
    cluster_mgrs, branch_contacts, area_mgrs, territory_to_area = mis.load_contacts(wb_shim)
    capital_log = mis.load_capital_log(wb_shim)

    db_raw = {}
    try:
        db_vals = sh.worksheet('DashBoard').get_all_values()
        for raw in db_vals[:11]:
            label = raw[0] if len(raw) > 0 else None
            val = raw[1] if len(raw) > 1 else None
            if label:
                db_raw[str(label).strip()] = val
    except gspread.exceptions.WorksheetNotFound:
        pass

    return rows, db_raw, mcoll, (cluster_mgrs, branch_contacts, area_mgrs, territory_to_area), capital_log

# ── Extraction helpers ────────────────────────────────────────────────────────

def parse_inr_amount(text):
    for p in [
        r'INR\s+([\d,]+(?:\.\d+)?)',
        r'Rs\.?\s*([\d,]+(?:\.\d+)?)',
        r'debited for Rs\.([\d,]+(?:\.\d+)?)',
        r'(?:amount|transferred|transfer)[^\d]*([\d,]+)(?:\s*/-)?',
        r'([\d,]+)\s*/-',
    ]:
        m = re.search(p, text, re.IGNORECASE)
        if m:
            try:
                val = float(m.group(1).replace(',', ''))
                # A genuine repayment can be well under Rs.1000 (part
                # payments, small collections) -- the old >= 1000 floor
                # silently rejected those instead of just accepting any
                # real positive amount matched by the patterns above.
                if val > 0:
                    return val
            except ValueError:
                pass
    return None

def parse_date_from_message(text):
    patterns = [
        # Explicit "Date:" / "Txn Date:" / "Value Date:" label -- the most
        # common real-world format (e.g. "a. Date: 28-07-26") and previously
        # not matched at all since every other pattern below requires either
        # an "on " prefix or a 3-letter month name.
        (r'Date[:\s]+(\d{2}[-/]\d{2}[-/]\d{2,4})',   ["%d-%m-%y", "%d-%m-%Y"]),
        (r'on\s+(\d{2}[-/][A-Za-z]{3}[-/]\d{2,4})', ["%d-%b-%y", "%d-%b-%Y"]),
        (r'On\s+(\d{2}[-/]\d{2}[-/]\d{2,4})',        ["%d-%m-%y", "%d-%m-%Y"]),
        (r'on\s+(\d{2}[-/]\d{2}[-/]\d{2,4})',        ["%d-%m-%y", "%d-%m-%Y"]),
        (r'(\d{2}-[A-Z]{3}-\d{2})\b',                ["%d-%b-%y"]),
        (r'(\d{4}-\d{2}-\d{2})',                      ["%Y-%m-%d"]),
    ]
    for pattern, fmts in patterns:
        m = re.search(pattern, text, re.IGNORECASE)
        if m:
            for fmt in fmts:
                try:
                    return datetime.strptime(m.group(1), fmt).strftime('%d-%m-%Y')
                except ValueError:
                    pass
    # No pattern matched -- return None rather than defaulting to today.
    # Silently returning today() made every failed parse indistinguishable
    # from a genuine "the message says today", so the frontend confidently
    # filled in a date that had nothing to do with the pasted message.
    return None

def extract_utr(text):
    t = text.strip()

    # Explicit UTR/Ref label in SMS-style messages
    # (?:No|Number|#) only — a bare "ID" alternative here would greedily
    # swallow the first two characters of any UTR that itself starts with
    # "ID" (e.g. IDFC's own "IDFBR..." RTGS codes), truncating the capture.
    m = re.search(r'(?:UTR|Ref(?:erence)?)\s*(?:No|Number|#)?\.?\s*[:\-]?\s*([A-Z0-9]{8,22})', t, re.IGNORECASE)
    if m:
        val = m.group(1).upper()
        if not re.match(r'^\d{1,6}$', val):
            return val

    # Transaction ID label (same truncation risk as above — "ID" dropped
    # from the optional suffix so codes starting with "ID" aren't clipped)
    m = re.search(r'(?:Transaction|Txn|Trans)[\s_]?(?:No|Ref)?\s*[:\-]?\s*([A-Z0-9]{8,22})', t, re.IGNORECASE)
    if m:
        val = m.group(1).upper()
        if not re.match(r'^\d{1,6}$', val):
            return val

    # Bank UTR codes: HDFCR (RTGS), HDFCH (NEFT), CNRBR, SBINR, KARBR, KARBN, SBIN4,
    # UTIBR, SIBLR, IOBAR, IOBAN, SUSBR, PKGBR, BDBLR, UBINR, BARBR, UJVNH, etc.
    # Pattern: 3-6 uppercase letters followed by any letter or digit, then 8+ digits
    m = re.search(r'\b([A-Z]{3,6}[A-Z0-9]\d{8,16})\b', t, re.IGNORECASE)
    if m:
        val = m.group(1).upper()
        # Reject pure IFSC codes (e.g. UBIN0905925 = 4 letters + 0 + 6 alphanum, no long digit block)
        if not re.match(r'^[A-Z]{4}0[A-Z0-9]{6}$', val):
            return val

    # IDFC's NEFT UTR format breaks the digit run with a single checksum
    # letter partway through (e.g. IDFB6196M1199949 = IDFB + 6196 + M +
    # 1199949) — the pattern above requires one unbroken 8-16 digit block
    # right after the bank code, so it never matches these.
    m = re.search(r'\b([A-Z]{3,6}\d{3,6}[A-Z]\d{6,12})\b', t, re.IGNORECASE)
    if m:
        return m.group(1).upper()

    # IMPS: IMPS-REFNUM-... or IMPS Ref No: REFNUM
    m = re.search(r'IMPS[\s\-](\d{10,15})\b', t, re.IGNORECASE)
    if m:
        return m.group(1)
    m = re.search(r'IMPS\s*Ref\s*(?:no|number|#)?\s*[:\-]?\s*(\d{10,15})', t, re.IGNORECASE)
    if m:
        return m.group(1)
    # IMPS with a name/company in between (e.g. "IMPS -TRIGOLDWYSE LLP- 617415701480
    # Avl bal INR") — find any standalone long digit run anywhere after IMPS
    m = re.search(r'IMPS\b.*?\b(\d{10,15})\b', t, re.IGNORECASE)
    if m:
        return m.group(1)

    # UPI: ref number is the 12-digit segment before the last -UPI or -PAYMENT suffix,
    # or before @VPA section. Pattern: UPI-...-DIGITS-WORD or UPI-...-DIGITS end-of-string
    m = re.search(r'UPI.*?-(\d{10,15})-?(?:[A-Z]*\s*$|UPI|PAYMENT)', t, re.IGNORECASE)
    if m:
        return m.group(1)
    # UPI with ref embedded anywhere as standalone 12-digit number
    if re.match(r'^UPI[\s\-]', t, re.IGNORECASE):
        m = re.search(r'\b(\d{12})\b', t)
        if m:
            return m.group(1)
    # "Credit Alert!" style: "... from VPA name@bank (UPI 615689900342)"
    m = re.search(r'\(UPI\s+(\d{10,15})\)', t, re.IGNORECASE)
    if m:
        return m.group(1)

    # Ref label
    m = re.search(r'\bref\b\s*[:\-#]?\s*([A-Z0-9]{8,22})', t, re.IGNORECASE)
    if m:
        val = m.group(1).upper()
        if not re.match(r'^\d{1,6}$', val):
            return val

    # 12-digit number at end of string (fallback)
    m = re.search(r'\b(\d{12})\s*\.?\s*$', t)
    if m:
        return m.group(1)

    return ''

def extract_sender_name(text):
    m = re.search(r'Cr-\w{8,}-(.+?)-(?:M[\s./]?S[\s./]?|Bridgeline|BRIDG)', text, re.IGNORECASE)
    if m:
        return m.group(1).strip().title()
    return ''

def extract_repayment(text):
    info = {
        'amount':  parse_inr_amount(text),
        'date':    parse_date_from_message(text),
        'utr':     extract_utr(text),
        'disb_id': '',
        'sender':  extract_sender_name(text),
    }
    m = re.search(r'BLP[-/]\d{6}[-/]\d{3}', text, re.IGNORECASE)
    if m:
        info['disb_id'] = m.group(0).upper()
    return info

def extract_disbursement(text):
    info = {
        'amount':   parse_inr_amount(text),
        'date':     parse_date_from_message(text),
        'customer': '',
        'company':  '',
        'cluster':  '',
        'branch':   '',
    }
    _exclude = {w.lower() for w in BRANCHES + CLUSTERS + COMPANIES} | {
        'name', 'branch', 'mobile', 'bank', 'ifsc', 'account', 'normal', 'bt',
        'pledged', 'value', 'transferred', 'charges', 'amount', 'net', 'wt',
    }

    def _valid_name(n):
        n = n.strip()
        if not (2 < len(n) < 60):
            return False
        words = re.split(r'\s+', n.lower())
        if any(w in _exclude for w in words):
            return False
        return True

    SEP = r'[\s:*\-]+'
    for p in [
        r'(?:customer\s+name|customer)' + SEP + r'([A-Za-z][A-Za-z\s\.]+?)(?:\n|,|/|$)',
        r'(?:^|\b)name' + SEP + r'([A-Za-z][A-Za-z\s\.]+?)(?:\n|,|/|$)',
        r'(?:borrower|client)' + SEP + r'([A-Za-z][A-Za-z\s\.]+?)(?:\n|,|/|$)',
    ]:
        m = re.search(p, text, re.IGNORECASE | re.MULTILINE)
        if m:
            name = m.group(1).strip()
            if _valid_name(name):
                info['customer'] = name.title()
                break

    if re.search(r'\bHDB\b', text, re.IGNORECASE):
        info['company'] = 'HDB'
    elif re.search(r'\bICICI\b', text, re.IGNORECASE):
        info['company'] = 'ICICI'
    elif re.search(r'\bESAF\b', text, re.IGNORECASE):
        info['company'] = 'ESAF'

    for c in CLUSTERS:
        if re.search(r'\b' + re.escape(c) + r'\b', text, re.IGNORECASE):
            info['cluster'] = c
            break

    for b in BRANCHES:
        if re.search(r'\b' + re.escape(b) + r'\b', text, re.IGNORECASE):
            info['branch'] = b
            break

    return info

# ── Data layer ────────────────────────────────────────────────────────────────

import pandas as pd

COL = {
    'disb_id': 1, 'date': 2, 'customer': 3, 'chq': 4, 'company': 5,
    'cluster': 6, 'branch': 7, 'amount': 8, 'charges': 9, 'gst': 10,
    'total': 11, 'coll_date': 12, 'coll_amount': 13, 'discount': 14,
    'balance': 15, 'tat': 16, 'current_date': 17, 'discrepancy': 18,
    'status': 19, 'srv_branch': 20, 'srv_cluster': 21, 'debit_note': 22,
    'credit_note': 23,
    'remarks': 24,
    'bank_account': 25,
    'charge_plan': 26,
    'kyc_folder': 27,
    'request_id': 28,
}

def _to_num(v):
    try:
        if pd.isna(v): return 0.0
        return float(str(v).replace(',','').replace('₹','').strip())
    except Exception:
        return 0.0

BIB_HDB_CLUSTER_NAME = "BIB HDB Karnataka"

def _norm_cluster(cluster):
    return str(cluster or '').strip().casefold()

def _is_bib_hdb_cluster(cluster):
    # Cluster is free-text (an <input list=...>, not a real <select>), so a
    # bare == would silently miss "bib hdb karnataka" / trailing-whitespace
    # variants and fall through to the flat formula for a case that should
    # be on the new schedule.
    return _norm_cluster(cluster) == _norm_cluster(BIB_HDB_CLUSTER_NAME)

def _bib_hdb_territory_names():
    """Branch/Area names registered under BIB HDB Karnataka in the Contact
    sheet (Area Manager rows like 'ROK'/'Bengaluru-1' + Territory Manager
    rows like 'SHIKARIPURA'), normalized. Sourced live from the sheet
    (data-driven, same convention as find_branch_contact()'s Area/Territory
    resolution) so a newly onboarded territory is covered automatically,
    with no code change. Returns an empty set on any read failure — this
    check must never block or error out a save just because Contact was
    briefly unreachable."""
    try:
        contacts = read_contacts()
    except Exception:
        return set()
    return {_norm_cluster(c['branch']) for c in contacts
            if c.get('branch') and _is_bib_hdb_cluster(c.get('cluster', ''))}

def _bib_territory_mismatch_warning(cluster, branch):
    """Confirmed real failure mode (Fayaz ahammed, BLP-040826-371,
    04-08-2026): BIB HDB Karnataka's own real place names (e.g.
    'Shikaripura') can be identical to an ordinary cluster's real branch
    name, so an officer picking the wrong cluster from the free-text
    dropdown looks completely valid to every other check -- nothing else
    would ever catch it. Non-blocking by design (a name collision is a
    signal, not proof; the ordinary cluster's branch might just happen to
    share the name for real) -- surfaced as a warning for a human to
    confirm, never silently auto-corrected."""
    if _is_bib_hdb_cluster(cluster):
        return None
    b = _norm_cluster(branch)
    if not b:
        return None
    if b in _bib_hdb_territory_names():
        return (f"Branch '{branch.strip()}' matches a known BIB HDB Karnataka "
                f"territory/area name, but Cluster is set to "
                f"'{cluster.strip() if cluster and cluster.strip() else '(blank)'}'. "
                f"If this is actually BIB HDB Karnataka business, change Cluster so the "
                f"correct day-based charge schedule applies.")
    return None

def _to_plain_date(v):
    """Accepts a date, a datetime, or a string in any format _parse_any_date
    understands; returns a plain date (or None)."""
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    parsed = _parse_any_date(v)
    return parsed.date() if parsed else None

# ── Charge-plan registry ─────────────────────────────────────────────────────
# Onboarding a counterparty with its own rate must be a CONFIG change, not a
# code change. Each plan declares what it matches on, when it starts, and the
# formula — so a new company needs one entry here, and every code path
# (disbursement, edit, closing, invoice, integrity check) picks it up at once.
#
# Matching is on company AND/OR cluster because the two real arrangements are
# shaped differently: ICICI is a whole company, while "BIB HDB Karnataka" is
# one programme inside HDB (394 of 404 cases are company HDB; only 3 of those
# are that cluster). Keying on either field alone cannot express both.
#
# 'id' is what gets stamped into COL['charge_plan'] and is thereafter
# authoritative for that row — the row's OWN plan, never re-derived from
# today's config. The legacy id 'NEW' is kept for BIB HDB so the ~3 rows
# already stamped need no migration; blank still means the flat default.
DEFAULT_CHARGE_PLANS = [
    {
        'id': 'NEW',                       # legacy id, do not rename
        'label': 'BIB HDB Karnataka day-based',
        'match': {'cluster': BIB_HDB_CLUSTER_NAME},
        'from': '2026-07-23',
        'formula': {'type': 'day_based_inclusive', 'base_pct': 0.4, 'per_day_pct': 0.2,
                    'gst_pct': 18, 'free_days': 1},
    },
]

FLAT_DEFAULT_FORMULA = {'type': 'flat_exclusive', 'pct': 0.5, 'gst_pct': 18}

def get_charge_plans():
    """Registered plans, Config-overridable. A malformed Config value must
    never silently disable a counterparty's rate — fall back to the built-in
    list rather than to 'no plans', which would quietly charge everyone the
    flat default (the exact failure this whole system exists to prevent)."""
    try:
        plans = load_config().get('charge_plans')
    except Exception:
        plans = None
    if not isinstance(plans, list) or not plans:
        return DEFAULT_CHARGE_PLANS
    good = [p for p in plans if isinstance(p, dict) and p.get('id') and p.get('formula')]
    return good or DEFAULT_CHARGE_PLANS

def _plan_matches(plan, company, cluster):
    """Every key present in the plan's 'match' must match. An empty match
    would apply to everything, so it never matches at all."""
    m = plan.get('match') or {}
    if not m:
        return False
    if 'cluster' in m and _norm_cluster(m['cluster']) != _norm_cluster(cluster or ''):
        return False
    if 'company' in m and _norm_cluster(m['company']) != _norm_cluster(company or ''):
        return False
    return True

def resolve_charge_plan(company, cluster, disbursement_date):
    """The plan id for a NEW row, or None for the flat default. Only ever
    called where no stamped marker yet exists (save_disbursement, and an edit
    that changes company/cluster/date). Every other caller must pass the
    row's already-stamped id — see calc_charges()'s docstring.

    More specific plans win: a rule naming both company and cluster beats one
    naming only a company, so a programme inside a company can override that
    company's own default rate."""
    d = _to_plain_date(disbursement_date)
    if d is None:
        return None
    best, best_specificity = None, -1
    for plan in get_charge_plans():
        if not _plan_matches(plan, company, cluster):
            continue
        start = _to_plain_date(plan.get('from')) if plan.get('from') else None
        if start and d < start:
            continue            # pre-cutover rows keep the flat default, forever
        until = _to_plain_date(plan.get('until')) if plan.get('until') else None
        if until and d > until:
            continue
        spec = len(plan.get('match') or {})
        if spec > best_specificity:
            best, best_specificity = plan.get('id'), spec
    return best

def get_plan_by_id(plan_id):
    if not plan_id:
        return None
    for plan in get_charge_plans():
        if str(plan.get('id')) == str(plan_id):
            return plan
    return None

def _resolve_bib_hdb_methodology(cluster, disbursement_date, company=None):
    """Back-compat shim over resolve_charge_plan(). Kept because several call
    sites and the 'OLD'/'NEW'/None vocabulary predate the registry."""
    return resolve_charge_plan(company, cluster, disbursement_date)

def calc_charges(amount, cluster=None, disbursement_date=None, collection_date=None,
                 methodology=None, company=None):
    """Single source of truth for Charges/GST/Total math — replaces the
    flat 0.5%/18%-on-top formula that used to be duplicated across
    calc_total(), save_disbursement(), and update_disbursement().

    For every cluster except BIB HDB Karnataka (and for BIB HDB Karnataka
    cases disbursed before BIB_HDB_TAT_CUTOVER_DATE), this is exactly the
    original flat formula — charges = 0.5% of amount, GST = 18% of charges,
    added on top.

    BIB HDB Karnataka cases disbursed on/after the cutover use a day-based,
    GST-INCLUSIVE schedule instead: 0.4% inclusive if collected by EOD of
    the day after disbursement, +0.2% inclusive per additional day late,
    uncapped. The inclusive total is rounded FIRST, then the GST portion is
    back-calculated as the exact remainder (charges = incl/1.18, gst =
    incl - charges) — never independently rounded — so charges+gst always
    reconstructs the inclusive total to the paisa (the invoice's combined
    "Charges (incl. GST)" line depends on this).

    `methodology` ('OLD'/'NEW') should normally be passed explicitly,
    sourced from the row's already-stamped charge_plan marker — this keeps
    every caller except save_disbursement() from ever re-deriving
    methodology from today's date after the fact (see COL['charge_plan']
    and _ensure_accounts_charge_plan_column()). Only save_disbursement()
    should leave it as None and let this function resolve it fresh, since
    that's the one moment no marker yet exists.

    `collection_date=None` means "not collected yet" — used for the
    provisional estimate shown at disbursement time (days_late=0, i.e. the
    best-case 0.4% rate). The real, final charge is only computed once a
    collection_date is supplied, at the moment a case actually closes.
    """
    amt = float(amount or 0)

    if methodology is None:
        methodology = resolve_charge_plan(company, cluster, disbursement_date)

    plan = get_plan_by_id(methodology)
    # 'OLD', a stale id, or None all mean the flat default. Resolving an
    # unknown id to flat rather than raising is deliberate: a plan deleted
    # from Config must not break historical rows that reference it.
    formula = (plan or {}).get('formula') or FLAT_DEFAULT_FORMULA
    gst_pct = float(formula.get('gst_pct', 18))

    if formula.get('type') != 'day_based_inclusive':
        pct     = float(formula.get('pct', 0.5))
        charges = round(amt * pct / 100, 2)
        gst     = round(charges * gst_pct / 100, 2)
        total   = round(amt + charges + gst, 2)
        return {'charges': charges, 'gst': gst, 'total': total,
                'methodology': methodology if plan else None,
                'plan_label': (plan or {}).get('label', 'Flat default'),
                'rate_pct': pct, 'days_late': None, 'gst_inclusive': False}

    d_date = _to_plain_date(disbursement_date)
    c_date = _to_plain_date(collection_date)
    free_days = int(formula.get('free_days', 1))
    days_late = 0 if c_date is None or d_date is None else max(0, (c_date - d_date).days - free_days)
    rate_pct  = float(formula.get('base_pct', 0.4)) + float(formula.get('per_day_pct', 0.2)) * days_late

    # Inclusive total rounded FIRST, GST then taken as the exact remainder —
    # never independently rounded — so charges+gst always reconstructs the
    # inclusive total to the paisa (the invoice's single combined line
    # depends on this).
    incl_total = round(amt * rate_pct / 100, 2)
    charges    = round(incl_total / (1 + gst_pct / 100), 2)
    gst        = round(incl_total - charges, 2)
    total      = round(amt + incl_total, 2)

    return {'charges': charges, 'gst': gst, 'total': total,
            'methodology': methodology, 'plan_label': plan.get('label', methodology),
            'rate_pct': rate_pct, 'days_late': days_late, 'gst_inclusive': True}

def calc_total(amount, charges=None, gst=None):
    if charges is not None or gst is not None:
        amt = float(amount or 0)
        ch  = round(float(charges if charges is not None else amt * 0.005), 2)
        g   = round(float(gst if gst is not None else ch * 0.18), 2)
        return round(amt + ch + g, 2)
    # No explicit override — delegate to the flat/legacy branch (cluster and
    # dates omitted resolves methodology to None, i.e. the original formula).
    # calc_total()'s 4 existing callers never pass cluster/dates and can't
    # participate in the BIB HDB Karnataka schedule, so this preserves their
    # exact prior behavior.
    return calc_charges(amount)['total']

FOLLOWUP_SHEET_NAME = 'Apr/May26'  # legacy follow-up tab, grandfathered into active_archive_tabs

def get_archive_tab_names():
    """Names of monthly archive tabs (e.g. 'June 26') still being merged into
    every Accounts read, sourced from the Config sheet's 'active_archive_tabs'
    list (JSON array) instead of a hardcoded name — so a brand-new tab the
    monthly rollover creates is picked up automatically, with no code change.
    Falls back to just the legacy FOLLOWUP_SHEET_NAME if the Config key is
    missing (e.g. first run before it's been seeded).
    """
    cfg = load_config()
    tabs = cfg.get('active_archive_tabs')
    if not tabs:
        return [FOLLOWUP_SHEET_NAME]
    return tabs

_accounts_cache = None  # (timestamp, rows)
_ACCOUNTS_CACHE_TTL = 15  # seconds
_config_cache = None  # (timestamp, cfg dict) — load_config()'s lenient-path cache

def read_accounts_from_gsheet():
    """Cached: this is the shared hot path for lookup_case(), get_open_cases(),
    _case_detail()'s fallback, get_recent_activity(), and more — each call
    re-scans Accounts + every active archive tab (4-5 Sheets reads). Left
    uncached, ordinary usage (dashboard load + opening Edit Case + a save)
    burns through the per-minute Sheets quota fast, and gspread's backoff
    then silently retries instead of failing, making saves take 20-40s.
    A warm serverless instance reuses this across requests within the TTL;
    trigger_ledger_rebuild() clears it after every write so a save's own
    follow-up read never sees stale data."""
    global _accounts_cache
    if _accounts_cache and (time.time() - _accounts_cache[0]) < _ACCOUNTS_CACHE_TTL:
        return _accounts_cache[1]
    rows = _read_accounts_from_gsheet_uncached()
    _accounts_cache = (time.time(), rows)
    return rows

def _read_accounts_from_gsheet_uncached():
    sh = get_gspread_client().open_by_key(SPREADSHEET_ID)
    ws = sh.worksheet(SHEET_NAME)
    all_vals = ws.get_all_values()
    # Find header row dynamically — it's the row containing "Disbursement ID"
    header_idx = next((i for i, r in enumerate(all_vals) if r and 'Disbursement ID' in r), 1)
    headers = all_vals[header_idx]
    rows = []
    for i, row in enumerate(all_vals[header_idx + 1:], start=header_idx + 2):
        if row and row[0].startswith('BLP-'):
            record = {'_row': i, '_sheet': SHEET_NAME}
            for j, h in enumerate(headers):
                record[h] = row[j] if j < len(row) else ''
            rows.append(record)

    # Monthly archive tabs (e.g. 'Apr/May26', 'June 26', ...). No header row —
    # columns are in the same order as Accounts, so reuse its header list.
    # Keep reading a tab until every case in it shows Overdue Status = Closed.
    for archive_name in get_archive_tab_names():
        try:
            ws2 = sh.worksheet(archive_name)
            vals2 = ws2.get_all_values()
            for i, row in enumerate(vals2, start=1):
                if row and row[0].startswith('BLP-'):
                    record = {'_row': i, '_sheet': archive_name}
                    for j, h in enumerate(headers):
                        record[h] = row[j] if j < len(row) else ''
                    rows.append(record)
        except gspread.exceptions.WorksheetNotFound:
            continue

    return rows

def read_mcoll_from_gsheet():
    """Each row here is one individual instalment/payment against a case,
    with its own date and amount - unlike Accounts.'Collected Amount', which
    is a running total across every payment ever made for that case."""
    sh = get_gspread_client().open_by_key(SPREADSHEET_ID)
    try:
        ws = sh.worksheet('M Coll')
    except gspread.exceptions.WorksheetNotFound:
        return []
    all_vals = ws.get_all_values()
    rows = []
    for row in all_vals[1:]:
        if row and row[0]:
            rows.append({
                'disb_id': row[0],
                'coll_date': row[1] if len(row) > 1 else '',
                'amount': row[2] if len(row) > 2 else '',
                'bank_account': row[5] if len(row) > 5 else '',
            })
    return rows

def get_payment_events(records, mcoll_rows):
    """Flattens collections into individual dated payment events, instead of
    the cumulative 'Collected Amount' column on the Accounts row. Filtering
    the cumulative column by date (e.g. 'collected today') is wrong whenever
    a case has more than one instalment, since that column already includes
    every prior instalment too. A case with M Coll history uses those rows
    (one per instalment); a case with no M Coll history was paid in a single
    one-shot payment, so its Accounts row IS that one event.
    """
    mcoll_by_disb = {}
    for m in mcoll_rows:
        mcoll_by_disb.setdefault(m['disb_id'], []).append(m)

    events = []
    for r in records:
        did = r.get('Disbursement ID', '')
        name = r.get('Customer Name', '')
        if did in mcoll_by_disb:
            for m in mcoll_by_disb[did]:
                d = parse_disb_date(m['coll_date'])
                amt = _to_num(m['amount'])
                if d and amt:
                    events.append({'disb_id': did, 'customer': name, 'date': d, 'amount': amt,
                                    'bank_account': (m.get('bank_account', '') or '').strip()})
        else:
            cd = r.get('Collected   Date', '') or r.get('Collected Date', '')
            d = parse_disb_date(cd)
            amt = _to_num(r.get('Collected Amount', 0))
            if d and amt:
                events.append({'disb_id': did, 'customer': name, 'date': d, 'amount': amt,
                                'bank_account': (r.get('Bank Account', '') or '').strip()})
    return events

def read_contacts():
    """Read staff directory from Contact sheet."""
    try:
        sh = get_gspread_client().open_by_key(SPREADSHEET_ID)
        ws = sh.worksheet('Contact')
        rows = ws.get_all_values()
        if not rows:
            return []
        result = []
        current_cluster = ''
        for row in rows[1:]:
            if not any(r.strip() for r in row):
                continue
            cluster = row[0].strip() if len(row) > 0 and row[0].strip() else current_cluster
            if row[0].strip():
                current_cluster = cluster
            record = {
                'cluster':     cluster,
                'name':        row[1].strip() if len(row) > 1 else '',
                'designation': row[2].strip() if len(row) > 2 else '',
                'branch':      row[3].strip() if len(row) > 3 else '',
                'phone':       row[4].strip().rstrip('.') if len(row) > 4 else '',
                'email':       row[5].strip() if len(row) > 5 else '',
            }
            if record['name']:
                result.append(record)
        return result
    except Exception as e:
        raise RuntimeError(f'Contact sheet error: {e}')

def save_contacts(contacts):
    """Write full contacts list back to Contact sheet.

    Sheets has no transactions, so clear()-then-update() has a window where a
    failed update leaves the tab wiped with no rollback. Writing the new data
    first (a plain update() is retriable/idempotent on its own) and only
    clearing now-unused trailing rows afterward means a failure before the
    write completes leaves the old data intact instead of blank."""
    sh = get_gspread_client().open_by_key(SPREADSHEET_ID)
    ws = sh.worksheet('Contact')
    prev_row_count = len(ws.get_all_values())
    rows = [['Cluster', 'Name', 'Designation', 'Branch', 'Phone', 'Email']]
    prev_cluster = ''
    for c in contacts:
        cluster = c.get('cluster', '').strip()
        rows.append([
            cluster if cluster != prev_cluster else '',
            c.get('name', '').strip(),
            c.get('designation', '').strip(),
            c.get('branch', '').strip(),
            c.get('phone', '').strip(),
            c.get('email', '').strip(),
        ])
        if cluster:
            prev_cluster = cluster
    ws.update('A1', rows)
    if prev_row_count > len(rows):
        ws.batch_clear([f'A{len(rows) + 1}:Z{prev_row_count}'])
    ws.format('A1:F1', {
        'textFormat': {'bold': True, 'foregroundColor': {'red':1,'green':1,'blue':1}},
        'backgroundColor': {'red': 0.1, 'green': 0.23, 'blue': 0.36}
    })

def _parse_case(r):
    amount    = _to_num(r.get('Amount', 0))
    total     = _to_num(r.get('Total', 0)) or calc_total(amount)
    collected = _to_num(r.get('Collected Amount', 0))
    balance   = max(0, total - collected)
    return amount, total, collected, balance

def get_open_cases():
    records = read_accounts_from_gsheet()
    cases = []
    for r in records:
        status = r.get('Overdue Status', '').strip()
        if status and status != 'Closed':
            amount, total, collected, balance = _parse_case(r)
            cases.append({
                'disb_id':   r.get('Disbursement ID', ''),
                'customer':  r.get('Customer Name', ''),
                'amount':    amount,
                'total':     total,
                'collected': collected,
                'balance':   balance,
                'status':    status,
            })
    return cases

def lookup_case(disb_id):
    records = read_accounts_from_gsheet()
    for r in records:
        if r.get('Disbursement ID', '').upper() == disb_id.upper():
            amount, total, collected, balance = _parse_case(r)
            return {
                'found':     True,
                'row':       r['_row'],
                'sheet':     r.get('_sheet', SHEET_NAME),
                'customer':  r.get('Customer Name', ''),
                'amount':    amount,
                'total':     total,
                'collected': collected,
                'balance':   balance,
                'status':    r.get('Overdue Status', '').strip(),
            }
    return {'found': False}

def get_next_seq(records=None):
    if records is None:
        records = read_accounts_from_gsheet()
    max_seq = 0
    for r in records:
        m = re.search(r'-(\d{3})$', r.get('Disbursement ID', ''))
        if m:
            max_seq = max(max_seq, int(m.group(1)))
    return max_seq + 1

def save_disbursement(data):
    # Server-side backstop — the widget's own form already requires this,
    # but a request hitting this route directly (API call, retry script)
    # must be blocked too, or a disbursement with no proof of transfer can
    # still land in the books.
    if not (data.get('utr') or '').strip():
        raise ValueError('UTR / Reference number is required — a disbursement cannot be recorded without proof of transfer.')
    ws  = get_sheet()
    _ensure_accounts_bank_account_column(ws)
    _ensure_accounts_charge_plan_column(ws)
    _ensure_accounts_kyc_folder_column(ws)
    _ensure_accounts_request_id_column(ws)
    # One read serves both the sequence number AND the insert row — the
    # latter used to cost a second, separate ws.get_all_values() call on
    # top of the read read_accounts_from_gsheet() already just did.
    records  = read_accounts_from_gsheet()
    seq      = get_next_seq(records)
    accounts_rows = [r['_row'] for r in records if r.get('_sheet', SHEET_NAME) == SHEET_NAME]
    next_row = max(accounts_rows, default=3) + 1
    try:
        d = datetime.strptime(data['date'], '%d-%m-%Y')
    except Exception:
        d = datetime.today()
    ddmmyy  = d.strftime('%d%m%y')
    disb_id = f"BLP-{ddmmyy}-{seq:03d}"

    amount = float(data['amount'])
    cluster_for_charges = data.get('cluster', '')
    charge_calc = calc_charges(amount, cluster=cluster_for_charges, disbursement_date=d.date(),
                               company=data.get('company', ''))
    charges = charge_calc['charges']
    gst     = charge_calc['gst']
    total   = charge_calc['total']
    charge_methodology = charge_calc['methodology']  # None, or a charge-plan id

    # Manual selection always wins; auto-resolve from the source Request's
    # export-time debit account only when the form field was left blank —
    # reliability (a Request-linked disbursement always gets SOME tag even
    # if forgotten) without overriding a deliberate manual choice.
    bank_account = data.get('bank_account', '').strip()
    request_id_for_lookup = data.get('request_id', '').strip()
    if not bank_account and request_id_for_lookup:
        try:
            bank_account = _lookup_request_bank_account(request_id_for_lookup)
        except Exception:
            pass

    # KYC Folder has no manual form field to prefer -- it's set purely by
    # whatever the field app's scan-upload flow stamped on the source
    # Request row (if any). A missing/unreadable value must never fail the
    # disbursement itself, same reasoning as the bank_account lookup above.
    kyc_folder = ''
    if request_id_for_lookup:
        try:
            kyc_folder = _lookup_request_kyc_folder(request_id_for_lookup)
        except Exception:
            pass

    row_data = [''] * 28
    row_data[COL['disb_id']-1]      = disb_id
    row_data[COL['date']-1]         = d.strftime('%d-%m-%Y')
    row_data[COL['customer']-1]     = data.get('customer', '')
    row_data[COL['chq']-1]          = data.get('chq', '')
    row_data[COL['company']-1]      = data.get('company', '')
    row_data[COL['cluster']-1]      = data.get('cluster', '')
    row_data[COL['branch']-1]       = data.get('branch', '')
    row_data[COL['amount']-1]       = amount
    row_data[COL['charges']-1]      = charges
    row_data[COL['gst']-1]          = gst
    row_data[COL['total']-1]        = total
    # Balance/TAT/Current Date (O/P/Q) are left blank here — written as
    # explicit formulas right after the insert below, not as plain values,
    # since the app's own balance figure is always computed fresh from
    # Total/Collected/Discount in Python (see _parse_case()/_case_detail())
    # and never reads column O itself.
    row_data[COL['status']-1]       = 'Follow Up!'
    row_data[COL['srv_branch']-1]   = data.get('serviced_branch', '')
    row_data[COL['srv_cluster']-1]  = data.get('serviced_cluster', '')
    row_data[COL['debit_note']-1]   = data.get('utr', '')
    # Only stamp when new-methodology actually applies — left blank for
    # every non-BIB-HDB-Karnataka row AND pre-cutover BIB HDB Karnataka
    # rows, so a blank cell always correctly means "flat legacy formula"
    # with no backfill ever needed. Stamped from the SAME methodology value
    # calc_charges() already resolved above, so the stamp and the number
    # charged can never disagree with each other.
    remarks_in = data.get('remarks', '')
    if charge_methodology:
        row_data[COL['charge_plan']-1] = charge_methodology
        note = f"Charge plan applied: {charge_calc.get('plan_label', charge_methodology)}"
        remarks_in = f"{remarks_in} | {note}" if remarks_in.strip() else note
    row_data[COL['remarks']-1]      = remarks_in
    row_data[COL['bank_account']-1] = bank_account
    row_data[COL['kyc_folder']-1]   = kyc_folder
    row_data[COL['request_id']-1]   = request_id_for_lookup

    ws.insert_row(row_data, next_row)
    # Columns O/P/Q (Balance, TAT, Current Date) need a live formula on
    # every row. Previously copied from the row directly above via a
    # copyPaste (PASTE_FORMULA, adjusting relative references like a manual
    # copy-down would) instead of writing plain values, specifically so a
    # legitimate formula's cell references would keep working. That copy-
    # from-above approach turned out to have a real blind spot: right after
    # monthEndRollover() empties Accounts down to just its header row, the
    # "row above" for that month's first new disbursement IS the header —
    # so PASTE_FORMULA copied the header's own text/blank cells instead of
    # a real formula (confirmed live, 03-08-2026: Balance came back as a
    # bare '=K{row}' self-reference with no Collected-Amount subtraction at
    # all, TAT as the literal string 'TAT', Current Date as a frozen
    # number). Writing the formulas out explicitly here removes the
    # dependency on whatever happens to be in the row above entirely, so it
    # can't break this way again. Reconstructed by cross-referencing the
    # archived "Jul 26" tab's frozen values against every input column:
    # Balance is negative-outstanding / positive-overpaid (blank Collected
    # Amount means nothing collected yet, so -(Total); otherwise Collected
    # minus Total), TAT is days elapsed (blank Collected Date means still
    # open, so TODAY()-Disbursement Date; otherwise Collected Date minus
    # Disbursement Date), Current Date is just TODAY().
    try:
        ws.batch_update([
            {'range': f'O{next_row}', 'values': [[f'=IF(M{next_row}="",-(K{next_row}),M{next_row}-K{next_row})']]},
            {'range': f'P{next_row}', 'values': [[f'=IF(L{next_row}="",TODAY()-B{next_row},L{next_row}-B{next_row})']]},
            {'range': f'Q{next_row}', 'values': [['=TODAY()']]},
        ], value_input_option='USER_ENTERED')
    except Exception as e:
        print(f'WARNING: could not write O/P/Q formulas to row {next_row}: {e}')
    trigger_ledger_rebuild()

    request_id = data.get('request_id', '').strip()
    if request_id:
        try:
            _mark_request_disbursed(request_id, disb_id)
        except Exception:
            pass  # don't fail the disbursement if Requests update errors
        _log_request_disb_link(request_id, disb_id, data.get('customer', ''),
                                d.strftime('%d-%m-%Y'))

    # Drop the initial (provisional) invoice+ledger into the same Drive
    # folder the field app already scanned KYC docs into, and rename that
    # folder from "<customer> - <request_id>" to "<customer> - <disb_id>"
    # now that a real disbursement ID exists. Only runs when this case is
    # actually linked to a field request with a KYC folder -- a manually
    # entered disbursement with no request_id has nothing to upload into.
    # Best-effort throughout: a Drive/PDF failure here must never undo or
    # block a disbursement that's already been recorded in the books.
    if kyc_folder:
        try:
            folder_id = _drive_folder_id_from_url(kyc_folder)
            if folder_id:
                case_for_invoice = {
                    'id': disb_id, 'date': d.date(), 'customer': data.get('customer', ''),
                    'company': data.get('company', ''), 'cluster': data.get('cluster', ''),
                    'branch': data.get('branch', ''), 'chq': data.get('chq', ''),
                    'debit_note': data.get('utr', ''), 'amount': amount, 'charges': charges,
                    'gst': gst, 'total': total, 'balance': total, 'status': 'Follow Up!',
                    'days_out': 0, 'charge_plan': 'NEW' if charge_methodology == 'NEW' else '',
                    'request_id': request_id_for_lookup, 'coll_amt': 0,
                }
                pdf_bytes = mis.generate_invoice_ledger(case_for_invoice, mcoll_entry=None, paid_in_full=False)
                drive = get_drive_client()
                _upsert_drive_file(drive, folder_id, 'Invoice (Initial).pdf', bytes(pdf_bytes), 'application/pdf')
                safe_customer = (data.get('customer', '') or 'Customer').replace('/', '-')
                _rename_drive_folder(drive, folder_id, f"{safe_customer} - {disb_id}")
        except Exception as e:
            notify_ops('save_disbursement_kyc_invoice', e)

    return disb_id

def _cell_num(ws, row, col):
    val = ws.cell(row, col).value or '0'
    try:
        return float(str(val).replace(',', '').replace('₹', '').strip())
    except Exception:
        return 0.0

def _plan_is_provisional(plan_id):
    """True when a plan's rate depends on the collection date, so the figure
    booked at disbursement is only an estimate until the case closes. Driven
    off the formula type, not a hardcoded id, so a new day-based counterparty
    finalizes at closing automatically."""
    plan = get_plan_by_id(plan_id)
    return bool(plan) and (plan.get('formula') or {}).get('type') == 'day_based_inclusive'

def _maybe_finalize_bib_hdb_charges(amount, cluster, disbursement_date_str,
                                     collection_date_str, current_charge_plan,
                                     was_closed, becomes_closed, new_coll, discount,
                                     company=None):
    """Shared by save_repayment() and _recompute_case_from_mcoll() — both
    reach the same 'a case just transitioned to Closed' moment via
    different code paths (main Repayment tab vs. the Edit Case tab's
    add/update/delete-payment actions). Charges for a new-methodology BIB
    HDB Karnataka case are only PROVISIONAL (best-case 0.4%) until the case
    actually closes; this finalizes them using the real collection date,
    but ONLY on the specific call that first transitions non-Closed ->
    Closed (never an earlier partial instalment, never repeated afterward).

    Since the escalating rate only ever increases with lateness, the
    recomputed total can only be >= the provisional one — so this can only
    ever REVEAL a shortfall, never an overpayment. Re-derives new_bal/
    new_status against the final total rather than blindly keeping
    'Closed', so a case that collected exactly the provisional estimate
    but closed a day later than assumed correctly stays open with the true
    outstanding balance, instead of being marked paid in full.

    Returns (final_charges, final_gst, final_total, new_bal, new_status),
    or all-None if no recompute applies here — callers keep whatever
    bal/status they already computed in that case.
    """
    if was_closed or not becomes_closed or not _plan_is_provisional(current_charge_plan):
        return None, None, None, None, None
    d_date = _to_plain_date(disbursement_date_str)
    c_date = _to_plain_date(collection_date_str)
    charge_calc = calc_charges(amount, cluster=cluster, disbursement_date=d_date,
                                collection_date=c_date, methodology=current_charge_plan,
                                company=company)
    final_charges = charge_calc['charges']
    final_gst     = charge_calc['gst']
    final_total   = charge_calc['total']
    new_bal    = max(0, final_total - new_coll - discount)
    new_status = 'Closed' if new_bal < 1 else 'Follow Up!'
    return final_charges, final_gst, final_total, new_bal, new_status

def _maybe_upload_final_invoice(disb_id, was_closed, becomes_closed):
    """Shared by save_repayment(), _recompute_case_from_mcoll(), and
    update_disbursement() -- every place a case can transition into
    'Closed' -- same was_closed/becomes_closed gating convention as
    _maybe_finalize_bib_hdb_charges() so this only fires on the specific
    call that first closes the case, never an earlier partial instalment
    or a later no-op re-save.

    MUST be called after the caller's own sheet write has already
    completed. Re-reads the case fresh via the exact same
    _load_for_invoice()/parse_cases() pipeline the manual "Generate
    Invoice" button itself uses, rather than threading a hand-built case
    dict through 3 call sites that each have different data already in
    scope -- fewer places for the final PDF's numbers to disagree with
    what a human clicking that same button would see (in particular, this
    means it automatically picks up whatever _maybe_finalize_bib_hdb_charges()
    just wrote, without duplicating that recompute here).

    Best-effort throughout: never raises, since a Drive/PDF failure here
    must never affect the real status/balance write that already
    happened."""
    if was_closed or not becomes_closed:
        return
    try:
        sh = get_gspread_client().open_by_key(SPREADSHEET_ID)
        rows, mcoll = _load_for_invoice(sh, disb_id)
        _, _, all_cases_full = mis.parse_cases(rows, mcoll)
        case = next((c for c in all_cases_full if c['id'].upper() == disb_id.upper()), None)
        if not case or not case.get('kyc_folder'):
            return  # not a field-request-linked case, or no KYC folder to file into
        folder_id = _drive_folder_id_from_url(case['kyc_folder'])
        if not folder_id:
            return
        mcoll_entry = mcoll.get(disb_id.upper())
        pdf_bytes = mis.generate_invoice_ledger(case, mcoll_entry=mcoll_entry, paid_in_full=True)
        drive = get_drive_client()
        _upsert_drive_file(drive, folder_id, 'Invoice (Final).pdf', bytes(pdf_bytes), 'application/pdf')
    except Exception as e:
        notify_ops('final_invoice_upload', e)

def save_repayment(data):
    disb_id = data['disb_id'].strip().upper()
    info    = lookup_case(disb_id)
    if not info['found']:
        raise ValueError(f"Disbursement ID '{disb_id}' not found.")

    utr = (data.get('utr') or '').strip()
    # Server-side backstop, same reasoning as save_disbursement(): the
    # widget's form already requires this, but a direct call to this route
    # must be blocked too.
    if not utr:
        raise ValueError('UTR / Reference number is required — a repayment cannot be recorded without proof of payment.')

    try:
        coll_date = datetime.strptime(data['date'], '%d-%m-%Y').strftime('%d-%m-%Y')
    except Exception:
        coll_date = datetime.today().strftime('%d-%m-%Y')

    amount   = float(data['amount'])
    discount = float(data.get('discount', 0) or 0)
    raw_msg  = data.get('raw_msg', '')

    sh  = get_gspread_client().open_by_key(SPREADSHEET_ID)
    ws  = sh.worksheet(info.get('sheet', SHEET_NAME))
    row = info['row']
    _ensure_accounts_charge_plan_column(ws)

    # One row read instead of up to 4 separate single-cell round trips
    # (coll_amount, total, amount, credit_note, remarks each used to be
    # their own ws.cell() call).
    row_vals = ws.row_values(row)
    def _rownum(col_key):
        i = COL[col_key] - 1
        v = row_vals[i] if i < len(row_vals) else ''
        try:
            return float(str(v).replace(',', '').replace('₹', '').strip() or 0)
        except Exception:
            return 0.0
    def _rowstr(col_key):
        i = COL[col_key] - 1
        return (row_vals[i] if i < len(row_vals) else '') or ''

    existing = _rownum('coll_amount')
    total    = _rownum('total') or calc_total(_rownum('amount'))
    new_coll = existing + amount
    new_bal  = max(0, total - new_coll - discount)
    # Sub-rupee residue (rounding leftovers from discount/instalment math)
    # counts as fully settled, matching the >= 1.0 "open" threshold
    # generate_mis.py already uses for its own reports.
    new_status = 'Closed' if new_bal < 1 else info['status']

    final_charges, final_gst, final_total, recomputed_bal, recomputed_status = \
        _maybe_finalize_bib_hdb_charges(
            amount=_rownum('amount'), cluster=_rowstr('cluster'),
            disbursement_date_str=_rowstr('date'), collection_date_str=coll_date,
            current_charge_plan=_rowstr('charge_plan'), company=_rowstr('company'),
            was_closed=(info['status'] == 'Closed'), becomes_closed=(new_status == 'Closed'),
            new_coll=new_coll, discount=discount)
    if final_charges is not None:
        new_bal, new_status = recomputed_bal, recomputed_status

    updates = [
        (row, COL['coll_date'],   coll_date),
        (row, COL['coll_amount'], new_coll),
        (row, COL['status'],      new_status),
        # Balance (col O) is intentionally never written here — it's a live
        # sheet formula (=IF(...,-(Total),Collected-Total)) that recomputes
        # automatically once Collected Amount changes above. Writing a
        # static number would permanently destroy that row's formula. The
        # app's own displayed balance is always computed fresh in Python
        # from Total/Collected/Discount (_parse_case()/_case_detail()), so
        # nothing here depends on what column O contains.
    ]
    if final_charges is not None:
        updates += [(row, COL['charges'], final_charges),
                    (row, COL['gst'],     final_gst),
                    (row, COL['total'],   final_total)]
    if discount:
        updates.append((row, COL['discount'], discount))

    if utr:
        existing_utr = _rowstr('credit_note').strip()
        combined_utr = f"{existing_utr}, {utr}" if existing_utr else utr
        updates.append((row, COL['credit_note'], combined_utr))

    remarks = data.get('remarks', '').strip()
    if remarks:
        existing_rem = _rowstr('remarks').strip()
        combined_rem = f"{existing_rem} | {remarks}" if existing_rem else remarks
        updates.append((row, COL['remarks'], combined_rem))

    ws.batch_update([{
        'range': gspread.utils.rowcol_to_a1(r, c),
        'values': [[v]]
    } for r, c, v in updates])

    # Record every repayment in M Coll, including single one-shot full
    # closures -- M Coll needs to be the one definitive payment history
    # everywhere it's read (the Customer Ledger, MIS reports, etc.), not
    # just for multi-instalment cases.
    mcoll_error = None
    try:
        mc = sh.worksheet('M Coll')
        mc_vals = mc.get_all_values()
        next_mc_row = len(mc_vals) + 1
        mc.insert_row([
            disb_id, coll_date, amount,
            utr or raw_msg, info.get('customer', ''),
            data.get('bank_account', '').strip(),
        ], next_mc_row)
    except Exception as e:
        # Previously a bare `except: pass` — Accounts had already been
        # updated above, so a failure here silently detached this payment
        # from M Coll (the definitive payment-history log used by the
        # Customer Ledger and MIS reports), undetected until reconciliation
        # didn't match. Surface it two ways: durably, in the Accounts row's
        # own Remarks cell (so anyone auditing this case later sees it even
        # if nobody was watching the API response at the time), and in the
        # API response so the officer sees a warning immediately.
        mcoll_error = str(e)
        try:
            existing_rem = (ws.cell(row, COL['remarks']).value or '').strip()
            warn = f"M Coll log FAILED for {coll_date} payment of Rs.{amount}: {mcoll_error}"
            combined = f"{existing_rem} | {warn}" if existing_rem else warn
            ws.update_cell(row, COL['remarks'], combined)
        except Exception:
            pass  # best-effort — the warning-write itself must not blow up the request

    _maybe_upload_final_invoice(disb_id, was_closed=(info['status'] == 'Closed'),
                                 becomes_closed=(new_status == 'Closed'))

    trigger_ledger_rebuild()
    result = {'new_collected': new_coll, 'new_balance': new_bal, 'status': new_status}
    if mcoll_error:
        result['mcoll_warning'] = (
            f"Payment recorded on the account but NOT logged to M Coll (payment "
            f"history): {mcoll_error}. It's noted in Remarks — check manually."
        )
    return result

# ── Case editing (fix mistakes in saved disbursements / repayments) ───────────

def _parse_any_date(s):
    for f in ('%d-%m-%Y', '%d/%m/%Y', '%d-%b-%Y', '%Y-%m-%d', '%d-%b-%y'):
        try:
            return datetime.strptime(str(s).strip(), f)
        except Exception:
            pass
    return None

_recent_activity_cache = {}  # {days: (timestamp, events)}
_RECENT_ACTIVITY_TTL = 25  # seconds

def get_recent_activity(days=7):
    """Cached wrapper: this view fires automatically every time the Edit
    Case tab opens and costs 5-6 Sheets reads (Accounts + every active
    archive tab + M Coll). Left uncached, repeated tab opens compete for
    the same per-minute Sheets quota as the actual Load Case / Save
    requests, making THOSE more likely to hit 429 and hang. A warm
    serverless instance reuses this module-level cache across requests
    within the TTL window; a cold start just recomputes once."""
    cached = _recent_activity_cache.get(days)
    if cached and (time.time() - cached[0]) < _RECENT_ACTIVITY_TTL:
        return cached[1]
    events = _get_recent_activity_uncached(days)
    _recent_activity_cache[days] = (time.time(), events)
    return events

def _get_recent_activity_uncached(days=7):
    """Every individual disbursement AND every individual repayment recorded
    in the last N days — a flat, unfiltered event log rather than one
    deduped row per case, so a case with several separate repayments today
    (e.g. Gunashekar with 2 repayments) shows each one instead of collapsing
    them into a single 'last activity' summary.

    Disbursements come from read_accounts_from_gsheet() (merges Accounts +
    every active monthly archive tab). Repayments come straight from M Coll,
    the definitive payment-event log — cluster/branch are looked up from the
    matching Accounts/archive record since M Coll doesn't carry them."""
    cutoff = datetime.today() - timedelta(days=days)
    records = read_accounts_from_gsheet()
    by_disb_id = {r.get('Disbursement ID', ''): r for r in records}

    events = []
    for r in records:
        d = _parse_any_date(r.get('Disbursement Date', ''))
        if not d or d < cutoff:
            continue
        events.append({
            'type': 'Disbursement',
            'disb_id':  r.get('Disbursement ID', ''),
            'customer': r.get('Customer Name', ''),
            'cluster':  r.get('Cluster', ''),
            'branch':   r.get('Branch', ''),
            'amount':   _to_num(r.get('Amount', 0)),
            'date':     r.get('Disbursement Date', ''),
            'utr':      r.get('Debit Note', ''),
            'status':   r.get('Overdue Status', '').strip(),
            'sheet':    r.get('_sheet', SHEET_NAME),
            '_sortkey': d,
        })

    try:
        mc_rows = get_gspread_client().open_by_key(SPREADSHEET_ID).worksheet('M Coll').get_all_values()[1:]
    except Exception:
        mc_rows = []
    for row in mc_rows:
        if not row or not row[0]:
            continue
        d = _parse_any_date(row[1] if len(row) > 1 else '')
        if not d or d < cutoff:
            continue
        disb_id = row[0].strip()
        case = by_disb_id.get(disb_id, {})
        events.append({
            'type': 'Repayment',
            'disb_id':  disb_id,
            'customer': row[4] if len(row) > 4 else case.get('Customer Name', ''),
            'cluster':  case.get('Cluster', ''),
            'branch':   case.get('Branch', ''),
            'amount':   _to_num(row[2]) if len(row) > 2 else 0,
            'date':     row[1] if len(row) > 1 else '',
            'utr':      row[3] if len(row) > 3 else '',
            'status':   case.get('Overdue Status', '').strip(),
            'sheet':    case.get('_sheet', SHEET_NAME),
            '_sortkey': d,
        })

    events.sort(key=lambda e: e['_sortkey'], reverse=True)
    for e in events:
        del e['_sortkey']
    return events

def _get_case_payments(sh, disb_id):
    """M Coll rows for one case, each carrying its live row number so edits/
    deletes can target it. Pulled into its own helper so callers that
    already read M Coll for another reason (recomputing collected/balance)
    can build this list from data they already have in hand, instead of
    _case_detail() re-reading the whole M Coll sheet from scratch."""
    payments = []
    try:
        mc = sh.worksheet('M Coll')
        for i, r in enumerate(mc.get_all_values(), start=1):
            if r and str(r[0]).strip().upper() == disb_id:
                payments.append({
                    'mc_row': i,
                    'date':   r[1] if len(r) > 1 else '',
                    'amount': _to_num(r[2]) if len(r) > 2 else 0,
                    'utr':    r[3] if len(r) > 3 else '',
                })
    except Exception:
        pass
    return payments

def _case_detail(disb_id, known=None):
    """Everything the Edit Case screen needs: the Accounts row's editable +
    computed fields, and the case's individual payments from M Coll.

    `known`, if given, is {'sheet', 'row', 'payments'} already resolved by
    the caller's own write (update_disbursement / _recompute_case_from_mcoll)
    — this skips a second full lookup_case() (which re-scans Accounts plus
    every archive tab) and, when payments are already known, a second full
    M Coll read. Without it, both are looked up fresh here as before."""
    disb_id = disb_id.strip().upper()
    sh = get_gspread_client().open_by_key(SPREADSHEET_ID)

    if known:
        sheet, row, payments = known['sheet'], known['row'], known.get('payments')
    else:
        info = lookup_case(disb_id)
        if not info['found']:
            return None
        sheet, row, payments = info['sheet'], info['row'], None

    ws = sh.worksheet(sheet)
    vals = ws.row_values(row)
    def cell(key):
        i = COL[key] - 1
        return vals[i] if i < len(vals) else ''

    if payments is None:
        payments = _get_case_payments(sh, disb_id)

    total     = _to_num(cell('total')) or calc_total(_to_num(cell('amount')))
    collected = _to_num(cell('coll_amount'))
    discount  = _to_num(cell('discount'))
    # Balance is computed live from Total/Collected/Discount, same as
    # _parse_case() everywhere else in the app — the sheet's stored Balance
    # column (col O) is a stale snapshot written once at creation and never
    # updated by save_repayment()/save_disbursement(), so it must not be
    # read directly here.
    balance = max(0, total - collected - discount)

    return {
        'disb_id': disb_id, 'sheet': sheet,
        'date': cell('date'), 'customer': cell('customer'), 'chq': cell('chq'),
        'company': cell('company'), 'cluster': cell('cluster'), 'branch': cell('branch'),
        'amount': _to_num(cell('amount')), 'charges': _to_num(cell('charges')),
        'gst': _to_num(cell('gst')), 'total': total,
        'discount': discount, 'collected': collected,
        'balance': balance, 'status': cell('status'),
        'srv_branch': cell('srv_branch'), 'srv_cluster': cell('srv_cluster'),
        'debit_note': cell('debit_note'), 'remarks': cell('remarks'),
        'kyc_folder': cell('kyc_folder'),
        'request_id': cell('request_id'),
        'payments': payments,
    }

EDITABLE_TEXT_FIELDS = ('date', 'customer', 'chq', 'company', 'cluster', 'branch',
                        'srv_branch', 'srv_cluster', 'debit_note', 'remarks')

def update_disbursement(data):
    """Writes corrected disbursement fields back to the case's Accounts row.
    An amount change recomputes the derived money chain via calc_charges()
    (flat 0.5%/18% for every cluster except BIB HDB Karnataka's day-based
    schedule) and re-derives status/balance against existing collections."""
    disb_id = data['disb_id'].strip().upper()
    info = lookup_case(disb_id)
    if not info['found']:
        raise ValueError(f"Disbursement ID '{disb_id}' not found.")
    sh  = get_gspread_client().open_by_key(SPREADSHEET_ID)
    ws  = sh.worksheet(info.get('sheet', SHEET_NAME))
    row = info['row']
    _ensure_accounts_charge_plan_column(ws)

    updates = [(row, COL[k], str(data[k]).strip())
               for k in EDITABLE_TEXT_FIELDS if k in data]

    cluster_changing = 'cluster' in data
    amount_changing  = str(data.get('amount', '')).strip() != ''

    # One row read (only when actually needed) instead of separate single-
    # cell round trips — same efficiency rationale save_repayment() already
    # documents. Needed whenever we must know this row's CURRENT cluster/
    # date/charge_plan marker (i.e. whichever of those aren't being
    # overwritten by this exact call).
    row_vals = None
    def _current(col_key):
        nonlocal row_vals
        if row_vals is None:
            row_vals = ws.row_values(row)
        i = COL[col_key] - 1
        return row_vals[i] if i < len(row_vals) else ''

    date_changing    = 'date' in data
    company_changing = 'company' in data
    # Anything that DETERMINES the charge forces a recompute, not just the
    # amount. Company, cluster and date all feed resolve_charge_plan(), so a
    # cluster-only correction used to re-stamp the marker and leave the money
    # untouched — the row then claimed one charge plan while carrying another
    # plan's figures, with nothing anywhere to notice. (Real case: the
    # 12-08-2026 BIB HDB disbursement of Vijaya K.)
    plan_inputs_changing = cluster_changing or date_changing or company_changing
    recompute = plan_inputs_changing or amount_changing

    if recompute:
        effective_cluster = str(data['cluster']).strip() if cluster_changing else _current('cluster')
        effective_company = str(data['company']).strip() if company_changing else _current('company')
        effective_date = _to_plain_date(str(data['date']).strip() if date_changing else _current('date'))

    # A mis-entered company/cluster/date fixed after the fact shouldn't
    # permanently miss (or wrongly keep stamped) a counterparty's schedule
    # just because the marker was only ever decided once, at disbursement.
    new_marker = None
    if plan_inputs_changing:
        new_marker = resolve_charge_plan(effective_company, effective_cluster, effective_date)
        # Blank means "flat default" — same convention save_disbursement() uses.
        updates.append((row, COL['charge_plan'], new_marker or ''))

    if recompute:
        amount = float(data['amount']) if amount_changing else _cell_num(ws, row, COL['amount'])
        # Prefer a plan just (re-)decided above if its inputs changed in this
        # same call; otherwise this row's already-stamped marker is
        # authoritative and untouched by this specific edit.
        methodology = new_marker if plan_inputs_changing else (_current('charge_plan') or None)
        charge_calc = calc_charges(amount, cluster=effective_cluster,
                                    disbursement_date=effective_date,
                                    methodology=methodology,
                                    company=effective_company)
        charges = charge_calc['charges']; gst = charge_calc['gst']; total = charge_calc['total']
        collected = _cell_num(ws, row, COL['coll_amount'])
        discount  = _cell_num(ws, row, COL['discount'])
        bal       = max(0, total - collected - discount)
        # Same sub-₹1 closed threshold as save_repayment; a previously Closed
        # case whose corrected amount leaves money owed reopens.
        status = 'Closed' if bal < 1 else (
            'Follow Up!' if info['status'] in ('', 'Closed') else info['status'])
        # Balance (col O) is never written — see save_repayment()'s comment;
        # it's a live formula that recomputes from Total/Collected on its own.
        updates += [(row, COL['amount'],  amount),
                    (row, COL['charges'], charges),
                    (row, COL['gst'],     gst),
                    (row, COL['total'],   total),
                    (row, COL['status'],  status)]

    if updates:
        ws.batch_update([{
            'range': gspread.utils.rowcol_to_a1(r, c),
            'values': [[v]]
        } for r, c, v in updates])
        trigger_ledger_rebuild()

    # Only the recompute branch above can transition status here (a
    # correction that leaves bal < 1) -- `status` only exists in that
    # branch's scope, so gate on the same flag rather than checking
    # `'status' in locals()`.
    if recompute:
        _maybe_upload_final_invoice(disb_id, was_closed=(info['status'] == 'Closed'),
                                     becomes_closed=(status == 'Closed'))

    # Handed back to the route so it can build the response via _case_detail()
    # without a second full lookup_case() (Accounts + every archive tab scan).
    return {'sheet': info['sheet'], 'row': row}

def _recompute_case_from_mcoll(disb_id):
    """After an M Coll payment row is added, edited, or deleted, re-derive
    the Accounts row's aggregates from the full M Coll history — M Coll is
    the one definitive payment record. Returns the resolved sheet/row plus
    the case's payments (built from the M Coll data already read here) so
    the caller's response can skip a second lookup_case() + M Coll read."""
    disb_id = disb_id.strip().upper()
    info = lookup_case(disb_id)
    if not info['found']:
        raise ValueError(f"Disbursement ID '{disb_id}' not found.")
    sh  = get_gspread_client().open_by_key(SPREADSHEET_ID)
    ws  = sh.worksheet(info.get('sheet', SHEET_NAME))
    row = info['row']
    _ensure_accounts_charge_plan_column(ws)

    mc_all = sh.worksheet('M Coll').get_all_values()
    payments = [{
        'mc_row': i, 'date': r[1] if len(r) > 1 else '',
        'amount': _to_num(r[2]) if len(r) > 2 else 0, 'utr': r[3] if len(r) > 3 else '',
    } for i, r in enumerate(mc_all, start=1) if r and str(r[0]).strip().upper() == disb_id]
    pays = [r for r in mc_all if r and str(r[0]).strip().upper() == disb_id]
    collected = sum(_to_num(r[2]) if len(r) > 2 else 0 for r in pays)

    def _pd(dstr):
        for f in ('%d-%m-%Y', '%d/%m/%Y', '%d-%b-%Y', '%Y-%m-%d'):
            try:
                return datetime.strptime(str(dstr).strip(), f)
            except Exception:
                pass
        return None
    dated = [(d, r[1]) for r in pays if len(r) > 1 for d in [_pd(r[1])] if d]
    coll_date = max(dated)[1] if dated else (pays[-1][1] if pays and len(pays[-1]) > 1 else '')

    # One row read — same efficiency rationale as save_repayment() — covers
    # amount/total/discount/cluster/date/charge_plan without extra round trips.
    row_vals = ws.row_values(row)
    def _rownum(col_key):
        i = COL[col_key] - 1
        v = row_vals[i] if i < len(row_vals) else ''
        try:
            return float(str(v).replace(',', '').replace('₹', '').strip() or 0)
        except Exception:
            return 0.0
    def _rowstr(col_key):
        i = COL[col_key] - 1
        return (row_vals[i] if i < len(row_vals) else '') or ''

    total    = _rownum('total') or calc_total(_rownum('amount'))
    discount = _rownum('discount')
    bal      = max(0, total - collected - discount)
    status   = 'Closed' if bal < 1 else 'Follow Up!'
    was_closed = _rowstr('status').strip() == 'Closed'

    final_charges, final_gst, final_total, recomputed_bal, recomputed_status = \
        _maybe_finalize_bib_hdb_charges(
            amount=_rownum('amount'), cluster=_rowstr('cluster'),
            disbursement_date_str=_rowstr('date'), collection_date_str=coll_date,
            current_charge_plan=_rowstr('charge_plan'), company=_rowstr('company'),
            was_closed=was_closed, becomes_closed=(status == 'Closed'),
            new_coll=collected, discount=discount)
    if final_charges is not None:
        bal, status = recomputed_bal, recomputed_status

    # Balance (col O) is never written — same reasoning as save_repayment()
    # and update_disbursement(): it's a live formula that recomputes on its
    # own once Collected Amount (written below) changes.
    batch = [(COL['coll_date'], coll_date), (COL['coll_amount'], collected),
             (COL['status'], status)]
    if final_charges is not None:
        batch += [(COL['charges'], final_charges), (COL['gst'], final_gst),
                  (COL['total'], final_total)]
    ws.batch_update([{
        'range': gspread.utils.rowcol_to_a1(row, c),
        'values': [[v]]
    } for c, v in batch])
    trigger_ledger_rebuild()

    _maybe_upload_final_invoice(disb_id, was_closed=was_closed, becomes_closed=(status == 'Closed'))

    return {'sheet': info['sheet'], 'row': row, 'payments': payments}

def _mcoll_row_checked(sh, disb_id, mc_row):
    """Guard against stale row numbers: the M Coll row being targeted must
    still belong to this case. Returns (worksheet, current_row_values) so
    callers that need the row's existing values don't re-fetch them."""
    mc = sh.worksheet('M Coll')
    current = mc.row_values(mc_row)
    if not current or str(current[0]).strip().upper() != disb_id.strip().upper():
        raise ValueError('That payment row has moved — reload the case and try again.')
    return mc, current

def update_repayment(data):
    disb_id = data['disb_id'].strip().upper()
    mc_row  = int(data['mc_row'])
    sh = get_gspread_client().open_by_key(SPREADSHEET_ID)
    mc, current = _mcoll_row_checked(sh, disb_id, mc_row)

    # Known gap: _recompute_case_from_mcoll()'s BIB HDB Karnataka finalize
    # hook only fires on a non-Closed -> Closed transition. Correcting the
    # DATE of a payment that already closed the case stays Closed -> Closed
    # throughout, so it's silently skipped — if that date correction would
    # have shifted the days-late bucket, the already-finalized charge could
    # go stale. Surfacing a warning (rather than silently doing nothing, or
    # auto-recomputing against the "only on transition" default) matches
    # this codebase's existing mcoll_warning precedent in save_repayment().
    new_date = str(data.get('date', '')).strip()
    warn = None
    try:
        info = lookup_case(disb_id)
        if info['found'] and info['status'] == 'Closed':
            ws_check = sh.worksheet(info.get('sheet', SHEET_NAME))
            _ensure_accounts_charge_plan_column(ws_check)
            if ws_check.cell(info['row'], COL['charge_plan']).value == 'NEW':
                old_date = current[1] if len(current) > 1 else ''
                if new_date and new_date != old_date:
                    warn = ('This case is Closed under the BIB HDB Karnataka day-based '
                            'schedule — correcting this payment\'s date does not '
                            'automatically re-finalize Charges/GST/Total. Recheck them '
                            'manually if the days-late bucket may have changed.')
    except Exception:
        pass  # best-effort warning only — must never block the actual correction

    mc.batch_update([
        {'range': gspread.utils.rowcol_to_a1(mc_row, 2), 'values': [[new_date]]},
        {'range': gspread.utils.rowcol_to_a1(mc_row, 3), 'values': [[float(data['amount'])]]},
        {'range': gspread.utils.rowcol_to_a1(mc_row, 4), 'values': [[str(data.get('utr', '')).strip()]]},
    ])
    result = _recompute_case_from_mcoll(disb_id)
    if warn:
        result['charge_plan_warning'] = warn
    return result

def delete_repayment(data):
    disb_id = data['disb_id'].strip().upper()
    mc_row  = int(data['mc_row'])
    sh = get_gspread_client().open_by_key(SPREADSHEET_ID)
    mc, _ = _mcoll_row_checked(sh, disb_id, mc_row)
    mc.delete_rows(mc_row)
    return _recompute_case_from_mcoll(disb_id)

def add_repayment(data):
    """Append a new M Coll payment row from the Edit Case tab — the same
    write save_repayment() does, minus the WhatsApp-extraction fields that
    only apply to the standalone Repayment tab's paste-a-message flow."""
    disb_id = data['disb_id'].strip().upper()
    info = lookup_case(disb_id)
    if not info['found']:
        raise ValueError(f"Disbursement ID '{disb_id}' not found.")
    sh = get_gspread_client().open_by_key(SPREADSHEET_ID)
    mc = sh.worksheet('M Coll')
    next_row = len(mc.get_all_values()) + 1
    mc.insert_row([
        disb_id,
        str(data.get('date', '')).strip(),
        float(data['amount']),
        str(data.get('utr', '')).strip(),
        info.get('customer', ''),
    ], next_row)
    return _recompute_case_from_mcoll(disb_id)

# ── Config (stored in a 'Config' tab of the same spreadsheet) ─────────────────
# No persistent local disk exists on a serverless host, so config lives in the
# Sheet (key/value rows, one per top-level key; dict/list values JSON-encoded
# into the cell) instead of a local bridgeline_config.json file.

CONFIG_SHEET_NAME = "Config"

DEFAULT_CONFIG = {
    "whatsapp_groups": {c: "" for c in CLUSTERS},
    "report_time": "09:00",
    "overdue_threshold_days": 7,
    "custom_types": [],
    "ledger_webhook_url": "",
    "ledger_webhook_token": "",
    # Registered bank accounts used for reconciliation, e.g.
    # [{"name": "HDFC xx0923", "account_number": "5010..."}, ...]. A
    # transfer narration containing another registered account's number is
    # an internal movement between our own accounts (Capital In/Out), not
    # a real disbursement/expense/collection.
    "bank_accounts": [],
    # Per-counterparty charge schedules. Empty here means "use the built-in
    # DEFAULT_CHARGE_PLANS"; set it to override or to onboard a new company
    # without a code change. See get_charge_plans()/resolve_charge_plan().
    "charge_plans": [],
    # Defaults for the Bank File tab's bulk-upload exports (overridable per
    # export in the UI).
    "bulk_debit_account": "",
    "bulk_narration": "BridgeLine Disbursement",
}

def _load_config_raw():
    """Actual Config sheet read, no fallback — raises on any failure. Retries
    a couple of times first since most failures here are a transient Sheets
    429/network blip, not a real outage."""
    last_err = None
    for attempt in range(3):
        try:
            ws = get_gspread_client().open_by_key(SPREADSHEET_ID).worksheet(CONFIG_SHEET_NAME)
            rows = ws.get_all_values()
            cfg = {}
            for row in rows[1:] if rows and rows[0] and rows[0][0].strip().lower() == 'key' else rows:
                if len(row) < 2 or not row[0].strip():
                    continue
                key, raw = row[0].strip(), row[1]
                try:
                    cfg[key] = json.loads(raw)
                except (json.JSONDecodeError, TypeError):
                    cfg[key] = raw  # plain strings (report_time, webhook url/token) stay as-is
            return cfg
        except Exception as e:
            last_err = e
            time.sleep(1.5 * (attempt + 1))
    raise last_err

def load_config():
    """Lenient reader for ordinary display/use throughout the app (dashboard,
    MIS generation, etc.) — a transient Sheets hiccup here should degrade to
    safe defaults rather than break the page. NEVER use this as the basis
    for a write-back (see load_config_strict + the save_config() docstring
    for why: this silently substituting DEFAULT_CONFIG on failure is exactly
    what turned one transient read error into permanently deleted
    ledger_webhook_url/token + active_archive_tabs + bank_accounts on
    2026-07-09 — every key not in that one request's payload and not in
    DEFAULT_CONFIG's 8 baked-in keys was gone the moment the merged,
    defaults-only dict got written back over the whole sheet).

    Cached briefly (15s) — this gets called multiple times within a single
    save (e.g. once via get_archive_tab_names() for the Accounts read),
    each of which used to cost its own fresh Sheets read. Cleared by
    trigger_ledger_rebuild()/save_config() same as _accounts_cache."""
    global _config_cache
    if _config_cache and (time.time() - _config_cache[0]) < _ACCOUNTS_CACHE_TTL:
        return _config_cache[1]
    try:
        cfg = _load_config_raw()
    except Exception:
        return DEFAULT_CONFIG.copy()
    for k, v in DEFAULT_CONFIG.items():
        cfg.setdefault(k, v)
    _config_cache = (time.time(), cfg)
    return cfg

def load_config_strict():
    """For any code path about to save_config() a merged result. Raises
    instead of silently returning DEFAULT_CONFIG on a failed read — a
    save built on top of a failed read must never proceed, since
    save_config() rewrites the ENTIRE Config sheet from exactly the dict
    it's given. Callers must let this exception abort the save (return an
    error to the user) rather than falling back to any default dict."""
    cfg = _load_config_raw()
    for k, v in DEFAULT_CONFIG.items():
        cfg.setdefault(k, v)
    return cfg

def save_config(cfg):
    # Write-then-clear-trailing (not clear-then-write) — see save_contacts()
    # for why: Sheets has no transactions, so this shrinks the failure
    # window to "the write itself fails" (old config intact) instead of
    # "clear succeeded, write failed" (config wiped, unrecoverable).
    #
    # Defence in depth for an audit-critical sheet: re-read whatever is
    # CURRENTLY on the sheet right before writing and merge `cfg` on top of
    # it, instead of writing `cfg` alone. A key already in Config that isn't
    # present in `cfg` is preserved rather than deleted — this is what
    # should have stopped the 2026-07-09 incident (ledger_webhook_url/token,
    # active_archive_tabs, bank_accounts silently wiped by a save built on
    # a failed read) even if some future caller ever builds an incomplete
    # dict again. The only way to actually remove a key now is to delete
    # its row directly in the sheet.
    global _config_cache
    _config_cache = None
    ws = get_gspread_client().open_by_key(SPREADSHEET_ID).worksheet(CONFIG_SHEET_NAME)
    current_rows = ws.get_all_values()
    prev_row_count = len(current_rows)
    merged = {}
    for row in current_rows[1:] if current_rows and current_rows[0] and current_rows[0][0].strip().lower() == 'key' else current_rows:
        if len(row) >= 2 and row[0].strip():
            merged[row[0].strip()] = row[1]
    for k, v in cfg.items():
        merged[k] = v

    rows = [['key', 'value']]
    for k, v in merged.items():
        val = json.dumps(v) if isinstance(v, (dict, list)) else str(v)
        rows.append([k, val])
    ws.update(range_name='A1', values=rows)
    if prev_row_count > len(rows):
        ws.batch_clear([f'A{len(rows) + 1}:Z{prev_row_count}'])

# ── Today summary ─────────────────────────────────────────────────────────────

def parse_disb_date(s):
    for fmt in ['%d-%b-%Y','%d-%m-%Y','%Y-%m-%d','%d/%m/%Y','%d-%b-%y','%-d-%b-%Y']:
        try:
            return datetime.strptime(str(s).strip(), fmt)
        except:
            pass
    return None

RECON_LOG_SHEET_NAME = "Recon Log"

def _parse_recon_date(s):
    if not s:
        return None
    d = parse_disb_date(s)
    if d:
        return d
    for fmt in ('%d/%m/%y', '%d/%m/%Y'):
        try:
            return datetime.strptime(str(s).strip(), fmt)
        except Exception:
            pass
    return None

def get_latest_bank_balance():
    """Read the most recent closing balance + date from the Recon Log sheet
    tab, per registered bank account, and combine them. Each /reconcile/save
    call appends one row here (tagged with an Account column) — this
    replaces reading back a persistent Daily Reconciliation.xlsx from a
    local/Drive path, since the hosted version no longer keeps that file
    anywhere central (each reconciliation save is now a one-off in-browser
    download).

    With multiple pooled bank accounts, "the bank balance" is the SUM of
    each account's own latest closing balance (last appended row for that
    account). The "since this date" cutoff used elsewhere to compute
    disbursed/collected-since-reconciliation uses the EARLIEST of the
    accounts' latest-reconciled dates — conservative, so we never miss
    counting a transaction that happened after one account's reconciliation
    but before another's.
    """
    try:
        ws = get_gspread_client().open_by_key(SPREADSHEET_ID).worksheet(RECON_LOG_SHEET_NAME)
        rows = ws.get_all_values()[1:]  # skip header
        if not rows:
            return None, None
        # Only count accounts currently registered in Config — a blank/
        # unrecognised Account tag is orphaned data (e.g. rows saved before
        # multi-account support existed, or a since-removed account) and
        # must not be summed in as if it were still a live account.
        registered = {a.get('name', '').strip() for a in (load_config().get('bank_accounts') or [])}
        latest_by_account = {}
        for row in rows:
            acct = row[4].strip() if len(row) > 4 else ''
            if registered and acct not in registered:
                continue
            latest_by_account[acct] = row  # last write wins -> most recent append for that account
        closings, dates = [], []
        for row in latest_by_account.values():
            if len(row) > 2 and row[2]:
                try:
                    closings.append(float(str(row[2]).replace(',', '').replace('₹', '').strip()))
                except ValueError:
                    pass
            d = _parse_recon_date(row[0] if row else None)
            if d:
                dates.append(d)
        if not closings:
            return None, None
        total_closing = sum(closings)
        earliest_date = min(dates) if dates else None
        return total_closing, earliest_date
    except Exception:
        return None, None

def _last_recon_closing_for_account(account):
    """Single-account counterpart to get_latest_bank_balance()'s sum-across-
    accounts figure — the starting point for that ONE account's book-balance
    tie-out. Reads the same Recon Log rows independently rather than
    refactoring get_latest_bank_balance() itself, so the existing dashboard
    balance figure can't regress from this change."""
    try:
        ws = get_gspread_client().open_by_key(SPREADSHEET_ID).worksheet(RECON_LOG_SHEET_NAME)
        rows = ws.get_all_values()[1:]
        last_row = None
        for row in rows:
            acct = row[4].strip() if len(row) > 4 else ''
            if acct == account:
                last_row = row  # last write wins -> most recent append for this account
        if not last_row:
            return None, None
        closing = None
        if len(last_row) > 2 and last_row[2]:
            try:
                closing = float(str(last_row[2]).replace(',', '').replace('₹', '').strip())
            except ValueError:
                closing = None
        d = _parse_recon_date(last_row[0] if last_row else None)
        return closing, d
    except Exception:
        return None, None

# FD balance is deliberately NOT derived from summing "FD Booking"-classified
# sweep-in/sweep-out transactions across reconciliation history (the old
# approach) — that sum has no opening baseline (implicitly assumes the FD
# balance was exactly ₹0 the day reconciliation tracking began) and is
# silently corrupted by any gap in day-by-day reconciliation coverage: a
# sweep-out captured on a reconciled day whose later matching sweep-in falls
# on an unreconciled day never gets offset, permanently inflating the
# figure. Confirmed live, 06-08-2026: applying a proper flow-balance check
# (Prem's own model) against real classified IDFC data showed a ~₹39.5L gap
# between what the transaction history implied and the real observed
# closing balance, traced to 8 unreconciled days within the supposedly-
# covered window. Available to Disburse was silently inflated by that same
# broken derivation.
#
# Fixed the same way bank balance already works: trust a directly-entered,
# real observed number (from an FD statement / netbanking FD summary) at
# each check-in, exactly like Opening/Closing Balance already are for the
# bank accounts, instead of reconstructing it from transaction arithmetic
# that a reconciliation gap can silently corrupt. Reuses the Recon Log
# sheet/mechanism verbatim via a pseudo-account label — no schema change,
# and get_latest_bank_balance()'s own registered-accounts filter already
# excludes anything not in Config's bank_accounts list, so this can never
# accidentally get summed into the real bank_balance figure.
FD_ACCOUNT_LABEL = 'FD (Total)'

def get_latest_fd_balance():
    """(balance, date) from the last directly-entered FD balance snapshot,
    or (None, None) if none has ever been entered yet."""
    return _last_recon_closing_for_account(FD_ACCOUNT_LABEL)

def save_fd_balance(date_str, balance):
    d = _parse_recon_date(date_str)
    if not d:
        raise ValueError(f"Could not parse date '{date_str}'.")
    bal = float(balance)
    ws = get_gspread_client().open_by_key(SPREADSHEET_ID).worksheet(RECON_LOG_SHEET_NAME)
    ws.append_row([d.strftime('%d-%m-%Y'), '', bal,
                    datetime.today().strftime('%d-%m-%Y %H:%M'), FD_ACCOUNT_LABEL])
    return bal, d

def _account_tagged_activity(records, mc_rows, account, since_date, through_date):
    """Tagged disbursement/collection totals for ONE account, dated strictly
    AFTER since_date (the last reconciled closing's own date — already
    reflected in that closing balance, so re-including it double-counts)
    through through_date inclusive (this statement's own recon date) — NOT
    the whole calendar month. Plus a count of book entries in that same
    window (any account, including untagged) — the untagged count is what
    decides whether the tie-out can be trusted at all (see _account_tieout)."""
    disb_total = disb_count = 0.0
    coll_total = coll_count = 0.0
    untagged = total = 0
    for r in records:
        ddate = str(r.get('Disbursement Date', '') or '')
        if not _date_in_range(ddate, since_date, through_date):
            continue
        total += 1
        tag = (r.get('Bank Account', '') or '').strip()
        if not tag:
            untagged += 1
        elif tag == account:
            disb_total += _to_num(r.get('Amount', 0)); disb_count += 1
    for mc in mc_rows:
        did   = (mc[0] if len(mc) > 0 else '').strip()
        pdate = mc[1] if len(mc) > 1 else ''
        if not did or not _date_in_range(str(pdate), since_date, through_date):
            continue
        total += 1
        tag = (mc[5] if len(mc) > 5 else '').strip()
        if not tag:
            untagged += 1
        elif tag == account:
            coll_total += _clean_amount(mc[2]) if len(mc) > 2 else 0; coll_count += 1
    return {'disb_total': round(disb_total, 2), 'disb_count': disb_count,
            'coll_total': round(coll_total, 2), 'coll_count': coll_count,
            'untagged_count': untagged, 'total_count': total}

def _account_tieout(records, mc_rows, account, period_month, classified_txns, statement_closing, recon_date=None):
    """Is this bank statement's own closing balance what the (tagged) books
    say it should be for this one account? Compares against book activity
    strictly AFTER the last reconciled closing's own date, through this
    statement's own recon date — NOT the whole calendar month. The last
    closing already reflects everything up to and including its own date,
    so summing "this period" as the full month double-counts every day
    between the 1st and the last reconciliation, which is exactly what
    produced an impossible deeply-negative expected closing the one time
    this went unnoticed (07-08-2026). Degrades to 'incomplete' rather than
    computing a number whenever any book entry in that window — for ANY
    account, including untagged ones — lacks a Bank Account tag, since an
    untagged row could belong to the account being reconciled and there's
    no way to know; treating it as "not this account's" would silently
    undercount. This can only ever apply to disbursements/collections saved
    after the Bank Account field existed — historical rows are untagged by
    definition, so old periods will read 'incomplete' until enough of a
    period's activity is tagged going forward."""
    if not account or not period_month:
        return {'status': 'no_account_selected'}

    last_closing, last_date = _last_recon_closing_for_account(account)
    if last_closing is None or last_date is None:
        return {'status': 'no_history',
                'message': f'No prior reconciliation found for "{account}" — nothing to tie out against yet.'}

    through_date = _parse_flex_date(recon_date) or date.today()
    activity = _account_tagged_activity(records, mc_rows, account, last_date.date(), through_date)
    if activity['untagged_count'] > 0:
        return {
            'status': 'incomplete',
            'untagged_count': activity['untagged_count'],
            'total_count': activity['total_count'],
            'message': (f"{activity['untagged_count']} of {activity['total_count']} book entries "
                        f"since the last reconciliation ({last_date.strftime('%d-%m-%Y')}) have no Bank "
                        f"Account tag — cannot reliably attribute book activity to \"{account}\" yet."),
        }

    expected_closing = round(last_closing + activity['coll_total'] - activity['disb_total'], 2)
    variance = round(statement_closing - expected_closing, 2)

    # Context, not part of the formula: this statement's own Expense/Capital
    # totals (already computed by _match_transactions on this same upload)
    # are surfaced alongside the variance so a legitimate non-book cash
    # movement (bank charges, FD booking, inter-account transfer) doesn't
    # read as an unexplained discrepancy.
    expense_total = sum(t['debit'] for t in classified_txns if t['type'] == 'Expense')
    capital_net = (sum(t['credit'] for t in classified_txns if t['type'] in ('Capital In', 'Contra'))
                   - sum(t['debit'] for t in classified_txns if t['type'] in ('Capital Out', 'FD Booking', 'Contra')))

    return {
        'status': 'ok',
        'last_closing': last_closing,
        'last_closing_date': last_date.strftime('%d-%m-%Y') if last_date else None,
        'tagged_disbursed': activity['disb_total'], 'tagged_disbursed_count': activity['disb_count'],
        'tagged_collected': activity['coll_total'], 'tagged_collected_count': activity['coll_count'],
        'expected_closing': expected_closing,
        'statement_closing': statement_closing,
        'variance': variance,
        'ok': abs(variance) <= 1.0,
        'this_statement_expense_total': round(expense_total, 2),
        'this_statement_capital_net': round(capital_net, 2),
        'note': ('Book-tagged view only — does not include expenses or internal transfers on this '
                 'account this period; cross-check against the Expense/Capital figures above if the '
                 'variance looks large.'),
    }

def get_today_summary():
    records = read_accounts_from_gsheet()
    mcoll_rows = read_mcoll_from_gsheet()
    payment_events = get_payment_events(records, mcoll_rows)

    bank_balance, bank_date = get_latest_bank_balance()
    today = datetime.today().date()

    # Per-account last-reconciled date, so "since bank" activity is measured
    # against the account it actually happened in -- NOT the single earliest
    # date across all accounts. That single-cutoff approach double-counts:
    # any disbursement/collection tagged to whichever account was
    # reconciled MORE recently is already reflected in THAT account's own
    # closing balance, but re-adding/re-subtracting it here too (because
    # it's more recent than the OTHER account's cutoff) throws off
    # Available to Disburse the moment two accounts' reconciliation dates
    # drift apart -- which is the normal case, not an edge case; accounts
    # are rarely reconciled on the exact same day. Untagged rows (no Bank
    # Account column value) and any account with no reconciliation history
    # yet fall back to the conservative shared earliest date, so they're
    # never silently dropped from the window.
    account_dates = {}
    try:
        for acc in (load_config().get('bank_accounts') or []):
            acc_name = acc.get('name', '').strip()
            if acc_name:
                _, acc_date = _last_recon_closing_for_account(acc_name)
                if acc_date:
                    account_dates[acc_name] = acc_date
    except Exception:
        pass

    def _cutoff_for_tag(tag):
        tag = (tag or '').strip()
        if tag and tag in account_dates:
            return account_dates[tag]
        return bank_date

    disbursed_today = 0
    collected_today = 0
    total_outstanding = 0
    disbursed_since_bank = 0
    collected_since_bank = 0
    disb_rows, coll_rows, out_rows = [], [], []

    for r in records:
        did  = r.get('Disbursement ID', '')
        name = r.get('Customer Name', '')
        dd = r.get('Disbursement Date','')
        d = parse_disb_date(dd)
        if d and d.date() == today:
            amt = _to_num(r.get('Amount', 0))
            disbursed_today += amt
            disb_rows.append({'disb_id': did, 'customer': name, 'amount': amt})
        cutoff = _cutoff_for_tag(r.get('Bank Account', ''))
        if d and cutoff and d.date() > cutoff.date():
            disbursed_since_bank += _to_num(r.get('Amount', 0))
        if r.get('Overdue Status','').strip() not in ('Closed',''):
            _, total, collected, balance = _parse_case(r)
            total_outstanding += balance
            if balance > 0:
                out_rows.append({'disb_id': did, 'customer': name, 'amount': balance})

    # 'Collected Today' / 'collected since bank reconciliation' must sum just
    # the individual payment events on/after the relevant date - not the
    # cumulative Accounts column, which mixes in every prior instalment too.
    for ev in payment_events:
        ev_date = ev['date'].date()
        if ev_date == today:
            collected_today += ev['amount']
            coll_rows.append({'disb_id': ev['disb_id'], 'customer': ev['customer'], 'amount': ev['amount']})
        cutoff = _cutoff_for_tag(ev.get('bank_account', ''))
        if cutoff and ev_date > cutoff.date():
            collected_since_bank += ev['amount']

    # Money currently parked in a sweep-in FD is still disbursable — it
    # sweeps back automatically the moment the account needs it — so it
    # must not read as unavailable just because it isn't sitting as raw
    # balance.
    #
    # Sourced from the PER-FD LEDGER (_fd_ledger()) — every sweep leg is in
    # the bank statement and carries its own FD account number, so this is
    # fully derivable with no manual entry at all, and it reconciled to the
    # real FD balance to the rupee on 11-08-2026 (see _fd_ledger()'s comment
    # for the proof and for why the old blind net-sum lost 2,14,528).
    #
    # A directly-entered balance (save_fd_balance()) still wins, but ONLY
    # when it is dated on/after the newest FD transaction the ledger has
    # seen — otherwise it is by definition stale (sweeps have happened since
    # a human last looked) and would freeze the dashboard at an old number,
    # which is precisely the failure this whole mechanism exists to end.
    #
    # And the rule that was missing until 14-08-2026: a figure DERIVED FROM
    # INCOMPLETE DATA IS NOT A FIGURE. If any FD shows money sweeping back in
    # that was never seen going out, the ledger cannot know what that deposit
    # still holds — its arithmetic goes negative and gets floored at zero. It
    # then reports ₹0 with total confidence, which is the single most
    # damaging thing this dashboard can do: ₹0 looks like an answer. On
    # 14-08-2026 that put Available to Disburse at ₹20.6 lakh against a real
    # position near ₹72 lakh. An orphaned sweep-in now disqualifies the
    # derived figure outright and hands over to the entered balance.
    fd_date = None
    fd_ledger_info = None
    fd_incomplete = False
    try:
        fd_ledger_info = _fd_ledger()
        orphans = fd_ledger_info.get('orphans_unexplained') or {}
        unparsed = fd_ledger_info.get('unparsed') or []
        fd_incomplete = bool(orphans or unparsed)
        fd_outstanding = max(0.0, fd_ledger_info['fd_total'])
        fd_source = 'fd_ledger'
        direct_bal, direct_date = get_latest_fd_balance()
        if fd_incomplete:
            # Trust the entered balance regardless of date — a stale figure is
            # far closer to the truth than a confident zero.
            if direct_bal is not None:
                fd_outstanding, fd_date = direct_bal, direct_date
                fd_source = 'entered_ledger_incomplete'
            else:
                fd_outstanding = None
                fd_source = 'unknown_ledger_incomplete'
        elif direct_bal is not None and direct_date is not None:
            latest_txn = fd_ledger_info.get('latest_fd_txn_date')
            if latest_txn is None or direct_date.date() >= latest_txn:
                fd_outstanding, fd_date = direct_bal, direct_date
                fd_source = 'direct_override'
    except Exception:
        # Last-resort only: the old blind net sum. Known to understate
        # whenever an FD predates the reconciliation history or interest
        # comes back with the principal — kept purely so the dashboard
        # degrades to a number rather than going blank, and labelled so it
        # is never mistaken for the trustworthy figure.
        try:
            fd_outstanding = max(0.0, _all_time_recon_totals()['fd_total'])
            fd_source = 'derived_fallback'
        except Exception:
            fd_outstanding = 0.0
            fd_source = 'unavailable'

    # Available = what's in the current accounts + what's in the FDs,
    # adjusted only for business booked since each account's own statement.
    # If the FD figure is unknown, Available is unknown too — a number that
    # silently omits the FDs is worse than no number, because it looks like
    # an answer and is short by tens of lakhs.
    available = None
    if bank_balance is not None and fd_outstanding is not None:
        available = bank_balance - disbursed_since_bank + collected_since_bank + fd_outstanding
    return {
        'disbursed_today':   disbursed_today,
        'collected_today':   collected_today,
        'total_outstanding': total_outstanding,
        'available_for_disbursement': available,
        'fd_outstanding': fd_outstanding,
        'fd_outstanding_date': fd_date.strftime('%d %b %Y') if fd_date else None,
        'fd_outstanding_source': fd_source,
        'fd_open_count': (fd_ledger_info or {}).get('open_fd_count'),
        'fd_orphan_total': (fd_ledger_info or {}).get('orphan_sweepin_total'),
        'fd_unparsed_count': len((fd_ledger_info or {}).get('unparsed') or []),
        'bank_balance': bank_balance,
        'bank_balance_date': bank_date.strftime('%d %b %Y') if bank_date else None,
        'disbursed_since_bank': disbursed_since_bank,
        'collected_since_bank': collected_since_bank,
        'date': datetime.today().strftime('%d %b %Y'),
        'disb_rows': disb_rows,
        'coll_rows': coll_rows,
        'out_rows': out_rows,
    }

# ── Bank Reconciliation ───────────────────────────────────────────────────────

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def _parse_pdf_bank_statement(filepath):
    """PDF 'Statement of Account' bank statements — confirmed against a real
    IDFC FIRST Bank PDF (06-08-2026). Every page repeats two bordered
    tables: a 4-column Opening/Total Debit/Total Credit/Closing summary,
    and the 7-column transaction table (Transaction Date, Value Date,
    Particulars, Cheque No, Debit, Credit, Balance) with its header row
    repeated on every page. Tables are told apart by their own header text,
    not position, so a page with only one of the two (or an extra
    boilerplate table, e.g. the abbreviations glossary on the last pages)
    is handled correctly either way.

    One real wrinkle, confirmed in the sample: a transaction's multi-line
    narration can be split across a page boundary — the tail of the
    narration (and the actual Debit/Credit/Balance values) then land in a
    separate 'row' at the very top of the next page, with both date cells
    blank. Detected purely on "both date cells blank" and merged into the
    previous row. The one other blank-dates case is the leading "Opening
    Balance" row itself, handled as a special first case so it's never
    mistaken for a continuation with nothing to attach to.

    Cheque No is folded into Particulars unconditionally rather than kept
    as its own column — in the real sample it never once carried a genuine
    cheque number, only stray line-wrap fragments (e.g. a trailing "/")
    that pdfplumber's grid extraction occasionally spills into that
    column. Left as its own field, one of those fragments could get read
    downstream as if it were a real credit reference — see the UTR
    comment further below in parse_bank_statement().

    Verified against the real sample: summed Debit/Credit across every
    parsed row matched the statement's own printed Total Debit/Total
    Credit to the paisa, and Opening − Debit + Credit reproduced the
    printed Closing Balance exactly — i.e. no transaction is lost,
    duplicated, or corrupted by the page-merge step.
    """
    import pdfplumber
    all_rows = []
    summary = {}
    with pdfplumber.open(filepath) as pdf:
        for page in pdf.pages:
            for table in page.extract_tables():
                if not table or not table[0]:
                    continue
                header_text = ' '.join((c or '') for c in table[0]).lower()
                if 'opening balance' in header_text and 'closing balance' in header_text:
                    if not summary and len(table) > 1:
                        vals = table[1]
                        summary['opening'] = _clean_amount(vals[0]) if len(vals) > 0 else None
                        summary['closing'] = _clean_amount(vals[3]) if len(vals) > 3 else None
                elif 'transaction' in header_text and 'particulars' in header_text:
                    for row in table[1:]:
                        cells = [(c or '').strip() for c in row]
                        cells += [''] * (7 - len(cells))
                        all_rows.append(cells[:7])
                # Any other table (e.g. the abbreviations glossary on the
                # final pages) is boilerplate — not transaction data, skip.

    if not all_rows:
        raise ValueError("Could not find a transaction table in the PDF.")

    merged = []
    for date, value_date, particulars, cheque, debit, credit, balance in all_rows:
        particulars = (particulars + ' ' + cheque).strip() if cheque else particulars
        if not date and not value_date:
            if particulars.strip().lower() == 'opening balance' and not merged:
                continue  # already have this from the summary table
            if merged:
                prev = merged[-1]
                if particulars: prev[1] = (prev[1] + ' ' + particulars).strip()
                if debit:        prev[2] = debit
                if credit:       prev[3] = credit
                if balance:      prev[4] = balance
                continue
        merged.append([date, particulars, debit, credit, balance])

    header = ['Date', 'Particulars', 'Debit', 'Credit', 'Balance']
    df = pd.DataFrame([header] + merged)
    return df, summary.get('opening'), summary.get('closing')

# The statement declares its own period, e.g.
#   'STATEMENT PERIOD    11-Aug-2026 TO 11-Aug-2026'
# so the reconciliation date should be READ FROM THE FILE, never hand-typed.
# Added 12-08-2026 after a mislabelled upload (an 11-Aug IDFC statement saved
# under 12-Aug) put the two bank accounts on different reconciliation dates —
# which is precisely the state in which a contra transfer between our own
# accounts silently vanishes from the bank total (see _solvency_check()'s
# divergence warning). Removing the manual step removes that whole class of
# error rather than warning about it after the fact.
_STMT_PERIOD_PATTERNS = (
    re.compile(r'statement\s*period\s*[:\-]?\s*(.+?)\s+(?:to|through|-)\s+([0-9]{1,2}[-/][A-Za-z0-9]{2,9}[-/][0-9]{2,4})',
               re.IGNORECASE),
    re.compile(r'(?:period|from)\s*[:\-]?\s*([0-9]{1,2}[-/][A-Za-z0-9]{2,9}[-/][0-9]{2,4})\s+(?:to|through)\s+([0-9]{1,2}[-/][A-Za-z0-9]{2,9}[-/][0-9]{2,4})',
               re.IGNORECASE),
)

def _extract_statement_period(text_lines):
    """(from_date, to_date) as date objects, or (None, None). Only accepts a
    period whose BOTH ends parse — a half-read period is worse than none,
    since it would silently mislabel the save."""
    for line in text_lines:
        s = ' '.join(str(line).split())
        if not s:
            continue
        for pat in _STMT_PERIOD_PATTERNS:
            m = pat.search(s)
            if not m:
                continue
            raw_from = m.group(1).strip().split()[-1] if m.group(1).strip() else ''
            d_from = _parse_flex_date(raw_from)
            d_to = _parse_flex_date(m.group(2).strip())
            if d_from and d_to:
                return d_from, d_to
    return None, None

# The statement also names its own account, e.g.
#   'Account Number   50200012345678'   /   'A/c No: 10012345678'
# Read it for the same reason the period is read: the account selector is the
# other hand-entered field in this flow with no source of truth behind it, and
# picking the wrong one files a whole day's transactions against the wrong
# bank. With two statements uploaded together (13-08-2026) an unnoticed
# mis-selection would additionally cross-pair the two accounts' transfers, so
# detection matters more than it did when files were parsed one at a time.
#
# Masked forms (XXXXXXXX0923, ****3437) are accepted too — most statements
# print the number that way, and the visible tail is enough to identify it.
_STMT_ACCOUNT_PATTERNS = (
    re.compile(r'(?:account|a\s*/?\s*c)\s*(?:number|no\.?|#)?\s*[:\-]?\s*([0-9X\*x]{2,20}[0-9]{4})\b',
               re.IGNORECASE),
    re.compile(r'\b([0-9]{9,18})\b\s*(?:current|savings|sb|ca)\s*(?:account|a/c)', re.IGNORECASE),
)

def _extract_statement_account(text_lines, bank_accounts):
    """(account_name, detected_number) matched against Config's registered
    bank_accounts, or (None, detected_number/None).

    Matches on the FULL number when the statement prints it in full, and
    otherwise on the last 4 digits — statements routinely mask the leading
    digits (XXXXXXXX3437), and the last 4 are what a human recognises anyway.
    A last-4 match is only accepted when exactly one registered account ends
    in those digits; two accounts sharing a suffix means we cannot tell them
    apart, and guessing would be worse than asking.
    """
    registered = [(a.get('name', '').strip(), re.sub(r'\D', '', str(a.get('account_number', '') or '')))
                  for a in (bank_accounts or []) if a.get('name', '').strip()]

    # Simplest rule first, and the one that actually works: if a registered
    # account's number appears ANYWHERE in the file, that is the account.
    # No dependence on the words "Account Number", on where in the header the
    # bank chose to print it, or on the 60-line window the pattern search
    # used — all of which failed on Prem's real August download even though,
    # in his words, "the statment dates and ac numbers were there".
    # Checked longest-number-first so a short number that happens to be a
    # substring of a longer one can never win.
    whole = ' '.join(str(l) for l in text_lines)
    digits_only = re.sub(r'\D', '', whole)
    for name, acct in sorted(registered, key=lambda x: -len(x[1])):
        if acct and len(acct) >= 9 and acct in digits_only:
            return name, acct

    # Same idea for a masked number printed without the word "account"
    # anywhere near it (XXXXXXXX0923, ****8167). Accepted only when exactly
    # one registered account ends in those four digits — two candidates means
    # we genuinely cannot tell, and asking beats guessing.
    for masked in re.findall(r'(?:[X\*x]\s*){3,}(\d[\s\d]{2,}\d)', whole):
        tail = re.sub(r'\D', '', masked)[-4:]
        if len(tail) < 4:
            continue
        hits = [n for n, a in registered if a and a[-4:] == tail]
        if len(hits) == 1:
            return hits[0], f'****{tail}'
        if len(hits) > 1:
            return None, f'****{tail}'

    for line in text_lines[:60]:   # header block only; txn rows can contain long refs
        s = ' '.join(str(line).split())
        if not s:
            continue
        for pat in _STMT_ACCOUNT_PATTERNS:
            for m in pat.finditer(s):
                raw = m.group(1)
                masked = bool(re.search(r'[X\*x]', raw))
                num = re.sub(r'\D', '', raw)
                if len(num) < 4:
                    continue
                # An unmasked value must still look like a real account
                # number, or every long reference in a header line qualifies.
                if not masked and not (9 <= len(num) <= 18):
                    continue
                if not masked:
                    exact = [n for n, a in registered if a and a == num]
                    if exact:
                        return exact[0], num
                tail = [n for n, a in registered if a and len(a) >= 4 and a[-4:] == num[-4:]]
                if len(tail) == 1:
                    return tail[0], raw
                if len(tail) > 1:
                    return None, raw
    return None, None

def parse_bank_statement(filepath, filename, bank_accounts=None):
    """Parse bank statement CSV/Excel/PDF. Returns transactions + opening/closing balance."""
    ext = filename.lower().rsplit('.', 1)[-1]
    try:
        if ext == 'csv':
            df = pd.read_csv(filepath, header=None, dtype=str, encoding='utf-8', on_bad_lines='skip')
        elif ext == 'xls':
            df = pd.read_excel(filepath, header=None, dtype=str, engine='xlrd')
        elif ext == 'pdf':
            df, pdf_opening, pdf_closing = _parse_pdf_bank_statement(filepath)
        else:
            df = pd.read_excel(filepath, header=None, dtype=str, engine='openpyxl')
    except Exception as e:
        raise ValueError(f"Could not read file: {e}")

    # Read the statement's own declared period before anything else — this is
    # the authoritative reconciliation date, in preference to whatever a human
    # typed into the date box.
    try:
        _period_lines = [' '.join(str(c).strip() for c in row if pd.notna(c) and str(c).strip())
                         for _, row in df.iterrows()]
    except Exception:
        _period_lines = []
    stmt_from, stmt_to = _extract_statement_period(_period_lines)
    stmt_account, stmt_account_no = _extract_statement_account(_period_lines, bank_accounts)

    opening_balance = closing_balance = None
    if ext == 'pdf':
        # The PDF's transaction table has no "Opening/Closing Balance"
        # labeled row for the scan below to find (that text lives only in
        # the separate summary table, already extracted above) — trust it
        # directly rather than running a scan that would find nothing.
        opening_balance, closing_balance = pdf_opening, pdf_closing
    else:
        # Scan for opening / closing balance in metadata rows
        for _, row in df.iterrows():
            cells = [str(c).strip() for c in row if pd.notna(c) and str(c).strip()]
            combined = ' '.join(cells).lower()
            if opening_balance is None and 'opening' in combined and 'balance' in combined:
                for c in cells:
                    v = _clean_amount(c)
                    if v: opening_balance = v; break
            if closing_balance is None and ('closing' in combined or 'available balance' in combined) and 'balance' in combined:
                for c in cells:
                    v = _clean_amount(c)
                    if v: closing_balance = v; break

    # Find transaction header row
    header_row = None
    for i, row in df.iterrows():
        cells = [str(c).lower() for c in row if pd.notna(c)]
        combined = ' '.join(cells)
        if ('debit' in combined or 'withdrawal' in combined or 'credit' in combined or 'deposit' in combined) \
                and ('date' in combined or 'narration' in combined or 'description' in combined):
            header_row = i; break

    if header_row is None:
        raise ValueError("Could not detect header row in the statement.")

    df.columns = df.iloc[header_row]
    df = df.iloc[header_row + 1:].reset_index(drop=True)
    df.columns = [str(c).strip().lower() if pd.notna(c) else f'col_{i}' for i, c in enumerate(df.columns)]

    col_map = {}
    for col in df.columns:
        cl = str(col).lower()
        if any(k in cl for k in ['value date', 'txn date', 'transaction date', 'posting date', 'date']):
            col_map.setdefault('date', col)
        if any(k in cl for k in ['narration', 'description', 'particulars', 'remarks', 'details']):
            col_map.setdefault('description', col)
        if any(k in cl for k in ['utr', 'ref no', 'chq/ref', 'cheque', 'reference']):
            col_map.setdefault('utr', col)
        if any(k in cl for k in ['debit', 'withdrawal', 'dr amount']):
            col_map.setdefault('debit', col)
        if any(k in cl for k in ['credit', 'deposit', 'cr amount']):
            col_map.setdefault('credit', col)
        if 'balance' in cl:
            col_map.setdefault('balance', col)

    # Keywords that identify summary/footer rows — skip these
    SKIP_KEYWORDS = {'opening balance', 'closing balance', 'total', 'available balance',
                     'ledger balance', 'brought forward', 'carried forward', 'statement summary',
                     'end of statement', 'generated on', 'dr count', 'cr count',
                     'debit count', 'credit count', 'page no', '****'}

    rows = []
    last_balance = None
    for _, row in df.iterrows():
        date_val = str(row.get(col_map.get('date', ''), '') or '').strip()
        desc     = str(row.get(col_map.get('description', ''), '') or '').strip()
        utr_raw  = str(row.get(col_map.get('utr', ''), '') or '').strip()
        debit    = _clean_amount(row.get(col_map.get('debit', ''), ''))
        credit   = _clean_amount(row.get(col_map.get('credit', ''), ''))
        balance  = _clean_amount(row.get(col_map.get('balance', ''), ''))

        if not date_val or date_val.lower() in ('nan', 'none', '', 'date'):
            continue
        # Skip separator/asterisk rows
        if set(date_val.replace('*','').replace('-','').strip()) == set():
            continue
        # Skip rows where "date" cell is a pure number (statement summary totals row e.g. 5704953.02)
        try:
            float(date_val.replace(',', ''))
            continue
        except ValueError:
            pass
        # Skip summary/footer rows by keyword anywhere in the row
        row_text = ' '.join(str(v).lower() for v in row.values if pd.notna(v))
        desc_l = desc.lower()
        if any(kw in desc_l or kw in date_val.lower() or kw in row_text for kw in SKIP_KEYWORDS):
            continue
        # Skip rows that look like count/summary lines (no real date format).
        # Middle group accepts either a numeric month (15-07-2026, 15/07/2026)
        # or a 3+ letter abbreviation (15-Jul-2026, the format IDFC's newer
        # statement download uses) — the numeric-only version silently
        # discarded every transaction row from any statement using the
        # letter-month format, with no error surfaced (0 transactions parsed).
        if not re.search(r'\d{1,2}[/\-](?:\d{1,2}|[A-Za-z]{3,9})[/\-]\d{2,4}', date_val):
            continue
        if debit == 0 and credit == 0:
            continue

        # UTR strategy:
        #   Credits  → Chq/Ref column IS the RTGS/NEFT UTR (e.g. KARBR52026...)
        #   Debits   → Chq/Ref column has HDFC sequential number (000...159) — useless
        #              Real UTR is at the END of the narration (e.g. -HDFCR52026060165094792)
        chq_clean = utr_raw if utr_raw.lower() not in ('nan','none','') else ''
        if credit > 0:
            # For credits, prefer the Chq/Ref col; fall back to narration extraction
            utr = chq_clean if (chq_clean and not re.match(r'^0+\d{0,6}$', chq_clean)) \
                  else extract_utr(desc)
        else:
            # For debits, always extract from narration (Chq col has sequential nums)
            utr = extract_utr(desc)
            # If narration extraction failed, try the chq col as last resort
            if not utr and chq_clean and not re.match(r'^0+\d{0,6}$', chq_clean):
                utr = chq_clean

        # Unconditional, not "if balance:" — a real transaction row's balance
        # can legitimately BE zero (an account fully drained/closed, as
        # happened with Saraswat), and treating 0 as falsy here silently
        # left last_balance stuck at the previous nonzero value, reporting
        # the wrong closing balance for exactly the accounts most likely to
        # actually reach zero. By this point in the loop the row has already
        # passed the debit-and-credit-both-zero skip check above, so it's a
        # genuine transaction and its balance cell should always be trusted.
        last_balance = balance

        # Collapse the bank's own line wrapping. IDFC wraps long narrations
        # mid-phrase, so the stored text is literally "Sweepout FD\n<acct>
        # booked" -- and `\bfd\b.*\bbooked\b` cannot match across a newline
        # because `.` excludes it. Six August sweep-outs worth Rs 1.23 crore
        # were therefore typed Expense instead of FD Booking, which zeroed the
        # FD balance and would have charged the lot to the P&L. Normalising
        # once here means no downstream pattern ever has to think about it.
        desc = ' '.join(desc.split())
        rows.append({'date': date_val, 'description': desc, 'utr': utr,
                     'debit': debit, 'credit': credit, 'balance': balance})

    if rows:
        # Not "and rows[0]['balance']"/"and last_balance" — same zero-is-
        # not-falsy fix as above. _clean_amount() always returns a float
        # (0.0 for both "genuinely zero" and "blank/unparseable"), so there
        # was never a way to tell those apart via truthiness anyway; trust
        # the parsed value once we know real rows exist.
        if opening_balance is None:
            first = rows[0]
            opening_balance = round(first['balance'] + first['debit'] - first['credit'], 2)
        if closing_balance is None:
            closing_balance = last_balance

    # If the header didn't declare a period, the transactions themselves are
    # the period — first row to last. Simpler than any header pattern, always
    # present, and impossible to get wrong: a statement cannot cover a day it
    # has no rows for. Only the header's own declaration outranks it.
    if not (stmt_from and stmt_to):
        row_dates = [d for d in (_parse_flex_date(r['date']) for r in rows) if d]
        if row_dates:
            stmt_from = stmt_from or min(row_dates)
            stmt_to = stmt_to or max(row_dates)

    return {'transactions': rows,
            'opening_balance': opening_balance or 0,
            'closing_balance': closing_balance or 0,
            'statement_from': stmt_from.strftime('%d-%m-%Y') if stmt_from else None,
            'statement_to': stmt_to.strftime('%d-%m-%Y') if stmt_to else None,
            'statement_period_source': 'header' if _extract_statement_period(_period_lines)[1] else 'transactions',
            'statement_account': stmt_account,
            'statement_account_no': stmt_account_no}

def _clean_amount(v):
    try:
        s = str(v or '').replace(',', '').replace('₹', '').replace(' ', '').strip()
        if not s or s.lower() in ('nan', 'none', '-', ''):
            return 0.0
        return float(s)
    except Exception:
        return 0.0

def _xl_styles():
    thin  = Side(style='thin',   color='D0DCE8')
    return {
        'thin_border': Border(left=thin, right=thin, top=thin, bottom=thin),
        'hdr_fill':    PatternFill("solid", fgColor="1A3A5C"),
        'hdr_font':    Font(bold=True, color="FFFFFF", size=9, name='Arial'),
        'grn_fill':    PatternFill("solid", fgColor="E2EFDA"),   # matched / closed
        'amb_fill':    PatternFill("solid", fgColor="FFF2CC"),   # near match / partial
        'red_fill':    PatternFill("solid", fgColor="FFE0E0"),   # unmatched / review
        'blu_fill':    PatternFill("solid", fgColor="DDEEFF"),   # capital / other
        'tot_fill':    PatternFill("solid", fgColor="F2F2F2"),
        'tot_font':    Font(bold=True, size=9, name='Arial'),
        'norm_font':   Font(size=9, name='Arial'),
        'bold_font':   Font(bold=True, size=9, name='Arial'),
        'title_font':  Font(bold=True, size=11, name='Arial', color='1A3A5C'),
        'sub_font':    Font(bold=True, size=9, name='Arial', color='555555'),
    }

def _hdr(ws, row, cols, s):
    for i, (label, width) in enumerate(cols, 1):
        c = ws.cell(row, i, label)
        c.fill = s['hdr_fill']; c.font = s['hdr_font']; c.border = s['thin_border']
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.row_dimensions[row].height = 26

def _row(ws, r, vals, s, right_cols=(), fill=None, font=None):
    f = fill or (s['grn_fill'] if r % 2 == 0 else PatternFill("solid", fgColor="FFFFFF"))
    for i, v in enumerate(vals, 1):
        c = ws.cell(r, i, v)
        c.fill = f; c.font = font or s['norm_font']; c.border = s['thin_border']
        if i in right_cols: c.alignment = Alignment(horizontal='right')

def _n(v):
    """Number → Indian-comma string, blank if zero/None."""
    try:
        f = float(v)
        return f"{f:,.2f}" if f else ''
    except Exception:
        return ''

def _title(ws, text, ncols, s):
    ws.merge_cells(f'A1:{get_column_letter(ncols)}1')
    c = ws.cell(1, 1, text)
    c.font = s['title_font']; c.alignment = Alignment(horizontal='left')
    ws.row_dimensions[1].height = 22

def _legend(ws, text, ncols, s):
    ws.merge_cells(f'A2:{get_column_letter(ncols)}2')
    c = ws.cell(2, 1, text)
    c.font = Font(italic=True, size=8, name='Arial', color='555555')
    c.alignment = Alignment(horizontal='left')

def _expense_category(desc):
    d = desc.lower()
    if 'pos' in d or 'swipe' in d:                                    return 'Card Purchase (POS)'
    if any(k in d for k in ['claude','openai','chatgpt','base44','lovable','subscription','me dc si']):
                                                                       return 'Software Subscription'
    if any(k in d for k in ['airtel','jio','telecom','broadband']):   return 'Telecom'
    if any(k in d for k in ['gst/bank','gst/gst','markup','dcc','dc intl','bank charg','service charge']):
                                                                       return 'Bank Charges / GST'
    if 'salary' in d or 'sal ' in d:                                  return 'Salary'
    # A bare 'partner' match is far too loose: the FIRM'S OWN NAME is
    # "BridgeLine Partners", so any cheque or transfer whose narration
    # carries the company name was booked as a partner drawing. Confirmed
    # live 12-08-2026 — two cheques totalling Rs 5,75,000
    # ("CHQ PAID-CTS S6-RK S-BRIDGELINE PARTNERS") were being reported as
    # Partner Share purely because our own name contains the word PARTNERS,
    # which alone turned a real profit into a reported loss. Strip the
    # company name out first, then require actual drawing phrasing.
    d_nc = d.replace('bridgeline partners', ' ').replace('bridgeline', ' ')
    if any(k in d_nc for k in ('partner share', 'partner salary', 'partners share',
                                'drawing', 'withdrawal', 'proprietor')):
        return 'Partner Share'
    if any(k in d for k in ['pradaan','pradan']):                     return 'Pradaan Routing'
    return 'Miscellaneous'

def _looks_like_own_transfer(desc):
    """True when the narration names BridgeLine itself as the SENDER/DRAWER —
    i.e. our own money arriving from another of our own accounts, never a
    customer payment.

    The distinction is the slot our name sits in, not merely that it appears:
    on a genuine customer payment we are the RECEIVING account holder (e.g.
    'FT -BRIDGELINE PARTNERS CR - ...'), which must NOT match here.

    Cheque deposits use 'BB/CHQ DEP/<no>/<date>/<drawer>/<drawer bank>', so the
    drawer occupies that sender slot. Confirmed live 12-08-2026: the ₹5,00,000
    that opened the IDFC account reads
    'BB/CHQ DEP/000107/14-07-2026/BRIDGELINE/HDFC BANK' — drawn by us, on our
    own HDFC account. It had been auto-matched to a customer case purely on
    amount, which is exactly what this prevents.
    """
    return bool(
        re.search(r'\bCR[-/][A-Z]{4}0[A-Z0-9]{6}[-/]\s*BRIDGELINE PARTNERS\b', desc, re.IGNORECASE)
        # Separators vary between statements for the SAME transfer type: IDFC
        # has emitted both 'RTGS/<utr>/BRIDGELINE PARTNERS/' and
        # 'RTGS/ <utr> BRIDGELINE PARTNERS/' (spaces, not slashes) — the latter
        # seen live 12-08-2026 on an ₹18,00,000 own-account transfer that got
        # typed Repayment because the old slash-only pattern missed it. Accept
        # either separator. The leading 'RTGS/' (slash) is still REQUIRED: a
        # customer payment reads 'RTGS CR-<ifsc>-<payer>-BRIDGELINE PARTNERS-…',
        # which has no slash there and so still cannot match.
        or re.search(r'(?:RTGS|NEFT|IMPS)\s*/\s*[A-Z0-9]+[\s/]+BRIDGELINE PARTNERS', desc, re.IGNORECASE)
        or re.search(r'chq\s*dep\b[^/]*(?:/[^/]*){0,2}/\s*bridgeline', desc, re.IGNORECASE)
    )

def _norm_ws(text):
    """Bank narration with its line wrapping collapsed. Any matcher that
    looks for two words in sequence must run on this, never on raw text."""
    return ' '.join(str(text or '').split())

def _is_fd_booking(desc):
    return bool(re.search(r'\bfd\b.*\bbooked\b', _norm_ws(desc), re.IGNORECASE))

def _is_expense_debit(desc, amt):
    """Return True if this debit is an operating expense (not a loan disbursement)."""
    d = desc.lower()
    # Fixed deposits are capital movements, not expenses
    if _is_fd_booking(desc):
        return False
    # RTGS/NEFT DR with large amounts are almost always disbursements
    if re.search(r'rtgs\s+dr|neft\s+dr', d) and amt >= 50000:
        return False
    # Partner share / owner withdrawals
    if any(k in d for k in ['partshare', 'part share', 'prem narayan', 's a prem']):
        return True
    # POS, card, subscriptions, bank charges, UPI small payments → expense
    if any(k in d for k in ['pos ', ' pos', 'me dc', 'dc intl', 'markup', 'gst/bank',
                              'subscription', 'claude', 'openai', 'chatgpt', 'base44',
                              'airtel', 'jio', 'salary', 'partner']):
        return True
    # Small UPI/IMPS debits (< 25000) not RTGS/NEFT DR → expense
    if amt < 25000 and re.search(r'^(upi|imps|ft\s)', d):
        return True
    return False

def _parse_flex_date(date_str):
    """Parse date_str (any format DD/MM/YY, DD-MM-YYYY, YYYY-MM-DD, etc. —
    whatever mix book rows and bank statements use) into a plain date, or
    None if unparseable. Shared by _date_in_period (month-wide) and
    _date_in_range (exact-window) so both agree on what counts as a date."""
    if not date_str:
        return None
    from datetime import datetime as _dt
    # %d-%b-%Y / %d-%b-%y (e.g. "7-Apr-2026", "17-Apr-26") -- the legacy
    # "Apr/May26" archive tab's own date format, missed here until
    # 11-08-2026 (a GST-figures cross-check against a real CA pivot table
    # came back with April/May missing entirely -- these dates were
    # silently failing to parse and getting dropped everywhere this
    # function is used, not just that one check).
    for fmt in ('%d/%m/%Y', '%d-%m-%Y', '%d/%m/%y', '%d-%m-%y',
                '%Y-%m-%d', '%m/%d/%Y', '%d %b %Y', '%d %B %Y',
                '%d-%b-%Y', '%d-%b-%y'):
        try:
            return _dt.strptime(str(date_str).strip(), fmt).date()
        except ValueError:
            continue
    return None

def _date_in_period(date_str, period_month):
    """Return True if date_str (any format DD/MM/YY, DD-MM-YYYY, etc.) falls in period_month (e.g. 'Jun 2026')."""
    if not period_month:
        return False
    try:
        from datetime import datetime as _dt
        pm = _dt.strptime(period_month, '%b %Y')
    except Exception:
        return False
    d = _parse_flex_date(date_str)
    return d is not None and d.month == pm.month and d.year == pm.year

def _date_in_range(date_str, since_date, through_date):
    """Return True if date_str falls strictly after since_date (exclusive —
    that date's own activity is already reflected in a prior closing
    balance) and on/before through_date (inclusive). since_date may be None
    for an unbounded lower end; through_date is required."""
    if through_date is None:
        return False
    d = _parse_flex_date(date_str)
    if d is None:
        return False
    if since_date is not None and d <= since_date:
        return False
    return d <= through_date

# ── Staged amount+date matcher ────────────────────────────────────────────────
# Adapted from a standalone bank-reconciliation engine (exact match, then
# date-tolerance match, run in that order of confidence). Slots in as a new
# tier inside _match_transactions() between UTR matching (already
# unambiguous, untouched) and the older amount-only fallback (kept as a
# backstop below, untouched, for anything this doesn't resolve). The old
# fallback picked the FIRST same-amount candidate by plain insertion order
# with no date awareness at all -- a real wrong-match risk whenever an
# amount recurs, e.g. two disbursements of the same round amount in one
# week. Many-to-one/one-to-many grouping from the original engine was
# deliberately left out of this integration: matched_ref is a single id
# string everywhere downstream (_book_completeness, _sheet_coll_recon,
# _account_tieout), and a comma-joined multi-id ref would silently break
# every one of those without a wider audit this pass didn't do.
@dataclass
class _ReconTxn:
    id: str
    date: Optional[datetime]
    amount: float


_NARRATIVE_STOPWORDS = {'THE','AND','FOR','FROM','WITH','BANK','NEFT','RTGS','IMPS','UPI','CR','DR',
                         'ACCOUNT','CHEQUE','DEPOSIT','DEPOSITS','SUBJECT','CLEARING','AVL','BAL','INR','UPDATE'}


def _narrative_signature(text):
    """Extract (trailing_digits, name_tokens) from a masked-account-style
    narrative reference (e.g. 'FT - Cr - XXXXXXXXXX3437 - K S MAHESHA.'),
    which the UTR regex extraction in _match_transactions() can never
    capture -- it requires an unbroken alphanumeric token, and narrative
    text with spaces/hyphens fails that \\b(...)\\b boundary entirely. A
    case whose ONLY recorded reference is narrative-style then has nothing
    to key UTR matching off, and previously fell straight through to the
    weak amount-only fallback -- confirmed a real bug via a full-book
    reference cross-reference audit (31-Jul-2026): K S Mahesh's repayment
    matched to an unrelated, differently-dated case that happened to share
    the same amount, when his own recorded reference (this narrative
    pattern) exactly matched the real bank line once checked directly.
    Returns None for a clean UTR-style reference (no name tokens in a bare
    numeric/alphanumeric code), so this only ever fires for genuinely
    narrative text."""
    text = (text or '').strip()
    digits = re.findall(r'\d', text)
    if len(digits) < 4:
        return None
    names = {w for w in re.split(r'[^A-Za-z]+', text.upper()) if len(w) >= 4 and w not in _NARRATIVE_STOPWORDS}
    if not names:
        return None
    return (''.join(digits[-4:]), names)


def _narrative_ref_match(book_items, bank_txns_indexed, exclude_book_ids=()):
    """book_items: [(book_id, raw_ref_text), ...]. Matches a bank txn to a
    book item when the narration contains BOTH the reference's trailing 4
    digits (typically the unmasked tail of an account number the book only
    stored masked) and at least one shared name token. Deliberately only a
    fallback consulted after UTR and the staged Amount+Date tier already
    had their shot, since this is fuzzier than either of those."""
    sigs = []
    for book_id, text in book_items:
        if book_id in exclude_book_ids:
            continue
        sig = _narrative_signature(text)
        if sig:
            sigs.append((book_id, sig[0], sig[1]))
    result = {}
    claimed = set()
    for idx, desc in bank_txns_indexed:
        desc_digits = re.sub(r'\D', '', desc)
        desc_names = {w for w in re.split(r'[^A-Za-z]+', desc.upper()) if len(w) >= 4}
        for book_id, tail, names in sigs:
            if book_id in claimed:
                continue
            if tail in desc_digits and (names & desc_names):
                result[idx] = book_id
                claimed.add(book_id)
                break
    return result


def _split_ref_tokens(value):
    """Split a Debit/Credit Note field into individual reference tokens.
    Only splits on a comma NOT immediately followed by a digit, so a
    thousand-separator comma inside a raw narrative/SMS reference (e.g.
    'Update! INR 61,360.00 deposited...') is never mistaken for a token
    boundary between two genuinely separate references (e.g. '621119299326,
    CNRBR52026073192673966', two real UTRs for a case with two payments) --
    confirmed both patterns are real via the 31-Jul-2026 cross-reference
    audit, and a naive re.split(',') breaks the first one."""
    value = (value or '').strip()
    if not value:
        return []
    return [p.strip() for p in re.split(r',(?!\d)', value) if p.strip()]


def _staged_amount_date_match(book_txns, bank_txns_indexed, date_tolerance_days=5):
    """book_txns: list of _ReconTxn (one per Accounts disbursement row, or
    per M Coll payment row). bank_txns_indexed: list of (original txns
    index, _ReconTxn) for the bank side. Returns {bank_index: (book_id,
    match_type)} for every bank txn this can confidently tie to exactly one
    still-unmatched book txn."""
    by_amount: dict = {}
    for b in book_txns:
        by_amount.setdefault(round(b.amount, 2), []).append(b)

    matched_book_ids = set()
    result = {}

    # Stage 1: exact same amount + same calendar date.
    for idx, t in bank_txns_indexed:
        if t.date is None:
            continue
        cands = [b for b in by_amount.get(round(t.amount, 2), [])
                  if b.id not in matched_book_ids and b.date == t.date]
        if cands:
            chosen = cands[0]
            matched_book_ids.add(chosen.id)
            result[idx] = (chosen.id, 'exact')

    # Stage 2: same amount, closest date within the tolerance window --
    # covers the ordinary case of a bank clearing delay (cheque/NEFT taking
    # a couple of days to reflect).
    for idx, t in bank_txns_indexed:
        if idx in result or t.date is None:
            continue
        cands = [b for b in by_amount.get(round(t.amount, 2), [])
                  if b.id not in matched_book_ids and b.date is not None
                  and abs((b.date - t.date).days) <= date_tolerance_days]
        if cands:
            chosen = min(cands, key=lambda b: abs((b.date - t.date).days))
            matched_book_ids.add(chosen.id)
            result[idx] = (chosen.id, 'date_tolerance')

    return result


# ── Transaction matcher ───────────────────────────────────────────────────────

def _match_transactions(txns, records, mc_rows):
    """Auto-classify + match each bank transaction against Accounts + M Coll."""

    # Registered bank accounts (Config 'bank_accounts') — a transfer
    # narration containing one of THESE account numbers is money moving
    # between our own accounts (Capital Out on the sending statement,
    # Capital In on the receiving one), not a real disbursement/expense/
    # collection. Distinct from acct_to_disb below, which indexes CUSTOMER
    # beneficiary account numbers for FT collection matching.
    own_account_numbers = set()
    try:
        for acc in (load_config().get('bank_accounts') or []):
            num = str(acc.get('account_number', '')).strip()
            if num:
                own_account_numbers.add(num)
    except Exception:
        pass

    # UTR → disb_id from debit note column (disbursements going out)
    utr_to_disb     = {}
    utr_to_coll_acct = {}  # UTR/ref in Credit Note of Accounts → disb_id (collection receipts recorded in sheet)
    amt_to_disb_id  = {}   # round(amt) → disb_id  (for amount-only fallback)
    for r in records:
        did      = r.get('Disbursement ID','').strip()
        deb_note = str(r.get('Debit Note','') or '').strip()
        crd_note = str(r.get('Credit Note','') or '').strip()
        amt      = _to_num(r.get('Amount', 0))
        coll_amt = _to_num(r.get('Collected Amount', 0) or r.get('Collected   Amount', 0))
        # Debit Note → disb UTR map
        for u in re.split(r'[,;\s]+', deb_note):
            u = u.strip().upper()
            if len(u) >= 8: utr_to_disb[u] = did
        # Also extract numeric/alphanumeric refs from any SMS-style debit note text
        for num in re.findall(r'\b([A-Z0-9]{10,22})\b', deb_note.upper()):
            utr_to_disb[num] = did
        # Credit Note → collection UTR map (IMPS/NEFT/UPI refs embedded in SMS)
        for u in re.split(r'[,;\s]+', crd_note):
            u = u.strip().upper()
            if len(u) >= 8: utr_to_coll_acct[u] = did
        for num in re.findall(r'\b([A-Z0-9]{10,22})\b', crd_note.upper()):
            utr_to_coll_acct[num] = did
        if amt: amt_to_disb_id.setdefault(round(amt), []).append(did)
        if coll_amt: amt_to_disb_id.setdefault(round(coll_amt), []).append(did)

    # UTR → disb_id from M Coll (collections coming in)
    utr_to_coll   = {}
    amt_to_coll   = {}   # round(amt) → disb_id
    for mc in mc_rows:
        did = (mc[0] if len(mc)>0 else '').strip()
        raw = (mc[3] if len(mc)>3 else '').strip()
        amt = _clean_amount(mc[2]) if len(mc)>2 else 0
        if did:
            # Map the full Credit Note value
            if raw: utr_to_coll[raw.upper()] = did
            # Also extract any IMPS/NEFT/RTGS ref numbers embedded in SMS text (10-15 digits)
            for num in re.findall(r'\b\d{10,15}\b', raw):
                utr_to_coll[num] = did
            # Extract bank UTR codes (e.g. KARBR5..., HDFCR5...)
            for code in re.findall(r'[A-Z]{4,}[A-Z0-9]{6,}', raw.upper()):
                utr_to_coll[code] = did
        if amt and did: amt_to_coll.setdefault(round(amt), []).append(did)

    # Customer name → disb_id (for Pradaan / name-based match)
    name_to_disb = {}   # full name → disb_id
    words_to_disb = []  # [(significant_words_set, disb_id, full_name)] for fuzzy matching
    STOP_WORDS = {'mr', 'mrs', 'ms', 'dr', 'the', 'and', 'of', 'to', 'in', 'for', 'a', 'an'}
    for r in records:
        did  = r.get('Disbursement ID','').strip()
        name = r.get('Customer Name','').strip().lower()
        if name:
            name_to_disb[name] = did
            words = {w for w in re.split(r'[\s\.\-]+', name) if len(w) >= 3 and w not in STOP_WORDS}
            if words:
                words_to_disb.append((words, did, name))

    def _fuzzy_name_match(desc):
        """Return (disb_id, score) for best name match in desc. Score = matched word count."""
        dl = desc.lower()
        best_did, best_score = '', 0
        for words, did, full_name in words_to_disb:
            score = sum(1 for w in words if w in dl)
            ratio = score / len(words) if words else 0
            # Need ≥2 words matched OR full name is single meaningful word with exact match
            if score >= 2 or (len(words) == 1 and ratio == 1.0):
                if score > best_score:
                    best_score, best_did = score, did
        return best_did, best_score

    # HDFC account number → disb_id (for FT/internal transfers)
    # Pattern: "PRANAV T P DR - 50100551677276 - PRANAV T P" — the 14-digit number is beneficiary acct
    # We index by account number so "FT -BRIDGELINE PARTNERS CR - 50100551677276 - PRANAV T P" can match
    acct_to_disb = {}
    for tx in txns:
        desc = tx.get('description', '')
        if tx['debit'] > 0:
            m = re.search(r'\b(5010\d{10})\b', desc)  # HDFC account numbers start with 5010
            if m:
                acct_to_disb[m.group(1)] = None  # placeholder; will resolve via amount below

    # Staged exact+date-tolerance matcher (see _staged_amount_date_match) --
    # built once here, consulted per-transaction below as a new tier between
    # UTR matching and the older amount-only fallback.
    disb_book_txns = []
    for r in records:
        did = r.get('Disbursement ID', '').strip()
        amt = _to_num(r.get('Amount', 0))
        if did and amt:
            disb_book_txns.append(_ReconTxn(id=did, date=parse_disb_date(r.get('Disbursement Date', '')), amount=amt))
    coll_book_txns = []
    for mc in mc_rows:
        did = (mc[0] if len(mc) > 0 else '').strip()
        amt = _clean_amount(mc[2]) if len(mc) > 2 else 0
        if did and amt:
            coll_book_txns.append(_ReconTxn(id=did, date=parse_disb_date(mc[1] if len(mc) > 1 else ''), amount=amt))

    debit_bank_txns = [(i, _ReconTxn(id=str(i), date=parse_disb_date(t.get('date', '')), amount=t['debit']))
                       for i, t in enumerate(txns) if t['debit'] > 0]
    credit_bank_txns = [(i, _ReconTxn(id=str(i), date=parse_disb_date(t.get('date', '')), amount=t['credit']))
                        for i, t in enumerate(txns) if t['credit'] > 0]

    staged_disb_matches = _staged_amount_date_match(disb_book_txns, debit_bank_txns)
    staged_coll_matches = _staged_amount_date_match(coll_book_txns, credit_bank_txns)

    # Narrative-reference matcher (see _narrative_ref_match) -- a further
    # tier for cases whose ONLY recorded reference is masked-narrative text
    # (e.g. Debit/Credit Note = 'FT - Cr - XXXXXXXXXX3437 - K S MAHESHA.'),
    # which contributes nothing to the UTR maps above and can't carry a
    # calendar date either, so the staged Amount+Date tier can't help it.
    # Consulted after both of those, before the plain amount-only fallback.
    disb_narrative_items = [(r.get('Disbursement ID', '').strip(), x)
                             for r in records for x in _split_ref_tokens(r.get('Debit Note', ''))]
    coll_narrative_items = [((mc[0] if len(mc) > 0 else '').strip(), x)
                             for mc in mc_rows for x in _split_ref_tokens(mc[3] if len(mc) > 3 else '')]
    debit_bank_desc = [(i, t.get('description', '')) for i, t in enumerate(txns) if t['debit'] > 0]
    credit_bank_desc = [(i, t.get('description', '')) for i, t in enumerate(txns) if t['credit'] > 0]
    staged_disb_ids = {book_id for book_id, _ in staged_disb_matches.values()}
    staged_coll_ids = {book_id for book_id, _ in staged_coll_matches.values()}
    narrative_disb_matches = _narrative_ref_match(disb_narrative_items, debit_bank_desc, exclude_book_ids=staged_disb_ids)
    narrative_coll_matches = _narrative_ref_match(coll_narrative_items, credit_bank_desc, exclude_book_ids=staged_coll_ids)

    result = []
    for i, tx in enumerate(txns):
        desc  = tx.get('description', '')
        utr   = tx.get('utr', '').strip().upper()
        dr, cr = tx['debit'], tx['credit']
        tx_type = tx_ref = tx_basis = tx_notes = ''

        own_acct_m = None
        if own_account_numbers:
            for num in own_account_numbers:
                if num in desc:
                    own_acct_m = num
                    break

        # Some banks label their own inter/intra-account transfers as
        # "CONTRA" directly in the narration (seen verbatim on a real
        # Saraswat statement: ".../CONTRA"). This is bank-provided ground
        # truth — more reliable than reconstructing "is this our own
        # transfer" from account numbers or narration shape, and it works
        # for any bank's own formatting without per-bank pattern-matching,
        # unlike the DR/CR-IFSC-BRIDGELINE-PARTNERS patterns below (those
        # are anchored to HDFC/IDFC's specific narration shapes and don't
        # generalize — a genuinely new bank's own truncation/format could
        # miss them entirely, as happened here).
        bank_says_contra = bool(re.search(r'\bcontra\b', desc, re.IGNORECASE))

        if dr > 0:
            if own_acct_m:
                tx_type = 'Contra'; tx_notes = f'Internal transfer to own account {own_acct_m}'
            elif bank_says_contra:
                tx_type = 'Contra'; tx_notes = 'Bank narration explicitly labels this Contra'
            # Own-name contra: a debit whose BENEFICIARY is BridgeLine itself
            # is money moving between our own accounts, never a customer
            # disbursement. Without this, the amount-matcher would grab the
            # nearest same-amount customer case and falsely mark it
            # bank-confirmed (happened with real Saraswat and HDFC→IDFC
            # transfers). Pattern anchored to the DR-IFSC-<beneficiary>
            # narration shape so customer names can never trigger it; still
            # lands in the review queue (basis '') for a human once-over.
            elif re.search(r'\bDR[-/][A-Z]{4}0[A-Z0-9]{6}[-/]\s*BRIDGELINE PARTNERS\b', desc, re.IGNORECASE):
                tx_type = 'Contra'; tx_notes = 'Own-account transfer (beneficiary is BridgeLine)'
            elif _is_fd_booking(desc):
                tx_type = 'FD Booking'; tx_notes = 'Fixed Deposit — internal capital movement'
            elif _is_expense_debit(desc, dr):
                tx_type  = 'Expense'
                tx_notes = _expense_category(desc)
            else:
                # 1. UTR match
                matched = utr_to_disb.get(utr)
                if not matched:
                    desc_utr = extract_utr(desc).upper()
                    matched  = utr_to_disb.get(desc_utr)
                # IFSC-in-narration: Debit Note may contain an IFSC (e.g. UBIN0905925)
                # that appears inside the bank narration (e.g. "NEFT DR-UBIN0905925-BASAVARAJ...")
                if not matched:
                    for r in records:
                        did2     = r.get('Disbursement ID','').strip()
                        dn2      = str(r.get('Debit Note','') or '').strip().upper()
                        ifsc_m   = re.match(r'^([A-Z]{4}0[A-Z0-9]{6})$', dn2)
                        if ifsc_m and ifsc_m.group(1) in desc.upper():
                            matched = did2
                            break

                if matched:
                    tx_type = 'Disbursement'; tx_ref = matched
                    tx_basis = 'UTR'; tx_notes = 'Accounts (Debit Note)'
                elif i in staged_disb_matches:
                    book_id, mtype = staged_disb_matches[i]
                    tx_type = 'Disbursement'; tx_ref = book_id
                    tx_basis = 'Amount+Date'
                    tx_notes = ('Accounts (Disb) — exact amount+date match' if mtype == 'exact'
                                else 'Accounts (Disb) — closest-date match within tolerance')
                elif i in narrative_disb_matches:
                    tx_type = 'Disbursement'; tx_ref = narrative_disb_matches[i]
                    tx_basis = 'Narrative'; tx_notes = 'Accounts (Debit Note) — masked-ref narrative match'
                else:
                    amt_cands = amt_to_disb_id.get(round(dr), [])
                    name_did, name_score = _fuzzy_name_match(desc)
                    # 2. Amount + name (strongest non-UTR match)
                    if amt_cands and name_did and name_did in amt_cands:
                        tx_type = 'Disbursement'; tx_ref = name_did
                        tx_basis = 'Amount+Name'; tx_notes = 'Accounts (Disb)'
                    # 3. Amount only
                    elif amt_cands:
                        tx_type = 'Disbursement'; tx_ref = amt_cands[0]
                        tx_basis = 'Amount'; tx_notes = 'Accounts (Disb)'
                    # 4. Name only (with decent score)
                    elif name_did and name_score >= 2:
                        tx_type = 'Disbursement'; tx_ref = name_did
                        tx_basis = 'Name'; tx_notes = 'Fuzzy name — confirm'
                    else:
                        tx_type = 'Expense'; tx_notes = _expense_category(desc)

        elif cr > 0:
            if own_acct_m:
                tx_type = 'Contra'; tx_notes = f'Internal transfer from own account {own_acct_m}'

            elif bank_says_contra:
                tx_type = 'Contra'; tx_notes = 'Bank narration explicitly labels this Contra'

            # ₹1 test credits — banks/borrowers send ₹1 to verify account before full payment
            elif cr == 1.0:
                tx_type = 'Test Credit'; tx_notes = 'Penny verification — ignore'

            # Own-name contra (credit side): a credit whose SENDER is
            # BridgeLine itself is our own money arriving from another of
            # our accounts, never a customer collection. Two narration
            # shapes, both with our name in the sender slot: HDFC's
            # 'RTGS CR-<IFSC>-BRIDGELINE PARTNERS-...' and IDFC's
            # 'RTGS/<UTR>/BRIDGELINE PARTNERS/...'. The FT collection format
            # ('FT -BRIDGELINE PARTNERS CR - ...') deliberately doesn't
            # match either pattern — there our name is the RECEIVING account
            # holder on a genuine customer payment.
            elif _looks_like_own_transfer(desc):
                tx_type = 'Contra'; tx_notes = 'Own-account transfer (sender is BridgeLine)'

            # Bank-paid interest (savings/FD interest credits) — income the
            # lending books never see, so it needs its own type to be added
            # back into the solvency check's expected balance.
            elif re.search(r'credit\s*int(?:erest)?|int(?:erest)?\s*(?:pd|paid|credit)|\bint\.?\s*pd\b|\bsb\s*int\b|fd\s*int', desc, re.IGNORECASE):
                tx_type = 'Interest Income'; tx_notes = 'Bank interest credit'

            # FD principal sweeping/maturing back into the account — typed
            # 'FD Booking' same as the outgoing booking so the net FD figure
            # (_all_time_recon_totals) self-corrects: these are sweep-in FDs
            # counted as available-to-disburse while parked, and must stop
            # being counted the moment the money is back in the account.
            elif re.search(r'\bfd\b.*(?:clos|matur|premat|redeem|sweep|liquidat)|(?:sweep|swp).*(?:trf|transfer|cr)|prin\s*\+\s*int|sweep\s*in\s*from', desc, re.IGNORECASE):
                tx_type = 'FD Booking'; tx_notes = 'FD maturity / sweep-in credit'

            # Pradaan routing = repayment routed via Pradaan account
            elif re.search(r'pradaan|pradan', desc, re.IGNORECASE):
                cands = amt_to_coll.get(round(cr), [])
                tx_type = 'Repayment'
                tx_ref  = cands[0] if cands else ''
                tx_basis= 'Amount' if cands else '—'
                tx_notes= 'Pradaan Routing'

            else:
                # Match to collection by UTR — check M Coll first, then Accounts Credit Note
                matched = utr_to_coll.get(utr)
                if not matched:
                    desc_utr = extract_utr(desc).upper()
                    matched  = utr_to_coll.get(desc_utr)
                if not matched:
                    matched = utr_to_coll_acct.get(utr)

                # FT (internal fund transfer) — match via HDFC account number in narration
                # e.g. "FT -BRIDGELINE PARTNERS CR - 50100551677276 - PRANAV T P"
                # The same account number appears in the debit narration for that disbursement
                if not matched and re.match(r'FT\s*[\-–]', desc, re.IGNORECASE):
                    acct_m = re.search(r'\b(5010\d{10})\b', desc)
                    if acct_m:
                        acct_num = acct_m.group(1)
                        # Find the debit transaction with this account number and match its Disb ID
                        for other in txns:
                            if other['debit'] > 0 and acct_num in other.get('description', ''):
                                cands = amt_to_disb_id.get(round(other['debit']), [])
                                if cands:
                                    matched = cands[0]
                                    break

                if matched:
                    tx_type = 'Repayment'; tx_ref = matched
                    tx_basis = 'UTR'; tx_notes = 'Matched Collections'
                elif i in staged_coll_matches:
                    book_id, mtype = staged_coll_matches[i]
                    tx_type = 'Repayment'; tx_ref = book_id
                    tx_basis = 'Amount+Date'
                    tx_notes = ('M Coll — exact amount+date match' if mtype == 'exact'
                                else 'M Coll — closest-date match within tolerance')
                elif i in narrative_coll_matches:
                    tx_type = 'Repayment'; tx_ref = narrative_coll_matches[i]
                    tx_basis = 'Narrative'; tx_notes = 'M Coll — masked-ref narrative match'
                else:
                    coll_cands = amt_to_coll.get(round(cr), [])
                    name_did, name_score = _fuzzy_name_match(desc)
                    # Amount + name
                    if coll_cands and name_did and name_did in coll_cands:
                        tx_type = 'Repayment'; tx_ref = name_did
                        tx_basis = 'Amount+Name'; tx_notes = 'M Coll + name'
                    # Amount only (M Coll)
                    elif coll_cands:
                        tx_type = 'Repayment'; tx_ref = coll_cands[0]
                        tx_basis = 'Amount'; tx_notes = 'Fuzzy — confirm'
                    # Amount match against Accounts disbursement amounts (repayment of full loan)
                    elif name_did and name_score >= 2:
                        # Check if amount also matches for higher confidence
                        disb_cands = amt_to_disb_id.get(round(cr), [])
                        if disb_cands and name_did in disb_cands:
                            tx_type = 'Repayment'; tx_ref = name_did
                            tx_basis = 'Amount+Name'; tx_notes = 'Name+amount match'
                        else:
                            tx_type = 'Repayment'; tx_ref = name_did
                            tx_basis = 'Name'; tx_notes = 'Fuzzy name — confirm'
                    else:
                        if cr < 5000 or re.search(r'ft\s+-\s+cr|capital|transfer from', desc, re.IGNORECASE):
                            tx_type = 'Other Income'; tx_notes = 'Unmatched credit — confirm source'
                        else:
                            tx_type = 'Repayment'; tx_ref = ''
                            tx_basis = '—'; tx_notes = 'Review — no match found'

        # Honour manual override from widget
        override = tx.get('type_override', '').strip()
        if override and override != 'Skip':
            tx_type  = override
            tx_basis = 'Manual'
        row_rem = tx.get('row_remarks', '').strip()
        if row_rem:
            tx_notes = f"{tx_notes} | {row_rem}" if tx_notes else row_rem

        result.append({**tx, 'type': tx_type, 'matched_ref': tx_ref,
                        'match_basis': tx_basis, 'match_notes': tx_notes})
    return result

def _bank_matched_refs(classified_txns):
    """Ref -> tx for every bank txn matched to a book record this period,
    split by direction. Single source of truth reused by the interactive
    preview (api_reconcile_parse) AND the xlsx builders (_sheet_disb_recon,
    _sheet_coll_recon) so the two views can never disagree on what counts
    as 'matched'."""
    disb_by_ref, coll_by_ref = {}, {}
    for tx in classified_txns:
        if tx['debit'] > 0 and tx['matched_ref']:
            disb_by_ref.setdefault(tx['matched_ref'], tx)
        if tx['credit'] > 0 and tx['matched_ref']:
            coll_by_ref.setdefault(tx['matched_ref'], tx)
    return disb_by_ref, coll_by_ref

def _ref_present_in_bank_text(ref_text, bank_descs_upper):
    """True if ref_text (a book-recorded Debit/Credit Note token) can be
    found in ANY bank transaction's narration this period, by exact
    substring or by the same trailing-digits+name narrative signature
    _narrative_ref_match uses. This is deliberately independent of
    tx['matched_ref'] -- a single bank transaction can only ever be
    matched_ref'd to ONE disb_id, but Prem's own books sometimes legitimately
    cite the SAME real payment as the reference for two different
    disbursements (confirmed real pattern, 31-Jul-2026 audit — e.g. one
    payment recorded against both Ashrath Banu's and P Shanawaz's cases).
    Without this, the second case would always show as a false "unmatched"
    no matter how good the underlying matching gets, since the model can
    only assign that transaction's matched_ref to one of the two."""
    ref_norm = ref_text.strip().upper()
    if not ref_norm:
        return False
    # An index prepared by _build_bank_ref_index() when the caller has many
    # refs to test. Same answers, but the per-description digit-stripping and
    # word-splitting happen once for the whole batch instead of once per ref.
    # A full month of two accounts is ~500 bank rows against ~400 book
    # entries; redoing that work per ref is 200,000 regex passes and can push
    # the request past its 60s limit.
    if isinstance(bank_descs_upper, dict):
        idx = bank_descs_upper
        if ref_norm in idx['blob']:
            return True
        sig = _narrative_signature(ref_text)
        if not sig:
            return False
        tail, names = sig
        for digits, words in idx['rows']:
            if tail in digits and (names & words):
                return True
        return False

    if any(ref_norm in d for d in bank_descs_upper):
        return True
    sig = _narrative_signature(ref_text)
    if not sig:
        return False
    tail, names = sig
    for d in bank_descs_upper:
        if tail in re.sub(r'\D', '', d) and (names & {w for w in re.split(r'[^A-Za-z]+', d) if len(w) >= 4}):
            return True
    return False

def _build_bank_ref_index(bank_descs_upper):
    """Precomputed form of the description list for _ref_present_in_bank_text.

    'blob' is every narration joined by a separator no reference can span, so
    one substring scan replaces a per-description loop. 'rows' caches the
    digits-only and long-word forms each narrative-signature comparison would
    otherwise recompute for every ref tested."""
    return {
        'blob': '\x00'.join(bank_descs_upper),
        'rows': [(re.sub(r'\D', '', d),
                  {w for w in re.split(r'[^A-Za-z]+', d) if len(w) >= 4})
                 for d in bank_descs_upper],
    }


def _book_completeness(records, mc_rows, classified_txns, period_month):
    """Book-side disbursements/collections dated in period_month with no
    matching bank transaction in classified_txns — the reverse-direction
    check the interactive preview never ran before (it only ever asked
    "does this bank line match a book record", never "does every book
    record this period have a matching bank line").

    Splits the result by WHY there's no bank evidence, because the two
    reasons could not be more different and only one is alarming:

      * pending  -- the entry is dated AFTER the last day any bank statement
                    covers. Today's business, banked and booked, simply not
                    reconciled yet. Completely normal, expected every day.
      * missing  -- the entry falls INSIDE the reconciled window and still
                    has no bank line. That is the real red flag.

    Reported as one undifferentiated "not found in bank" list until
    13-08-2026, when six collections and a disbursement all dated the same
    day the books were written showed up as a Rs 15.3 lakh alarm purely
    because that day's statement hadn't been uploaded yet. The check was
    right every time; the wording made correct output look like a crisis,
    which costs more trust than a missed entry would.
    """
    disb_by_ref, coll_by_ref = _bank_matched_refs(classified_txns)
    bank_descs_upper = [tx.get('description', '').upper() for tx in classified_txns]
    # Built once for the whole book scan below — see _build_bank_ref_index().
    bank_idx = _build_bank_ref_index(bank_descs_upper)

    # The last day ANY bank statement in play actually reaches. Derived from
    # transaction dates, not the period label — a statement saved under one
    # date routinely carries several days of rows.
    bank_dates = {d for d in (_parse_flex_date(tx.get('date', '')) for tx in classified_txns) if d}
    last_bank_date = max(bank_dates) if bank_dates else None

    def _bucket(date_str):
        d = _parse_flex_date(date_str)
        if last_bank_date is None or d is None:
            return 'missing'
        return 'pending' if d > last_bank_date else 'missing'

    unmatched_disb, pending_disb = [], []
    for r in records:
        did   = r.get('Disbursement ID', '')
        ddate = str(r.get('Disbursement Date', '') or '')
        if not did or not _date_in_period(ddate, period_month) or did in disb_by_ref:
            continue
        debit_note = str(r.get('Debit Note', '') or '').strip()
        if any(_ref_present_in_bank_text(tok, bank_idx) for tok in _split_ref_tokens(debit_note)):
            continue
        entry = {'disb_id': did, 'date': ddate,
                 'customer': r.get('Customer Name', ''),
                 'amount': _to_num(r.get('Amount', 0)),
                 'utr': debit_note}
        (pending_disb if _bucket(ddate) == 'pending' else unmatched_disb).append(entry)

    unmatched_coll, pending_coll = [], []
    for mc in mc_rows:
        did   = (mc[0] if len(mc) > 0 else '').strip()
        pdate = mc[1] if len(mc) > 1 else ''
        if not did or not _date_in_period(str(pdate), period_month) or did in coll_by_ref:
            continue
        credit_note = mc[3] if len(mc) > 3 else ''
        if any(_ref_present_in_bank_text(tok, bank_idx) for tok in _split_ref_tokens(credit_note)):
            continue
        entry = {'disb_id': did, 'date': pdate,
                 'customer': mc[4] if len(mc) > 4 else '',
                 'amount': _clean_amount(mc[2]) if len(mc) > 2 else 0,
                 'utr_or_note': credit_note}
        (pending_coll if _bucket(str(pdate)) == 'pending' else unmatched_coll).append(entry)

    return {
        'reconciled_through': last_bank_date.strftime('%d-%m-%Y') if last_bank_date else None,
        'unmatched_disbursements': unmatched_disb,
        'unmatched_disbursements_count': len(unmatched_disb),
        'unmatched_disbursements_total': round(sum(d['amount'] for d in unmatched_disb), 2),
        'unmatched_collections': unmatched_coll,
        'unmatched_collections_count': len(unmatched_coll),
        'unmatched_collections_total': round(sum(c['amount'] for c in unmatched_coll), 2),
        # Dated past the reconciled window — awaiting a statement, not missing.
        'pending_disbursements': pending_disb,
        'pending_disbursements_count': len(pending_disb),
        'pending_disbursements_total': round(sum(d['amount'] for d in pending_disb), 2),
        'pending_collections': pending_coll,
        'pending_collections_count': len(pending_coll),
        'pending_collections_total': round(sum(c['amount'] for c in pending_coll), 2),
    }

def _account_day_gaps(periods, upto=None):
    """Days each account has NO bank transaction while another account does.

    A day genuinely without activity looks identical to a day never
    uploaded — unless a different account moved money that day, which makes
    it a business day and the silence suspicious. That comparison is what
    isolates a real hole: on 13-08-2026 it showed IDFC had no rows for
    02-Aug or 11-Aug while HDFC had both, and the 11-Aug one was exactly
    why a real disbursement (BLP-110826-398, an IDFC UTR) had no bank
    evidence. Reported per account so the fix is obvious: re-upload that
    account's statement for those days.
    """
    by_acct = {}
    for p in periods:
        for tx in p['txns']:
            d = _parse_flex_date(tx.get('date', ''))
            if d and (upto is None or d <= upto):
                by_acct.setdefault(p['account'], set()).add(d)
    if len(by_acct) < 2:
        return {}
    all_days = set().union(*by_acct.values())
    gaps = {}
    for acct, days in by_acct.items():
        if not days:
            continue
        # Only inside the window this account actually covers — before its
        # first day or after its last is coverage, not a gap.
        lo, hi = min(days), max(days)
        missing = sorted(d for d in all_days if lo <= d <= hi and d not in days)
        if missing:
            gaps[acct] = [d.strftime('%d-%m-%Y') for d in missing]
    return gaps

# ── Sheet 1: Statement ────────────────────────────────────────────────────────

def _sheet_statement(wb, period_label, remarks, opening, closing, classified_txns, s, tieout=None):
    sname = 'Statement'
    if sname not in wb.sheetnames:
        ws = wb.create_sheet(sname)
        _title(ws, f'BridgeLine Partners — {remarks or "HDFC Bank A/c"} | {period_label} Reconciliation', 10, s)
        _legend(ws, 'Capital In | Disbursement | Collection | Expense  —  Type auto-classified. Update Matched Book Ref / Match Basis manually where needed.', 10, s)
        cols = [('Date',12),('Narration',50),('UTR / Ref No.',24),
                ('Withdrawal (₹)',16),('Deposit (₹)',16),('Closing Bal (₹)',16),
                ('Type',24),('Matched Book Ref',18),('Match Basis',14),('Match Source / Notes',28)]
        _hdr(ws, 3, cols, s)
        next_row = 4
    else:
        ws = wb[sname]
        next_row = ws.max_row + 2

    # tieout only ever passed for the period currently being saved (not the
    # historical periods this function also gets called for while rebuilding
    # the full workbook) — recomputing "book balance as of that past date"
    # for every prior period isn't meaningful the same way "as of right now"
    # is, so this stays a current-save-only annotation.
    tieout_suffix = ''
    if tieout:
        if tieout.get('status') == 'ok':
            tieout_suffix = (f"  |  Book Tie-out: {'✓' if tieout['ok'] else '✗'} "
                              f"(Δ₹{abs(tieout['variance']):,.2f})")
        elif tieout.get('status') == 'incomplete':
            tieout_suffix = f"  |  Book Tie-out: Incomplete ({tieout['untagged_count']} untagged)"

    ws.merge_cells(f'A{next_row}:J{next_row}')
    c = ws.cell(next_row, 1,
        f"▶  {period_label}  —  Opening: ₹{opening:,.2f}  |  Closing: ₹{closing:,.2f}  |  {remarks or ''}{tieout_suffix}")
    c.fill = s['hdr_fill']; c.font = Font(bold=True, size=9, name='Arial', color='FFFFFF')
    c.alignment = Alignment(horizontal='left')
    next_row += 1

    type_fill = {
        'Disbursement':           s['amb_fill'],
        'Repayment':              s['grn_fill'],
        'Other Income':           s['grn_fill'],
        'Interest Income':        s['grn_fill'],
        'Expense':                s['red_fill'],
        'Contra':                 s['blu_fill'],
        'FD Booking':             s['blu_fill'],
        # Legacy labels — older stored periods still carry these type
        # strings and get rebuilt into every downloaded workbook.
        'Collection':             s['grn_fill'],
        'Collection (via Pradaan)':s['grn_fill'],
        'Capital In':             s['blu_fill'],
        'Capital Out':            s['blu_fill'],
    }
    total_dr = total_cr = 0.0
    for tx in classified_txns:
        total_dr += tx['debit']; total_cr += tx['credit']
        fill = type_fill.get(tx['type'], PatternFill("solid", fgColor="FFFFFF"))
        for ci, val in enumerate([
            tx['date'], tx['description'], tx.get('utr',''),
            _n(tx['debit']), _n(tx['credit']), _n(tx['balance']),
            tx['type'], tx['matched_ref'], tx['match_basis'], tx['match_notes']
        ], 1):
            c = ws.cell(next_row, ci, val)
            c.fill = fill; c.font = s['norm_font']; c.border = s['thin_border']
            if ci in (4,5,6): c.alignment = Alignment(horizontal='right')
        next_row += 1

    for ci, val in enumerate(['','TOTALS','',_n(total_dr),_n(total_cr),'','','','',''], 1):
        c = ws.cell(next_row, ci, val)
        c.fill = s['tot_fill']; c.font = s['tot_font']; c.border = s['thin_border']
        if ci in (4,5): c.alignment = Alignment(horizontal='right')

    ws.freeze_panes = 'A4'
    return total_dr, total_cr

# ── Sheet 2: Disbursement Recon ───────────────────────────────────────────────

def _sheet_disb_recon(wb, period_label, records, classified_txns, s, period_month, remarks_map=None):
    """Only disbursements from the current period month."""
    sname = 'Disbursement Recon'
    cols = [('Disb ID',18),('Disb Date',14),('Customer',26),('Book Amount (₹)',16),
            ('Charges (₹)',12),('GST (₹)',10),('Bank Date',12),('Bank Withdrawal (₹)',18),
            ('Disb UTR',26),('Match Basis',14),('Status',18),('Remarks',34)]
    ncols = len(cols)
    if sname not in wb.sheetnames:
        ws = wb.create_sheet(sname)
        _title(ws, 'Disbursement Reconciliation — Books vs Bank Debits', ncols, s)
        _legend(ws, 'Green = UTR matched  |  Amber = amount/date matched  |  Red = not found in bank', ncols, s)
        _hdr(ws, 3, cols, s)
        row = 4
    else:
        ws = wb[sname]
        row = ws.max_row + 2

    # Period separator
    ws.merge_cells(f'A{row}:{get_column_letter(ncols)}{row}')
    c = ws.cell(row, 1, f'▶  {period_label}')
    c.fill = s['hdr_fill']; c.font = Font(bold=True, size=9, name='Arial', color='FFFFFF')
    c.alignment = Alignment(horizontal='left')
    row += 1

    bank_by_ref, _ = _bank_matched_refs(classified_txns)

    for r in records:
        ddate = str(r.get('Disbursement Date','') or '')
        did   = r.get('Disbursement ID','')
        # Only show this period's disbursements (or any that matched in bank this period)
        if not _date_in_period(ddate, period_month) and did not in bank_by_ref:
            continue

        amt     = _to_num(r.get('Amount', 0))
        charges = _to_num(r.get('Charges', 0) or r.get('Processing Charges', 0))
        gst     = _to_num(r.get('GST(18%)', 0) or r.get('GST', 0))
        utr     = str(r.get('Debit Note','') or '').strip()
        bank    = bank_by_ref.get(did)
        if bank:
            basis  = bank['match_basis']
            bdelta = abs(amt - bank['debit'])
            status = 'Matched' if bdelta <= 1 else f'Matched (Δ₹{bdelta:,.0f})'
            fill   = s['grn_fill'] if basis == 'UTR' else s['amb_fill']
            bdate  = bank['date']; bamt = _n(bank['debit'])
        else:
            basis = 'Not found'; status = 'Not found in bank'; fill = s['red_fill']
            bdate = bamt = ''

        remark = (remarks_map or {}).get(did, '')
        _row(ws, row, [did, ddate, r.get('Customer Name',''),
                        _n(amt), _n(charges), _n(gst),
                        bdate, bamt, utr, basis, status, remark],
             s, right_cols=(4,5,6,8), fill=fill)
        row += 1

    ws.freeze_panes = 'A4'


# ── Sheet 3: Collection Recon ─────────────────────────────────────────────────

def _sheet_coll_recon(wb, period_label, records, mc_rows, classified_txns, s, period_month, remarks_map=None):
    sname = 'Collection Recon'
    cols = [('Disb ID',18),('Customer',24),('Branch',14),('Disbursed (₹)',16),
            ('Status',10),('Pay Date',12),('Payment (₹)',16),('Collection UTR',24),
            ('Narration (short)',36),('Collected (₹)',16),('Outstanding (₹)',16),('Remarks',34)]
    if sname not in wb.sheetnames:
        ws = wb.create_sheet(sname)
        _title(ws, f'Collection Reconciliation — Payment level', 12, s)
        _legend(ws, 'Green = Closed  |  Amber = Partial.  One row per payment.', 12, s)
        _hdr(ws, 3, cols, s)
        row = 4
    else:
        ws = wb[sname]
        row = ws.max_row + 2

    # Period separator
    ws.merge_cells(f'A{row}:L{row}')
    c = ws.cell(row, 1, f'\u25b6  {period_label}')
    c.fill = s['hdr_fill']; c.font = Font(bold=True, size=9, name='Arial', color='FFFFFF')
    c.alignment = Alignment(horizontal='left')
    row += 1

    # Bank credits indexed by UTR for narration lookup; also by amount for
    # backfilling a UTR when the M Coll note has none embedded.
    bank_cr_by_utr = {}
    bank_cr_by_amt = {}
    for tx in classified_txns:
        if tx['credit'] > 0:
            if tx.get('utr'):
                bank_cr_by_utr[tx['utr'].upper()] = tx
            bank_cr_by_amt.setdefault(round(tx['credit']), []).append(tx)

    # M Coll grouped by disb_id — filter to payments made this period
    mc_map = {}
    for mc in mc_rows:
        did  = (mc[0] if len(mc)>0 else '').strip()
        pdate= (mc[1] if len(mc)>1 else '')
        if did and (not period_month or _date_in_period(str(pdate), period_month)):
            mc_map.setdefault(did, []).append(mc)

    # Also include any disb that has a bank match this period even without M Coll entry
    _, coll_by_ref = _bank_matched_refs(classified_txns)
    coll_disb_ids = set(mc_map.keys()) | set(coll_by_ref.keys())

    for r in records:
        did = r.get('Disbursement ID','')
        if did not in coll_disb_ids:
            continue

        amt, total, collected, balance = _parse_case(r)
        status   = r.get('Overdue Status','').strip()
        sfill    = s['grn_fill'] if status == 'Closed' else s['amb_fill']
        payments = mc_map.get(did, [])

        if not payments:
            # Matched via bank credit only — add a single row
            bank_tx = next((tx for tx in classified_txns
                            if tx['credit'] > 0 and tx['matched_ref'] == did), None)
            if bank_tx:
                _row(ws, row, [
                    did, r.get('Customer Name',''), r.get('Branch',''),
                    _n(total), status,
                    bank_tx['date'], _n(bank_tx['credit']), bank_tx.get('utr',''),
                    bank_tx['description'][:50], _n(collected), _n(balance),
                    (remarks_map or {}).get(did, '')
                ], s, right_cols=(4,7,10,11), fill=sfill)
                row += 1
            continue

        first = True
        for mc in payments:
            # M Coll col D holds `utr or raw_msg` (see save_repayment) — pull
            # the UTR out of SMS text when that's what landed there, and if
            # there's still none, backfill from a unique same-amount bank
            # credit in this period.
            mc_raw  = str(mc[3] if len(mc)>3 else '').strip()
            mc_amt  = _clean_amount(mc[2]) if len(mc)>2 else 0
            mc_utr  = extract_utr(mc_raw)
            if not mc_utr and re.fullmatch(r'[A-Za-z0-9\-/]{6,25}', mc_raw or ''):
                mc_utr = mc_raw  # already a bare ref, just not SMS-shaped
            bank_tx = bank_cr_by_utr.get(mc_utr.upper()) if mc_utr else None
            if not mc_utr:
                amt_matches = bank_cr_by_amt.get(round(mc_amt), [])
                if len(amt_matches) == 1:
                    bank_tx = amt_matches[0]
                    mc_utr  = bank_tx.get('utr', '')
            # Parity with _sheet_disb_recon's red 'Not found in bank' flag —
            # a book-recorded payment with no matching bank credit at all
            # used to just render with a blank narration and the book's
            # own Closed/Partial color, with no signal that it was never
            # actually confirmed against the bank statement.
            narr      = bank_tx['description'][:50] + '…' if bank_tx else 'Not found in bank'
            row_fill  = sfill if bank_tx else s['red_fill']

            _row(ws, row, [
                did if first else '',
                r.get('Customer Name','') if first else '',
                r.get('Branch','') if first else '',
                _n(total) if first else '',
                status if first else '',
                mc[1] if len(mc)>1 else '', _n(mc_amt), mc_utr, narr,
                _n(collected) if first else '',
                _n(balance) if first else '',
                (remarks_map or {}).get(did, '') if first else '',
            ], s, right_cols=(4,7,10,11), fill=row_fill)
            first = False; row += 1

    ws.freeze_panes = 'A4'

# ── Sheet 4: Expenses ─────────────────────────────────────────────────────────

def _sheet_expenses(wb, period_label, classified_txns, s):
    sname = 'Expenses'
    cols = [('Date',12),('Narration',54),('UTR / Ref No.',24),('Amount (₹)',16),('Category',28)]
    if sname not in wb.sheetnames:
        ws = wb.create_sheet(sname)
        _title(ws, f'Operational Expenses', 5, s)
        _legend(ws, 'Auto-extracted — POS, subscriptions, bank charges, salaries, partner share.', 5, s)
        _hdr(ws, 3, cols, s)
        row = 4
    else:
        ws = wb[sname]
        row = ws.max_row + 2

    # Period separator
    ws.merge_cells(f'A{row}:E{row}')
    c = ws.cell(row, 1, f'\u25b6  {period_label}')
    c.fill = s['hdr_fill']; c.font = Font(bold=True, size=9, name='Arial', color='FFFFFF')
    c.alignment = Alignment(horizontal='left')
    row += 1

    expenses = [tx for tx in classified_txns if tx['type'] == 'Expense' and tx['debit'] > 0]
    total = 0.0
    for tx in expenses:
        total += tx['debit']
        _row(ws, row, [
            tx['date'], tx['description'],
            tx.get('utr','') or extract_utr(tx['description']),
            _n(tx['debit']), tx['match_notes']
        ], s, right_cols=(4,))
        row += 1

    for ci, val in enumerate(['','TOTAL','',_n(total),''], 1):
        c = ws.cell(row, ci, val)
        c.fill = s['tot_fill']; c.font = s['tot_font']; c.border = s['thin_border']
        if ci == 4: c.alignment = Alignment(horizontal='right')

    ws.freeze_panes = 'A4'
    return expenses, total

# ── Sheet 5: Mapped (Disb ID ↔ Disb UTR ↔ Collection UTR) ───────────────

def _sheet_mapped(wb, period_label, records, mc_rows, classified_txns, s, period_month=None):
    """Only cases with activity in the period month: disbursed in-month, an
    M Coll payment in-month, or a bank debit/credit match in this period's
    statement. Without this filter every case ever (incl. archive tabs)
    repeats in each period section."""
    sname = 'Mapped'
    cols = [('Disb ID',18),('Customer',26),('Disb Date',12),('Disb Amount (₹)',16),
            ('Disb UTR (Bank)',26),('Coll Date',12),('Coll Amount (₹)',16),
            ('Collection UTR',26),('Match Basis',14),('Status',18)]
    if sname not in wb.sheetnames:
        ws = wb.create_sheet(sname)
        _title(ws, 'UTR Mapping — Disbursements & Collections', 10, s)
        _legend(ws, 'One row per payment event. Disb UTR from bank debits; Coll UTR from bank credits.', 10, s)
        _hdr(ws, 3, cols, s)
        row = 4
    else:
        ws = wb[sname]
        row = ws.max_row + 2

    ws.merge_cells(f'A{row}:J{row}')
    c = ws.cell(row, 1, f'▶  {period_label}')
    c.fill = s['hdr_fill']; c.font = Font(bold=True, size=9, name='Arial', color='FFFFFF')
    c.alignment = Alignment(horizontal='left')
    row += 1

    bank_debit_by_ref  = {}
    bank_credit_by_utr = {}
    bank_credit_refs   = set()
    bank_cr_by_amt     = {}
    for tx in classified_txns:
        if tx['debit'] > 0 and tx['matched_ref']:
            bank_debit_by_ref[tx['matched_ref']] = tx
        if tx['credit'] > 0:
            if tx.get('utr'):
                bank_credit_by_utr[tx['utr'].upper()] = tx
            if tx['matched_ref']:
                bank_credit_refs.add(tx['matched_ref'])
            bank_cr_by_amt.setdefault(round(tx['credit']), []).append(tx)

    # Payments filtered to the period month — pre-period instalments don't
    # belong in this period's mapping.
    mc_map = {}
    for mc in mc_rows:
        did   = (mc[0] if len(mc) > 0 else '').strip()
        pdate = str(mc[1] if len(mc) > 1 else '')
        if did and (not period_month or _date_in_period(pdate, period_month)):
            mc_map.setdefault(did, []).append(mc)

    for r in records:
        did   = r.get('Disbursement ID', '')
        ddate = str(r.get('Disbursement Date', '') or '')
        # Strictly this period: disbursed in-month, paid in-month, or matched
        # against this period's bank statement.
        if period_month and not (_date_in_period(ddate, period_month)
                                 or did in mc_map
                                 or did in bank_debit_by_ref
                                 or did in bank_credit_refs):
            continue
        amt   = _to_num(r.get('Amount', 0))
        books_utr = str(r.get('Debit Note', '') or '').strip()

        bank_dr   = bank_debit_by_ref.get(did)
        bank_utr  = bank_dr.get('utr', '') if bank_dr else books_utr
        bank_ddate = bank_dr['date'] if bank_dr else ddate
        dr_basis  = bank_dr['match_basis'] if bank_dr else ('Books' if books_utr else 'Not found')
        dr_fill   = s['grn_fill'] if bank_dr and bank_dr['match_basis'] == 'UTR' else (
                    s['amb_fill'] if bank_dr else s['red_fill'])

        payments = mc_map.get(did, [])
        if not payments:
            _row(ws, row, [did, r.get('Customer Name',''), bank_ddate, _n(amt),
                           bank_utr or books_utr, '', '', '', dr_basis,
                           r.get('Overdue Status','')], s, right_cols=(4,7,8), fill=dr_fill)
            row += 1
        else:
            first = True
            for mc in payments:
                # Same UTR handling as Collection Recon: col D may hold raw
                # SMS text — extract, else backfill from a unique same-amount
                # bank credit.
                mc_raw  = str(mc[3] if len(mc) > 3 else '').strip()
                mc_amt  = _clean_amount(mc[2]) if len(mc) > 2 else 0
                mc_date = mc[1] if len(mc) > 1 else ''
                mc_utr  = extract_utr(mc_raw)
                if not mc_utr and re.fullmatch(r'[A-Za-z0-9\-/]{6,25}', mc_raw or ''):
                    mc_utr = mc_raw  # already a bare ref, just not SMS-shaped
                bank_cr = bank_credit_by_utr.get(mc_utr.upper()) if mc_utr else None
                if not mc_utr:
                    amt_matches = bank_cr_by_amt.get(round(mc_amt), [])
                    if len(amt_matches) == 1:
                        bank_cr = amt_matches[0]
                        mc_utr  = bank_cr.get('utr', '')
                row_fill = (dr_fill if first else (s['grn_fill'] if bank_cr else s['amb_fill']))
                _row(ws, row, [
                    did if first else '',
                    r.get('Customer Name','') if first else '',
                    bank_ddate if first else '',
                    _n(amt) if first else '',
                    (bank_utr or books_utr) if first else '',
                    mc_date, _n(mc_amt), mc_utr,
                    dr_basis if first else '',
                    r.get('Overdue Status','') if first else '',
                ], s, right_cols=(4,7,8), fill=row_fill)
                first = False; row += 1

    ws.freeze_panes = 'A4'

# ── Main save function ────────────────────────────────────────────────────────

RECON_TXNS_SHEET_NAME = "Recon Txns"
RECON_TXNS_HEADERS = [
    "Batch ID", "Recon Date", "Account", "Period Label", "Opening", "Closing",
    "Remarks", "Txn Date", "Description", "UTR", "Debit", "Credit", "Balance",
    "Type", "Matched Ref", "Match Basis", "Match Notes"
]

def get_recon_txns_sheet(sh):
    try:
        return sh.worksheet(RECON_TXNS_SHEET_NAME)
    except Exception:
        ws = sh.add_worksheet(title=RECON_TXNS_SHEET_NAME, rows=5000, cols=len(RECON_TXNS_HEADERS))
        ws.append_row(RECON_TXNS_HEADERS)
        return ws

# ── Capital Log (company solvency check) ─────────────────────────────────────

CAPITAL_LOG_SHEET_NAME = "Capital Log"
CAPITAL_LOG_HEADERS = ["DATE", "TYPE", "PARTNER", "AMOUNT", "RUNNING BALANCE", "REFERENCE", "REMARKS"]

def get_capital_log_sheet(sh):
    """The real tab already exists with a title row + blank row + header at
    row 3 (manually created) — this create-branch only fires defensively if
    the tab is ever deleted, and deliberately doesn't try to reproduce that
    formatting, just a plain header row."""
    try:
        return sh.worksheet(CAPITAL_LOG_SHEET_NAME)
    except Exception:
        ws = sh.add_worksheet(title=CAPITAL_LOG_SHEET_NAME, rows=1000, cols=len(CAPITAL_LOG_HEADERS))
        ws.append_row(CAPITAL_LOG_HEADERS)
        return ws

def read_capital_log(sh):
    """Live-Sheets wrapper around generate_mis.load_capital_log() — same
    single-sheet shim pattern used elsewhere to reuse a wb-agnostic parser
    against gspread data instead of an openpyxl Workbook."""
    try:
        vals = sh.worksheet(CAPITAL_LOG_SHEET_NAME).get_all_values()
    except gspread.exceptions.WorksheetNotFound:
        return None
    wb_shim = _WorkbookShim({'Capital Log': _SheetShim(vals)})
    return mis.load_capital_log(wb_shim)

def append_capital_log_entry(data):
    """Single-row append — mirrors save_reconciliation()'s Recon Log
    log_ws.append_row([...]) pattern (not the Contacts full-sheet-rewrite
    pattern), since Capital Log is meant to be an append-only audit trail
    (the sheet's own note: 'NOT overwritten by refreshAllDashboards() —
    safe to edit')."""
    sh = get_gspread_client().open_by_key(SPREADSHEET_ID)
    ws = get_capital_log_sheet(sh)
    before = read_capital_log(sh)
    net_before = before['net_capital'] if before and before.get('available') else 0.0

    type_ = data['type'].strip().upper()
    partner = data.get('partner', '').strip()
    amount = mis.clean_capital_amount(data.get('amount', 0))
    delta = mis.capital_log_signed_delta(type_, amount)
    new_balance = round(net_before + delta, 2)

    try:
        d = datetime.strptime(data['date'], '%d-%m-%Y')
    except Exception:
        d = datetime.today()
    ws.append_row([
        d.strftime('%d-%b-%Y'), type_, partner,
        f"Rs {mis.inr(amount)}", f"Rs {mis.inr(new_balance)}",
        data.get('reference', '').strip(), data.get('remarks', '').strip(),
    ])
    return new_balance

def _recon_num(v):
    try:
        return float(str(v).replace(',', '').strip() or 0)
    except Exception:
        return 0.0

def _load_recon_periods(ws):
    """Read the Recon Txns tab and return the stored periods sorted
    chronologically. Where the same (recon_date, account) was saved more than
    once, only the latest Batch ID wins — re-running a day's recon replaces
    that period instead of duplicating it."""
    all_vals = ws.get_all_values()
    # batches[(recon_date, account)] = {batch_id, meta, txns}
    batches = {}
    for row in all_vals[1:]:
        if len(row) < 14 or not row[0]:
            continue
        key = (row[1], row[2])
        batch_id = row[0]
        if key not in batches or batch_id > batches[key]['batch_id']:
            batches[key] = {'batch_id': batch_id,
                            'recon_date': row[1], 'account': row[2],
                            'period_label': row[3],
                            'opening': _recon_num(row[4]), 'closing': _recon_num(row[5]),
                            'remarks': row[6], 'txns': []}
        # Placeholder rows (period saved with zero transactions) carry the
        # metadata only — don't surface them as ledger lines.
        if batches[key]['batch_id'] == batch_id and (row[8] or _recon_num(row[10]) or _recon_num(row[11])):
            batches[key]['txns'].append({
                'date':        row[7],
                'description': row[8],
                'utr':         row[9],
                'debit':       _recon_num(row[10]),
                'credit':      _recon_num(row[11]),
                'balance':     _recon_num(row[12]),
                'type':        row[13],
                'matched_ref': row[14] if len(row) > 14 else '',
                'match_basis': row[15] if len(row) > 15 else '',
                'match_notes': row[16] if len(row) > 16 else '',
            })

    # A period whose transaction dates fall entirely inside a LATER-SAVED
    # period for the same account is superseded by it, and must not be read
    # alongside it.
    #
    # "Latest Batch ID wins" above only replaces a period saved under the
    # exact same (recon_date, account). Uploading a whole month's statement
    # creates a NEW key — so every day already reconciled inside that month
    # would be counted twice: twice in the downloaded workbook, twice in the
    # per-account tie-out, twice everywhere that doesn't happen to carry its
    # own content-key dedup. Four aggregate functions do carry that dedup and
    # would have survived; the workbook and the tie-out would not.
    #
    # Fixed at the read layer on purpose: nothing is deleted, a wider upload
    # simply outranks the narrower ones it contains, and re-uploading a
    # single day afterwards (a later batch, narrower range) is still honoured
    # for the days it covers because containment is checked both ways.
    def _range(p):
        ds = [d for d in (_parse_flex_date(t.get('date', '')) for t in p['txns']) if d]
        return (min(ds), max(ds)) if ds else None

    ranged = [(p, _range(p)) for p in batches.values()]
    superseded = set()
    for i, (p, rp) in enumerate(ranged):
        if rp is None:
            continue
        for j, (q, rq) in enumerate(ranged):
            if i == j or rq is None or q['account'] != p['account']:
                continue
            # Day-set containment, not span containment: a wider period only
            # supersedes a narrower one when it actually carries every day the
            # narrower one has. A span test hides real days the wider upload
            # was silent about (13-08-2026: three FD bookings vanished).
            qd = {d for d in (_parse_flex_date(t.get('date', '')) for t in q['txns']) if d}
            pd = {d for d in (_parse_flex_date(t.get('date', '')) for t in p['txns']) if d}
            if not pd <= qd:
                continue
            covers = True
            # Identical ranges: only the later batch survives. Strict
            # containment: the wider one wins regardless of which is newer,
            # since it is by definition the more complete record.
            same = pd == qd
            if (same and q['batch_id'] > p['batch_id']) or (not same):
                superseded.add(id(p))
                break

    kept = [p for p in batches.values() if id(p) not in superseded]

    def _period_sort_key(p):
        try:
            return datetime.strptime(p['recon_date'], '%d-%m-%Y')
        except Exception:
            return datetime.min
    return sorted(kept, key=_period_sort_key)

def _all_time_recon_totals(periods=None):
    """All-time Expense/FD totals across every saved reconciliation period
    and account. No new sheet-scan needed — _load_recon_periods() already
    reads the ENTIRE Recon Txns sheet with latest-Batch-ID de-duplication
    built in; periods=None triggers a fresh load for standalone use,
    otherwise reuses an already-loaded list to avoid a second Sheets
    round-trip within one request (see _solvency_check()).

    _load_recon_periods()'s "latest Batch ID wins" only replaces a period
    sharing the exact same (recon_date, account) key -- it does NOT detect
    when a LATER period's own date range overlaps or fully subsumes an
    EARLIER one under a different recon_date (confirmed real, 01-08-2026:
    July was reconciled incrementally in small date-specific batches
    throughout the month, e.g. recon_date 20-07-2026, 21-07-2026, ... each
    covering only its own few-day slice; a subsequent full-month save under
    31-07-2026 then sat ALONGSIDE those instead of replacing them, since its
    key differs, and every transaction in the overlap got summed twice --
    inflated fd_total by ~15.7L and expense_total by ~60k in exactly this
    way). Fixed by deduplicating individual transactions per account on a
    content key (date+description+debit+credit+balance -- the same real
    bank ledger line re-saved under a different period will always match
    all five; two genuinely different transactions effectively never will),
    processing periods NEWEST-first so a transaction's most recently-saved
    classification wins if it was ever re-typed/corrected on a later save."""
    if periods is None:
        sh = get_gspread_client().open_by_key(SPREADSHEET_ID)
        periods = _load_recon_periods(get_recon_txns_sheet(sh))
    def _recon_date_key(p):
        try:
            return datetime.strptime(p['recon_date'], '%d-%m-%Y')
        except Exception:
            return datetime.min

    expense_total = fd_total = income_total = 0.0
    periods_seen = []
    seen_txn_keys = set()
    for p in sorted(periods, key=_recon_date_key, reverse=True):
        periods_seen.append((p['recon_date'], p['account']))
        for tx in p['txns']:
            key = (p['account'], tx['date'], tx['description'], tx['debit'], tx['credit'], tx['balance'])
            if key in seen_txn_keys:
                continue
            seen_txn_keys.add(key)
            # Money-out types beyond the literal 'Expense' label must count
            # here too, or they vanish from every downstream figure. Found
            # live 12-08-2026: 'Partner Salary' (₹1,10,000) and
            # 'Employee Salary' (₹60,000) are custom types invented at
            # reconciliation time, carried real money out of the bank, and
            # were counted NOWHERE — reading as unexplained missing money.
            # Anything with no known cash role is treated by DIRECTION (a net
            # debit is money out) so an unrecognised type can never again
            # silently swallow cash; _money_integrity_check() reports it by
            # name so it gets a proper role.
            _role = CASH_ROLE.get(_effective_type(tx))
            if _role in ('expense', 'withdrawal'):
                expense_total += tx['debit'] - tx['credit']
            elif _role is None and tx['debit'] > tx['credit']:
                expense_total += tx['debit'] - tx['credit']
            # 'FD Booking' is a NET figure: debits are money parked into an
            # FD, credits are the FD sweeping/maturing back into the account
            # — so fd_total is always the amount CURRENTLY sitting in FDs,
            # and a matured FD can't be double-counted (once in the bank
            # balance and again as a parked FD).
            elif tx['type'] == 'FD Booking':
                fd_total += tx['debit'] - tx['credit']
            # Non-lending income arriving in the bank (interest credits,
            # misc receipts) — cash the lending books never see, so it must
            # be added back into the expected balance or it reads as an
            # unexplained surplus.
            elif tx['type'] in ('Interest Income', 'Other Income'):
                income_total += tx['credit']
    return {'expense_total': round(expense_total, 2), 'fd_total': round(fd_total, 2),
            'income_total': round(income_total, 2),
            'periods_seen': periods_seen}

# ── Per-FD-account sweep ledger ───────────────────────────────────────────────
# Both banks put the FD's OWN account number in the narration, on BOTH legs:
#   IDFC out : 'Sweepout FD 10297960466 booked'
#   IDFC in  : 'Auto sweep in from 10297960466/Principal:444734/ BLKRTGS/...'
#   HDFC out : 'FD BOOKED - 50301379044082 : BRIDGELINE PARTNERS'
#   HDFC in  : 'SWEEP-IN CREDIT - 50301382223321'
# so every booking can be matched to its OWN sweep-ins instead of blindly
# netting one big total. That distinction is not cosmetic -- it is the whole
# difference between a right and a wrong number, proven against real data on
# 11-08-2026:
#
#   old blind net sum ............ 38,91,266
#   per-FD ledger ................ 41,05,794   <- matches the real FD balance
#                                                 to the rupee
#
# The 2,14,528 the old approach lost breaks down as:
#   * 2,13,765 -- HDFC FD 50301355033761 swept back in on 20-Jul, but was
#     BOOKED before reconciliation history begins (HDFC data starts 01-Jul).
#     A blind net sum subtracts money returning from an FD it never saw
#     created, and can never recover it.
#   * 763 -- FD interest. Three IDFC FDs swept back slightly MORE than their
#     principal (206 + 233 + 324). A blind sum lets that excess push the FD
#     balance below zero.
# Both are fixed by the same rule: track each FD separately and clamp it at
# zero -- a fully-swept FD holds nothing, and the excess that came back IS
# interest, which is already sitting in (and counted via) the bank balance.
FD_OUT_PATTERNS = (
    re.compile(r'sweepout\s+fd\s+(\d+)\s+booked', re.IGNORECASE),
    re.compile(r'fd\s+booked\s*-\s*(\d+)', re.IGNORECASE),
)
FD_IN_PATTERNS = (
    re.compile(r'auto\s+sweep\s+in\s+from\s+(\d+)\s*/\s*principal\s*:\s*([\d.]+)', re.IGNORECASE),
    re.compile(r'sweep[-\s]*in\s+credit\s*-\s*(\d+)', re.IGNORECASE),
)

def _fd_ledger(periods=None):
    """Money currently parked in FDs, derived per FD account number.

    Returns total plus the diagnostics that make it self-validating:
      * orphan_sweepins  -- a sweep-in citing an FD never seen booked. Proof
        that FD predates the reconciliation history (or a day is missing).
        Historical/closed orphans don't affect today's balance, but they are
        surfaced rather than silently absorbed.
      * unparsed         -- an 'FD Booking' txn whose narration matched no
        known pattern. A new bank or changed phrasing lands here instead of
        silently corrupting the total, which is exactly how the old
        derivation went wrong unnoticed for weeks.
    """
    if periods is None:
        sh = get_gspread_client().open_by_key(SPREADSHEET_ID)
        periods = _load_recon_periods(get_recon_txns_sheet(sh))

    def _rdk(p):
        try:
            return datetime.strptime(p['recon_date'], '%d-%m-%Y')
        except Exception:
            return datetime.min

    fds = {}
    unparsed = []
    latest_fd_txn_date = None
    seen_txn_keys = set()
    # Newest-first + content-key dedup: identical to _all_time_recon_totals(),
    # and load-bearing here for the same reason -- the same bank line re-saved
    # under a different recon_date would otherwise double every FD's booked
    # amount (this was missed on a first pass and inflated the total ~2.5x).
    for p in sorted(periods, key=_rdk, reverse=True):
        acct = p['account']
        for tx in p['txns']:
            # The NARRATION decides, not the stored type. A row is an FD leg
            # because it says "Sweepout FD <acct> booked" or "Auto sweep in
            # from <acct>" -- reading a type that some earlier classification
            # pass happened to write is one indirection too many, and it is
            # what hid Rs 1.23 crore of August sweep-outs stored as Expense.
            # Keying off the evidence means already-saved rows heal
            # themselves with no re-upload.
            desc_ws = _norm_ws(tx.get('description', ''))
            is_fd_leg = (any(pat.search(desc_ws) for pat in FD_OUT_PATTERNS)
                         or any(pat.search(desc_ws) for pat in FD_IN_PATTERNS)
                         or (tx.get('type') or '').strip() == 'FD Booking')
            if not is_fd_leg:
                continue
            desc = desc_ws
            dr, cr = tx['debit'], tx['credit']
            key = (acct, tx['date'], desc, dr, cr, tx['balance'])
            if key in seen_txn_keys:
                continue
            seen_txn_keys.add(key)

            d = _parse_flex_date(tx.get('date', ''))
            if d and (latest_fd_txn_date is None or d > latest_fd_txn_date):
                latest_fd_txn_date = d

            fd_no = None
            is_out = False
            for pat in FD_OUT_PATTERNS:
                m = pat.search(desc)
                if m and dr > 0:
                    fd_no, is_out = m.group(1), True
                    break
            if fd_no is None:
                for pat in FD_IN_PATTERNS:
                    m = pat.search(desc)
                    if m and cr > 0:
                        fd_no = m.group(1)
                        break
            if fd_no is None:
                unparsed.append({'account': acct, 'date': tx.get('date'),
                                  'debit': dr, 'credit': cr, 'description': desc[:120]})
                continue

            e = fds.setdefault(fd_no, {'account': acct, 'booked': 0.0, 'swept_in': 0.0,
                                        'book_dates': [], 'sweepin_dates': []})
            if is_out:
                e['booked'] += dr
                e['book_dates'].append(tx.get('date'))
            else:
                e['swept_in'] += cr
                e['sweepin_dates'].append(tx.get('date'))

    # Earliest day we hold for each account. An FD that swept back shortly
    # after that day was booked before our records begin -- we could never
    # have seen it, so it is explained, not a gap. Anything sweeping back
    # later than that means a booking we SHOULD have and don't.
    first_seen = {}
    for p in periods:
        for tx in p['txns']:
            d = _parse_flex_date(tx.get('date', ''))
            if d and (p['account'] not in first_seen or d < first_seen[p['account']]):
                first_seen[p['account']] = d

    open_fds, orphans = {}, {}
    total = 0.0
    for fd_no, e in fds.items():
        e['booked'] = round(e['booked'], 2)
        e['swept_in'] = round(e['swept_in'], 2)
        e['remaining'] = round(e['booked'] - e['swept_in'], 2)
        if e['booked'] <= 0 and e['swept_in'] > 0:
            start = first_seen.get(e['account'])
            first_in = min((d for d in (_parse_flex_date(x) for x in e['sweepin_dates']) if d),
                           default=None)
            # 30 days: these are short-term sweep deposits, so a booking more
            # than a month before its sweep-in is not plausible. The real
            # case this covers is HDFC FD 50301355033761 -- Rs 2,13,765 that
            # swept back 20-Jul-2026 against history starting 01-Jul.
            e['pre_history'] = bool(start and first_in and (first_in - start).days <= 30)
            orphans[fd_no] = e
            continue
        # Clamp at zero: a fully-swept FD holds nothing, and any excess that
        # came back is interest already reflected in the bank balance.
        if e['remaining'] > 0.5:
            open_fds[fd_no] = e
            total += e['remaining']

    return {
        'fd_total': round(total, 2),
        'open_fds': open_fds,
        'open_fd_count': len(open_fds),
        'orphan_sweepins': orphans,
        'orphan_sweepin_total': round(sum(o['swept_in'] for o in orphans.values()), 2),
        # Orphans that predate our records are permanent and understood;
        # only the rest mean data actually went missing, so only the rest
        # may disqualify the derived FD figure.
        'orphans_unexplained': {k: v for k, v in orphans.items() if not v.get('pre_history')},
        'orphans_unexplained_total': round(
            sum(o['swept_in'] for o in orphans.values() if not o.get('pre_history')), 2),
        'unparsed': unparsed,
        'unparsed_net': round(sum(u['debit'] - u['credit'] for u in unparsed), 2),
        'latest_fd_txn_date': latest_fd_txn_date,
    }

# Hardcoded per Prem's explicit instruction (06-08-2026): "just concentrate
# on august" -- pre-August history has real, numerous day-level gaps (see
# FD_ACCOUNT_LABEL's comment) that aren't worth excavating retroactively.
# This draws a clean line and only holds August-onward to a day-by-day
# standard, so a fresh, trustworthy baseline doesn't require first
# resolving every historical gap. Revisit if this should ever roll forward
# past August instead of staying fixed.
# ── Money Integrity: what every transaction TYPE means for cash ──────────────
# Single source of truth. Anything NOT in here is treated as unclassified and
# loudly reported rather than silently dropped — which is exactly how
# 'Partner Salary' (₹1,10,000) and 'Employee Salary' (₹60,000) went uncounted:
# every totaliser matched type strings exactly, so a custom type invented at
# reconciliation time carried real money out of the bank and landed nowhere.
CASH_ROLE = {
    'Disbursement':             'lending_out',
    'Repayment':                'lending_in',
    'Collection':               'lending_in',   # legacy label
    'Collection (via Pradaan)': 'lending_in',   # legacy label
    'Expense':                  'expense',
    'Employee Salary':          'expense',
    # A partner drawing is a CAPITAL withdrawal, not an operating expense —
    # given its own role so it never distorts the expense/profit picture,
    # while still being counted as money out so it can't go missing.
    # ⚠ If partner drawings are ALSO recorded in the Capital Log, net_capital
    # already reflects them and counting them here as well would double-count.
    # Confirm with Prem before treating this as settled.
    'Partner Salary':           'withdrawal',
    'Partner Withdrawal':       'withdrawal',
    'Interest Income':          'income',
    'Other Income':             'income',
    'FD Booking':               'fd',
    'Contra':                   'contra',
    'Capital In':               'contra',       # legacy label
    'Capital Out':              'contra',       # legacy label
    'Suspense':                 'suspense',
    'Test Credit':              'ignore',
    'Skip':                     'ignore',
}

# A transfer between our own accounts must show as a debit in one and a credit
# in another. When only ONE leg gets typed Contra, the pair no longer nets to
# zero and the difference reads as missing money.
CONTRA_PAIR_MAX_DAYS = 3      # cheque/IMPS settlement slack between the legs
CONTRA_PAIR_MIN_AMOUNT = 1000 # below this, coincidental same-amount pairs are likelier than real transfers

def _cross_account_contra_pairs(periods):
    """Own-account transfers where only one leg was typed Contra.

    A single statement can never reveal these: the parser sees ONE account at
    a time, so a cheque leaving HDFC and landing in Saraswat looks like an
    ordinary payment on the HDFC side. Only the merged all-accounts view can
    pair them — which is why this lives here and not in _match_transactions().

    Confirmed live 12-08-2026: a ₹75,000 cheque out of HDFC on 07-Jul
    ('CHQ PAID-CTS S6-RK S-BRIDGELINE PARTNERS') and a ₹75,000 branch deposit
    into Saraswat the same day ('OW/570240002/29/000094/Dep Br 404') were the
    two legs of one transfer. The Saraswat leg was typed Contra, the HDFC leg
    was typed Expense, and that single unpaired leg WAS the entire −₹75,000
    contra imbalance.

    Derived at READ time rather than by re-typing stored rows: needs no
    re-save (so it cannot drop a manual type_override — see the Jeevitha Bai K
    precedent, where Suspense typing has to be re-applied on every save), and
    it self-corrects historical data as well as future uploads.

    Guards against false pairing: both legs must be in DIFFERENT own accounts,
    neither already typed Contra, and — critically — neither tied to a customer
    case (a real disbursement or collection always carries a matched_ref), so
    an unrelated same-amount disbursement and collection can't pair up.
    Matched one-to-one, nearest date first.
    """
    def _rdk(p):
        try:
            return datetime.strptime(p['recon_date'], '%d-%m-%Y')
        except Exception:
            return datetime.min

    debits, credits, seen = [], [], set()
    for p in sorted(periods, key=_rdk, reverse=True):
        for tx in p['txns']:
            key = (p['account'], tx['date'], tx['description'], tx['debit'], tx['credit'], tx['balance'])
            if key in seen:
                continue
            seen.add(key)
            # Already-Contra legs MUST stay in the pool: the whole point is to
            # find the counter-leg of a transfer whose OTHER side was already
            # typed Contra. Excluding them makes the unpaired leg unmatchable —
            # which is exactly how a first attempt at this found nothing.
            already = (tx.get('type') or '').strip() == 'Contra'
            # The matched_ref guard stops a real disbursement pairing with an
            # unrelated same-amount collection — but it must NOT apply to a leg
            # explicitly typed Contra. Re-typing a transaction to Contra does
            # not clear the customer ref an earlier auto-match had already
            # stamped on it, so a genuine transfer can carry a stale ref
            # forever. An explicit Contra type outranks a leftover auto-match.
            # (Real case: the Saraswat 07-Jul ₹75,000 deposit is typed Contra
            # yet still refs BLP-050626-179, and that stale ref alone was
            # hiding the counter-leg of the −₹75,000 imbalance.)
            # ...and likewise a leg whose narration names US as the sender is
            # our own money by definition, however the amount-matcher happened
            # to label it. That is what frees the IDFC opening deposit, which
            # had been auto-matched to a customer case purely on amount.
            if (not already and not _looks_like_own_transfer(tx.get('description', '') or '')
                    and (tx.get('matched_ref') or '').strip()):
                continue
            d = _parse_flex_date(tx.get('date', ''))
            if not d:
                continue
            rec = {'key': key, 'acct': p['account'], 'date': d,
                    'desc': (tx.get('description') or '')[:110], 'already_contra': already}
            if tx['debit'] >= CONTRA_PAIR_MIN_AMOUNT:
                debits.append({**rec, 'amt': tx['debit']})
            elif tx['credit'] >= CONTRA_PAIR_MIN_AMOUNT:
                credits.append({**rec, 'amt': tx['credit']})

    paired_keys, pairs, used = set(), [], set()
    for dr in sorted(debits, key=lambda x: x['date']):
        best = None
        for i, cr in enumerate(credits):
            if i in used or cr['acct'] == dr['acct']:
                continue
            if abs(cr['amt'] - dr['amt']) > 1:
                continue
            gap = abs((cr['date'] - dr['date']).days)
            if gap > CONTRA_PAIR_MAX_DAYS:
                continue
            if best is None or gap < best[1]:
                best = (i, gap)
        if best is not None:
            i = best[0]
            used.add(i)
            cr = credits[i]
            # Only the side that ISN'T already Contra needs correcting. A pair
            # already typed Contra on both sides is a no-op and isn't reported —
            # otherwise every correct transfer would show up as a "correction".
            corrected = []
            if not dr['already_contra']:
                paired_keys.add(dr['key']); corrected.append('outgoing leg')
            if not cr['already_contra']:
                paired_keys.add(cr['key']); corrected.append('incoming leg')
            if corrected:
                pairs.append({'amount': round(dr['amt'], 2), 'gap_days': best[1],
                               'corrected': corrected,
                               'out': {'account': dr['acct'], 'date': dr['date'].strftime('%d-%m-%Y'),
                                        'description': dr['desc'],
                                        'already_contra': dr['already_contra']},
                               'in': {'account': cr['acct'], 'date': cr['date'].strftime('%d-%m-%Y'),
                                       'description': cr['desc'],
                                       'already_contra': cr['already_contra']}})
    return {'paired_keys': paired_keys, 'pairs': pairs,
            'paired_total': round(sum(p['amount'] for p in pairs), 2)}

def _pair_uploaded_statements(statements):
    """Pair own-account transfers ACROSS statements uploaded together, and
    re-type both legs Contra in place. Returns the list of pairs found.

    This is _cross_account_contra_pairs()'s job moved forward in time. That
    one repairs history at read time and cannot be removed — it is the only
    thing that fixes rows already saved. But repairing at read time means the
    reconciliation screen still SHOWS the wrong classification on the day, and
    every wrong classification lands in the review queue for a human to fix by
    hand. Uploading both accounts together removes the guessing entirely:
    ₹18,00,000 leaving HDFC and ₹18,00,000 arriving in IDFC on the same day is
    a transfer by structure, whatever either narration happens to say.

    That distinction matters because narration is not stable. Three separate
    regex patches were needed in _looks_like_own_transfer() alone, the last
    (13-08-2026) because the bank swapped slashes for spaces in an RTGS
    reference. Structural matching does not care.

    Guards are deliberately identical to the read-time version — a leg tied to
    a customer case (matched_ref) can't be re-typed unless it is already Contra
    or its narration names us as the sender — so the two passes can never
    disagree about what is or isn't a transfer.
    """
    legs = []
    for si, st in enumerate(statements):
        for ti, tx in enumerate(st['transactions']):
            already = (tx.get('type') or '').strip() == 'Contra'
            if (not already and not _looks_like_own_transfer(tx.get('description', '') or '')
                    and (tx.get('matched_ref') or '').strip()):
                continue
            d = _parse_flex_date(tx.get('date', ''))
            if not d:
                continue
            legs.append({'si': si, 'ti': ti, 'date': d, 'already': already,
                         'acct': st['account'] or f'File {si + 1}',
                         'desc': (tx.get('description') or '')[:110],
                         'utr': tx.get('utr', '') or '',
                         'debit': tx.get('debit', 0) or 0, 'credit': tx.get('credit', 0) or 0})

    debits  = [l for l in legs if l['debit']  >= CONTRA_PAIR_MIN_AMOUNT]
    credits = [l for l in legs if l['credit'] >= CONTRA_PAIR_MIN_AMOUNT]

    pairs, used = [], set()
    for dr in sorted(debits, key=lambda x: x['date']):
        best = None
        for i, cr in enumerate(credits):
            # Different STATEMENTS, not merely different account labels: two
            # files could arrive before either account is named.
            if i in used or cr['si'] == dr['si']:
                continue
            if abs(cr['credit'] - dr['debit']) > 1:
                continue
            gap = abs((cr['date'] - dr['date']).days)
            if gap > CONTRA_PAIR_MAX_DAYS:
                continue
            if best is None or gap < best[1]:
                best = (i, gap)
        if best is None:
            continue
        i, gap = best
        cr = credits[i]
        # Amount + direction + date alone is not enough evidence at month
        # scale. Across one day's statements a same-amount debit and credit
        # in different accounts is almost certainly a transfer; across a
        # whole month of ~500 rows per account, a disbursement and an
        # unrelated collection of the same size a day apart is ordinary.
        # So require one leg to positively say "this is our own money":
        # narration naming us, a leg already typed Contra, or a reference
        # shared by both legs. All three real cases fixed this week satisfy
        # this — the Rs 18,00,000 RTGS (narration names BRIDGELINE PARTNERS),
        # the Rs 75,000 cheque (Saraswat leg already Contra), and the IDFC
        # opening deposit. A transfer with no corroboration at all is left
        # for the review queue and for _cross_account_contra_pairs() at read
        # time, where a wrong guess costs nothing.
        def _shared_ref(a, b):
            ta = {t for t in _split_ref_tokens(a.get('utr', '') or '') if len(t) >= 6}
            tb = {t for t in _split_ref_tokens(b.get('utr', '') or '') if len(t) >= 6}
            if ta & tb:
                return True
            da, db = (a.get('desc') or '').upper(), (b.get('desc') or '').upper()
            return any(t in db for t in ta) or any(t in da for t in tb)

        corroborated = (dr['already'] or cr['already']
                        or _looks_like_own_transfer(dr.get('desc', '') or '')
                        or _looks_like_own_transfer(cr.get('desc', '') or '')
                        or _shared_ref(dr, cr))
        if not corroborated:
            continue
        used.add(i)
        retyped = []
        for leg, side in ((dr, 'outgoing leg'), (cr, 'incoming leg')):
            if leg['already']:
                continue
            tx = statements[leg['si']]['transactions'][leg['ti']]
            other = cr['acct'] if leg is dr else dr['acct']
            tx['type'] = 'Contra'
            tx['match_basis'] = 'Cross-account'
            tx['matched_ref'] = ''
            tx['match_notes'] = (f"Own-account transfer — paired with the {other} leg "
                                 f"of the same amount uploaded in this batch").strip()
            retyped.append(side)
        if retyped:
            pairs.append({'amount': round(dr['debit'], 2), 'gap_days': gap, 'retyped': retyped,
                          'out': {'account': dr['acct'], 'date': dr['date'].strftime('%d-%m-%Y'),
                                  'description': dr['desc']},
                          'in':  {'account': cr['acct'], 'date': cr['date'].strftime('%d-%m-%Y'),
                                  'description': cr['desc']}})
    return pairs

def _money_integrity_check(periods=None, records=None, mc_rows=None, sh=None):
    """Every distinct way a rupee could go missing, each as its own number
    that should be zero. Deliberately NOT a single variance: the bank
    statement's own arithmetic always balances, so one aggregate figure
    proves nothing. What actually matters is whether every rupee of bank
    movement is explained AND every book entry is backed by bank movement.

    Leaks reported (each should be 0):
      1. unclassified   -- a type with no cash role: money moved, meaning unknown
      2. bank_unmatched -- a bank Disbursement/Repayment not tied to any case
      3. book_unmatched -- a book disbursement/collection with no bank evidence
      4. contra_imbalance -- own-account transfers must net to zero
      5. fd_orphans / fd_unparsed -- sweeps that can't be attributed
    """
    if sh is None:
        sh = get_gspread_client().open_by_key(SPREADSHEET_ID)
    if periods is None:
        periods = _load_recon_periods(get_recon_txns_sheet(sh))
    if records is None:
        records = read_accounts_from_gsheet()
    if mc_rows is None:
        mc_rows = sh.worksheet('M Coll').get_all_values()[1:]

    def _rdk(p):
        try:
            return datetime.strptime(p['recon_date'], '%d-%m-%Y')
        except Exception:
            return datetime.min

    # Own-account transfers whose two legs were typed inconsistently. Treated
    # as Contra here so a half-typed transfer stops reading as missing money.
    contra_pairing = _cross_account_contra_pairs(periods)
    contra_paired_keys = contra_pairing['paired_keys']

    role_totals = {}
    unclassified = {}
    bank_unmatched = []
    all_txns = []
    seen = set()
    first_date = last_date = None
    for p in sorted(periods, key=_rdk, reverse=True):
        for tx in p['txns']:
            k = (p['account'], tx['date'], tx['description'], tx['debit'], tx['credit'], tx['balance'])
            if k in seen:
                continue
            seen.add(k)
            all_txns.append(tx)
            d = _parse_flex_date(tx.get('date', ''))
            if d:
                first_date = d if first_date is None or d < first_date else first_date
                last_date = d if last_date is None or d > last_date else last_date

            t = _effective_type(tx) or '(blank)'
            if k in contra_paired_keys:
                t = 'Contra'          # matched to its counter-leg in another own account
            role = CASH_ROLE.get(t)
            net_out = tx['debit'] - tx['credit']
            if role is None:
                e = unclassified.setdefault(t, {'count': 0, 'net_out_of_bank': 0.0})
                e['count'] += 1
                e['net_out_of_bank'] += net_out
                continue
            r = role_totals.setdefault(role, {'count': 0, 'debit': 0.0, 'credit': 0.0})
            r['count'] += 1
            r['debit'] += tx['debit']
            r['credit'] += tx['credit']
            # A bank line the classifier called lending but couldn't tie to a
            # case: real money in/out with no identified counterparty.
            if role in ('lending_out', 'lending_in') and not (tx.get('matched_ref') or '').strip():
                bank_unmatched.append({'date': tx.get('date'), 'type': t,
                                        'debit': tx['debit'], 'credit': tx['credit'],
                                        'description': (tx.get('description') or '')[:100]})

    for r in role_totals.values():
        r['debit'] = round(r['debit'], 2)
        r['credit'] = round(r['credit'], 2)
        r['net_out_of_bank'] = round(r['debit'] - r['credit'], 2)
    for e in unclassified.values():
        e['net_out_of_bank'] = round(e['net_out_of_bank'], 2)

    # Book side: entries dated inside the reconciled window with no bank
    # evidence at all. Range-scoped (not month-scoped like _book_completeness)
    # so it covers exactly the window reconciliation actually proves.
    disb_by_ref, coll_by_ref = _bank_matched_refs(all_txns)
    bank_descs_upper = [(tx.get('description') or '').upper() for tx in all_txns]
    def _in_window(dstr):
        d = _parse_flex_date(str(dstr or ''))
        return bool(d and first_date and last_date and first_date <= d <= last_date)

    book_unmatched_disb, book_unmatched_coll = [], []
    for r in records:
        did = r.get('Disbursement ID', '')
        if not did or not _in_window(r.get('Disbursement Date', '')) or did in disb_by_ref:
            continue
        note = str(r.get('Debit Note', '') or '').strip()
        if any(_ref_present_in_bank_text(tok, bank_idx) for tok in _split_ref_tokens(note)):
            continue
        book_unmatched_disb.append({'disb_id': did, 'customer': r.get('Customer Name', ''),
                                     'date': r.get('Disbursement Date', ''),
                                     'amount': _to_num(r.get('Amount', 0))})
    for mc in mc_rows:
        did = (mc[0] if len(mc) > 0 else '').strip()
        if not did or not _in_window(mc[1] if len(mc) > 1 else '') or did in coll_by_ref:
            continue
        note = mc[3] if len(mc) > 3 else ''
        if any(_ref_present_in_bank_text(tok, bank_idx) for tok in _split_ref_tokens(note)):
            continue
        book_unmatched_coll.append({'disb_id': did, 'date': mc[1] if len(mc) > 1 else '',
                                     'amount': _clean_amount(mc[2]) if len(mc) > 2 else 0})

    fd = _fd_ledger(periods)
    contra = role_totals.get('contra', {'net_out_of_bank': 0.0})
    suspense = role_totals.get('suspense', {'net_out_of_bank': 0.0})

    leaks = {
        'unclassified_types': {
            'total_net_out_of_bank': round(sum(e['net_out_of_bank'] for e in unclassified.values()), 2),
            'types': unclassified},
        'bank_entries_not_tied_to_a_case': {
            'count': len(bank_unmatched),
            'total_debit': round(sum(b['debit'] for b in bank_unmatched), 2),
            'total_credit': round(sum(b['credit'] for b in bank_unmatched), 2),
            'entries': bank_unmatched[:50]},
        'book_entries_with_no_bank_evidence': {
            'disbursement_count': len(book_unmatched_disb),
            'disbursement_total': round(sum(d['amount'] for d in book_unmatched_disb), 2),
            'collection_count': len(book_unmatched_coll),
            'collection_total': round(sum(c['amount'] for c in book_unmatched_coll), 2),
            'disbursements': book_unmatched_disb[:50],
            'collections': book_unmatched_coll[:50]},
        'contra_imbalance': round(contra.get('net_out_of_bank', 0.0), 2),
        'suspense_net': round(suspense.get('net_out_of_bank', 0.0), 2),
        'fd_orphan_total': fd['orphan_sweepin_total'],
        'fd_unparsed_net': fd['unparsed_net'],
    }
    clean = (abs(leaks['unclassified_types']['total_net_out_of_bank']) < 1
             and leaks['bank_entries_not_tied_to_a_case']['count'] == 0
             and leaks['book_entries_with_no_bank_evidence']['disbursement_count'] == 0
             and leaks['book_entries_with_no_bank_evidence']['collection_count'] == 0
             and abs(leaks['contra_imbalance']) < 1
             and abs(leaks['fd_orphan_total']) < 1
             and abs(leaks['fd_unparsed_net']) < 1)

    return {
        'reconciled_window': {
            'from': first_date.strftime('%d-%m-%Y') if first_date else None,
            'to': last_date.strftime('%d-%m-%Y') if last_date else None},
        'role_totals': role_totals,
        'leaks': leaks,
        'all_clear': clean,
        'fd_total': fd['fd_total'],
        # Auto-paired own-account transfers — shown so the correction is
        # visible and auditable, never applied silently.
        'contra_auto_paired': contra_pairing['pairs'],
        'contra_auto_paired_total': contra_pairing['paired_total'],
    }

# Sub-rupee residue is rounding, not a defect — the same tolerance the rest
# of this codebase already treats as fully settled.
CHARGE_MATERIALITY = 1.0

def _charge_integrity_check(records=None):
    """Every case whose STORED Charges/GST/Total disagree with what its own
    company/cluster/date imply. Should always be empty.

    This exists because no amount of care inside save_disbursement() or
    update_disbursement() can protect a row that is edited BY HAND in the
    Google Sheet — which is exactly how BLP-120826-401 went wrong on
    12-08-2026: disbursed under an ordinary cluster and charged the flat
    0.5%, then the Cluster cell was corrected to BIB HDB Karnataka directly
    in the sheet. The charge plan marker stayed blank, the money stayed flat,
    and nothing anywhere compared the two. A rule can be bypassed; a detector
    cannot, so the guarantee has to be a detector.

    Two distinct faults are reported, because they need different fixes:
      * plan_mismatch  -- the row's stamped plan isn't what its own
                          company/cluster/date resolve to (a marker never
                          stamped, or stale after a hand-edit)
      * amount_mismatch -- stored money doesn't match the row's own stamped
                          plan (charged wrong, or the amount was edited in
                          the sheet without recomputing)

    Provisional (day-based) plans are compared at their best-case rate while
    the case is still open, since the final figure legitimately isn't known
    until it closes — comparing against the final rate would flag every
    healthy open case.
    """
    if records is None:
        records = read_accounts_from_gsheet()
    issues = []
    for r in records:
        disb_id = (r.get('Disbursement ID', '') or '').strip()
        if not disb_id:
            continue
        amount = _to_num(r.get('Amount', 0))
        if amount <= 0:
            continue
        cluster = (r.get('Cluster', '') or '').strip()
        company = (r.get('Company', '') or '').strip()
        # Only a value that names a REGISTERED plan counts as a stamp. The
        # older archive tabs don't share the live sheet's column layout, so
        # column 26 there holds unrelated leftovers ('96', '0', even the
        # literal header text) — reading those as plan ids reported six
        # perfectly healthy April rows as mismatched on the first run.
        stamped = (r.get('Charge Plan', '') or '').strip() or None
        if stamped and not get_plan_by_id(stamped):
            stamped = None
        d = _to_plain_date(r.get('Disbursement Date', ''))
        if d is None:
            continue
        expected_plan = resolve_charge_plan(company, cluster, d)

        stored_charges = _to_num(r.get('Charges', 0))
        stored_gst     = _gst_of(r)
        stored_total   = _to_num(r.get('Total', 0))

        status  = (r.get('Overdue Status', '') or '').strip()
        is_open = status != 'Closed'
        # A closed provisional case is compared at its ACTUAL collection date;
        # an open one at the best case, which is all that's knowable yet.
        coll_date = None if is_open else _to_plain_date(r.get('Collected Date', ''))
        by_stamped = calc_charges(amount, cluster=cluster, company=company,
                                  disbursement_date=d, collection_date=coll_date,
                                  methodology=stamped)

        faults = []
        if (stamped or None) != (expected_plan or None):
            faults.append('plan_mismatch')
        # An amount mismatch is only a DEFECT on a plan-governed row, where
        # the rate is contractual. On a flat-default row a different figure
        # is usually a deliberate commercial decision — a negotiated rate or
        # a discount, which the Edit Case screen exists to allow — and
        # reporting those forever would bury the real faults. (First live run
        # flagged 13 such historical rows alongside the one true defect.)
        on_a_plan = bool(stamped or expected_plan)
        if on_a_plan and abs(stored_total - by_stamped['total']) >= CHARGE_MATERIALITY:
            faults.append('amount_mismatch')
        if not faults:
            continue

        by_expected = calc_charges(amount, cluster=cluster, company=company,
                                   disbursement_date=d, collection_date=coll_date,
                                   methodology=expected_plan)
        issues.append({
            'disb_id': disb_id, 'customer': (r.get('Customer Name', '') or '').strip(),
            'date': (r.get('Disbursement Date', '') or '').strip(),
            'company': company, 'cluster': cluster, 'branch': (r.get('Branch', '') or '').strip(),
            'status': status, 'amount': amount, 'faults': faults,
            'stamped_plan': stamped or '(flat default)',
            'expected_plan': expected_plan or '(flat default)',
            'stored':   {'charges': stored_charges, 'gst': stored_gst, 'total': stored_total},
            'expected': {k: by_expected[k] for k in ('charges', 'gst', 'total')},
            'expected_label': by_expected.get('plan_label', ''),
            'difference': round(by_expected['total'] - stored_total, 2),
        })
    over  = round(sum(i['difference'] for i in issues if i['difference'] < 0), 2)
    under = round(sum(i['difference'] for i in issues if i['difference'] > 0), 2)
    return {'all_clear': not issues, 'count': len(issues), 'issues': issues,
            'overcharged_total': abs(over), 'undercharged_total': under,
            'plans': [{'id': p.get('id'), 'label': p.get('label'),
                       'match': p.get('match'), 'from': p.get('from'),
                       'formula': p.get('formula')} for p in get_charge_plans()]}

REPORT_GST_RATE = 0.18
# A debit at or above this with no recognisable expense category is treated as
# unidentified (held in suspense, excluded from profit, reported by name)
# rather than silently charged to the P&L. Materiality threshold, not a rule
# about any particular transaction.
UNIDENTIFIED_MATERIALITY = 50000

def _effective_type(tx):
    """A stored row's type, corrected where its own narration contradicts it.

    Only ever upgrades an FD sweep leg that was mis-typed because the bank
    wrapped the narration across lines (see _norm_ws). Read-time repair, so
    already-saved rows are right without a re-save -- same approach as
    _cross_account_contra_pairs(). Without this the six August sweep-outs
    sit in the P&L as Rs 1.23 crore of operating expenses."""
    stored = (tx.get('type') or '').strip()
    if stored == 'FD Booking':
        return stored
    d = _norm_ws(tx.get('description', ''))
    if any(pat.search(d) for pat in FD_OUT_PATTERNS) or any(pat.search(d) for pat in FD_IN_PATTERNS):
        return 'FD Booking'
    return stored

def _gst_of(record):
    """The Accounts sheet's GST header carries a narrow no-break space, so a
    plain 'GST (18%)' string literal never matches it. Find it by prefix."""
    k = next((k for k in record if str(k).strip().upper().startswith('GST')), None)
    return _to_num(record.get(k, 0)) if k else 0.0

def _period_financials(from_date=None, to_date=None, periods=None, records=None, sh=None):
    """Cash-basis P&L inputs for a date range — the single source of truth
    behind every financial report, so the P&L, GST register and position
    statement can never quote different numbers for the same period.

    Revenue is recognised on COLLECTION, using the SAME rule as the GST
    register (Prem's decision, 12-08-2026): a case contributes when it is
    Closed AND its collection date falls in range. Charges invoiced on a
    still-open case are future income, not cash earned, so they're excluded.

    Excess collected — anything received beyond Total, even Re 1 — counts as
    additional charges earned (Prem's explicit instruction, 11-08-2026),
    decomposed GST-inclusive at 18%: the excess IS the cash received, so it's
    split the same way Total = Amount + Charges + Charges×18% already is,
    rather than grossed up on top.

    GST is returned as its own LIABILITY figure and is deliberately NOT part
    of net_profit — it's collected on the government's behalf, never earnings.

    Partner drawings ('withdrawal' role) are returned separately too: a
    capital appropriation belongs below the profit line, not inside expenses.
    """
    if sh is None:
        sh = get_gspread_client().open_by_key(SPREADSHEET_ID)
    if records is None:
        records = read_accounts_from_gsheet()
    if periods is None:
        periods = _load_recon_periods(get_recon_txns_sheet(sh))
    if to_date is None:
        to_date = date.today()

    def _in_range(d):
        if d is None:
            return False
        if from_date and d < from_date:
            return False
        return d <= to_date

    charges_earned = gst_collected = excess_total = 0.0
    cases_closed = 0
    for r in records:
        if (r.get('Overdue Status', '') or '').strip() != 'Closed':
            continue
        if not _in_range(_parse_flex_date(r.get('Collected Date', ''))):
            continue
        _amt, total, collected, _bal = _parse_case(r)
        excess = max(0.0, round(collected - total, 2))
        ex_charges = excess / (1 + REPORT_GST_RATE)
        charges_earned += _to_num(r.get('Charges', 0)) + ex_charges
        gst_collected += _gst_of(r) + (excess - ex_charges)
        excess_total += excess
        cases_closed += 1

    def _rdk(p):
        try:
            return datetime.strptime(p['recon_date'], '%d-%m-%Y')
        except Exception:
            return datetime.min

    # A leg of an own-account transfer is not a cost — exclude both sides so a
    # half-typed transfer can never land in expenses or in suspense.
    contra_paired_keys = _cross_account_contra_pairs(periods)['paired_keys']

    expenses_by_category = {}
    unidentified = []
    withdrawals = other_income = 0.0
    recon_first = recon_last = None
    seen = set()
    for p in sorted(periods, key=_rdk, reverse=True):
        for tx in p['txns']:
            k = (p['account'], tx['date'], tx['description'], tx['debit'], tx['credit'], tx['balance'])
            if k in seen:
                continue
            seen.add(k)
            d = _parse_flex_date(tx.get('date', ''))
            if d:
                recon_first = d if recon_first is None or d < recon_first else recon_first
                recon_last = d if recon_last is None or d > recon_last else recon_last
            if not _in_range(d):
                continue
            if k in contra_paired_keys:
                continue          # own-account transfer, not income or cost
            role = CASH_ROLE.get(_effective_type(tx))
            net_out = tx['debit'] - tx['credit']
            if role == 'expense':
                # Re-derive from the narration rather than trusting the label
                # stored at save time: a stored label can carry a since-fixed
                # classification bug forward forever. Exactly that happened —
                # the firm's own name ("BridgeLine Partners") matched the
                # partner rule and misbooked ₹5,75,000 of cheques as drawings.
                cat = _expense_category(tx.get('description', '') or '') or 'Uncategorised'
                if cat == 'Partner Share':
                    # A partner drawing is a CAPITAL appropriation, not an
                    # operating cost — it belongs below the profit line even
                    # when the transaction is typed 'Expense'. Without this the
                    # P&L charges real drawings against profit and can report a
                    # loss on a genuinely profitable period.
                    withdrawals += net_out
                elif cat == 'Miscellaneous' and net_out >= UNIDENTIFIED_MATERIALITY:
                    # A large debit with no recognisable category is UNIDENTIFIED,
                    # not proven to be an expense. Standard practice is to hold it
                    # in suspense until identified rather than let a guess hit the
                    # P&L — so it is listed separately and excluded from profit,
                    # and the report says so on its face.
                    unidentified.append({
                        'date': tx.get('date'), 'account': p['account'],
                        'amount': round(net_out, 2),
                        'description': (tx.get('description') or '')[:140]})
                else:
                    expenses_by_category[cat] = round(expenses_by_category.get(cat, 0.0) + net_out, 2)
            elif role == 'withdrawal':
                withdrawals += net_out
            elif role == 'income':
                other_income += tx['credit'] - tx['debit']
            elif role is None and net_out > 0:
                # Never let an unrecognised type silently swallow cash — it
                # lands in its own visible category instead.
                cat = f"Unclassified: {(tx.get('type') or '').strip() or '(blank)'}"
                expenses_by_category[cat] = round(expenses_by_category.get(cat, 0.0) + net_out, 2)

    total_expenses = round(sum(expenses_by_category.values()), 2)
    charges_earned = round(charges_earned, 2)
    other_income = round(other_income, 2)
    return {
        'from': from_date.strftime('%d-%m-%Y') if from_date else None,
        'to': to_date.strftime('%d-%m-%Y'),
        'basis': 'cash',
        'charges_earned': charges_earned,
        'gst_collected': round(gst_collected, 2),
        'excess_collected': round(excess_total, 2),
        'cases_closed': cases_closed,
        'other_income': other_income,
        'expenses_by_category': dict(sorted(expenses_by_category.items(),
                                             key=lambda kv: -kv[1])),
        'total_expenses': total_expenses,
        # Stated BEFORE unidentified items — see `unidentified` below.
        'net_profit': round(charges_earned + other_income - total_expenses, 2),
        'partner_withdrawals': round(withdrawals, 2),
        'unidentified': sorted(unidentified, key=lambda u: -u['amount']),
        'unidentified_total': round(sum(u['amount'] for u in unidentified), 2),
        # So every report can print its own data window and disclose, rather
        # than imply, that expenses only exist from the first reconciled day.
        'recon_data_from': recon_first.strftime('%d-%m-%Y') if recon_first else None,
        'recon_data_to': recon_last.strftime('%d-%m-%Y') if recon_last else None,
        'expenses_incomplete_before': recon_first.strftime('%d-%m-%Y') if recon_first else None,
    }

RECON_DAY_LEVEL_START = date(2026, 8, 1)

def _day_level_coverage_gaps(periods, since=RECON_DAY_LEVEL_START):
    """Unlike _recon_coverage_gaps() (month-level, all-time), this checks
    EVERY CALENDAR DAY from `since` to today for each registered account --
    the granularity that actually matters for fd_outstanding's accuracy
    (see FD_ACCOUNT_LABEL). A single missing day is exactly what silently
    breaks the sweep-out/sweep-in netting, so surfacing gaps at day
    resolution (not month resolution) is the whole point of this check."""
    registered = [a.get('name', '').strip() for a in (load_config().get('bank_accounts') or [])
                  if a.get('name', '').strip()]
    today = date.today()
    gaps = {}
    for acct in registered:
        covered = set()
        for p in periods:
            if p['account'] != acct:
                continue
            d = _parse_recon_date(p['recon_date'])
            if d:
                covered.add(d.date())
        missing = []
        d = since
        while d <= today:
            if d not in covered:
                missing.append(d.strftime('%d-%b'))
            d += timedelta(days=1)
        gaps[acct] = {'status': 'complete' if not missing else 'incomplete', 'missing_days': missing}
    return gaps

def _recon_coverage_gaps(periods):
    """Mirrors _account_tieout()'s 'incomplete' degradation pattern, but at
    the period-coverage level: for each registered bank account, is there a
    saved period for every calendar month between that account's OWN
    earliest reconciliation and today? Scoped per-account (not a single
    company-wide start date) since accounts open at different times."""
    registered = [a.get('name', '').strip() for a in (load_config().get('bank_accounts') or [])
                  if a.get('name', '').strip()]
    today = datetime.today()
    gaps = {}
    for acct in registered:
        dates = []
        for p in periods:
            if p['account'] != acct:
                continue
            try:
                dates.append(datetime.strptime(p['recon_date'], '%d-%m-%Y'))
            except Exception:
                continue
        if not dates:
            gaps[acct] = {'status': 'no_periods', 'missing_months': []}
            continue
        covered = {(d.year, d.month) for d in dates}
        earliest = min(dates)
        missing, y, m = [], earliest.year, earliest.month
        while (y, m) <= (today.year, today.month):
            if (y, m) not in covered:
                missing.append(f"{y:04d}-{m:02d}")
            m += 1
            if m > 12:
                m = 1; y += 1
        gaps[acct] = {'status': 'complete' if not missing else 'incomplete',
                      'earliest_month': f"{earliest.year:04d}-{earliest.month:02d}",
                      'missing_months': missing}
    return gaps

def _solvency_check():
    """The actual "is money missing" check: does Net Capital (Capital Log,
    all-time) plus net cash generated by lending (all-time Collected minus
    all-time Disbursed) minus Expenses (all-time, from reconciliation
    history) equal the real bank balance?

    Deliberately NOT "Net Capital − Outstanding − Expenses": Outstanding
    (for open cases) includes charges/GST accrued but not yet collected —
    future income, not cash that has left the bank — so subtracting it
    understates expected cash, and separately never adds back charges that
    HAVE already been collected (on both open and closed cases). Verified
    with a worked example (₹100 principal + ₹10 charge, disbursed then
    fully repaid): Net Capital − Outstanding − Expenses gives the wrong
    answer at both the mid-loan and fully-repaid checkpoints; Net Capital +
    (Collected_all − Disbursed_all) − Expenses gives the right answer at
    both. Total Disbursed/Collected are summed across EVERY case (open and
    closed), not just open ones — closed cases' fully-collected charges are
    exactly the margin this business exists to earn, and must be reflected
    in the expected cash position.

    FD money is deliberately NOT subtracted here — it's surfaced as its own
    explained line ('₹X currently in FDs') rather than folded into the main
    variance, so parking money in a Fixed Deposit never reads as unexplained
    missing money. Same reasoning _account_tieout() already uses for this
    statement's own Expense/Capital totals: context, not part of the formula.
    That line now shows the DIRECTLY-ENTERED FD balance (see
    FD_ACCOUNT_LABEL) as the trusted figure, alongside the old derived-from-
    transaction-history total purely for comparison — a material gap between
    the two is itself a signal something in the reconciled data is off
    (exactly what surfaced the ~₹39.5L IDFC gap, 06-08-2026), not something
    to silently paper over.

    Degrades to 'incomplete' (not a confident-looking wrong number) whenever
    reconciliation coverage has gaps, since Expenses sourced from that
    history are then undercounted for the missing months. Also checked at
    DAY resolution for August onward (see _day_level_coverage_gaps) — the
    month-level check alone missed real gaps that were silently corrupting
    fd_total.
    """
    sh = get_gspread_client().open_by_key(SPREADSHEET_ID)
    capital_log = read_capital_log(sh)
    if not capital_log or not capital_log.get('available'):
        return {'status': 'no_capital_log_data',
                'message': (capital_log or {}).get('error', 'Capital Log tab not found or unreadable.')}

    net_capital = capital_log['net_capital']
    records = read_accounts_from_gsheet()
    total_disbursed_all = round(sum(_parse_case(r)[0] for r in records), 2)
    total_collected_all = round(sum(_parse_case(r)[2] for r in records), 2)
    total_outstanding = get_today_summary()['total_outstanding']  # display context only, not in the formula

    periods = _load_recon_periods(get_recon_txns_sheet(sh))
    recon_totals = _all_time_recon_totals(periods)
    total_expenses = recon_totals['expense_total']
    fd_total_derived = recon_totals['fd_total']
    other_income_total = recon_totals['income_total']
    fd_balance_direct, fd_balance_date = get_latest_fd_balance()
    fd_ledger = _fd_ledger(periods)
    fd_total = fd_ledger['fd_total']

    bank_balance, bank_date = get_latest_bank_balance()
    coverage_gaps = _recon_coverage_gaps(periods)
    incomplete_accounts = [a for a, g in coverage_gaps.items() if g['status'] != 'complete']
    day_gaps = _day_level_coverage_gaps(periods)
    day_gap_accounts = [a for a, g in day_gaps.items() if g['status'] != 'complete']

    # A contra transfer dated inside the window between two accounts' own
    # reconciliation dates is money that has LEFT the earlier-dated account's
    # closing balance but not yet ARRIVED in the later-dated one — it simply
    # vanishes from bank_balance, and no disbursement/collection adjustment
    # compensates for it (contra isn't in Accounts or M Coll at all). Harmless
    # only while every account shares one reconciliation date.
    recon_dates_by_account = {}
    for acct in [a.get('name', '').strip() for a in (load_config().get('bank_accounts') or [])
                 if a.get('name', '').strip()]:
        _, d = _last_recon_closing_for_account(acct)
        if d:
            recon_dates_by_account[acct] = d.strftime('%d-%m-%Y')
    distinct_recon_dates = set(recon_dates_by_account.values())

    expected_bank_balance = variance = None
    if bank_balance is not None:
        net_lending_cash = total_collected_all - total_disbursed_all
        expected_bank_balance = round(net_capital + net_lending_cash - total_expenses
                                       + other_income_total, 2)
        variance = round(bank_balance - expected_bank_balance, 2)

    status, reasons = 'ok', []
    if incomplete_accounts:
        status = 'incomplete'
        reasons.append(f"Reconciliation coverage gaps for: {', '.join(incomplete_accounts)}")
    if day_gap_accounts:
        status = 'incomplete'
        for a in day_gap_accounts:
            reasons.append(f"{a} missing day-level reconciliation for: {', '.join(day_gaps[a]['missing_days'])}")
    if fd_ledger['unparsed']:
        status = 'incomplete'
        reasons.append(
            f"{len(fd_ledger['unparsed'])} FD sweep transaction(s) have a narration this system "
            f"doesn't recognise, so they're excluded from the FD balance (net "
            f"₹{fd_ledger['unparsed_net']:,.2f}). Usually means a bank changed its wording or a new "
            f"account was added — the FD figure can't be fully trusted until these are handled.")
    if fd_ledger['orphan_sweepins']:
        reasons.append(
            f"₹{fd_ledger['orphan_sweepin_total']:,.2f} swept back in from "
            f"{len(fd_ledger['orphan_sweepins'])} FD(s) that were booked before reconciliation history "
            f"begins ({', '.join(fd_ledger['orphan_sweepins'].keys())}). Those FDs are already fully "
            f"swept back, so today's FD balance is unaffected — but if any pre-history FD is still open, "
            f"its balance is invisible here.")
    if len(distinct_recon_dates) > 1:
        status = 'incomplete'
        reasons.append(
            'Bank accounts are reconciled to DIFFERENT dates ('
            + '; '.join(f'{a}: {d}' for a, d in sorted(recon_dates_by_account.items()))
            + '). Any transfer between your own accounts dated inside that window leaves one balance '
              'but has not yet landed in the other, so the bank total silently understates. '
              'Reconcile both to the same date.')
    if bank_balance is None:
        status = 'incomplete'
        reasons.append('No bank reconciliation history yet — cannot compare against a real bank balance.')

    return {
        'status': status, 'reasons': reasons,
        'net_capital': net_capital,
        'total_disbursed_all': total_disbursed_all, 'total_collected_all': total_collected_all,
        'total_outstanding': total_outstanding,  # display context only — money currently with customers
        'total_expenses': total_expenses, 'fd_total': fd_total,
        'fd_total_derived': fd_total_derived, 'fd_balance_direct': fd_balance_direct,
        'fd_balance_date': fd_balance_date.strftime('%d-%m-%Y') if fd_balance_date else None,
        'fd_open_count': fd_ledger['open_fd_count'],
        'fd_open_accounts': {k: v['remaining'] for k, v in fd_ledger['open_fds'].items()},
        'fd_orphan_total': fd_ledger['orphan_sweepin_total'],
        'fd_unparsed_count': len(fd_ledger['unparsed']),
        'other_income_total': other_income_total,
        'bank_balance': bank_balance,
        'bank_balance_date': bank_date.strftime('%d-%m-%Y') if bank_date else None,
        'recon_dates_by_account': recon_dates_by_account,
        'expected_bank_balance': expected_bank_balance, 'variance': variance,
        'ok': (variance is not None and abs(variance) <= 1.0),
        'coverage_gaps': coverage_gaps,
        'day_level_gaps': day_gaps,
    }

def _supersede_covered_periods(txns_ws, account, classified):
    """Delete Recon Txns rows for `account` that the statement being saved
    now covers, so a wider upload REWRITES the days inside it.

    Only rows whose own transaction date falls inside the new statement's
    date span are removed, and only for this account — a day outside the
    span, or another account's row, is never touched. Deletes bottom-up in
    contiguous blocks (rows are appended per batch, so a superseded period is
    almost always one block) to keep this to a couple of API calls.

    Returns the number of rows removed, for reporting. A failure here is
    swallowed: it degrades to the read-time containment rule in
    _load_recon_periods(), which produces the same numbers without touching
    the sheet, so a Sheets hiccup can never cost a save.
    """
    # Rule: replace a DAY only if the new statement actually has that day.
    #
    # Not "every day in the statement's span". A span rule deletes days the
    # new file is silent about, and silence is not the same as "no activity"
    # — it is usually a day the download simply left out. That is exactly
    # what went wrong on 13-08-2026: an August IDFC upload spanning 01–13 Aug
    # deleted every IDFC row in that span, but the file had no rows for the
    # days carrying three FD sweep-out bookings. Their sweep-ins survived, so
    # the FD ledger saw ₹74.4 lakh coming back from deposits it had never
    # seen created and reported the FD balance as zero.
    #
    # Day-present is the safe version of the same intent: re-uploading August
    # still rewrites every August day the file covers, and a day it does not
    # cover keeps whatever was already reconciled for it.
    days = {d for d in (_parse_flex_date(tx.get('date', '')) for tx in classified) if d}
    if not days or not account:
        return 0
    try:
        vals = txns_ws.get_all_values()
    except Exception:
        return 0
    doomed = []
    for i, row in enumerate(vals[1:], start=2):     # 1-based, skipping header
        if len(row) < 14 or not row[0] or row[2] != account:
            continue
        d = _parse_flex_date(row[7])
        if d and d in days:
            doomed.append(i)
    if not doomed:
        return 0
    blocks, start, prev = [], doomed[0], doomed[0]
    for r in doomed[1:]:
        if r == prev + 1:
            prev = r; continue
        blocks.append((start, prev)); start = prev = r
    blocks.append((start, prev))
    removed = 0
    for a, b in sorted(blocks, reverse=True):       # bottom-up: indices above stay valid
        try:
            txns_ws.delete_rows(a, b)
            removed += b - a + 1
        except Exception:
            pass
    return removed

def save_reconciliation(recon_date, opening_balance, closing_balance, transactions,
                         remarks='', remarks_map=None, account=''):
    """Persists this period's classified transactions to the 'Recon Txns'
    sheet tab, then rebuilds the FULL cumulative reconciliation workbook from
    every stored period and returns its bytes for an in-browser download.

    The Google Sheet is the system of record — the user never re-uploads a
    previous file to keep history; every download contains all periods saved
    so far. Saving the same (recon_date, account) again replaces that period
    (latest Batch ID wins in _load_recon_periods), so corrections are safe.

    Each call also appends one summary row to the 'Recon Log' tab so
    get_latest_bank_balance() / the dashboard's bank-balance figure keeps
    working without reading back any saved file.
    """
    try:
        records = read_accounts_from_gsheet()
    except Exception:
        records = []
    try:
        sh      = get_gspread_client().open_by_key(SPREADSHEET_ID)
        mc_ws   = sh.worksheet('M Coll')
        mc_rows = mc_ws.get_all_values()[1:]
    except Exception:
        sh      = get_gspread_client().open_by_key(SPREADSHEET_ID)
        mc_rows = []

    # Derive period label from date  e.g. "Jun 2026"
    try:
        period_label = datetime.strptime(recon_date, '%d-%m-%Y').strftime('%b %Y')
    except Exception:
        period_label = recon_date

    # Classify and match bank transactions; drop any manually marked Skip
    classified = [tx for tx in _match_transactions(transactions, records, mc_rows)
                  if tx.get('type_override', '') != 'Skip']
    rm = remarks_map or {}

    # Persist this period's classified rows to the Recon Txns tab (one batch
    # write), then read everything back so current + historical periods flow
    # through one rebuild path.
    txns_ws = get_recon_txns_sheet(sh)
    batch_id = datetime.now().strftime('%Y%m%d%H%M%S')
    # A wider statement REWRITES the days it covers (Prem, 13-08-2026: "u
    # just rewrite the entire month"). Uploading August in one file must
    # replace the daily August periods already saved for that account, not
    # sit alongside them — otherwise every day already reconciled is counted
    # twice in the workbook and the tie-out. Physically removed rather than
    # only outranked at read time so the Recon Txns tab stays true to what
    # the books actually use.
    superseded_rows = _supersede_covered_periods(txns_ws, account, classified)
    txns_ws.append_rows([[
        batch_id, recon_date, account, period_label,
        opening_balance, closing_balance, remarks,
        tx['date'], tx['description'], tx.get('utr', ''),
        tx['debit'], tx['credit'], tx['balance'],
        tx['type'], tx['matched_ref'], tx['match_basis'], tx['match_notes'],
    ] for tx in classified] or [[batch_id, recon_date, account, period_label,
                                 opening_balance, closing_balance, remarks,
                                 '', '', '', 0, 0, 0, '', '', '', '']])
    periods = _load_recon_periods(txns_ws)

    s = _xl_styles()
    wb = openpyxl.Workbook()
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']

    # Rebuild every stored period in chronological order. The _sheet_*
    # builders append below existing content (ws.max_row + 2) when the sheet
    # already exists, so periods stack naturally. Historical periods use their
    # STORED classified values — never re-matched, so past results don't shift
    # as live sheet data changes. Mapped is built once (it's a full-book UTR
    # map with no period filter — per-period calls would duplicate the book).
    total_dr = total_cr = 0.0
    for p in periods:
        p_label = p['period_label']  # month scope, e.g. "Jul 2026" — drives filtering
        # Daily sections need the day in the header, otherwise consecutive
        # saves within a month are indistinguishable.
        day_label = f"{p['recon_date']} ({p_label})"
        sheet_label = f"{day_label} — {p['account']}" if p['account'] else day_label
        is_current = (p['recon_date'] == recon_date and p['account'] == account)
        # Computed before the Recon Log append below, so "last closing"
        # correctly reflects the state BEFORE this save, not after.
        current_tieout = (_account_tieout(records, mc_rows, account, p_label, p['txns'], p['closing'],
                                          recon_date=p['recon_date'])
                           if is_current else None)
        dr, cr = _sheet_statement(wb, sheet_label, p['remarks'], p['opening'],
                                  p['closing'], p['txns'], s, tieout=current_tieout)
        if is_current:
            total_dr, total_cr = dr, cr
        _sheet_disb_recon(wb, day_label, records, p['txns'], s, p_label,
                          rm if is_current else {})
        _sheet_coll_recon(wb, day_label, records, mc_rows, p['txns'], s, p_label,
                          rm if is_current else {})
        _sheet_expenses(wb, day_label, p['txns'], s)
        _sheet_mapped(wb, day_label, records, mc_rows, p['txns'], s, p_label)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    try:
        log_ws = sh.worksheet(RECON_LOG_SHEET_NAME)
        log_ws.append_row([recon_date, opening_balance, closing_balance,
                            datetime.today().strftime('%d-%m-%Y %H:%M'), account])
    except Exception as e:
        print(f"[Recon Log] append failed: {e}")

    # Named by month, not by account — the workbook already combines every
    # account's periods into one file (the rebuild loop above has no account
    # filter), so an account-based filename was purely cosmetic and made two
    # accounts' reconciliations look like two separate files when they were
    # never actually split apart.
    filename = f"Daily Reconciliation - {period_label}.xlsx"

    return {
        'total_debit':  total_dr,
        'total_credit': total_cr,
        'closing':      closing_balance,
        'rows_saved':   len(transactions),
        # Rows this upload rewrote, so a wider statement replacing narrower
        # ones is reported rather than happening silently.
        'rows_superseded': superseded_rows,
        'filename':     filename,
        'file_bytes':   buf.getvalue(),
    }

# ── Flask app ─────────────────────────────────────────────────────────────────

app = Flask(__name__)
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16 MB upload limit

ICONS_DIR = os.path.join(os.path.dirname(__file__), 'assets', 'icons')

@app.route('/manifest.json')
def api_manifest():
    return jsonify({
        "name": "BridgeLine Accounts",
        "short_name": "BridgeLine",
        "start_url": "/",
        "display": "standalone",
        "background_color": "#1a3a5c",
        "theme_color": "#1a3a5c",
        "icons": [
            {"src": "/icons/icon-192.png", "sizes": "192x192", "type": "image/png", "purpose": "any maskable"},
            {"src": "/icons/icon-512.png", "sizes": "512x512", "type": "image/png", "purpose": "any maskable"},
        ],
    })

@app.route('/icons/<path:filename>')
def api_icons(filename):
    return send_from_directory(ICONS_DIR, filename)

SETUP_HTML = """<!DOCTYPE html><html><head><meta charset='UTF-8'>
<title>BridgeLine — Setup</title>
<style>body{font-family:Arial;max-width:680px;margin:40px auto;padding:20px;background:#eef2f7}
h1{color:#1a3a5c}
.step{background:white;border-radius:8px;padding:18px;margin:12px 0;box-shadow:0 1px 4px rgba(0,0,0,.1)}
a{color:#2d5986}
</style></head><body>
<h1>BridgeLine Widget</h1>
<div class='step'>This widget is hosted and always-on — there's nothing to set up.
Google Sheets access is configured once via environment variables on the server,
not per-device. If you're seeing an authorization error, contact the admin to
check the hosting environment's Google credentials rather than re-running any
local setup.</div>
<p><a href='/'>← Back to widget</a></p>
</body></html>"""

HTML = """<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>BridgeLine Accounts</title>
<link rel="manifest" href="/manifest.json">
<meta name="theme-color" content="#1a3a5c">
<link rel="icon" type="image/png" sizes="32x32" href="/icons/favicon-32.png">
<link rel="icon" type="image/png" sizes="192x192" href="/icons/icon-192.png">
<link rel="apple-touch-icon" href="/icons/apple-touch-icon.png">
<meta name="apple-mobile-web-app-capable" content="yes">
<meta name="apple-mobile-web-app-status-bar-style" content="black-translucent">
<meta name="apple-mobile-web-app-title" content="BridgeLine">
<style>
  *{box-sizing:border-box;margin:0;padding:0}
  body{font-family:Arial,sans-serif;background:#eef2f7;color:#222}
  header{background:#1a3a5c;color:white;padding:18px 24px;display:flex;justify-content:space-between;align-items:center}
  header h1{font-size:1.4rem}
  header p{font-size:.85rem;color:#a8c4e0;margin-top:2px}
  header a{color:#a8c4e0;font-size:.8rem;text-decoration:none}
  .container{max-width:1360px;margin:24px auto;padding:0 16px}
  .tabs{display:flex;gap:6px;margin-bottom:0;flex-wrap:wrap}
  .tab-btn{padding:10px 18px;border:none;border-radius:8px 8px 0 0;background:#c8d8ec;color:#1a3a5c;font-size:.88rem;cursor:pointer;font-weight:600}
  .tab-btn.active{background:white}
  .tab-content{display:none;background:white;border-radius:0 8px 8px 8px;padding:24px;box-shadow:0 2px 8px rgba(0,0,0,.1)}
  .tab-content.active{display:block}
  .section{border:1px solid #d0dce8;border-radius:6px;padding:14px 16px;margin-bottom:16px;background:#f8fbff}
  .section h3{font-size:.82rem;color:#2d5986;font-weight:700;text-transform:uppercase;letter-spacing:.5px;margin-bottom:10px}
  textarea{width:100%;border:1px solid #b0c4d8;border-radius:4px;padding:8px;font-size:.88rem;resize:vertical;min-height:80px;font-family:Arial}
  .grid{display:grid;grid-template-columns:1fr 1fr;gap:12px}
  .field{display:flex;flex-direction:column;gap:4px}
  .field label{font-size:.8rem;font-weight:600;color:#444}
  input,select{width:100%;padding:7px 10px;border:1px solid #b0c4d8;border-radius:4px;font-size:.9rem;background:white}
  input:focus,select:focus{outline:none;border-color:#2d5986;box-shadow:0 0 0 2px #d0dce8}
  input[readonly]{background:#f0f4f8;color:#555}
  .btn{padding:9px 16px;border:none;border-radius:5px;font-size:.88rem;cursor:pointer;font-weight:600}
  .btn-extract{background:#e8f0fe;color:#1a3a5c;margin-top:8px}
  .btn-extract:hover{background:#c8d8f8}
  .btn-save{background:#1a5c3a;color:white;width:100%;padding:13px;font-size:1rem;margin-top:8px;border-radius:6px}
  .btn-save:hover{background:#14472d}
  .btn-lookup{background:#2d5986;color:white;padding:7px 14px}
  .lookup-row{display:flex;gap:8px;align-items:center}
  .lookup-row input{flex:1}
  .info-box{background:#e8f4ec;border:1px solid #a8d4b8;border-radius:4px;padding:8px 12px;font-size:.85rem;color:#1a5c3a;margin-top:8px;display:none}
  .info-box.error{background:#fde8e8;border-color:#f5b0b0;color:#a00}
  .status{margin-top:10px;padding:10px 14px;border-radius:5px;font-size:.88rem;display:none}
  .status.success{background:#e8f4ec;color:#1a5c3a;border:1px solid #a8d4b8}
  .status.error{background:#fde8e8;color:#a00;border:1px solid #f5b0b0}
  .hint{font-size:.75rem;color:#888;margin-top:2px}
  .spinner{display:none;color:#888;font-size:.85rem;margin-top:6px}
  /* Summary banner */
  .summary-bar{display:grid;grid-template-columns:repeat(5,1fr);gap:8px;margin-bottom:14px}
  .sum-card{background:white;border-radius:8px;padding:12px 14px;box-shadow:0 1px 4px rgba(0,0,0,.1);text-align:center}
  .sum-card .val{font-size:1.1rem;font-weight:700;color:#1a3a5c}
  .sum-card .lbl{font-size:.72rem;color:#888;margin-top:2px}
  .sum-card.clickable{cursor:pointer;transition:transform .08s,box-shadow .08s}
  .sum-card.clickable:hover{transform:translateY(-2px);box-shadow:0 4px 10px rgba(0,0,0,.15)}
  .sum-modal-bg{display:none;position:fixed;inset:0;background:rgba(0,0,0,.45);z-index:999;align-items:center;justify-content:center}
  .sum-modal-bg.show{display:flex}
  .sum-modal{background:#fff;border-radius:10px;max-width:480px;width:92%;max-height:75vh;overflow-y:auto;box-shadow:0 8px 30px rgba(0,0,0,.3)}
  .sum-modal-head{background:#1a3a5c;color:#fff;padding:14px 18px;border-radius:10px 10px 0 0;display:flex;justify-content:space-between;align-items:center}
  .sum-modal-head h3{margin:0;font-size:1rem}
  .sum-modal-head button{background:none;border:none;color:#fff;font-size:1.3rem;cursor:pointer;line-height:1}
  .sum-modal table{width:100%;font-size:.85rem;border-collapse:collapse}
  .sum-modal th{text-align:left;padding:8px 14px;background:#f0f4f8;border-bottom:1px solid #dde}
  .sum-modal td{padding:7px 14px;border-bottom:1px solid #eef}
  .sum-modal tr:hover td{background:#f8fbff}
  .sum-modal .empty{padding:20px;text-align:center;color:#888;font-size:.85rem}
  .pdf-modal-bg{display:none;position:fixed;inset:0;background:rgba(0,0,0,.55);z-index:1000;align-items:center;justify-content:center}
  .pdf-modal-bg.show{display:flex}
  .pdf-modal{background:#fff;border-radius:10px;width:95%;max-width:1040px;height:92vh;overflow:hidden;box-shadow:0 8px 30px rgba(0,0,0,.35);display:flex;flex-direction:column}
  .pdf-modal-head{background:#1a3a5c;color:#fff;padding:12px 16px;display:flex;justify-content:space-between;align-items:center;flex-shrink:0}
  .pdf-modal-head h3{margin:0;font-size:.95rem;font-weight:600}
  .pdf-modal-head .acts{display:flex;gap:8px;align-items:center}
  .pdf-modal-head button{background:rgba(255,255,255,.15);border:none;color:#fff;font-size:.82rem;cursor:pointer;padding:6px 12px;border-radius:6px}
  .pdf-modal-head button:hover{background:rgba(255,255,255,.28)}
  .pdf-modal-head .x{background:none;font-size:1.3rem;padding:0 6px;line-height:1}
  .pdf-modal iframe{flex:1;width:100%;border:0}
  /* Calculator */
  .calc-display{background:#1a3a5c;color:#fff;font-size:2rem;text-align:right;padding:16px 20px;border-radius:8px;margin-bottom:12px;min-height:70px;word-break:break-all;line-height:1.2}
  .calc-display .calc-expr{font-size:.9rem;color:#a8c4e0;min-height:22px;margin-bottom:4px}
  .calc-keys{display:grid;grid-template-columns:repeat(4,1fr);gap:8px}
  .calc-key{padding:16px 8px;border:none;border-radius:6px;font-size:1.1rem;font-weight:700;cursor:pointer;transition:opacity .1s}
  .calc-key:active{opacity:.7}
  .k-num{background:#e8f0fe;color:#1a3a5c}
  .k-op{background:#2d5986;color:white}
  .k-eq{background:#1a5c3a;color:white}
  .k-clear{background:#c00;color:white}
  .k-back{background:#b8860b;color:white}
  .k-zero{grid-column:span 2}
  /* Reconciliation */
  .recon-summary{display:grid;grid-template-columns:repeat(4,1fr);gap:8px;margin:12px 0}
  .recon-card{background:#1a3a5c;color:white;border-radius:8px;padding:12px;text-align:center}
  .recon-card .rv{font-size:1.1rem;font-weight:700;margin-bottom:2px}
  .recon-card .rl{font-size:.72rem;color:#a8c4e0}
  .recon-table{width:100%;font-size:.82rem;border-collapse:collapse;margin-top:10px}
  .recon-table th{background:#1a3a5c;color:white;padding:7px 10px;text-align:left}
  .recon-table td{padding:6px 10px;border-bottom:1px solid #eef2f7}
  .recon-table tr:hover td{background:#f8fbff}
  .upload-area{border:2px dashed #b0c4d8;border-radius:8px;padding:30px;text-align:center;color:#888;cursor:pointer;margin-bottom:12px}
  .upload-area:hover{border-color:#2d5986;color:#2d5986;background:#f0f4fc}
  .upload-area input[type=file]{display:none}
  /* Settings */
  .settings-grid{display:grid;grid-template-columns:1fr 1fr;gap:10px}
  /* Date picker */
  .date-wrap{position:relative}
  .date-wrap input[type=text]{padding-right:32px;box-sizing:border-box;width:100%}
  .date-cal-btn{position:absolute;right:4px;top:50%;transform:translateY(-50%);background:none;border:none;cursor:pointer;font-size:1.05rem;color:#2d5986;padding:4px;line-height:1}
  .date-cal-btn:hover{color:#1a3a5c}
  .date-native{position:absolute;right:4px;top:50%;transform:translateY(-50%);width:24px;height:24px;opacity:0;border:none;padding:0}
</style>
</head>
<body>
<header>
  <div><h1>BridgeLine Partners</h1><p>Accounts Entry Widget</p></div>
  <a href="/setup" target="_blank">⚙ Setup / Help</a>
</header>
<div class="container">
  <!-- Summary Banner -->
  <div class="summary-bar" id="summary-bar">
    <div class="sum-card clickable" onclick="openSumModal('disb')"><div class="val" id="s-disb">—</div><div class="lbl">Disbursed Today</div></div>
    <div class="sum-card clickable" onclick="openSumModal('coll')"><div class="val" id="s-coll">—</div><div class="lbl">Collected Today</div></div>
    <div class="sum-card clickable" onclick="openSumModal('out')"><div class="val" id="s-out">—</div><div class="lbl">Total Outstanding</div></div>
    <div class="sum-card clickable" onclick="openSumModal('avail')"><div class="val" id="s-avail">—</div><div class="lbl">Available to Disburse</div></div>
    <div class="sum-card"><div class="val" id="s-date">—</div><div class="lbl">As of</div></div>
  </div>

  <div class="sum-modal-bg" id="sum-modal-bg" onclick="if(event.target===this) closeSumModal()">
    <div class="sum-modal">
      <div class="sum-modal-head"><h3 id="sum-modal-title">Details</h3><button onclick="closeSumModal()">&times;</button></div>
      <div id="sum-modal-body"></div>
    </div>
  </div>

  <!-- Report preview. Top-level sibling, NOT nested inside the summary
       modal — nesting it there would inherit that modal's display:none and
       it could never open. -->
  <div class="pdf-modal-bg" id="pdf-modal-bg" onclick="if(event.target===this) closePdfModal()">
    <div class="pdf-modal">
      <div class="pdf-modal-head">
        <h3 id="pdf-modal-title">Report</h3>
        <div class="acts">
          <!-- Inline iframe PDF rendering is unreliable on some browsers
               (notably iOS Safari, which may show a blank frame), so there
               is always an escape hatch to the browser's own PDF viewer. -->
          <button onclick="openViewedPdfInTab()">↗ Open in tab</button>
          <button onclick="downloadViewedPdf()">⬇ Download</button>
          <button class="x" onclick="closePdfModal()">&times;</button>
        </div>
      </div>
      <iframe id="pdf-frame" title="Report preview"></iframe>
    </div>
  </div>

  <div class="tabs">
    <button class="tab-btn active" onclick="showTab('disb',this); loadPendingRequests()">➕ New Disbursement</button>
    <button class="tab-btn" onclick="showTab('repa',this); loadOpenCases()">💰 Repayment</button>
    <button class="tab-btn" onclick="showTab('bulk',this); loadBulkRequests()">💸 Bank File</button>
    <button class="tab-btn" onclick="showTab('calc',this)">🧮 Calculator</button>
    <button class="tab-btn" onclick="showTab('recon',this)">🏦 Bank Reconciliation</button>
    <button class="tab-btn" onclick="showTab('invoice',this); loadInvoiceCases()">📄 Invoice</button>
    <button class="tab-btn" onclick="showTab('reports',this); loadReports()">📊 Reports</button>
    <button class="tab-btn" onclick="showTab('edit',this); loadRecentCases()">✏️ Edit Case</button>
    <button class="tab-btn" onclick="showTab('contacts',this); loadContacts()">👥 Contacts</button>
    <button class="tab-btn" onclick="showTab('settings',this); loadSettings()">⚙ Settings</button>
    <button class="tab-btn" id="mis-btn" onclick="generateMis()" style="background:#1a3a5c;color:#fff">📦 Generate MIS Package</button>
  </div>
  <div id="mis-status" style="padding:0 18px;font-size:.85rem"></div>

  <!-- DISBURSEMENT TAB -->
  <div id="disb" class="tab-content active">
    <div class="section" id="pending-requests-section">
      <h3 style="cursor:pointer;user-select:none;" onclick="togglePendingReqs()">
        📥 Pending Field Requests <span id="pending-count" style="font-size:13px;color:#888;font-weight:400;"></span>
        <span id="pending-chevron" style="float:right;font-size:13px;">▼</span>
      </h3>
      <div id="pending-requests-body">
        <div id="pending-requests-list" style="margin-top:8px;"></div>
        <div id="pending-selected-info" style="display:none;margin-top:10px;padding:10px 12px;background:#f0f7ff;border:1px solid #b3d4f5;border-radius:6px;font-size:13px;line-height:1.6;"></div>
        <input type="hidden" id="d-request-id">
      </div>
    </div>
    <div class="section">
      <h3>1 — Paste Customer / Disbursement Note</h3>
      <textarea id="d-msg" placeholder="Paste customer note, WhatsApp message, or disbursement details here..."></textarea>
      <button class="btn btn-extract" onclick="extractDisb()">⚡ Extract Details</button>
    </div>
    <div class="section">
      <h3>2 — Paste Bank Confirmation (for UTR)</h3>
      <textarea id="d-utr-msg" placeholder="Paste your bank's outgoing transfer SMS here to extract UTR..."></textarea>
      <button class="btn btn-extract" onclick="extractDisbUTR()">⚡ Extract UTR</button>
    </div>
    <div class="section">
      <h3>Disbursement Details</h3>
      <div class="grid">
        <div class="field"><label>Date (DD-MM-YYYY) *</label>
          <div class="date-wrap">
            <input type="text" id="d-date">
            <input type="date" class="date-native" id="d-date-native" onchange="_pickDate('d-date', this.value)">
            <button type="button" class="date-cal-btn" onclick="_openDatePicker('d-date-native')">📅</button>
          </div>
        </div>
        <div class="field"><label>Customer Name *</label><input type="text" id="d-customer" placeholder="Full name"></div>
        <div class="field"><label>Company *</label>
          <input type="text" id="d-company" list="company-list" placeholder="Select or type new...">
          <datalist id="company-list">""" + \
          "".join(f'<option value="{c}">' for c in COMPANIES) + """</datalist></div>
        <div class="field"><label>Cluster *</label>
          <input type="text" id="d-cluster" list="cluster-list" placeholder="Select or type new...">
          <datalist id="cluster-list">""" + \
          "".join(f'<option value="{c}">' for c in CLUSTERS) + """</datalist></div>
        <div class="field"><label>Branch *</label>
          <input type="text" id="d-branch" list="branch-list" placeholder="Select or type new...">
          <datalist id="branch-list">""" + \
          "".join(f'<option value="{b}">' for b in BRANCHES) + """</datalist></div>
        <div class="field"><label>Cheque No. (optional)</label><input type="text" id="d-chq"></div>
        <div class="field"><label>Amount (₹) *</label>
          <input type="number" id="d-amount" placeholder="e.g. 500000" oninput="calcCharges()"></div>
        <div class="field"><label>Charges</label>
          <input type="text" id="d-charges" readonly><span class="hint">Auto-calculated</span></div>
        <div class="field"><label>GST</label>
          <input type="text" id="d-gst" readonly><span class="hint">Auto-calculated</span></div>
        <div class="field"><label>Total Receivable (₹)</label>
          <input type="text" id="d-total" readonly><span class="hint">Auto-calculated</span></div>
        <div class="field"><label>Serviced Branch</label>
          <input type="text" id="d-srv-branch" list="srv-branch-list" placeholder="Select or type new...">
          <datalist id="srv-branch-list">""" + \
          "".join(f'<option value="{b}">' for b in BRANCHES) + """</datalist></div>
        <div class="field"><label>Serviced Cluster</label>
          <input type="text" id="d-srv-cluster" list="srv-cluster-list" placeholder="Select or type new...">
          <datalist id="srv-cluster-list">""" + \
          "".join(f'<option value="{c}">' for c in CLUSTERS) + """</datalist></div>
        <div class="field" style="grid-column:1/-1"><label>Disbursement UTR / Debit Note</label>
          <input type="text" id="d-utr" placeholder="Auto-extracted or enter manually"></div>
        <div class="field" style="grid-column:1/-1"><label>Bank Account (debited from)</label>
          <input type="text" id="d-bank-account" list="d-bank-account-list" placeholder="Auto-filled from Request, or select/type...">
          <datalist id="d-bank-account-list"></datalist></div>
        <div class="field" style="grid-column:1/-1"><label>Remarks</label>
          <input type="text" id="d-remarks" placeholder="e.g. Urgent case, referred by X, special rate approved…"></div>
      </div>
    </div>
    <button class="btn btn-save" onclick="saveDisb()">✅ Save Disbursement to Google Sheet</button>
    <div id="d-status" class="status"></div>
  </div>

  <!-- REPAYMENT TAB -->
  <div id="repa" class="tab-content">
    <div class="section">
      <h3>Paste Bank SMS / Payment Message</h3>
      <textarea id="r-msg" placeholder="Paste HDFC/SBI/IMPS bank message here..."></textarea>
      <button class="btn btn-extract" onclick="extractRepa()">⚡ Extract from Message</button>
    </div>
    <div class="section">
      <h3>Select Open Case</h3>
      <select id="r-open-cases" onchange="onCaseSelect()" style="width:100%;font-size:.85rem">
        <option value="">Loading open cases...</option>
      </select>
      <div class="spinner" id="cases-spinner">Loading...</div>
      <div id="r-info" class="info-box"></div>
    </div>
    <div class="section">
      <h3>Repayment Details</h3>
      <div class="grid">
        <div class="field"><label>Collection Date (DD-MM-YYYY) *</label>
          <div class="date-wrap">
            <input type="text" id="r-date">
            <input type="date" class="date-native" id="r-date-native" onchange="_pickDate('r-date', this.value)">
            <button type="button" class="date-cal-btn" onclick="_openDatePicker('r-date-native')">📅</button>
          </div>
        </div>
        <div class="field"><label>Amount Received (₹) *</label>
          <input type="number" id="r-amount" placeholder="e.g. 603540"></div>
        <div class="field"><label>UTR / Reference No.</label>
          <input type="text" id="r-utr" placeholder="e.g. BKIDR52026..."></div>
        <div class="field"><label>Discount (₹)</label>
          <input type="number" id="r-discount" placeholder="0" value="0"></div>
        <div class="field" style="grid-column:1/-1"><label>Bank Account (credited to)</label>
          <input type="text" id="r-bank-account" list="r-bank-account-list" placeholder="Select or type...">
          <datalist id="r-bank-account-list"></datalist></div>
        <div class="field" style="grid-column:1/-1"><label>Remarks</label>
          <input type="text" id="r-remarks" placeholder="e.g. Part payment, cheque cleared, customer requested receipt…"></div>
      </div>
    </div>
    <button class="btn btn-save" onclick="saveRepa()">✅ Save Repayment to Google Sheet</button>
    <div id="r-status" class="status"></div>
  </div>

  <!-- CALCULATOR TAB -->
  <div id="calc" class="tab-content">
    <div style="max-width:360px;margin:0 auto">
      <div class="calc-display">
        <div class="calc-expr" id="calc-expr"></div>
        <div id="calc-disp">0</div>
      </div>
      <div class="calc-keys">
        <button class="calc-key k-clear" onclick="calcClear()">C</button>
        <button class="calc-key k-back"  onclick="calcBack()">⌫</button>
        <button class="calc-key k-op"    onclick="calcOp('%')">%</button>
        <button class="calc-key k-op"    onclick="calcOp('/')">÷</button>

        <button class="calc-key k-num"   onclick="calcNum('7')">7</button>
        <button class="calc-key k-num"   onclick="calcNum('8')">8</button>
        <button class="calc-key k-num"   onclick="calcNum('9')">9</button>
        <button class="calc-key k-op"    onclick="calcOp('*')">×</button>

        <button class="calc-key k-num"   onclick="calcNum('4')">4</button>
        <button class="calc-key k-num"   onclick="calcNum('5')">5</button>
        <button class="calc-key k-num"   onclick="calcNum('6')">6</button>
        <button class="calc-key k-op"    onclick="calcOp('-')">−</button>

        <button class="calc-key k-num"   onclick="calcNum('1')">1</button>
        <button class="calc-key k-num"   onclick="calcNum('2')">2</button>
        <button class="calc-key k-num"   onclick="calcNum('3')">3</button>
        <button class="calc-key k-op"    onclick="calcOp('+')">+</button>

        <button class="calc-key k-num k-zero" onclick="calcNum('0')">0</button>
        <button class="calc-key k-num"   onclick="calcNum('.')">.</button>
        <button class="calc-key k-eq"    onclick="calcEq()">=</button>
      </div>
      <div style="margin-top:14px;text-align:center;font-size:.8rem;color:#888">
        Tip: results are auto-formatted in Indian numbering (₹ lakhs / crores)
      </div>
    </div>
  </div>

  <!-- BANK RECONCILIATION TAB -->
  <div id="recon" class="tab-content">
    <div class="section">
      <h3>Bank Statement Upload</h3>
      <p style="font-size:.82rem;color:#555;margin:-4px 0 12px;line-height:1.6">
        Upload <b>every account's statement for the same period together</b>. Transfers between your own
        accounts are then matched by structure — same amount, opposite direction, different account — instead
        of being guessed from the narration text, which the banks keep rewording.
        Each file's <b>account and date are read from the statement itself</b>.
      </p>
      <div class="field" style="margin-bottom:12px">
        <label>Remarks <span style="color:#888;font-weight:400">(optional, applies to every statement in this upload)</span></label>
        <input type="text" id="rec-remarks" placeholder="e.g. month-end close">
      </div>
      <datalist id="rec-account-list"></datalist>
      <div class="upload-area" id="upload-area" onclick="document.getElementById('rec-file').click()">
        <input type="file" id="rec-file" accept=".csv,.xlsx,.xls,.pdf" multiple onchange="onFileSelect(this)">
        <div id="upload-label">📂 Click to upload bank statements<br><span style="font-size:.78rem">Select all accounts at once — CSV, Excel (.xlsx/.xls) or PDF</span></div>
      </div>
      <div id="selected-files" style="display:none;margin:10px 0 0;font-size:.82rem"></div>
      <button class="btn btn-extract" style="margin:0" onclick="parseStatement()">⚡ Parse Statements</button>
    </div>

    <!-- Reconciliation output -->
    <div id="recon-preview-section" style="display:none">

      <!-- Own-account transfers paired across the uploaded statements -->
      <div class="section" id="cross-account-section" style="display:none;background:#eef5fd;border-color:#a9c8e8">
        <h3 style="color:#12467e">🔄 Own-Account Transfers Matched <span id="cross-badge" style="font-weight:400;color:#555"></span></h3>
        <p style="font-size:.8rem;color:#555;margin:0 0 10px">Both legs found in this upload and typed <b>Contra</b> automatically — no review needed, and they can't distort profit or the contra balance.</p>
        <div id="cross-account-body" style="font-size:.83rem"></div>
      </div>

      <!-- Statements are reconciled to different dates -->
      <div id="account-gap-note" style="display:none;margin-bottom:14px;padding:10px 12px;border:1px solid #e0a0a0;background:#fdecea;color:#8a2020;border-radius:6px;font-size:.82rem"></div>
      <div id="date-mismatch-note" style="display:none;margin-bottom:14px;padding:10px 12px;border:1px solid #e0c060;background:#fff3cd;color:#7a5c00;border-radius:6px;font-size:.82rem"></div>

      <!-- One block per uploaded statement, built by renderReconResult() -->
      <div id="recon-statements"></div>

      <!-- Review queue — uncertain entries across every statement -->
      <div class="section" id="review-section" style="display:none">
        <h3>⚠️ Needs Your Input <span id="review-badge" style="font-weight:400;color:#888"></span></h3>
        <p style="font-size:.8rem;color:#666;margin:0 0 10px">These couldn't be confidently matched. Set the correct <b>Type</b> and optionally add <b>Remarks</b> — then save.</p>
        <div style="overflow-x:auto">
          <table class="recon-table">
            <thead><tr>
              <th style="width:110px">Account</th>
              <th>Date</th><th>Description</th><th>UTR</th>
              <th style="text-align:right">Amount</th>
              <th style="width:60px">Dr/Cr</th>
              <th style="width:80px">Auto Type</th>
              <th style="width:190px">Correct Type</th>
              <th style="min-width:260px">Remarks</th>
            </tr></thead>
            <tbody id="review-tbody"></tbody>
          </table>
        </div>
      </div>

      <p style="font-size:.8rem;color:#666;margin:8px 0 4px">History is stored automatically — every download contains <b>all periods saved so far</b>. No need to re-upload previous files.</p>
      <button class="btn btn-save" onclick="saveRecon()" style="margin-top:4px">💾 Complete Reconciliation &amp; Save Excel</button>
      <div id="recon-status" class="status"></div>
    </div>
  </div>

  <!-- SETTINGS TAB -->
  <div id="contacts" class="tab-content">
    <div class="section">
      <div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:14px;flex-wrap:wrap;gap:8px">
        <h3 style="margin:0">Staff Directory</h3>
        <div style="display:flex;gap:8px;align-items:center;flex-wrap:wrap">
          <input type="text" id="contact-search" placeholder="Search…"
            oninput="filterContacts(this.value)"
            style="padding:6px 10px;border:1px solid #ccd;border-radius:6px;font-size:.85rem;width:180px">
          <button class="btn" onclick="addContactRow()" style="padding:6px 14px;font-size:.82rem;background:#1a3a5c;color:#fff;border:none;border-radius:6px;cursor:pointer">+ Add Person</button>
          <button class="btn" onclick="addCluster()" style="padding:6px 14px;font-size:.82rem;background:#2e7d32;color:#fff;border:none;border-radius:6px;cursor:pointer">+ Add Cluster</button>
          <button class="btn btn-save" id="contacts-save-btn" onclick="saveContacts()" style="padding:6px 14px;font-size:.82rem;display:none">💾 Save</button>
        </div>
      </div>
      <div id="contacts-status" class="status"></div>
      <div id="contacts-body"></div>
    </div>
  </div>

  <!-- BANK FILE TAB -->
  <div id="bulk" class="tab-content">
    <div class="section">
      <h3>📥 Pending Requests <span id="bulk-count" style="font-size:13px;color:#888;font-weight:400;"></span>
        <button onclick="loadBulkRequests()" style="float:right;font-size:12px;padding:3px 10px;cursor:pointer;border:1px solid #ccc;border-radius:4px;background:#fff;">↻ Refresh</button></h3>
      <div style="overflow-x:auto;">
        <table id="bulk-table" style="width:100%;border-collapse:collapse;font-size:13px;display:none;">
          <thead><tr style="text-align:left;border-bottom:2px solid #dce8f5;">
            <th style="padding:6px 8px;"><input type="checkbox" id="bulk-select-all" onchange="toggleBulkSelectAll(this.checked)"></th>
            <th style="padding:6px 8px;">Request ID</th><th style="padding:6px 8px;">Submitted</th>
            <th style="padding:6px 8px;">Customer</th><th style="padding:6px 8px;">Cluster / Branch</th>
            <th style="padding:6px 8px;">Amount</th><th style="padding:6px 8px;">Account</th>
            <th style="padding:6px 8px;">IFSC</th><th style="padding:6px 8px;">Bank</th>
            <th style="padding:6px 8px;">Phone</th><th style="padding:6px 8px;">SO</th>
            <th style="padding:6px 8px;">Gold g</th>
          </tr></thead>
          <tbody id="bulk-tbody"></tbody>
        </table>
      </div>
      <div id="bulk-empty" style="color:#aaa;font-size:13px;">Loading...</div>
      <div id="bulk-footer" style="display:none;margin-top:10px;padding:8px 12px;background:#f0f7ff;border:1px solid #b3d4f5;border-radius:6px;font-size:13px;font-weight:600;"></div>
      <button class="btn" style="margin-top:10px;width:auto;padding:8px 18px;background:#fff;color:#c0392b;border:1px solid #e0b4b4;" onclick="deleteBulkRequests()">🗑 Delete Selected</button>
      <div id="bulk-delete-status" class="status"></div>
    </div>
    <div class="section">
      <h3>⚙ Export Bank File</h3>
      <div class="grid">
        <div class="field"><label>Bank Format *</label>
          <select id="bulk-bank" style="width:100%;padding:9px;border:1px solid #ccc;border-radius:6px;font-size:.9rem;">""" + \
          "".join(f'<option value="{k}">{v["label"]} (.{v["filetype"]})</option>' for k, v in BANK_TEMPLATES.items()) + """</select></div>
        <div class="field"><label>Debit Account</label>
          <input type="text" id="bulk-debit" placeholder="Your BridgeLine account no."></div>
        <div class="field"><label>Narration / Reference</label>
          <input type="text" id="bulk-narration" maxlength="20" placeholder="Shows on beneficiary statement"></div>
        <div class="field"><label>Value Date (DD/MM/YYYY)</label>
          <input type="text" id="bulk-value-date"></div>
      </div>
      <button class="btn btn-save" style="margin-top:12px" onclick="exportBulk()">💸 Generate &amp; Download</button>
      <div id="bulk-status" class="status"></div>
      <p style="font-size:12px;color:#888;margin-top:8px;">Selected requests are marked <b>Exported</b> in the sheet after download — they leave this list but stay visible in New Disbursement for recording. RTGS is auto-assigned for amounts ≥ ₹2,00,000, NEFT below.</p>
    </div>
    <div class="section">
      <h3>📋 Paste WhatsApp Request</h3>
      <p style="font-size:12px;color:#888;">For requests that arrive as a DISBURSEMENT REQUEST message instead of through the field app. Parsed requests join the same pending queue.</p>
      <textarea id="bulk-paste" placeholder="Paste the DISBURSEMENT REQUEST message here..."></textarea>
      <button class="btn btn-extract" onclick="parseBulkMsg()">⚡ Parse Message</button>
      <div id="bulk-parse-preview" style="display:none;margin-top:12px;">
        <div class="grid">
          <div class="field"><label>Customer Name *</label><input type="text" id="bp-customer"></div>
          <div class="field"><label>Amount (₹) *</label><input type="number" id="bp-amount"></div>
          <div class="field"><label>Account No *</label><input type="text" id="bp-account"></div>
          <div class="field"><label>IFSC *</label><input type="text" id="bp-ifsc" style="text-transform:uppercase"></div>
          <div class="field"><label>Bank Name</label><input type="text" id="bp-bank"></div>
          <div class="field"><label>Phone</label><input type="text" id="bp-phone"></div>
          <div class="field"><label>Cluster</label><input type="text" id="bp-cluster" list="cluster-list"></div>
          <div class="field"><label>Branch</label><input type="text" id="bp-branch" list="branch-list"></div>
          <div class="field"><label>SO Name</label><input type="text" id="bp-so"></div>
          <div class="field"><label>Gold Wt (gms)</label><input type="text" id="bp-gold"></div>
        </div>
        <div id="bulk-parse-warnings" style="margin-top:8px;font-size:12px;color:#b8860b;"></div>
        <button class="btn btn-save" style="margin-top:10px" onclick="addParsedRequest()">➕ Add to Queue</button>
      </div>
      <div id="bulk-paste-status" class="status"></div>
    </div>
  </div>

  <div id="settings" class="tab-content">
    <div class="section">
      <h3>General Settings</h3>
      <div style="display:flex;gap:12px;align-items:center;margin-bottom:12px">
        <label style="font-size:.85rem;font-weight:600">Daily Report Time:</label>
        <input type="time" id="report-time" style="width:120px">
      </div>
      <div style="display:flex;gap:12px;align-items:center;margin-bottom:12px">
        <label style="font-size:.85rem;font-weight:600">Bulk Export Debit Account:</label>
        <input type="text" id="set-bulk-debit" style="width:220px" placeholder="Default debit account no.">
      </div>
      <div style="display:flex;gap:12px;align-items:center;margin-bottom:12px">
        <label style="font-size:.85rem;font-weight:600">Bulk Export Narration:</label>
        <input type="text" id="set-bulk-narration" style="width:220px" maxlength="20">
      </div>
      <button class="btn btn-save" onclick="saveSettings()" style="width:auto;padding:10px 24px">💾 Save Settings</button>
      <div id="settings-status" class="status"></div>
    </div>
    <div class="section">
      <h3>Reconciliation &amp; MIS Reports</h3>
      <p style="font-size:.85rem;color:#555">The Daily Reconciliation Excel and the MIS PDF package are generated fresh on each run and download directly to your browser — keep your own copy if you want a running archive.</p>
    </div>

    <div class="section" id="solvency-section">
      <h3>🧮 Company Solvency Check <span id="solvency-badge" style="font-weight:400;color:#888"></span></h3>
      <p style="font-size:.8rem;color:#666;margin:0 0 10px">Does Net Capital plus net cash from lending minus Expenses actually equal what's really in the bank?</p>
      <div id="solvency-body" style="font-size:.85rem;line-height:1.7"></div>
      <div style="margin-top:14px;padding-top:12px;border-top:1px solid #eee">
        <label style="font-size:.85rem;font-weight:600">Update FD Balance</label>
        <p style="font-size:.78rem;color:#666;margin:2px 0 8px">Enter the real current total from your FD/netbanking summary — trusted directly, same as a bank closing balance, not derived from transaction history.</p>
        <div style="display:flex;gap:8px;align-items:flex-end;flex-wrap:wrap">
          <div class="field" style="flex:0 0 140px;margin:0">
            <label style="font-size:.75rem">Date</label>
            <div class="date-wrap">
              <input type="text" id="fd-date" style="width:100%">
              <input type="date" class="date-native" id="fd-date-native" onchange="_pickDate('fd-date', this.value)">
              <button type="button" class="date-cal-btn" onclick="_openDatePicker('fd-date-native')">📅</button>
            </div>
          </div>
          <div class="field" style="flex:0 0 160px;margin:0">
            <label style="font-size:.75rem">Total FD Balance (₹)</label>
            <input type="number" id="fd-balance" step="0.01">
          </div>
          <button class="btn btn-save" style="width:auto;padding:9px 18px" onclick="saveFdBalance()">💾 Save</button>
        </div>
        <div id="fd-balance-status" class="status"></div>
      </div>
    </div>

    <div class="section" id="integrity-section">
      <h3>🔒 Money Integrity <span id="integrity-badge" style="font-weight:400;color:#888"></span></h3>
      <p style="font-size:.8rem;color:#666;margin:0 0 10px">Every distinct way a rupee could go missing, each proven separately. A single variance figure proves nothing — a bank statement's own arithmetic always balances.</p>
      <div id="integrity-body" style="font-size:.85rem;line-height:1.6"></div>
    </div>

    <div class="section" id="capital-log-section">
      <h3>💰 Capital Log <span id="capital-log-badge" style="font-weight:400;color:#888"></span></h3>
      <p style="font-size:.8rem;color:#666;margin:0 0 10px">Record a partner capital contribution, withdrawal, or correction — appended to the "Capital Log" sheet tab.</p>
      <div class="grid">
        <div class="field"><label>Date (DD-MM-YYYY) *</label>
          <div class="date-wrap">
            <input type="text" id="cl-date">
            <input type="date" class="date-native" id="cl-date-native" onchange="_pickDate('cl-date', this.value)">
            <button type="button" class="date-cal-btn" onclick="_openDatePicker('cl-date-native')">📅</button>
          </div>
        </div>
        <div class="field"><label>Type *</label>
          <select id="cl-type" style="width:100%;padding:9px;border:1px solid #ccc;border-radius:6px;font-size:.9rem;">
            <option value="INFLOW">INFLOW</option>
            <option value="WITHDRAWAL">WITHDRAWAL</option>
            <option value="RECYCLE">RECYCLE</option>
            <option value="ADJUSTMENT">ADJUSTMENT</option>
          </select>
        </div>
        <div class="field"><label>Partner</label>
          <input type="text" id="cl-partner" placeholder="e.g. Harsha, Shivu, Prem"></div>
        <div class="field"><label>Amount (₹) *</label>
          <input type="number" id="cl-amount" step="0.01"></div>
        <div class="field"><label>Reference</label>
          <input type="text" id="cl-reference" placeholder="e.g. bank UTR, cheque no."></div>
        <div class="field" style="grid-column:1/-1"><label>Remarks</label>
          <input type="text" id="cl-remarks"></div>
      </div>
      <button class="btn btn-save" onclick="saveCapitalLogEntry()" style="width:auto;padding:10px 24px">💾 Add Capital Log Entry</button>
      <div id="capital-log-status" class="status"></div>
      <div style="overflow-x:auto;margin-top:14px">
        <table class="recon-table">
          <thead><tr><th>Date</th><th>Type</th><th>Partner</th><th style="text-align:right">Amount</th><th>Reference</th><th>Remarks</th></tr></thead>
          <tbody id="capital-log-tbody"></tbody>
        </table>
      </div>
    </div>
  </div>

  <!-- EDIT CASE TAB -->
  <div id="edit" class="tab-content">
    <div class="section">
      <h3>Load a Case</h3>
      <div style="display:flex;gap:10px;align-items:flex-end">
        <div class="field" style="flex:1"><label>Disbursement ID</label>
          <input type="text" id="e-disb-id" placeholder="e.g. BLP-010726-001" style="text-transform:uppercase" oninput="this.value=this.value.toUpperCase()"></div>
        <button class="btn btn-extract" style="width:auto;padding:10px 24px" onclick="loadEditCase()">🔍 Load Case</button>
      </div>
      <div id="e-load-status" class="status"></div>
    </div>
    <div class="section" id="e-recent-section">
      <h3 style="cursor:pointer;user-select:none;" onclick="toggleRecentCases()">
        🕐 Recently Recorded (last 7 days) <span id="e-recent-count" style="font-size:13px;color:#888;font-weight:400;"></span>
        <span id="e-recent-chevron" style="float:right;font-size:13px;">▼</span>
      </h3>
      <div id="e-recent-body"><div id="e-recent-list" style="margin-top:8px;"></div></div>
    </div>
    <div id="e-body" style="display:none">
      <div class="section">
        <h3>Disbursement Details <span id="e-case-id" style="font-weight:400;color:#888;font-size:13px"></span></h3>
        <div class="grid">
          <div class="field"><label>Date (DD-MM-YYYY)</label><input type="text" id="e-date"></div>
          <div class="field"><label>Customer Name</label><input type="text" id="e-customer"></div>
          <div class="field"><label>Company</label>
            <input type="text" id="e-company" list="company-list"></div>
          <div class="field"><label>Cluster</label>
            <input type="text" id="e-cluster" list="cluster-list"></div>
          <div class="field"><label>Branch</label>
            <input type="text" id="e-branch" list="branch-list"></div>
          <div class="field"><label>Cheque No.</label><input type="text" id="e-chq"></div>
          <div class="field"><label>Amount (₹)</label>
            <input type="number" id="e-amount"><span class="hint">Changing this recomputes charges, GST, total &amp; status</span></div>
          <div class="field"><label>Serviced Branch</label>
            <input type="text" id="e-srv-branch" list="srv-branch-list"></div>
          <div class="field"><label>Serviced Cluster</label>
            <input type="text" id="e-srv-cluster" list="srv-cluster-list"></div>
          <div class="field" style="grid-column:1/-1"><label>Disbursement UTR / Debit Note</label>
            <input type="text" id="e-debit-note"></div>
          <div class="field" style="grid-column:1/-1"><label>Remarks</label>
            <input type="text" id="e-remarks"></div>
        </div>
        <div id="e-computed" style="margin-top:10px;padding:10px 12px;background:#f0f7ff;border:1px solid #b3d4f5;border-radius:6px;font-size:13px;line-height:1.7"></div>
        <button class="btn btn-save" style="margin-top:12px" onclick="saveEditDisb()">💾 Save Disbursement Changes</button>
        <div id="e-disb-status" class="status"></div>
      </div>
      <div class="section">
        <h3>Repayments</h3>
        <div style="overflow-x:auto">
          <table class="recon-table" style="width:100%">
            <thead><tr><th>Date</th><th style="text-align:right">Amount (₹)</th><th>UTR / Note</th><th style="width:130px">Actions</th></tr></thead>
            <tbody id="e-pay-tbody"></tbody>
          </table>
        </div>
        <div style="display:flex;gap:10px;align-items:flex-end;margin-top:14px;padding-top:14px;border-top:1px solid #e0e6ed">
          <div class="field" style="flex:0 0 130px;margin:0"><label>Date</label>
            <input type="text" id="e-new-pay-date" placeholder="DD-MM-YYYY"></div>
          <div class="field" style="flex:0 0 140px;margin:0"><label>Amount (₹)</label>
            <input type="number" id="e-new-pay-amount"></div>
          <div class="field" style="flex:1;margin:0"><label>UTR / Note</label>
            <input type="text" id="e-new-pay-utr"></div>
          <button class="btn btn-save" style="width:auto;padding:10px 20px" onclick="addPayRow()">➕ Add Payment</button>
        </div>
        <div id="e-pay-status" class="status"></div>
      </div>
    </div>
  </div>

  <!-- INVOICE TAB -->
  <div id="invoice" class="tab-content">
    <div class="section">
      <h3>Generate Invoice &amp; Ledger</h3>
      <div class="field" style="margin-bottom:14px">
        <label>Select Open Case</label>
        <select id="inv-open-cases" onchange="onInvCaseSelect()" style="width:100%;font-size:.85rem">
          <option value="">Loading open cases...</option>
        </select>
      </div>
      <div class="field" style="margin-bottom:14px">
        <label>Or enter Disbursement ID manually <span style="font-weight:400;color:#888">(for any case incl. closed)</span></label>
        <input type="text" id="inv-manual-id" placeholder="e.g. BLP-010726-001" style="text-transform:uppercase" oninput="this.value=this.value.toUpperCase()">
      </div>
      <div id="inv-case-info" style="display:none;padding:10px 12px;background:#f0f7ff;border:1px solid #b3d4f5;border-radius:6px;font-size:13px;margin-bottom:14px;line-height:1.6;"></div>
      <button class="btn btn-save" onclick="generateInvoice()" id="inv-btn" style="width:auto;padding:10px 28px">📄 Generate &amp; Download PDF</button>
      <div id="inv-status" class="status" style="margin-top:10px"></div>
    </div>
  </div>

  <div id="reports" class="tab-content">
    <div class="section">
      <h3>📊 Reports <span id="reports-badge" style="font-weight:400;color:#888"></span></h3>
      <p style="font-size:.8rem;color:#666;margin:0 0 12px">Every report downloads as a PDF. Financial reports are cash-basis — charges count when collected, not when invoiced.</p>
      <div class="field" style="max-width:420px;margin-bottom:14px">
        <label>Period <span style="font-weight:400;color:#888">(used by reports that need one)</span></label>
        <div style="display:flex;gap:8px;align-items:center">
          <div class="date-wrap" style="flex:1">
            <input type="text" id="rep-from" placeholder="DD-MM-YYYY">
            <input type="date" id="rep-from-native" class="date-native" onchange="_pickDate('rep-from',this.value)">
            <button type="button" class="date-btn" onclick="_openDatePicker('rep-from-native')">📅</button>
          </div>
          <span style="color:#888;font-size:.8rem">to</span>
          <div class="date-wrap" style="flex:1">
            <input type="text" id="rep-to" placeholder="DD-MM-YYYY">
            <input type="date" id="rep-to-native" class="date-native" onchange="_pickDate('rep-to',this.value)">
            <button type="button" class="date-btn" onclick="_openDatePicker('rep-to-native')">📅</button>
          </div>
        </div>
      </div>
      <div id="reports-list"></div>
      <div id="reports-status" class="status" style="margin-top:10px"></div>
    </div>
  </div>

</div>

<script>
let openCasesData = [];
let _pendingReqsOpen = true;

async function loadPendingRequests() {
  const list = document.getElementById('pending-requests-list');
  const countEl = document.getElementById('pending-count');
  list.innerHTML = '<span style="color:#888;font-size:13px;">Loading...</span>';
  try {
    const items = await (await fetch('/requests/pending?include=exported')).json();
    if (!Array.isArray(items) || items.length === 0) {
      list.innerHTML = '<span style="color:#aaa;font-size:13px;">No pending requests</span>';
      countEl.textContent = '';
      return;
    }
    countEl.textContent = `(${items.length})`;
    list.innerHTML = '';
    items.forEach(r => {
      const card = document.createElement('div');
      card.style.cssText = 'border:1px solid #dce8f5;border-radius:6px;padding:10px 12px;margin-bottom:8px;cursor:pointer;background:#fff;transition:background 0.15s;';
      card.onmouseenter = () => card.style.background = '#f0f7ff';
      card.onmouseleave = () => card.style.background = '#fff';
      const amt = Number(r.amount) > 0 ? '₹' + Number(r.amount).toLocaleString('en-IN', {maximumFractionDigits:0}) : r.amount;
      const badge = r.status === 'Exported'
        ? ' <span style="font-size:11px;background:#fff3cd;color:#856404;border:1px solid #ffeeba;border-radius:4px;padding:1px 6px;font-weight:400;">Exported</span>'
        : '';
      const kycLine = r.kyc_folder
        ? `<a href="${r.kyc_folder}" target="_blank" rel="noopener" class="req-kyc-link" style="font-size:11px;color:#1a7a3a;">📁 KYC Docs</a>`
        : `<span style="font-size:11px;color:#c0392b;">⚠ No KYC docs</span>`;
      card.innerHTML = `<button class="req-del-btn" title="Delete request" style="float:right;font-size:12px;padding:2px 8px;cursor:pointer;border:1px solid #e0b4b4;border-radius:4px;background:#fff;color:#c0392b;">🗑</button>
        <div style="font-weight:600;font-size:14px;">${r.customer}${badge}</div>
        <div style="font-size:12px;color:#555;margin-top:3px;">${r.cluster} · ${r.branch} · ${amt}</div>
        <div style="font-size:11px;color:#999;margin-top:2px;">${r.request_id} · ${r.submitted_at} · ${kycLine}</div>`;
      card.addEventListener('click', () => selectPendingRequest(r));
      card.querySelector('.req-del-btn').addEventListener('click', (ev) => {
        ev.stopPropagation();
        deletePendingRequest(r);
      });
      const kycLinkEl = card.querySelector('.req-kyc-link');
      if (kycLinkEl) kycLinkEl.addEventListener('click', (ev) => ev.stopPropagation());
      list.appendChild(card);
    });
  } catch(e) {
    list.innerHTML = `<span style="color:#c00;font-size:13px;">Error loading requests: ${e.message}</span>`;
  }
}

async function deletePendingRequest(r) {
  const reason = prompt(`Delete ${r.request_id} (${r.customer}, ${r.cluster}/${r.branch}) from the queue?\n\nReason (optional — saved to the sheet):`);
  if (reason === null) return;
  if (!confirm(`Confirm: remove ${r.request_id} from the queue? The row stays in the sheet marked Deleted.`)) return;
  try {
    const res = await (await fetch('/requests/delete', {method:'POST',
      headers:{'Content-Type':'application/json'},
      body: JSON.stringify({request_ids: [r.request_id], reason: reason.trim()})})).json();
    if (!res.ok) throw new Error(res.error || 'Failed');
    // If the deleted request was selected into the form, clear the link so
    // saving the disbursement doesn't try to mark a Deleted row Disbursed.
    if (document.getElementById('d-request-id').value === r.request_id) {
      document.getElementById('d-request-id').value = '';
      document.getElementById('pending-selected-info').style.display = 'none';
    }
    loadPendingRequests();
  } catch(e) {
    alert('Delete failed: ' + e.message);
  }
}

function selectPendingRequest(r) {
  document.getElementById('d-customer').value = r.customer || '';
  document.getElementById('d-company').value  = r.company  || '';
  document.getElementById('d-cluster').value  = r.cluster  || '';
  document.getElementById('d-branch').value   = r.branch   || '';
  if (r.amount) { document.getElementById('d-amount').value = r.amount; calcCharges(); }
  document.getElementById('d-request-id').value = r.request_id;
  // Pure UX prefill — the server-side auto-resolve in save_disbursement()
  // is the actual reliability backstop and works even if this lookup finds
  // no match (e.g. datalist not yet loaded) or the officer changes it.
  const matchedAcct = (_bankAccounts||[]).find(a => a.account_number && a.account_number === r.debit_account);
  document.getElementById('d-bank-account').value = matchedAcct ? matchedAcct.name : '';

  const info = document.getElementById('pending-selected-info');
  const amt = Number(r.amount) > 0 ? '₹' + Number(r.amount).toLocaleString('en-IN', {maximumFractionDigits:0}) : r.amount;
  const kycInfoLine = r.kyc_folder
    ? `<a href="${r.kyc_folder}" target="_blank" rel="noopener" style="color:#1a7a3a;">📁 KYC Docs</a>`
    : `<span style="color:#c0392b;">⚠ No KYC docs</span>`;
  info.innerHTML = `<b>Selected:</b> ${r.customer} &nbsp;·&nbsp; ${r.cluster} / ${r.branch} &nbsp;·&nbsp; ${amt}<br>
    <b>Account:</b> ${r.account_no || '—'} &nbsp; <b>IFSC:</b> ${r.ifsc || '—'}<br>
    <b>Phone:</b> ${r.phone || '—'} &nbsp; <b>SO:</b> ${r.so_name || '—'} &nbsp; <b>Gold:</b> ${r.gold_weight || '—'} gms<br>
    <b>KYC:</b> ${kycInfoLine}<br>
    <span style="color:#2a7ae2;font-size:11px;">${r.request_id}</span>
    <button onclick="clearDisb()" style="float:right;font-size:11px;padding:2px 8px;cursor:pointer;border:1px solid #ccc;border-radius:4px;background:#fff;">✕ Clear</button>`;
  info.style.display = 'block';
  document.getElementById('d-customer').scrollIntoView({ behavior: 'smooth', block: 'center' });
}

// ── Edit Case tab ────────────────────────────────────────────────────────────
let _editCase = null;
let _recentCasesOpen = true;

async function loadRecentCases() {
  const list = document.getElementById('e-recent-list');
  const countEl = document.getElementById('e-recent-count');
  list.innerHTML = '<span style="color:#888;font-size:13px;">Loading...</span>';
  try {
    const r = await _fetchJsonWithTimeout('/case/recent?days=7', {});
    if (!r.ok) throw new Error(r.error || 'Failed to load');
    const events = r.events;
    if (!events.length) {
      list.innerHTML = '<span style="color:#aaa;font-size:13px;">No disbursements or repayments recorded in the last 7 days</span>';
      countEl.textContent = '';
      return;
    }
    countEl.textContent = `(${events.length})`;
    list.innerHTML = '';
    events.forEach(e => {
      const card = document.createElement('div');
      card.style.cssText = 'border:1px solid #dce8f5;border-radius:6px;padding:10px 12px;margin-bottom:8px;cursor:pointer;background:#fff;transition:background 0.15s;';
      card.onmouseenter = () => card.style.background = '#f0f7ff';
      card.onmouseleave = () => card.style.background = '#fff';
      const isDisb = e.type === 'Disbursement';
      const typeTag = isDisb
        ? '<span style="background:#e8f0fe;color:#1a3a5c;font-size:10px;padding:1px 6px;border-radius:3px;">➕ DISBURSEMENT</span>'
        : '<span style="background:#e8f4ec;color:#155724;font-size:10px;padding:1px 6px;border-radius:3px;">💰 REPAYMENT</span>';
      const archTag = e.sheet && e.sheet !== 'Accounts'
        ? ` <span style="background:#fff3cd;color:#8a6500;font-size:10px;padding:1px 6px;border-radius:3px;">📁 ${e.sheet}</span>` : '';
      card.innerHTML = `<div style="display:flex;justify-content:space-between;align-items:flex-start;gap:8px">
          <div style="font-weight:600;font-size:14px;">${e.customer} <span style="font-weight:400;color:#888;font-size:12px">${e.disb_id}</span></div>
          <div style="font-weight:700;font-size:14px;color:${isDisb ? '#1a3a5c' : '#155724'}">₹${fmt(e.amount)}</div>
        </div>
        <div style="margin-top:3px">${typeTag}${archTag}</div>
        <div style="font-size:12px;color:#555;margin-top:5px;">${e.cluster} · ${e.branch} &nbsp;·&nbsp; ${e.status || '—'}${e.utr ? ' &nbsp;·&nbsp; ' + e.utr : ''}</div>
        <div style="font-size:11px;color:#999;margin-top:2px;">${e.date}</div>`;
      card.addEventListener('click', () => {
        document.getElementById('e-disb-id').value = e.disb_id;
        loadEditCase();
      });
      list.appendChild(card);
    });
  } catch(e) {
    list.innerHTML = `<span style="color:#c00;font-size:13px;">Error: ${e.message}</span>`;
  }
}

function toggleRecentCases() {
  _recentCasesOpen = !_recentCasesOpen;
  document.getElementById('e-recent-body').style.display = _recentCasesOpen ? '' : 'none';
  document.getElementById('e-recent-chevron').textContent = _recentCasesOpen ? '▼' : '▶';
}

async function loadEditCase() {
  const id = document.getElementById('e-disb-id').value.trim().toUpperCase();
  if (!id) return showStatus('e-load-status','error','Enter a Disbursement ID.');
  showStatus('e-load-status','success','Loading...');
  try {
    const r = await _fetchJsonWithTimeout(`/case/${encodeURIComponent(id)}/detail`, {});
    if (!r.ok) throw new Error(r.error || 'Not found');
    renderEditCase(r.case);
    showStatus('e-load-status','success',`✅ Loaded ${id}`);
  } catch(e) {
    document.getElementById('e-body').style.display = 'none';
    showStatus('e-load-status','error','❌ ' + e.message);
  }
}

function renderEditCase(c) {
  _editCase = c;
  document.getElementById('e-body').style.display = '';
  document.getElementById('e-case-id').textContent = `— ${c.disb_id} (${c.sheet})`;
  document.getElementById('e-body').scrollIntoView({ behavior: 'smooth', block: 'start' });
  const set = (id,v) => document.getElementById(id).value = v ?? '';
  set('e-date', c.date); set('e-customer', c.customer); set('e-chq', c.chq);
  set('e-company', c.company); set('e-cluster', c.cluster); set('e-branch', c.branch);
  set('e-amount', c.amount); set('e-srv-branch', c.srv_branch); set('e-srv-cluster', c.srv_cluster);
  set('e-debit-note', c.debit_note); set('e-remarks', c.remarks);
  document.getElementById('e-new-pay-date').value = new Date().toLocaleDateString('en-GB').replace(/\//g,'-');
  const kycDocsLine = c.kyc_folder
    ? `<a href="${c.kyc_folder}" target="_blank" rel="noopener">📁 View</a>`
    : `<span style="color:#a00">Not uploaded</span>`;
  document.getElementById('e-computed').innerHTML =
    `<b>Computed:</b> Charges ₹${fmtDec(c.charges)} &nbsp;|&nbsp; GST ₹${fmtDec(c.gst)} &nbsp;|&nbsp; Total ₹${fmtDec(c.total)}` +
    ` &nbsp;|&nbsp; Collected ₹${fmtDec(c.collected)} &nbsp;|&nbsp; Discount ₹${fmtDec(c.discount)}` +
    ` &nbsp;|&nbsp; <b>Balance ₹${fmtDec(c.balance)}</b> &nbsp;|&nbsp; Status: <b>${c.status || '—'}</b>` +
    ` &nbsp;|&nbsp; KYC Docs: ${kycDocsLine}`;
  renderEditPayments();
}

function renderEditPayments() {
  const tb = document.getElementById('e-pay-tbody');
  tb.innerHTML = '';
  if (!_editCase.payments.length) {
    tb.innerHTML = '<tr><td colspan="4" style="color:#aaa">No payments recorded in M Coll for this case.</td></tr>';
    return;
  }
  _editCase.payments.forEach((p, i) => {
    const tr = document.createElement('tr');
    tr.innerHTML = `<td>${p.date}</td><td style="text-align:right">${fmtDec(p.amount)}</td>` +
      `<td style="max-width:340px;overflow:hidden;text-overflow:ellipsis;white-space:nowrap" title="${(p.utr||'').replace(/"/g,'&quot;')}">${p.utr || '—'}</td>` +
      `<td><button onclick="editPayRow(${i})" style="padding:3px 10px;font-size:12px;cursor:pointer">✏️ Edit</button> ` +
      `<button onclick="delPayRow(${i})" style="padding:3px 10px;font-size:12px;cursor:pointer;color:#a00">🗑</button></td>`;
    tb.appendChild(tr);
  });
}

function editPayRow(i) {
  const p = _editCase.payments[i];
  const tr = document.getElementById('e-pay-tbody').children[i];
  tr.innerHTML = `<td><input type="text" id="ep-date-${i}" value="${p.date}" style="width:110px"></td>` +
    `<td style="text-align:right"><input type="number" id="ep-amt-${i}" value="${p.amount}" style="width:120px;text-align:right"></td>` +
    `<td><input type="text" id="ep-utr-${i}" value="${(p.utr||'').replace(/"/g,'&quot;')}" style="width:100%"></td>` +
    `<td><button onclick="savePayRow(${i})" style="padding:3px 10px;font-size:12px;cursor:pointer;color:#155724">✅ Save</button> ` +
    `<button onclick="renderEditPayments()" style="padding:3px 10px;font-size:12px;cursor:pointer">✕</button></td>`;
}

// 55s client-side ceiling so a slow/hung request (e.g. Sheets API retrying
// a rate-limit under the hood) always resolves to a visible error instead
// of leaving the button/status stuck on "Saving..." forever. Set above the
// server's own 60s function timeout isn't possible client-side, but 55s
// gives real Sheets calls (which can legitimately take 20-40s under
// quota pressure) room to finish rather than aborting a write that would
// have succeeded a few seconds later.
async function _fetchJsonWithTimeout(url, opts, ms) {
  const ctrl = new AbortController();
  const timer = setTimeout(() => ctrl.abort(), ms || 55000);
  try {
    const res = await fetch(url, {...opts, signal: ctrl.signal});
    return await res.json();
  } catch (e) {
    if (e.name === 'AbortError') throw new Error('Request timed out — the sheet may be busy, please try again.');
    throw e;
  } finally {
    clearTimeout(timer);
  }
}

// Fire-and-forget — deliberately not awaited. The ledger rebuild takes
// 10-15s on the Apps Script side; blocking the save's own "Saved!" status
// on it is what made every save visibly slow. This call runs independently
// in the background while the user already sees success.
function kickLedgerRebuild() {
  fetch('/rebuild-ledger', {method: 'POST'}).catch(() => {});
}

async function savePayRow(i) {
  const p = _editCase.payments[i];
  const body = {
    disb_id: _editCase.disb_id, mc_row: p.mc_row,
    date:   document.getElementById(`ep-date-${i}`).value.trim(),
    amount: document.getElementById(`ep-amt-${i}`).value,
    utr:    document.getElementById(`ep-utr-${i}`).value.trim(),
  };
  if (!body.amount) return showStatus('e-pay-status','error','Amount is required.');
  showStatus('e-pay-status','success','Saving...');
  try {
    const r = await _fetchJsonWithTimeout('/case/update-repayment', {method:'POST',
      headers:{'Content-Type':'application/json'}, body: JSON.stringify(body)});
    if (r.ok) { renderEditCase(r.case); loadRecentCases(); kickLedgerRebuild(); showStatus('e-pay-status','success','✅ Payment updated — balance & status recomputed.'); }
    else showStatus('e-pay-status','error','❌ ' + r.error);
  } catch (e) {
    showStatus('e-pay-status','error','❌ ' + e.message);
  }
}

async function delPayRow(i) {
  const p = _editCase.payments[i];
  if (!confirm(`Delete this payment?\n\n${p.date}  ₹${fmtDec(p.amount)}\n${p.utr || ''}\n\nThe case balance and status will be recomputed.`)) return;
  showStatus('e-pay-status','success','Deleting...');
  try {
    const r = await _fetchJsonWithTimeout('/case/delete-repayment', {method:'POST',
      headers:{'Content-Type':'application/json'},
      body: JSON.stringify({disb_id: _editCase.disb_id, mc_row: p.mc_row})});
    if (r.ok) { renderEditCase(r.case); loadRecentCases(); kickLedgerRebuild(); showStatus('e-pay-status','success','✅ Payment deleted — balance & status recomputed.'); }
    else showStatus('e-pay-status','error','❌ ' + r.error);
  } catch (e) {
    showStatus('e-pay-status','error','❌ ' + e.message);
  }
}

async function addPayRow() {
  if (!_editCase) return;
  const body = {
    disb_id: _editCase.disb_id,
    date:   document.getElementById('e-new-pay-date').value.trim(),
    amount: document.getElementById('e-new-pay-amount').value,
    utr:    document.getElementById('e-new-pay-utr').value.trim(),
  };
  if (!body.date || !body.amount) return showStatus('e-pay-status','error','Date and Amount are required.');
  showStatus('e-pay-status','success','Adding...');
  try {
    const r = await _fetchJsonWithTimeout('/case/add-repayment', {method:'POST',
      headers:{'Content-Type':'application/json'}, body: JSON.stringify(body)});
    if (r.ok) {
      renderEditCase(r.case); loadRecentCases(); kickLedgerRebuild();
      document.getElementById('e-new-pay-amount').value = '';
      document.getElementById('e-new-pay-utr').value = '';
      showStatus('e-pay-status','success','✅ Payment added — balance & status recomputed.');
    } else showStatus('e-pay-status','error','❌ ' + r.error);
  } catch (e) {
    showStatus('e-pay-status','error','❌ ' + e.message);
  }
}

async function saveEditDisb() {
  if (!_editCase) return;
  const g = id => document.getElementById(id).value.trim();
  const body = {
    disb_id: _editCase.disb_id,
    date: g('e-date'), customer: g('e-customer'), chq: g('e-chq'),
    company: g('e-company'), cluster: g('e-cluster'), branch: g('e-branch'),
    amount: g('e-amount'), srv_branch: g('e-srv-branch'), srv_cluster: g('e-srv-cluster'),
    debit_note: g('e-debit-note'), remarks: g('e-remarks'),
  };
  if (!body.customer || !body.amount) return showStatus('e-disb-status','error','Customer and Amount are required.');
  const btn = event.target; btn.disabled = true; btn.textContent = 'Saving...';
  try {
    const r = await _fetchJsonWithTimeout('/case/update-disbursement', {method:'POST',
      headers:{'Content-Type':'application/json'}, body: JSON.stringify(body)});
    if (r.ok) { renderEditCase(r.case); loadRecentCases(); kickLedgerRebuild(); showStatus('e-disb-status','success','✅ Saved — derived values recomputed.'); }
    else showStatus('e-disb-status','error','❌ ' + r.error);
  } catch (e) {
    showStatus('e-disb-status','error','❌ ' + e.message);
  } finally {
    btn.disabled = false; btn.textContent = '💾 Save Disbursement Changes';
  }
}

// ── Invoice tab ──────────────────────────────────────────────────────────────
async function loadInvoiceCases() {
  const sel = document.getElementById('inv-open-cases');
  sel.innerHTML = '<option value="">Loading...</option>';
  try {
    const cases = await (await fetch('/open-cases')).json();
    sel.innerHTML = '<option value="">— Select an open case —</option>';
    cases.forEach(c => {
      const opt = document.createElement('option');
      opt.value = c.disb_id;
      opt.textContent = `${c.disb_id}  |  ${c.customer}  |  ₹${fmt(c.balance)} due`;
      sel.appendChild(opt);
    });
  } catch(e) {
    sel.innerHTML = '<option value="">Error loading cases</option>';
  }
}

function onInvCaseSelect() {
  const id = document.getElementById('inv-open-cases').value;
  document.getElementById('inv-manual-id').value = '';
  const info = document.getElementById('inv-case-info');
  if (!id) { info.style.display = 'none'; return; }
  const c = openCasesData.find(x => x.disb_id === id);
  if (c) {
    info.innerHTML = `<b>${c.customer}</b> &nbsp;|&nbsp; Disbursed: ₹${fmt(c.amount)} &nbsp;|&nbsp; Total Due: ₹${fmt(c.total)} &nbsp;|&nbsp; <b>Balance: ₹${fmt(c.balance)}</b> &nbsp;|&nbsp; ${c.status}`;
    info.style.display = 'block';
  }
}

async function generateInvoice() {
  const fromDrop = document.getElementById('inv-open-cases').value;
  const fromText = document.getElementById('inv-manual-id').value.trim().toUpperCase();
  const disb_id = fromText || fromDrop;
  if (!disb_id) return showStatus('inv-status', 'error', 'Select a case or enter a Disbursement ID.');
  const btn = document.getElementById('inv-btn');
  btn.disabled = true; btn.textContent = '⏳ Generating...';
  showStatus('inv-status', 'success', 'Fetching data and building PDF...');
  try {
    const r = await fetch('/generate-invoice', {
      method: 'POST',
      headers: {'Content-Type': 'application/json'},
      body: JSON.stringify({disb_id}),
    });
    if (!r.ok) {
      const err = await r.json().catch(() => ({error: `HTTP ${r.status}`}));
      throw new Error(err.error || `HTTP ${r.status}`);
    }
    const blob = await r.blob();
    const disposition = r.headers.get('Content-Disposition') || '';
    const match = disposition.match(/filename="(.+)"/);
    const filename = match ? match[1] : `${disb_id} Invoice and Ledger.pdf`;
    const url = URL.createObjectURL(blob);
    const a = document.createElement('a');
    a.href = url; a.download = filename;
    document.body.appendChild(a); a.click(); a.remove();
    URL.revokeObjectURL(url);
    showStatus('inv-status', 'success', `✅ Downloaded: ${filename}`);
  } catch(e) {
    showStatus('inv-status', 'error', '❌ ' + e.message);
  } finally {
    btn.disabled = false; btn.textContent = '📄 Generate & Download PDF';
  }
}

function togglePendingReqs() {
  _pendingReqsOpen = !_pendingReqsOpen;
  document.getElementById('pending-requests-body').style.display = _pendingReqsOpen ? '' : 'none';
  document.getElementById('pending-chevron').textContent = _pendingReqsOpen ? '▼' : '▶';
}

// ── Date picker helper ────────────────────────────────────────────────────
function _openDatePicker(nativeId) {
  const inp = document.getElementById(nativeId);
  const textId = nativeId.replace('-native', '');
  const txtVal = (document.getElementById(textId).value || '').trim();
  const m = txtVal.match(/^(\d{1,2})-(\d{1,2})-(\d{4})$/);
  if (m) inp.value = `${m[3]}-${m[2].padStart(2,'0')}-${m[1].padStart(2,'0')}`;
  if (inp.showPicker) { try { inp.showPicker(); return; } catch(e) {} }
  inp.click();
}
function _pickDate(textId, isoVal) {
  if (!isoVal) return;
  const [y,m,d] = isoVal.split('-');
  document.getElementById(textId).value = `${d}-${m}-${y}`;
}

function showTab(id, btn) {
  document.querySelectorAll('.tab-content').forEach(t => t.classList.remove('active'));
  document.querySelectorAll('.tab-btn').forEach(b => b.classList.remove('active'));
  document.getElementById(id).classList.add('active');
  btn.classList.add('active');
}

function calcCharges() {
  const amt = parseFloat(document.getElementById('d-amount').value) || 0;
  const ch  = Math.round(amt * 0.005 * 100) / 100;
  const gst = Math.round(ch * 0.18 * 100) / 100;
  const tot = amt + ch + gst;
  document.getElementById('d-charges').value = ch.toLocaleString('en-IN', {minimumFractionDigits:2});
  document.getElementById('d-gst').value     = gst.toLocaleString('en-IN', {minimumFractionDigits:2});
  document.getElementById('d-total').value   = tot.toLocaleString('en-IN', {minimumFractionDigits:2});
}

async function extractDisb() {
  const msg = document.getElementById('d-msg').value.trim();
  if (!msg) return;
  const d = await (await fetch('/extract/disbursement', {method:'POST',
    headers:{'Content-Type':'application/json'}, body: JSON.stringify({message:msg})})).json();
  if (d.date)     document.getElementById('d-date').value = d.date;
  if (d.customer) document.getElementById('d-customer').value = d.customer;
  if (d.company)  document.getElementById('d-company').value = d.company;
  if (d.cluster)  document.getElementById('d-cluster').value = d.cluster;
  if (d.branch)   document.getElementById('d-branch').value = d.branch;
  if (d.amount)  { document.getElementById('d-amount').value = d.amount; calcCharges(); }
}

async function extractDisbUTR() {
  const msg = document.getElementById('d-utr-msg').value.trim();
  if (!msg) return;
  const d = await (await fetch('/extract/repayment', {method:'POST',
    headers:{'Content-Type':'application/json'}, body: JSON.stringify({message:msg})})).json();
  if (d.utr) {
    document.getElementById('d-utr').value = d.utr;
    if (d.date && !document.getElementById('d-date').value)
      document.getElementById('d-date').value = d.date;
  } else {
    document.getElementById('d-utr').value = '';
    alert('No UTR found in that message.');
  }
}

async function extractRepa() {
  const msg = document.getElementById('r-msg').value.trim();
  if (!msg) return;
  const d = await (await fetch('/extract/repayment', {method:'POST',
    headers:{'Content-Type':'application/json'}, body: JSON.stringify({message:msg})})).json();
  if (d.date)   document.getElementById('r-date').value = d.date;
  if (d.amount) document.getElementById('r-amount').value = d.amount;
  if (d.utr)    document.getElementById('r-utr').value = d.utr;
  if (d.disb_id) {
    const sel = document.getElementById('r-open-cases');
    for (let i = 0; i < sel.options.length; i++) {
      if (sel.options[i].value === d.disb_id) { sel.value = d.disb_id; break; }
    }
    onCaseSelect();
  }
  if (d.sender) showInfo(`Sender: <b>${d.sender}</b> — select the matching case above.`, false);
}

async function loadOpenCases() {
  const sel = document.getElementById('r-open-cases');
  const spinner = document.getElementById('cases-spinner');
  sel.innerHTML = '<option value="">Loading...</option>';
  spinner.style.display = 'block';
  try {
    const cases = await (await fetch('/open-cases')).json();
    openCasesData = cases;
    sel.innerHTML = '<option value="">— Select an open case —</option>';
    cases.forEach(c => {
      const opt = document.createElement('option');
      opt.value = c.disb_id;
      opt.textContent = `${c.disb_id}  |  ${c.customer}  |  ₹${fmt(c.balance)} due`;
      sel.appendChild(opt);
    });
  } catch(e) {
    sel.innerHTML = '<option value="">Error loading — check credentials</option>';
  }
  spinner.style.display = 'none';
}

function onCaseSelect() {
  const id = document.getElementById('r-open-cases').value;
  const c  = openCasesData.find(x => x.disb_id === id);
  if (!c) { document.getElementById('r-info').style.display='none'; return; }
  showInfo(`<b>${c.customer}</b> &nbsp;|&nbsp; Disbursed: ₹${fmt(c.amount)} &nbsp;|&nbsp; Total Due: ₹${fmt(c.total)} &nbsp;|&nbsp; Collected: ₹${fmt(c.collected)} &nbsp;|&nbsp; <b>Balance: ₹${fmt(c.balance)}</b> &nbsp;|&nbsp; ${c.status}`, false);
}

function showInfo(msg, isError) {
  const box = document.getElementById('r-info');
  box.style.display = 'block';
  box.className = 'info-box' + (isError ? ' error' : '');
  box.innerHTML = msg;
}

function fmt(n) { return Number(n).toLocaleString('en-IN', {maximumFractionDigits:0}); }
function fmtDec(n) { return Number(n).toLocaleString('en-IN', {minimumFractionDigits:2, maximumFractionDigits:2}); }

async function generateMis() {
  const btn = document.getElementById('mis-btn');
  btn.disabled = true; btn.textContent = '⏳ Generating...';
  try {
    const r = await fetch('/generate-mis', {method: 'POST'});
    if (!r.ok) {
      const err = await r.json().catch(() => ({error: 'Unknown error'}));
      throw new Error(err.error || `HTTP ${r.status}`);
    }
    const blob = await r.blob();
    const disposition = r.headers.get('Content-Disposition') || '';
    const match = disposition.match(/filename="(.+)"/);
    const filename = match ? match[1] : 'BridgeLine MIS Package.zip';
    const url = URL.createObjectURL(blob);
    const a = document.createElement('a');
    a.href = url; a.download = filename;
    document.body.appendChild(a); a.click(); a.remove();
    URL.revokeObjectURL(url);
    showStatus('mis-status', 'success', `✅ Downloaded ${filename}`);
  } catch (e) {
    showStatus('mis-status', 'error', '❌ ' + e.message);
  } finally {
    btn.disabled = false; btn.textContent = '📦 Generate MIS Package';
  }
}

async function saveDisb() {
  const data = {
    date: document.getElementById('d-date').value.trim(),
    customer: document.getElementById('d-customer').value.trim(),
    chq: document.getElementById('d-chq').value.trim(),
    company: document.getElementById('d-company').value.trim(),
    cluster: document.getElementById('d-cluster').value.trim(),
    branch: document.getElementById('d-branch').value.trim(),
    amount: document.getElementById('d-amount').value,
    serviced_branch: document.getElementById('d-srv-branch').value.trim(),
    serviced_cluster: document.getElementById('d-srv-cluster').value.trim(),
    utr:        document.getElementById('d-utr').value.trim(),
    remarks:    document.getElementById('d-remarks').value.trim(),
    request_id: document.getElementById('d-request-id').value.trim(),
    bank_account: document.getElementById('d-bank-account').value.trim(),
  };
  if (!data.customer || !data.amount || !data.company || !data.cluster || !data.branch)
    return showStatus('d-status','error','Please fill all required (*) fields.');
  if (!data.utr)
    return showStatus('d-status','error','❌ UTR / Reference number is required — a disbursement cannot be recorded without proof of transfer.');
  const btn = event.target; btn.disabled = true; btn.textContent = 'Saving...';
  const r = await (await fetch('/save/disbursement', {method:'POST',
    headers:{'Content-Type':'application/json'}, body: JSON.stringify(data)})).json();
  btn.disabled = false; btn.textContent = '✅ Save Disbursement to Google Sheet';
  if (r.ok) {
    let msg = `✅ Saved! ID: ${r.disb_id}`;
    if (r.warning) msg += `  ⚠ ${r.warning}`;
    showStatus('d-status','success',msg);
    clearDisb(); loadPendingRequests(); kickLedgerRebuild();
  }
  else       showStatus('d-status','error','❌ ' + r.error);
}

async function saveRepa() {
  const disb_id = document.getElementById('r-open-cases').value;
  if (!disb_id) return showStatus('r-status','error','Please select an open case.');
  const amount = document.getElementById('r-amount').value;
  if (!amount)  return showStatus('r-status','error','Amount is required.');
  const utr = document.getElementById('r-utr').value.trim();
  if (!utr) return showStatus('r-status','error','❌ UTR / Reference number is required — a repayment cannot be recorded without proof of payment.');
  const rDate = document.getElementById('r-date').value.trim();
  if (!rDate) return showStatus('r-status','error','Collection Date is required — the pasted message did not contain a date, please enter it manually.');
  const data = {
    disb_id,
    date:     rDate,
    amount,
    utr,
    discount: document.getElementById('r-discount').value || 0,
    raw_msg:  document.getElementById('r-msg').value.trim(),
    remarks:  document.getElementById('r-remarks').value.trim(),
    bank_account: document.getElementById('r-bank-account').value.trim(),
  };
  const btn = event.target; btn.disabled = true; btn.textContent = 'Saving...';
  const r = await (await fetch('/save/repayment', {method:'POST',
    headers:{'Content-Type':'application/json'}, body: JSON.stringify(data)})).json();
  btn.disabled = false; btn.textContent = '✅ Save Repayment to Google Sheet';
  if (r.ok) {
    if (r.mcoll_warning) {
      showStatus('r-status','error',`⚠️ ₹${fmt(amount)} recorded (balance ₹${fmt(r.new_balance)}) but NOT logged to payment history — ${r.mcoll_warning}`);
    } else {
      showStatus('r-status','success',`✅ ₹${fmt(amount)} recorded. New balance: ₹${fmt(r.new_balance)}. Status: ${r.status}`);
    }
    document.getElementById('r-amount').value = '';
    document.getElementById('r-utr').value = '';
    document.getElementById('r-discount').value = '0';
    document.getElementById('r-msg').value = '';
    document.getElementById('r-remarks').value = '';
    document.getElementById('r-bank-account').value = '';
    loadOpenCases();
    kickLedgerRebuild();
  } else showStatus('r-status','error','❌ ' + r.error);
}

function showStatus(id, type, msg) {
  const el = document.getElementById(id);
  el.className = 'status ' + type;
  el.textContent = msg;
  el.style.display = 'block';
  setTimeout(() => el.style.display = 'none', 9000);
}

function clearDisb() {
  ['d-date','d-customer','d-chq','d-amount','d-charges','d-gst','d-total','d-utr','d-remarks','d-bank-account'].forEach(id =>
    document.getElementById(id).value = '');
  ['d-company','d-cluster','d-branch'].forEach(id =>
    document.getElementById(id).value = '');
  ['d-srv-branch','d-srv-cluster'].forEach(id =>
    document.getElementById(id).value = '');
  document.getElementById('d-msg').value = '';
  document.getElementById('d-utr-msg').value = '';
  document.getElementById('d-date').value = new Date().toLocaleDateString('en-GB').replace(/\\//g,'-');
  document.getElementById('d-request-id').value = '';
  document.getElementById('pending-selected-info').style.display = 'none';
}

// ── Summary banner ──────────────────────────────────────────────────────────
let _summaryData = {};
async function loadSummary() {
  try {
    const d = await (await fetch('/summary')).json();
    _summaryData = d;
    document.getElementById('s-disb').textContent = (d.disbursed_today||0) > 0 ? '₹'+fmt(d.disbursed_today) : '—';
    document.getElementById('s-coll').textContent = (d.collected_today||0) > 0 ? '₹'+fmt(d.collected_today) : '—';
    document.getElementById('s-out').textContent  = d.total_outstanding != null ? '₹'+fmt(d.total_outstanding) : '—';
    document.getElementById('s-avail').textContent = d.available_for_disbursement != null
      ? '₹'+fmt(d.available_for_disbursement)
      : (d.bank_balance != null ? 'FD unknown' : '—');
    document.getElementById('s-date').textContent = d.date;
  } catch(e) {}
}

function openSumModal(kind) {
  const d = _summaryData;
  const cfgs = {
    disb: {title: 'Disbursed Today', rows: d.disb_rows || [], emptyMsg: 'No disbursements recorded today.'},
    coll: {title: 'Collected Today', rows: d.coll_rows || [], emptyMsg: 'No collections recorded today.'},
    out:  {title: 'Total Outstanding', rows: d.out_rows || [], emptyMsg: 'No open balances.'},
    avail:{title: 'Available to Disburse', rows: null, emptyMsg: ''},
  };
  const cfg = cfgs[kind];
  document.getElementById('sum-modal-title').textContent = cfg.title;
  const body = document.getElementById('sum-modal-body');
  if (kind === 'avail') {
    if (d.available_for_disbursement == null && d.bank_balance != null) {
      // Bank is known, FD is not. Say exactly that, show what IS known, and
      // give the two ways out — never a blank and never a confident wrong
      // total that quietly leaves the FDs out.
      body.innerHTML = `<div style="padding:18px">
        <div style="padding:10px 12px;background:#fdecea;border:1px solid #e0a0a0;border-radius:6px;font-size:.85rem;color:#8a2020;margin-bottom:14px">
          <b>FD balance can't be worked out from the statements right now,</b> so this total is deliberately not shown.
          ₹${fmt(d.fd_orphan_total||0)} was seen sweeping back IN from deposits whose matching sweep-OUT is missing from the
          uploaded statements — so how much those deposits still hold is unknown. Showing ₹0 for the FDs would understate
          your position by tens of lakhs, which is worse than showing nothing.
        </div>
        <table style="width:100%;font-size:.88rem">
          <tr><td style="padding:6px 0;color:#555">Bank Closing Balance (${d.bank_balance_date||''})</td><td style="padding:6px 0;text-align:right;font-weight:600">₹${fmt(d.bank_balance||0)}</td></tr>
          <tr><td style="padding:6px 0;color:#555">Plus: Parked in FDs</td><td style="padding:6px 0;text-align:right;font-weight:700;color:#c00">unknown</td></tr>
          <tr style="border-top:2px solid #1a3a5c"><td style="padding:8px 0;font-weight:700">Available to Disburse</td><td style="padding:8px 0;text-align:right;font-weight:700;color:#c00">unknown</td></tr>
        </table>
        <div style="margin-top:12px;font-size:.82rem;color:#333;line-height:1.7">
          <b>Two ways to fix it:</b><br>
          1. <b>Fastest</b> — enter today's FD balance under Settings → FD Balance. It is used immediately.<br>
          2. <b>Proper</b> — re-upload the bank statement covering the days holding the missing sweep-out lines.
             A re-upload now only rewrites the days the file actually contains, so nothing else is disturbed.
        </div>
      </div>`;
    } else if (d.available_for_disbursement == null) {
      body.innerHTML = `<div class="empty">No bank reconciliation found yet — upload a bank statement under "Bank Reconciliation" to enable this.</div>`;
    } else {
      body.innerHTML = `<div style="padding:18px">
        <table style="width:100%;font-size:.88rem">
          <tr><td style="padding:6px 0;color:#555">Bank Closing Balance (${d.bank_balance_date||''})</td><td style="padding:6px 0;text-align:right;font-weight:600">₹${fmt(d.bank_balance||0)}</td></tr>
          <tr><td style="padding:6px 0;color:#555">Plus: Collected since then</td><td style="padding:6px 0;text-align:right;font-weight:600;color:#1a5c3a">+₹${fmt(d.collected_since_bank||0)}</td></tr>
          <tr><td style="padding:6px 0;color:#555">Less: Disbursed since then</td><td style="padding:6px 0;text-align:right;font-weight:600;color:#c00">−₹${fmt(d.disbursed_since_bank||0)}</td></tr>
          <tr><td style="padding:6px 0;color:#555">Plus: Parked in FDs${d.fd_outstanding_date ? ' (as of '+d.fd_outstanding_date+')' : (d.fd_open_count ? ' ('+d.fd_open_count+' open FD'+(d.fd_open_count===1?'':'s')+')' : '')}</td><td style="padding:6px 0;text-align:right;font-weight:600;color:#1a5c3a">+₹${fmt(d.fd_outstanding||0)}</td></tr>
          <tr style="border-top:2px solid #1a3a5c"><td style="padding:8px 0;font-weight:700">Available to Disburse</td><td style="padding:8px 0;text-align:right;font-weight:700;color:#1a5c3a">₹${fmt(d.available_for_disbursement||0)}</td></tr>
        </table>
        ${d.fd_outstanding_source === 'fd_ledger' ? `<div style="margin-top:10px;padding:8px 10px;background:#eaf6ee;border:1px solid #9ccfae;border-radius:6px;font-size:.78rem;color:#1a5c3a">FD balance derived automatically from the sweep-out / sweep-in lines in your bank statements, tracked per FD account number. No manual entry needed.</div>` : ''}
        ${d.fd_outstanding_source === 'direct_override' ? `<div style="margin-top:10px;padding:8px 10px;background:#eef3fa;border:1px solid #a9c2e0;border-radius:6px;font-size:.78rem;color:#1a3a5c">Using a manually entered FD balance, which is dated on/after the latest sweep activity. It will hand back to the automatic per-FD figure once newer sweeps come in.</div>` : ''}
        ${d.fd_unparsed_count ? `<div style="margin-top:10px;padding:8px 10px;background:#fdecea;border:1px solid #e0a0a0;border-radius:6px;font-size:.78rem;color:#8a2020">⚠️ ${d.fd_unparsed_count} FD sweep transaction(s) have an unrecognised narration and are excluded from this FD figure. See Settings → Company Solvency Check.</div>` : ''}
        ${d.fd_outstanding_source === 'derived_fallback' ? `<div style="margin-top:10px;padding:8px 10px;background:#fff3cd;border:1px solid #e0c060;border-radius:6px;font-size:.78rem;color:#7a5c00">⚠️ Falling back to the old derived-from-history FD figure, which is known to understate. Check Settings → Company Solvency Check.</div>` : ''}
        </div>`;
    }
  } else if (!cfg.rows.length) {
    body.innerHTML = `<div class="empty">${cfg.emptyMsg}</div>`;
  } else {
    let total = cfg.rows.reduce((s,r) => s + (r.amount||0), 0);
    body.innerHTML = `<table>
      <thead><tr><th>Disb ID</th><th>Customer</th><th style="text-align:right">Amount</th></tr></thead>
      <tbody>${cfg.rows.map(r => `<tr><td>${r.disb_id||''}</td><td>${r.customer||''}</td><td style="text-align:right">₹${fmt(r.amount||0)}</td></tr>`).join('')}</tbody>
      <tfoot><tr style="font-weight:700;background:#f0f4f8"><td colspan="2" style="padding:8px 14px">Total</td><td style="padding:8px 14px;text-align:right">₹${fmt(total)}</td></tr></tfoot>
    </table>`;
  }
  document.getElementById('sum-modal-bg').classList.add('show');
}
function closeSumModal() {
  document.getElementById('sum-modal-bg').classList.remove('show');
}

// ── Calculator ───────────────────────────────────────────────────────────────
let calcExpr = '';
let calcCurrent = '0';
let calcJustEq  = false;

function calcUpdate() {
  document.getElementById('calc-disp').textContent =
    parseFloat(calcCurrent).toLocaleString('en-IN', {maximumFractionDigits:8});
  document.getElementById('calc-expr').textContent = calcExpr;
}

function calcNum(ch) {
  if (calcJustEq) { calcCurrent = ''; calcExpr = ''; calcJustEq = false; }
  if (ch === '.' && calcCurrent.includes('.')) return;
  if (calcCurrent === '0' && ch !== '.') calcCurrent = ch;
  else calcCurrent += ch;
  calcUpdate();
}

function calcOp(op) {
  calcJustEq = false;
  calcExpr = calcCurrent + ' ' + op + ' ';
  calcCurrent = '0';
  calcUpdate();
}

function calcEq() {
  if (!calcExpr) return;
  try {
    const expr = calcExpr + calcCurrent;
    const result = Function('"use strict"; return (' + expr + ')')();
    calcExpr = expr + ' =';
    calcCurrent = String(parseFloat(result.toFixed(8)));
    calcJustEq = true;
    calcUpdate();
  } catch(e) {}
}

function calcClear() {
  calcExpr = ''; calcCurrent = '0'; calcJustEq = false; calcUpdate();
}

function calcBack() {
  if (calcJustEq) { calcClear(); return; }
  calcCurrent = calcCurrent.length > 1 ? calcCurrent.slice(0,-1) : '0';
  calcUpdate();
}

// Keyboard support for calculator
document.addEventListener('keydown', e => {
  const calcTab = document.getElementById('calc');
  if (!calcTab.classList.contains('active')) return;
  if (e.key >= '0' && e.key <= '9') calcNum(e.key);
  else if (e.key === '.') calcNum('.');
  else if (e.key === '+') calcOp('+');
  else if (e.key === '-') calcOp('-');
  else if (e.key === '*') calcOp('*');
  else if (e.key === '/') { e.preventDefault(); calcOp('/'); }
  else if (e.key === '%') calcOp('%');
  else if (e.key === 'Enter' || e.key === '=') calcEq();
  else if (e.key === 'Backspace') calcBack();
  else if (e.key === 'Escape') calcClear();
});

// ── Reconciliation ───────────────────────────────────────────────────────────
// Captured here instead of re-reading input.files[0] later — browsers don't
// fire 'change' when the SAME filename is re-selected (common with bank
// portals that always export under one generic name), which used to make a
// second upload in the same session silently no-op until the page was
// reloaded. Resetting input.value immediately guarantees the NEXT selection
// (same name or not) always fires 'change' again.
// A LIST now — every account's statement for the period goes up together, so
// transfers between our own accounts pair structurally instead of being read
// out of narration text (see _pair_uploaded_statements() on the server).
let _selectedReconFiles = [];

// Accumulates across picks rather than replacing: banks export one account at
// a time, so HDFC and IDFC are usually two separate trips to the file dialog.
// Resetting input.value immediately guarantees the NEXT selection always fires
// 'change' again — browsers don't fire it when the SAME filename is re-picked,
// which used to make a second upload silently no-op until a page reload.
function onFileSelect(input) {
  for (const f of input.files) {
    if (!_selectedReconFiles.some(x => x.name === f.name && x.size === f.size)) {
      _selectedReconFiles.push(f);
    }
  }
  input.value = '';
  renderSelectedFiles();
}

function removeReconFile(i) {
  _selectedReconFiles.splice(i, 1);
  renderSelectedFiles();
}

function renderSelectedFiles() {
  const box = document.getElementById('selected-files');
  const label = document.getElementById('upload-label');
  if (!_selectedReconFiles.length) {
    box.style.display = 'none'; box.innerHTML = '';
    label.innerHTML = '📂 Click to upload bank statements<br><span style="font-size:.78rem">Select all accounts at once — CSV, Excel (.xlsx/.xls) or PDF</span>';
    return;
  }
  label.innerHTML = `📂 ${_selectedReconFiles.length} file${_selectedReconFiles.length>1?'s':''} selected — click to add another`;
  box.style.display = 'block';
  box.innerHTML = _selectedReconFiles.map((f, i) =>
    `<div style="display:flex;align-items:center;gap:8px;padding:5px 8px;background:#f4f8fc;border:1px solid #d0dce8;border-radius:5px;margin-bottom:5px">
       <span style="flex:1">📄 ${f.name}</span>
       <button type="button" onclick="removeReconFile(${i})"
         style="border:none;background:none;color:#c00;cursor:pointer;font-size:1rem;line-height:1">×</button>
     </div>`).join('');
}

async function parseStatement() {
  if (!_selectedReconFiles.length) return alert('Please upload at least one bank statement file first.');

  const formData = new FormData();
  _selectedReconFiles.forEach(f => formData.append('files', f, f.name));

  const btn = event.target; btn.disabled = true; btn.textContent = 'Parsing...';
  try {
    const r = await (await fetch('/reconcile/parse', {method:'POST', body: formData})).json();
    btn.disabled = false; btn.textContent = '⚡ Parse Statements';
    if (!r.ok) return alert('❌ ' + r.error);
    renderReconResult(r);
  } catch(e) {
    btn.disabled = false; btn.textContent = '⚡ Parse Statements';
    alert('Error parsing file: ' + e);
  }
}

const BASE_TYPES = ['Disbursement','Repayment','Other Income','Expense','Contra','Interest Income','Skip'];
const TYPE_COLOR = {
  'Disbursement':'#fff8e1','Repayment':'#e8f5e9','Other Income':'#e8f5e9','Interest Income':'#e8f5e9',
  'Expense':'#fce4ec','Contra':'#e3f2fd','Skip':'#eeeeee',
  // legacy labels still present on older stored rows
  'Collection':'#e8f5e9','Collection (via Pradaan)':'#e8f5e9','Capital In':'#e3f2fd','Capital Out':'#e3f2fd'
};

let _customTypes = [];

async function loadCustomTypes() {
  try {
    const cfg = await (await fetch('/config')).json();
    _customTypes = cfg.custom_types || [];
  } catch { _customTypes = []; }
}

async function saveCustomType(t) {
  if (!t || BASE_TYPES.includes(t) || _customTypes.includes(t)) return;
  _customTypes.push(t);
  await fetch('/config', {method:'POST', headers:{'Content-Type':'application/json'},
    body: JSON.stringify({custom_types: _customTypes})});
}

let _bankAccounts = [];

async function loadBankAccounts() {
  try {
    const cfg = await (await fetch('/config')).json();
    _bankAccounts = cfg.bank_accounts || [];
  } catch { _bankAccounts = []; }
  const opts = _bankAccounts.map(a => `<option value="${a.name}">`).join('');
  ['rec-account-list', 'd-bank-account-list', 'r-bank-account-list'].forEach(id => {
    const dl = document.getElementById(id);
    if (dl) dl.innerHTML = opts;
  });
}

// If the typed account name isn't already registered, prompt for its
// account number (used for Capital In/Out transfer matching between our
// own accounts) and save it to Config so it shows up in the dropdown and
// transfer-matching going forward.
async function registerAccountIfNew(name) {
  if (_bankAccounts.some(a => a.name === name)) return;
  const acctNum = (prompt(`New bank account "${name}" — enter its account number (used to auto-detect transfers between your own accounts). Leave blank to skip.`) || '').trim();
  _bankAccounts.push({name, account_number: acctNum});
  await fetch('/config', {method:'POST', headers:{'Content-Type':'application/json'},
    body: JSON.stringify({bank_accounts: _bankAccounts})});
  await loadBankAccounts();
}

function allTypeOpts() {
  return ['', ...BASE_TYPES, ..._customTypes];
}
function buildTypeSelect(i, val='') {
  // The stored value may be a legacy/auto-only label (Collection, Capital
  // In, FD Booking, ...) no longer offered in the list — keep it selectable
  // for THIS row so re-saving doesn't silently change it.
  const opts = allTypeOpts();
  if (val && !opts.includes(val)) opts.push(val);
  const optHtml = opts.map(t =>
    `<option value="${t}" ${t===val?'selected':''}>${t||'— auto —'}</option>`).join('')
    + `<option value="__add__">＋ Add new type…</option>`;
  return `<select data-row="${i}" onchange="onTypeChange(this)"
    style="width:100%;font-size:.82rem;padding:4px 6px;border:1px solid #b0c4d8;border-radius:5px;background:white">${optHtml}</select>`;
}
function rowBg(type) { return TYPE_COLOR[type] || 'white'; }

async function onTypeChange(sel) {
  const i = sel.dataset.row;
  if (sel.value === '__add__') {
    const t = (prompt('New transaction type name:') || '').trim();
    if (t && !BASE_TYPES.includes(t) && !_customTypes.includes(t)) {
      await saveCustomType(t);
    }
    // Rebuild every type select so the new type shows up everywhere,
    // preserving each row's current selection.
    document.querySelectorAll('#review-tbody select[data-row]').forEach(s => {
      const cur = (s === sel) ? (t || '') : s.value;
      s.outerHTML = buildTypeSelect(s.dataset.row, cur === '__add__' ? '' : cur);
    });
    const mine = document.querySelector(`#review-tbody select[data-row="${i}"]`);
    if (mine) mine.value = t || '';
  }
  const val = (document.querySelector(`#review-tbody select[data-row="${i}"]`) || sel).value;
  const rowEl = document.getElementById('txrow-'+i);
  if (rowEl) rowEl.style.background = rowBg(val);
}

function _fmtSigned(n) {
  return (n < 0 ? '-\u20b9' : '\u20b9') + fmtDec(Math.abs(n));
}

// Per-statement now \u2014 each uploaded file has its own editable opening/closing
// and its own arithmetic badge, since one upload can carry several accounts.
function recomputeStatementCheck(si) {
  const st = (window._reconStatements || [])[si];
  if (!st) return;
  const opEl = document.querySelector(`input[data-open="${si}"]`);
  const clEl = document.querySelector(`input[data-close="${si}"]`);
  if (opEl) st.opening_balance = parseFloat(opEl.value) || 0;
  if (clEl) st.closing_balance = parseFloat(clEl.value) || 0;
  const dr = (st.transactions||[]).reduce((s,t)=>s+t.debit,0);
  const cr = (st.transactions||[]).reduce((s,t)=>s+t.credit,0);
  const expected = st.opening_balance + cr - dr;
  const variance = st.closing_balance - expected;
  const el = document.querySelector(`[data-check="${si}"]`);
  if (!el) return;
  el.style.color = Math.abs(variance) <= 1 ? '#1a5c3a' : '#c00';
  el.textContent = Math.abs(variance) <= 1
    ? `\u2713 Statement is internally consistent (Opening + Credits \u2212 Debits = Closing)`
    : `\u2717 Mismatch: expected closing \u20b9${fmtDec(expected)}, statement says \u20b9${fmtDec(st.closing_balance)} \u2014 \u0394${_fmtSigned(variance)}`;
}

// The account a statement will be SAVED under. Auto-detected from the file
// where possible; this is the manual fallback and override.
function setStatementAccount(si, value) {
  const st = (window._reconStatements || [])[si];
  if (!st) return;
  st.account = (value || '').trim();
  const note = document.querySelector(`[data-acct-note="${si}"]`);
  if (note) note.textContent = st.account ? '(you set this)' : '';
}

function buildUnmatchedTable(rows, refLabel) {
  return `<div style="overflow-x:auto"><table class="recon-table"><thead><tr>
      <th>Date</th><th>Disb ID</th><th>Customer</th><th style="text-align:right">Amount</th><th>${refLabel}</th>
    </tr></thead><tbody>${rows.map(x => `<tr>
      <td style="white-space:nowrap;font-size:.82rem">${x.date||'\u2014'}</td>
      <td style="font-size:.82rem">${x.disb_id||'\u2014'}</td>
      <td style="font-size:.82rem">${x.customer||'\u2014'}</td>
      <td style="text-align:right;font-size:.82rem;font-weight:600">\u20b9${fmtDec(x.amount||0)}</td>
      <td style="font-size:.78rem;color:#666">${x.utr||x.utr_or_note||'\u2014'}</td>
    </tr>`).join('')}</tbody></table></div>`;
}

// One uploaded statement rendered as a self-contained block: its own balances,
// its own date/account provenance, its own arithmetic check, completeness and
// tie-out. Returns HTML; the caller joins them.
function _statementBlock(st, si) {
  if (st.error) {
    return `<div class="section" style="background:#fdecea;border-color:#e0a0a0">
      <h3 style="color:#8a2020">\u26a0\ufe0f ${st.filename} \u2014 could not be read</h3>
      <div style="font-size:.83rem;color:#8a2020">${st.error}</div>
      <div style="font-size:.78rem;color:#666;margin-top:6px">The other statements in this upload were parsed normally and can still be saved.</div>
    </div>`;
  }

  const totalDr = st.transactions.reduce((s,t) => s+t.debit, 0);
  const totalCr = st.transactions.reduce((s,t) => s+t.credit, 0);

  // Date provenance \u2014 never silently change the save date underneath the user.
  let dateNote;
  if (st.date_corrected) {
    dateNote = `<div style="margin:10px 0;padding:8px 10px;border:1px solid #e0c060;background:#fff3cd;color:#7a5c00;border-radius:6px;font-size:.82rem">\u26a0\ufe0f Date corrected from the statement: you entered <b>${st.typed_date}</b>, but the statement's own period is <b>${st.statement_from||''} to ${st.statement_to}</b>. This will be saved under <b>${st.recon_date}</b>.</div>`;
  } else if (st.statement_to) {
    dateNote = `<div style="margin:10px 0;padding:8px 10px;border:1px solid #9ccfae;background:#eaf6ee;color:#1a5c3a;border-radius:6px;font-size:.82rem">Statement period read from the file: <b>${st.statement_from||''} to ${st.statement_to}</b> \u2014 saving under <b>${st.recon_date}</b>.</div>`;
  } else {
    dateNote = `<div style="margin:10px 0;padding:8px 10px;border:1px solid #e0a0a0;background:#fdecea;color:#8a2020;border-radius:6px;font-size:.82rem">\u26a0\ufe0f This statement does not declare a STATEMENT PERIOD. Enter the reconciliation date below before saving.</div>`;
  }

  const acctKnown = !!st.account;
  const acctNote = acctKnown
    ? `<span style="color:#1a5c3a">\u2713 ${st.account_source}${st.account_number ? ' (a/c ' + st.account_number + ')' : ''}</span>`
    : `<span style="color:#c00">\u26a0\ufe0f could not be identified from the file \u2014 set it here</span>`;

  // Auto-reconciled summary, grouped by type.
  const byType = {};
  (st.confident || []).forEach(tx => {
    byType[tx.type] = byType[tx.type] || {count:0, dr:0, cr:0};
    byType[tx.type].count++; byType[tx.type].dr += tx.debit; byType[tx.type].cr += tx.credit;
  });
  const confHtml = Object.entries(byType).map(([t,v]) =>
    `<span style="display:inline-block;margin-right:24px"><b>${t}</b>: ${v.count} txns`
    + (v.dr ? ` &nbsp;Dr \u20b9${fmtDec(v.dr)}` : "")
    + (v.cr ? ` &nbsp;Cr \u20b9${fmtDec(v.cr)}` : "") + `</span>`).join("");

  // Two very different findings, deliberately styled apart: entries dated
  // PAST the reconciled window are simply awaiting a statement (blue,
  // informational), while entries INSIDE the window with no bank line are
  // the real red flag (red). Merging them made a normal day's business look
  // like a 15.3 lakh hole on 13-08-2026 and cost an hour chasing it.
  const bc = st.book_completeness;
  const realMissing = bc && (bc.unmatched_disbursements_count || bc.unmatched_collections_count);
  const pending     = bc && (bc.pending_disbursements_count || bc.pending_collections_count);
  const bcHtml =
    (realMissing
      ? `<div class="section" style="background:#fff5f5;border-color:#e0b0b0;margin-top:12px">
           <h3 style="color:#a00">\u26a0\ufe0f In the reconciled window but no bank line <span style="font-weight:400;color:#888">\u2014 ${bc.unmatched_disbursements_count} disb. + ${bc.unmatched_collections_count} coll.</span></h3>
           <p style="font-size:.8rem;color:#666;margin:0 0 10px">Dated <b>on or before ${bc.reconciled_through||'\u2014'}</b>, which the uploaded statements cover \u2014 so a matching bank entry should exist and doesn't. Worth investigating.</p>
           <div style="font-size:.83rem">`
        + (bc.unmatched_disbursements_count
            ? `<p><b>Disbursements (\u20b9${fmtDec(bc.unmatched_disbursements_total)}):</b></p>`
              + buildUnmatchedTable(bc.unmatched_disbursements, 'Debit Note') : '')
        + (bc.unmatched_collections_count
            ? `<p style="margin-top:12px"><b>Collections (\u20b9${fmtDec(bc.unmatched_collections_total)}):</b></p>`
              + buildUnmatchedTable(bc.unmatched_collections, 'UTR / Note') : '')
        + `</div></div>` : '')
    + (pending
      ? `<div class="section" style="background:#f2f7fd;border-color:#b8d0ea;margin-top:12px">
           <h3 style="color:#12467e">🕓 Awaiting a bank statement <span style="font-weight:400;color:#555">\u2014 ${bc.pending_disbursements_count} disb. + ${bc.pending_collections_count} coll.</span></h3>
           <p style="font-size:.8rem;color:#555;margin:0 0 10px">Booked <b>after ${bc.reconciled_through||'\u2014'}</b>, the last day any uploaded statement reaches. Nothing is wrong \u2014 these reconcile the moment that day's statement goes up. <b>This is not missing money.</b></p>
           <div style="font-size:.83rem">`
        + (bc.pending_disbursements_count
            ? `<p><b>Disbursements (\u20b9${fmtDec(bc.pending_disbursements_total)}):</b></p>`
              + buildUnmatchedTable(bc.pending_disbursements, 'Debit Note') : '')
        + (bc.pending_collections_count
            ? `<p style="margin-top:12px"><b>Collections (\u20b9${fmtDec(bc.pending_collections_total)}):</b></p>`
              + buildUnmatchedTable(bc.pending_collections, 'UTR / Note') : '')
        + `</div></div>` : '');

  const to = st.account_tieout || {};
  let toHtml = '';
  if (to.status && to.status !== 'no_account_selected') {
    let badge = '', body = '';
    if (to.status === 'incomplete') {
      badge = '\u2014 Incomplete';
      body = `<p style="color:#b8860b">\u26a0\ufe0f ${to.message}</p>`;
    } else if (to.status === 'no_history') {
      badge = '\u2014 No history yet';
      body = `<p style="color:#666">${to.message}</p>`;
    } else if (to.status === 'ok') {
      badge = `<span style="color:${to.ok?'#1a5c3a':'#c00'}">${to.ok ? '\u2014 \u2713 Matches' : '\u2014 \u2717 Mismatch'}</span>`;
      body = `
        <div>Last reconciled closing (${to.last_closing_date||'\u2014'}): <b>\u20b9${fmtDec(to.last_closing)}</b></div>
        <div>+ Tagged collections this period (${to.tagged_collected_count}): <b>\u20b9${fmtDec(to.tagged_collected)}</b></div>
        <div>\u2212 Tagged disbursements this period (${to.tagged_disbursed_count}): <b>\u20b9${fmtDec(to.tagged_disbursed)}</b></div>
        <div>= Expected closing: <b>\u20b9${fmtDec(to.expected_closing)}</b></div>
        <div>Statement's own closing: <b>\u20b9${fmtDec(to.statement_closing)}</b></div>
        <div style="color:${to.ok?'#1a5c3a':'#c00'};font-weight:700">Variance: ${_fmtSigned(to.variance)}</div>
        <div style="margin-top:8px;font-size:.78rem;color:#888">
          This statement's own Expense total: \u20b9${fmtDec(to.this_statement_expense_total)} &nbsp;|&nbsp;
          Capital In/Out net: ${_fmtSigned(to.this_statement_capital_net)}<br>${to.note||''}
        </div>`;
    }
    toHtml = `<div class="section" style="margin-top:12px">
      <h3>🏦 Per-Account Book Tie-Out <span style="font-weight:400;color:#888">${badge}</span></h3>
      <div style="font-size:.85rem;line-height:1.7">${body}</div></div>`;
  }

  return `<div class="section" style="border-left:4px solid #1A3A5C">
    <h3 style="margin-bottom:2px">🏦 ${st.account || 'Unidentified account'}</h3>
    <div style="font-size:.78rem;color:#666;margin-bottom:12px">📄 ${st.filename} &nbsp;\u00b7&nbsp; ${st.transactions.length} transactions</div>

    <div class="grid" style="margin-bottom:12px">
      <div class="field">
        <label>Bank Account * <span style="font-weight:400;font-size:.75rem">${acctNote}</span> <span data-acct-note="${si}" style="font-weight:400;color:#888;font-size:.75rem"></span></label>
        <input type="text" list="rec-account-list" value="${st.account||''}" placeholder="Select or type..."
               oninput="setStatementAccount(${si}, this.value)">
      </div>
      <div class="field">
        <label>Reconciliation Date *</label>
        <input type="text" value="${st.recon_date||''}" placeholder="DD-MM-YYYY"
               oninput="window._reconStatements[${si}].recon_date = this.value.trim()">
      </div>
      <div class="field">
        <label>Opening Balance (\u20b9)</label>
        <input type="number" data-open="${si}" value="${st.opening_balance}" step="0.01" oninput="recomputeStatementCheck(${si})">
      </div>
      <div class="field">
        <label>Closing Balance (\u20b9)</label>
        <input type="number" data-close="${si}" value="${st.closing_balance}" step="0.01" oninput="recomputeStatementCheck(${si})">
      </div>
    </div>

    <div class="recon-summary" style="margin-bottom:10px">
      <div class="recon-card"><div class="rv">\u20b9${fmtDec(st.opening_balance)}</div><div class="rl">Opening</div></div>
      <div class="recon-card" style="background:#c00"><div class="rv">\u20b9${fmtDec(totalDr)}</div><div class="rl">Total Debits</div></div>
      <div class="recon-card" style="background:#1a5c3a"><div class="rv">\u20b9${fmtDec(totalCr)}</div><div class="rl">Total Credits</div></div>
      <div class="recon-card" style="background:#b8860b"><div class="rv">\u20b9${fmtDec(st.closing_balance)}</div><div class="rl">Closing</div></div>
    </div>
    ${dateNote}
    <div data-check="${si}" style="margin:0 0 12px;font-size:.85rem;font-weight:600"></div>

    <div class="section" style="background:#f0faf4;border-color:#a8d5b5;margin:0">
      <h3 style="color:#1a5c3a">\u2705 Auto-Reconciled <span style="font-weight:400;color:#555">\u2014 ${(st.confident||[]).length} of ${st.transactions.length} entries matched automatically</span></h3>
      <div style="font-size:.83rem;color:#333;line-height:1.8">${confHtml || '<span style="color:#888">Nothing matched automatically.</span>'}</div>
    </div>
    ${bcHtml}${toHtml}
  </div>`;
}

function renderReconResult(r) {
  // The server always sends 'statements'; the fallback keeps an older cached
  // page from blanking if it ever receives the flat single-file shape.
  const statements = r.statements || [{
    filename: 'statement', transactions: r.transactions || [],
    opening_balance: r.opening_balance || 0, closing_balance: r.closing_balance || 0,
    account: '', account_source: 'not identified', recon_date: r.recon_date || '',
    statement_from: r.statement_from, statement_to: r.statement_to,
    date_corrected: r.date_corrected, typed_date: r.typed_date,
    confident: r.confident || [], review: r.review || [],
    book_completeness: r.book_completeness, account_tieout: r.account_tieout,
  }];
  window._reconStatements = statements;

  document.getElementById("recon-statements").innerHTML =
    statements.map((st, si) => _statementBlock(st, si)).join('');
  statements.forEach((st, si) => { if (!st.error) recomputeStatementCheck(si); });

  // Own-account transfers paired across the uploaded files.
  const pairs = r.cross_account_pairs || [];
  const xSec = document.getElementById("cross-account-section");
  if (pairs.length) {
    xSec.style.display = "block";
    document.getElementById("cross-badge").textContent =
      `\u2014 ${pairs.length} transfer${pairs.length>1?'s':''}, \u20b9${fmtDec(r.cross_account_total||0)}`;
    document.getElementById("cross-account-body").innerHTML =
      `<div style="overflow-x:auto"><table class="recon-table"><thead><tr>
         <th style="text-align:right">Amount</th><th>Out of</th><th>Into</th><th>Dates</th><th>Re-typed</th>
       </tr></thead><tbody>` + pairs.map(p => `<tr>
         <td style="text-align:right;font-weight:600">\u20b9${fmtDec(p.amount)}</td>
         <td style="font-size:.8rem"><b>${p.out.account}</b><br><span style="color:#666">${(p.out.description||'').replace(/</g,"&lt;")}</span></td>
         <td style="font-size:.8rem"><b>${p.in.account}</b><br><span style="color:#666">${(p.in.description||'').replace(/</g,"&lt;")}</span></td>
         <td style="font-size:.8rem;white-space:nowrap">${p.out.date} \u2192 ${p.in.date}${p.gap_days ? ` (${p.gap_days}d)` : ' (same day)'}</td>
         <td style="font-size:.78rem;color:#12467e">${p.retyped.join(' + ')}</td>
       </tr>`).join('') + `</tbody></table></div>`;
  } else {
    xSec.style.display = "none";
  }

  // Accounts reconciled to different dates is the one state in which a
  // transfer between them is missing from the combined bank total.
  // A day with no rows for one account, while another account moved money
  // that same day, is the signature of a statement never uploaded — and it
  // is exactly why a real IDFC disbursement on 11-Aug had no bank evidence.
  const gaps = r.account_day_gaps || {};
  const gapNames = Object.keys(gaps);
  const gapEl = document.getElementById("account-gap-note");
  if (gapEl) {
    if (gapNames.length) {
      gapEl.style.display = "block";
      gapEl.innerHTML = `⚠️ <b>Days missing from an account's history.</b> These days have transactions in another account but none in this one — usually a statement that was never uploaded, which leaves real entries with no bank evidence:<br>`
        + gapNames.map(a => `&nbsp;&nbsp;<b>${a}</b>: ${gaps[a].join(', ')}`).join('<br>')
        + `<br><span style="font-size:.78rem">Re-upload those days for that account to close the gap.</span>`;
    } else {
      gapEl.style.display = "none";
    }
  }

  const dm = document.getElementById("date-mismatch-note");
  if (r.date_mismatch) {
    dm.style.display = "block";
    dm.innerHTML = `\u26a0\ufe0f These statements cover <b>different dates</b> (${r.date_mismatch.join(', ')}). A transfer between the accounts dated in that gap has left one account but not yet arrived in the other, so the combined bank total will understate until both are reconciled to the same date. Re-download them for the same period if you can.`;
  } else {
    dm.style.display = "none";
  }

  // One review queue across every statement. Each row remembers WHICH
  // statement and WHICH row it came from, so a correction can never land on
  // the same-looking transaction in the other account.
  const rev = [];
  statements.forEach((st, si) => {
    if (st.error) return;
    (st.review || []).forEach(tx => {
      const ti = st.transactions.indexOf(tx);
      if (ti >= 0) rev.push({si, ti, tx});
    });
  });
  window._reviewTxns = rev;
  if (rev.length) {
    document.getElementById("review-section").style.display = "block";
    document.getElementById("review-badge").textContent = `\u2014 ${rev.length} entries need review`;
    document.getElementById("review-tbody").innerHTML = rev.map((row, i) => {
      const tx = row.tx;
      const desc = (tx.description||"").replace(/</g,"&lt;");
      const amt  = tx.debit ? "\u20b9"+fmtDec(tx.debit) : "\u20b9"+fmtDec(tx.credit);
      const drCr = tx.debit ? `<span style="color:#c00;font-weight:600">Dr</span>`
                            : `<span style="color:#1a5c3a;font-weight:600">Cr</span>`;
      const autoColor = ({Disbursement:"#b8860b",Repayment:"#1a5c3a",Collection:"#1a5c3a",
        "Other Income":"#1a5c3a","Interest Income":"#1a5c3a",Expense:"#c00",
        "FD Booking":"#1565c0","Capital In":"#1565c0","Capital Out":"#1565c0","Contra":"#1565c0"})[tx.type] || "#555";
      return `<tr style="background:${i%2?"#f8fbff":"white"}">
        <td style="font-size:.75rem;color:#12467e;font-weight:600">${statements[row.si].account || statements[row.si].filename}</td>
        <td style="white-space:nowrap;font-size:.82rem">${tx.date}</td>
        <td style="font-size:.8rem;max-width:220px;overflow:hidden;text-overflow:ellipsis;white-space:nowrap" title="${desc}">${desc}</td>
        <td style="font-size:.78rem;color:#666;max-width:130px;overflow:hidden;text-overflow:ellipsis;white-space:nowrap">${tx.utr||"\u2014"}</td>
        <td style="text-align:right;font-size:.82rem;font-weight:600">${amt}</td>
        <td style="text-align:center">${drCr}</td>
        <td style="font-size:.75rem;color:${autoColor};font-weight:600;white-space:nowrap">${tx.type||"\u2014"}</td>
        <td>${buildTypeSelect(i, tx.type)}</td>
        <td><input type="text" data-review-row="${i}" placeholder="remarks\u2026"
            style="width:100%;font-size:.82rem;padding:4px 6px;border:1px solid #b0c4d8;border-radius:5px;box-sizing:border-box"></td>
      </tr>`;
    }).join("");
  } else {
    document.getElementById("review-section").style.display = "none";
  }
  document.getElementById("recon-preview-section").style.display = "block";
  document.getElementById("recon-status").style.display = "none";
}


async function saveRecon() {
  const statements = (window._reconStatements || []).filter(s => !s.error && s.transactions.length);
  if (!statements.length) return alert('No transactions to save.');

  // Every statement needs an account and a date before anything is written —
  // checked up front, because the save is per-account and a half-done batch
  // is far more annoying to reason about than a refused one.
  for (const st of statements) {
    if (!st.account) return alert(`Please set the bank account for ${st.filename} before saving.`);
    if (!st.recon_date) return alert(`Please set the reconciliation date for ${st.filename} before saving.`);
  }
  for (const st of statements) await registerAccountIfNew(st.account);

  // Corrections from the review queue, applied back by (statement, row)
  // index. The old code matched on date+description+amount, which two
  // accounts can share — with several statements in one upload that key
  // could put a correction on the wrong account's transaction.
  // Remarks are remarks ONLY — they used to also get saved as new dropdown
  // "types", which is how the type list filled up with raw SMS text and
  // case IDs. New types come exclusively from "+ Add new type…" now.
  const fixes = {};
  (window._reviewTxns || []).forEach((row, i) => {
    const sel = document.querySelector(`#review-tbody select[data-row="${i}"]`);
    const inp = document.querySelector(`input[data-review-row="${i}"]`);
    fixes[`${row.si}|${row.ti}`] = {
      type_override: sel && sel.value !== '__add__' ? sel.value.trim() : '',
      row_remarks:   inp ? inp.value.trim() : '',
    };
  });

  const remarks = document.getElementById('rec-remarks').value.trim();
  const data = {
    statements: statements.map(st => {
      const si = window._reconStatements.indexOf(st);
      return {
        date:        st.recon_date,
        opening:     st.opening_balance || 0,
        closing:     st.closing_balance || 0,
        remarks:     remarks,
        account:     st.account,
        remarks_map: {},
        transactions: st.transactions.map((tx, ti) => {
          const fix = fixes[`${si}|${ti}`] || {};
          return {...tx, type_override: fix.type_override||'', row_remarks: fix.row_remarks||''};
        }),
      };
    }),
  };
  const btn = event.target; btn.disabled = true; btn.textContent = 'Saving...';
  try {
    const resp = await fetch('/reconcile/save', {method:'POST',
      headers:{'Content-Type':'application/json'}, body: JSON.stringify(data)});
    if (!resp.ok) {
      const err = await resp.json().catch(() => ({error: 'Unknown error'}));
      throw new Error(err.error || `HTTP ${resp.status}`);
    }
    const rowsSaved = resp.headers.get('X-Rows-Saved') || '';
    let perAccount = [];
    try { perAccount = JSON.parse(resp.headers.get('X-Accounts-Saved') || '[]'); } catch {}
    const blob = await resp.blob();
    const disposition = resp.headers.get('Content-Disposition') || '';
    const match = disposition.match(/filename="(.+)"/);
    const filename = match ? match[1] : 'Daily Reconciliation.xlsx';
    const url = URL.createObjectURL(blob);
    const a = document.createElement('a');
    a.href = url; a.download = filename;
    document.body.appendChild(a); a.click(); a.remove();
    URL.revokeObjectURL(url);
    const detail = perAccount.length
      ? perAccount.map(s => `${s.account} ${s.date}: ${s.rows} txns${s.replaced ? ', rewrote ' + s.replaced + ' earlier row(s)' : ''}, closing ₹${fmtDec(s.closing)}`).join(' · ')
      : `${rowsSaved} transactions`;
    showStatus('recon-status','success',
      `✅ Downloaded ${filename} (full history) — ${detail}`);

    // Full reset so the next upload (same session, no reload) starts clean —
    // otherwise a leftover account/date/balance could silently carry into the
    // next statement. Only on SUCCESS — a failed save should leave everything
    // intact so the user can just retry.
    window._reviewTxns = [];
    window._reconStatements = [];
    _selectedReconFiles = [];
    document.getElementById('rec-file').value = '';
    renderSelectedFiles();
    document.getElementById('recon-statements').innerHTML = '';
    document.getElementById('recon-preview-section').style.display = 'none';
    document.getElementById('rec-remarks').value = '';
  } catch (e) {
    showStatus('recon-status','error','❌ ' + e.message);
  } finally {
    btn.disabled = false; btn.textContent = '💾 Complete Reconciliation & Save Excel';
  }
}

// ── Settings tab ─────────────────────────────────────────────────────────────
let _allContacts = [];
let _contactsDirty = false;

async function loadContacts() {
  document.getElementById('contacts-body').innerHTML = '<p style="color:#888;font-size:.85rem">Loading…</p>';
  const d = await (await fetch('/contacts')).json();
  if (!d.ok && d.error) {
    document.getElementById('contacts-body').innerHTML =
      `<p style="color:#c00;font-size:.85rem">⚠️ Could not load contacts: ${d.error}</p>`;
    return;
  }
  _allContacts = (d.contacts || []).map((c,i) => ({...c, _id: i}));
  _contactsDirty = false;
  renderContactsTable(_allContacts);
}

function filterContacts(q) {
  const ql = q.toLowerCase();
  const filtered = ql ? _allContacts.filter(c =>
    [c.cluster, c.name, c.branch, c.designation, c.phone, c.email].some(v => (v||'').toLowerCase().includes(ql))
  ) : _allContacts;
  renderContactsTable(filtered);
}

function _markDirty() {
  _contactsDirty = true;
  document.getElementById('contacts-save-btn').style.display = '';
}

function _cellInput(val, field, id, placeholder) {
  return `<input data-id="${id}" data-field="${field}" value="${(val||'').replace(/"/g,'&quot;')}"
    placeholder="${placeholder}" oninput="_editContact(this)"
    style="width:100%;border:none;background:transparent;font-size:.83rem;font-family:inherit;padding:0;outline:none;min-width:60px">`;
}

function renderContactsTable(list) {
  const clusters = {};
  list.forEach(c => { (clusters[c.cluster] = clusters[c.cluster]||[]).push(c); });
  const DESIG_ORDER = {'CLUSTER MANAGER':0,'BRANCH HEAD':1,'BRANCH MANAGER':1};
  const cols = ['Name','Designation','Branch','Phone','Email',''];
  let html = '';
  for (const [cluster, members] of Object.entries(clusters)) {
    members.sort((a,b) => (DESIG_ORDER[(a.designation||'').toUpperCase()]??9) - (DESIG_ORDER[(b.designation||'').toUpperCase()]??9));
    html += `<div style="margin-bottom:20px">
      <div style="display:flex;align-items:center;gap:8px;background:#1a3a5c;color:#fff;padding:6px 12px;border-radius:6px 6px 0 0">
        <input value="${cluster}" data-cluster-old="${cluster}" onchange="_renameCluster(this)"
          style="background:transparent;border:none;color:#fff;font-weight:700;font-size:.85rem;letter-spacing:.5px;flex:1;outline:none;cursor:pointer"
          title="Click to rename cluster">
        <button onclick="addContactRow('${cluster}')" title="Add person to this cluster"
          style="background:rgba(255,255,255,.2);border:none;color:#fff;border-radius:4px;padding:2px 8px;cursor:pointer;font-size:.8rem">+ Person</button>
      </div>
      <table style="width:100%;border-collapse:collapse;font-size:.83rem">
        <thead><tr style="background:#f0f4f8">
          ${cols.map(c => `<th style="padding:6px 10px;text-align:left;border-bottom:1px solid #dde;font-size:.8rem">${c}</th>`).join('')}
        </tr></thead><tbody>`;
    members.forEach((c,i) => {
      const isManager = (c.designation||'').toUpperCase().includes('CLUSTER');
      const bg = isManager ? '#fffbe6' : (i%2===0?'#fff':'#f9fbfd');
      html += `<tr style="background:${bg}" data-id="${c._id}">
        <td style="padding:4px 8px;border-bottom:1px solid #eef;font-weight:${isManager?700:400}">${_cellInput(c.name,'name',c._id,'Name')}</td>
        <td style="padding:4px 8px;border-bottom:1px solid #eef">${_cellInput(c.designation,'designation',c._id,'e.g. Branch Manager')}</td>
        <td style="padding:4px 8px;border-bottom:1px solid #eef">${_cellInput(c.branch,'branch',c._id,'Branch')}</td>
        <td style="padding:4px 8px;border-bottom:1px solid #eef">${_cellInput(c.phone,'phone',c._id,'Phone')}</td>
        <td style="padding:4px 8px;border-bottom:1px solid #eef">${_cellInput(c.email,'email',c._id,'Email')}</td>
        <td style="padding:4px 8px;border-bottom:1px solid #eef;text-align:center">
          <button onclick="_deleteContact(${c._id})" title="Remove"
            style="background:none;border:none;color:#c00;cursor:pointer;font-size:1rem;line-height:1">🗑</button>
        </td></tr>`;
    });
    html += '</tbody></table></div>';
  }
  document.getElementById('contacts-body').innerHTML = html || '<p style="color:#888">No contacts found.</p>';
}

function _editContact(inp) {
  const id = parseInt(inp.dataset.id), field = inp.dataset.field;
  const c = _allContacts.find(x => x._id === id);
  if (c) { c[field] = inp.value; _markDirty(); }
}

function _renameCluster(inp) {
  const oldName = inp.dataset.clusterOld, newName = inp.value.trim();
  if (!newName || newName === oldName) return;
  _allContacts.forEach(c => { if (c.cluster === oldName) c.cluster = newName; });
  inp.dataset.clusterOld = newName;
  _markDirty();
}

function _deleteContact(id) {
  _allContacts = _allContacts.filter(c => c._id !== id);
  _markDirty();
  filterContacts(document.getElementById('contact-search').value);
}

function addContactRow(cluster) {
  const clusters = [...new Set(_allContacts.map(c => c.cluster))];
  const targetCluster = cluster || clusters[0] || 'New Cluster';
  const newId = Math.max(0, ..._allContacts.map(c => c._id)) + 1;
  // Insert after last member of that cluster
  const idx = _allContacts.map(c=>c.cluster).lastIndexOf(targetCluster);
  const newRow = {_id: newId, cluster: targetCluster, name:'', designation:'Branch Manager', branch:'', phone:'', email:''};
  if (idx >= 0) _allContacts.splice(idx + 1, 0, newRow);
  else _allContacts.push(newRow);
  _markDirty();
  filterContacts(document.getElementById('contact-search').value);
  // Focus the new row's name input
  setTimeout(() => {
    const inp = document.querySelector(`input[data-id="${newId}"][data-field="name"]`);
    if (inp) inp.focus();
  }, 50);
}

function addCluster() {
  const name = prompt('New cluster name:');
  if (!name || !name.trim()) return;
  const newId = Math.max(0, ..._allContacts.map(c => c._id)) + 1;
  _allContacts.push({_id: newId, cluster: name.trim(), name:'', designation:'CLUSTER MANAGER', branch:'', phone:'', email:''});
  _markDirty();
  filterContacts(document.getElementById('contact-search').value);
}

async function saveContacts() {
  const btn = document.getElementById('contacts-save-btn');
  btn.textContent = '⏳ Saving…'; btn.disabled = true;
  const payload = _allContacts.filter(c => c.name.trim()).map(({_id, ...rest}) => rest);
  const r = await (await fetch('/contacts', {method:'POST',
    headers:{'Content-Type':'application/json'}, body: JSON.stringify({contacts: payload})})).json();
  btn.textContent = '💾 Save'; btn.disabled = false;
  if (r.ok) {
    _contactsDirty = false; btn.style.display = 'none';
    showStatus('contacts-status', 'success', '✅ Contacts saved to Google Sheet');
  } else {
    showStatus('contacts-status', 'error', '❌ ' + r.error);
  }
}

async function loadSettings() {
  const cfg = await (await fetch('/config')).json();
  document.getElementById('report-time').value = cfg.report_time || '09:00';
  document.getElementById('set-bulk-debit').value = cfg.bulk_debit_account || '';
  document.getElementById('set-bulk-narration').value = cfg.bulk_narration || '';
  document.getElementById('cl-date').value = new Date().toLocaleDateString('en-GB').replace(/\//g,'-');
  loadSolvencyCheck();
  loadMoneyIntegrity();
  loadCapitalLog();
}

// ── Money Integrity panel ────────────────────────────────────────────────────
async function loadMoneyIntegrity() {
  const badge = document.getElementById('integrity-badge');
  const body  = document.getElementById('integrity-body');
  if (!body) return;
  body.innerHTML = '<span style="color:#888">Checking every rupee…</span>';
  try {
    const r = await (await fetch('/money-integrity')).json();
    if (r.error) {
      badge.textContent = '— error';
      body.innerHTML = '<span style="color:#c00">' + r.error + '</span>';
      return;
    }
    const L = r.leaks || {};
    const bank = L.bank_entries_not_tied_to_a_case || {};
    const book = L.book_entries_with_no_bank_evidence || {};
    const bookAmt = (book.disbursement_total||0) + (book.collection_total||0);
    const internal = Math.abs(L.contra_imbalance||0) + Math.abs(L.fd_unparsed_net||0);
    const rows = [
      ['Bank movement with no explanation', (L.unclassified_types||{}).total_net_out_of_bank||0,
       'Every debit and credit is classified.'],
      ['Bank entry not tied to any case', (bank.total_debit||0)+(bank.total_credit||0),
       (bank.count||0) + ' entry(s).'],
      ['Book entry with no bank evidence', bookAmt,
       (book.disbursement_count||0) + ' disbursement(s), ' + (book.collection_count||0) + ' collection(s).'],
      ['Internal transfers not netting to zero', internal,
       'Own-account transfers and FD sweeps must cancel out.'],
      ['FD swept in from an untracked deposit', L.fd_orphan_total||0,
       'Deposit booked before bank records begin.'],
    ];
    // Charges are the sixth way money can quietly go wrong: a case billed on
    // the wrong counterparty's rate. Fetched separately so a failure here
    // can't blank the bank-side lines, which are the older, load-bearing ones.
    let ci = null;
    try { ci = await (await fetch('/charge-integrity')).json(); } catch {}
    const ciAmt = ci ? (ci.overcharged_total||0) + (ci.undercharged_total||0) : 0;
    if (ci && !ci.error) {
      rows.push(['Case billed on the wrong charge plan', ciAmt,
        (ci.count||0) + ' case(s) whose stored charges disagree with their own company/cluster/date.']);
    }

    const clear = !!r.all_clear && !!(ci && ci.all_clear);
    badge.textContent = clear ? '— ✅ all clear' : '— ⚠️ items to review';
    badge.style.color = clear ? '#1a5c3a' : '#c00';
    const w = r.reconciled_window || {};
    body.innerHTML =
      rows.map(([label, amt, sub]) => {
        const ok = Math.abs(amt) < 1;
        return `<div style="display:flex;justify-content:space-between;gap:10px;padding:5px 0;border-bottom:1px solid #eee">
            <div><b>${label}</b><br><span style="font-size:.76rem;color:#888">${sub}</span></div>
            <div style="white-space:nowrap;font-weight:700;color:${ok?'#1a5c3a':'#c00'}">${ok?'✅ CLEAR':'⚠️ ₹'+fmtDec(Math.abs(amt))}</div>
          </div>`;
      }).join('') +
      ((ci && ci.count) ? `<div style="margin-top:10px;padding:8px 10px;background:#fff5f5;border:1px solid #e0b0b0;border-radius:6px;font-size:.8rem">
          <b>Charge-plan mismatches</b>
          ${ci.issues.map(i => `<div style="margin-top:6px;padding-top:6px;border-top:1px solid #f0dede">
             <b>${i.disb_id}</b> ${i.customer} &nbsp;<span style="color:#666">${i.company} / ${i.cluster}, ${i.date}</span><br>
             Stamped <b>${i.stamped_plan}</b>, should be <b>${i.expected_plan}</b>${i.expected_label?' ('+i.expected_label+')':''}<br>
             Stored ₹${fmtDec(i.stored.total)} vs correct ₹${fmtDec(i.expected.total)} &nbsp;
             <span style="font-weight:700;color:${i.difference<0?'#c00':'#b8860b'}">
               ${i.difference<0 ? 'over-charged ₹'+fmtDec(-i.difference) : 'under-charged ₹'+fmtDec(i.difference)}</span><br>
             <span style="font-size:.75rem;color:#666">Fix: open the case in Edit Case and re-save the Cluster/Company — that recomputes it.</span>
           </div>`).join('')}
        </div>` : '') +
      ((ci && ci.plans) ? `<div style="margin-top:10px;font-size:.76rem;color:#666">
          <b>Charge plans in force:</b> ${ci.plans.map(p =>
            `${p.label} (${Object.entries(p.match||{}).map(([k,v])=>k+'='+v).join(', ')}${p.from?', from '+p.from:''})`
          ).join(' &nbsp;·&nbsp; ') || 'flat 0.5% + GST only'}
          <br>Everything else: flat 0.5% + 18% GST. Add a counterparty rate in the Config sheet's <code>charge_plans</code> — no code change needed.
        </div>` : '') +
      `<div style="margin-top:8px;font-size:.78rem;color:#888">Bank data verified ${w.from||'—'} to ${w.to||'—'}. Activity before that date is not covered.</div>
       <button class="btn" style="width:auto;padding:8px 16px;margin-top:10px;background:#2d5986;color:#fff" onclick="viewReport('integrity')">👁 View full report</button>
       <button class="btn btn-save" style="width:auto;padding:8px 16px;margin-top:10px;margin-left:8px" onclick="downloadReport('integrity')">⬇ Download</button>`;
  } catch (e) {
    badge.textContent = '— error';
    body.innerHTML = '<span style="color:#c00">' + e.message + '</span>';
  }
}

// ── Reports tab ──────────────────────────────────────────────────────────────
let _reportsCatalogue = null;

async function loadReports() {
  const list = document.getElementById('reports-list');
  if (_reportsCatalogue) return;             // catalogue is static; fetch once
  list.innerHTML = '<span style="color:#888">Loading…</span>';
  try {
    const r = await (await fetch('/reports/catalogue')).json();
    if (!r.ok) throw new Error(r.error || 'Could not load reports');
    _reportsCatalogue = r;
    // Default period: current financial year to date.
    const now = new Date();
    const fyStart = new Date(now.getMonth() >= 3 ? now.getFullYear() : now.getFullYear()-1, 3, 1);
    const dmy = d => String(d.getDate()).padStart(2,'0')+'-'+String(d.getMonth()+1).padStart(2,'0')+'-'+d.getFullYear();
    if (!document.getElementById('rep-from').value) document.getElementById('rep-from').value = dmy(fyStart);
    if (!document.getElementById('rep-to').value)   document.getElementById('rep-to').value   = dmy(now);

    const groups = {};
    r.reports.forEach(rep => { (groups[rep.group] = groups[rep.group] || []).push(rep); });
    list.innerHTML = Object.entries(groups).map(([group, reps]) => `
      <div style="margin-bottom:16px">
        <div style="font-size:.78rem;font-weight:700;color:#1a3a5c;text-transform:uppercase;letter-spacing:.04em;margin-bottom:6px">${group}</div>
        ${reps.map(rep => `
          <div style="display:flex;justify-content:space-between;align-items:center;gap:12px;padding:10px 12px;border:1px solid #e2e8f0;border-radius:8px;margin-bottom:6px">
            <div style="flex:1">
              <b style="font-size:.88rem">${rep.label}</b>
              ${rep.needs_period ? '<span style="font-size:.72rem;color:#b8860b;margin-left:6px">uses period above</span>' : ''}
              <br><span style="font-size:.78rem;color:#666">${rep.desc}</span>
            </div>
            ${rep.needs_cluster ? `<select id="rep-cluster-${rep.id}" style="font-size:.8rem;max-width:130px">
                ${(r.clusters||[]).map(c => `<option value="${c}">${c}</option>`).join('')}
              </select>` : ''}
            <button class="btn" style="width:auto;padding:8px 14px;white-space:nowrap;background:#2d5986;color:#fff"
                    onclick="viewReport('${rep.id}')">👁 View</button>
            <button class="btn btn-save" style="width:auto;padding:8px 14px;white-space:nowrap"
                    onclick="downloadReport('${rep.id}')">⬇ PDF</button>
          </div>`).join('')}
      </div>`).join('');
  } catch (e) {
    list.innerHTML = '<span style="color:#c00">' + e.message + '</span>';
  }
}

// One fetch shared by View and Download, so the two can never disagree
// about which period/cluster was asked for.
async function _fetchReport(reportId) {
  const body = {report: reportId};
  const f = document.getElementById('rep-from'), t = document.getElementById('rep-to');
  if (f && f.value.trim()) body.from = f.value.trim();
  if (t && t.value.trim()) body.to = t.value.trim();
  const cl = document.getElementById('rep-cluster-' + reportId);
  if (cl) body.cluster = cl.value;

  const r = await fetch('/reports/generate', {
    method: 'POST', headers: {'Content-Type': 'application/json'}, body: JSON.stringify(body)});
  if (!r.ok) {
    const err = await r.json().catch(() => ({error: 'Unknown error'}));
    throw new Error(err.error || ('HTTP ' + r.status));
  }
  const blob = await r.blob();
  const disposition = r.headers.get('Content-Disposition') || '';
  const m = disposition.match(/filename="(.+)"/);
  return {blob: blob, filename: m ? m[1] : (reportId + '.pdf')};
}

function _saveBlob(blob, filename) {
  const url = URL.createObjectURL(blob);
  const a = document.createElement('a');
  a.href = url; a.download = filename;
  document.body.appendChild(a); a.click(); a.remove();
  URL.revokeObjectURL(url);
}

function _statusTarget() {
  return document.getElementById('reports-status') ? 'reports-status' : 'settings-status';
}

async function _withBusy(btn, label, fn) {
  const orig = btn ? btn.textContent : '';
  if (btn) { btn.disabled = true; btn.textContent = label; }
  try { await fn(); }
  catch (e) { showStatus(_statusTarget(), 'error', '❌ ' + e.message); }
  finally { if (btn) { btn.disabled = false; btn.textContent = orig; } }
}

async function downloadReport(reportId) {
  const btn = event && event.target ? event.target : null;
  await _withBusy(btn, '⏳', async () => {
    const {blob, filename} = await _fetchReport(reportId);
    _saveBlob(blob, filename);
    showStatus(_statusTarget(), 'success', `✅ Downloaded ${filename}`);
  });
}

// ── Report preview ───────────────────────────────────────────────────────────
let _viewedPdfBlob = null, _viewedPdfName = '', _viewedPdfUrl = null;

async function viewReport(reportId) {
  const btn = event && event.target ? event.target : null;
  await _withBusy(btn, '⏳', async () => {
    const {blob, filename} = await _fetchReport(reportId);
    if (_viewedPdfUrl) URL.revokeObjectURL(_viewedPdfUrl);
    _viewedPdfBlob = blob;
    _viewedPdfName = filename;
    _viewedPdfUrl = URL.createObjectURL(blob);
    document.getElementById('pdf-modal-title').textContent = filename;
    document.getElementById('pdf-frame').src = _viewedPdfUrl;
    document.getElementById('pdf-modal-bg').classList.add('show');
  });
}

function openViewedPdfInTab() {
  // Keep the object URL alive — closePdfModal() revokes it, which would
  // break the newly opened tab, so the modal is intentionally left open.
  if (_viewedPdfUrl) window.open(_viewedPdfUrl, '_blank');
}

function downloadViewedPdf() {
  if (_viewedPdfBlob) {
    _saveBlob(_viewedPdfBlob, _viewedPdfName);
    showStatus(_statusTarget(), 'success', `✅ Downloaded ${_viewedPdfName}`);
  }
}

function closePdfModal() {
  document.getElementById('pdf-modal-bg').classList.remove('show');
  document.getElementById('pdf-frame').src = 'about:blank';
  // Release the object URL — without this every preview leaks a blob for
  // the life of the page.
  if (_viewedPdfUrl) { URL.revokeObjectURL(_viewedPdfUrl); _viewedPdfUrl = null; }
  _viewedPdfBlob = null; _viewedPdfName = '';
}

async function saveSettings() {
  const cfg = {
    report_time: document.getElementById('report-time').value,
    bulk_debit_account: document.getElementById('set-bulk-debit').value.trim(),
    bulk_narration: document.getElementById('set-bulk-narration').value.trim(),
  };
  const r = await (await fetch('/config', {method:'POST',
    headers:{'Content-Type':'application/json'}, body: JSON.stringify(cfg)})).json();
  showStatus('settings-status', r.ok ? 'success' : 'error', r.ok ? '✅ Settings saved!' : '❌ '+r.error);
}

async function loadSolvencyCheck() {
  const badge = document.getElementById('solvency-badge');
  const body = document.getElementById('solvency-body');
  badge.textContent = '— loading…';
  body.innerHTML = '';
  try {
    const r = await (await fetch('/solvency-check')).json();
    if (r.status === 'no_capital_log_data' || r.status === 'error') {
      badge.textContent = '';
      body.innerHTML = `<p style="color:#c00">⚠️ ${r.message || r.error || 'Could not load solvency check.'}</p>`;
      return;
    }
    if (r.status === 'incomplete') {
      badge.textContent = '— Incomplete';
      badge.style.color = '#b8860b';
    } else {
      badge.textContent = r.ok ? '— ✓ Matches' : '— ✗ Mismatch';
      badge.style.color = r.ok ? '#1a5c3a' : '#c00';
    }
    const reasonsHtml = (r.reasons || []).length
      ? `<div style="color:#b8860b;margin-bottom:10px">⚠️ ${(r.reasons||[]).join('<br>⚠️ ')}</div>` : '';
    const fdSourceNote = `<span style="color:#1a5c3a">auto-derived per FD account from statement sweep lines, ${r.fd_open_count||0} open FD${(r.fd_open_count===1)?'':'s'}</span>`;
    const fdBreakdown = (r.fd_open_accounts && Object.keys(r.fd_open_accounts).length)
      ? `<div style="font-size:.78rem;color:#666;margin-left:12px">`
        + Object.entries(r.fd_open_accounts).sort((a,b)=>b[1]-a[1])
            .map(([no, amt]) => `FD ${no}: ₹${fmtDec(amt)}`).join('<br>')
        + `</div>`
      : '';
    // The old blind net-sum is kept purely as a cross-check. A gap between
    // it and the per-FD figure is EXPECTED, not alarming — it's exactly the
    // pre-history FDs and returned interest the per-FD ledger handles
    // correctly and the old sum never could.
    const fdCompareNote = (r.fd_total_derived != null && Math.abs(r.fd_total - r.fd_total_derived) > 1)
      ? `<div style="font-size:.78rem;color:#888">Old net-sum method would say ₹${fmtDec(r.fd_total_derived)} (differs by ₹${fmtDec(Math.abs(r.fd_total - r.fd_total_derived))} — the per-FD figure above is the correct one).</div>`
      : '';
    const varHtml = (r.expected_bank_balance != null) ? `
        <div>Net Capital: <b>₹${fmtDec(r.net_capital)}</b></div>
        <div>+ Collected all-time: <b>₹${fmtDec(r.total_collected_all)}</b></div>
        <div>− Disbursed all-time: <b>₹${fmtDec(r.total_disbursed_all)}</b></div>
        <div>− Expenses all-time: <b>₹${fmtDec(r.total_expenses)}</b></div>
        <div>+ Interest / Other income all-time: <b>₹${fmtDec(r.other_income_total||0)}</b></div>
        <div>= Expected Bank Balance: <b>₹${fmtDec(r.expected_bank_balance)}</b></div>
        <div>Actual Bank Balance (${r.bank_balance_date||'—'}): <b>₹${fmtDec(r.bank_balance)}</b></div>
        <div style="color:${r.ok?'#1a5c3a':'#c00'};font-weight:700">Variance: ${r.variance<0?'-':''}₹${fmtDec(Math.abs(r.variance))}</div>
        <div style="margin-top:6px;color:#555">₹${fmtDec(r.fd_total)} currently in FDs (${fdSourceNote}) — not counted in bank balance above.</div>
        ${fdBreakdown}
        ${fdCompareNote}
        <div style="font-size:.78rem;color:#888">Outstanding receivables (money currently with customers, context only): ₹${fmtDec(r.total_outstanding)}</div>`
      : '<p style="color:#666">No bank reconciliation history yet.</p>';
    body.innerHTML = reasonsHtml + varHtml;
    if (r.fd_balance_direct != null) {
      document.getElementById('fd-balance').value = r.fd_balance_direct;
    }
  } catch (e) {
    badge.textContent = '';
    body.innerHTML = `<p style="color:#c00">⚠️ ${e.message}</p>`;
  }
}

async function saveFdBalance() {
  const date = document.getElementById('fd-date').value.trim();
  const balance = document.getElementById('fd-balance').value;
  if (!date || balance === '') return showStatus('fd-balance-status', 'error', 'Date and balance are required.');
  const btn = event.target; btn.disabled = true; btn.textContent = 'Saving…';
  try {
    const r = await (await fetch('/fd-balance', {method:'POST',
      headers:{'Content-Type':'application/json'}, body: JSON.stringify({date, balance})})).json();
    if (r.ok) {
      showStatus('fd-balance-status', 'success', `✅ Saved — ₹${fmtDec(r.balance)} as of ${r.date}`);
      loadSolvencyCheck();
      loadSummary();
    } else {
      showStatus('fd-balance-status', 'error', '❌ ' + (r.error || 'Failed'));
    }
  } catch (e) {
    showStatus('fd-balance-status', 'error', '❌ ' + e.message);
  } finally {
    btn.disabled = false; btn.textContent = '💾 Save';
  }
}

async function loadCapitalLog() {
  const badge = document.getElementById('capital-log-badge');
  const tbody = document.getElementById('capital-log-tbody');
  try {
    const r = await (await fetch('/capital-log')).json();
    if (!r.ok || !r.available) {
      badge.textContent = '';
      tbody.innerHTML = `<tr><td colspan="6" style="color:#c00">${r.error || 'Capital Log unavailable'}</td></tr>`;
      return;
    }
    badge.textContent = `— Net Capital: ₹${fmtDec(r.net_capital)}`;
    tbody.innerHTML = (r.entries || []).slice().reverse().map(e => `
      <tr>
        <td style="white-space:nowrap;font-size:.82rem">${e.date_str}</td>
        <td style="font-size:.82rem">${e.type}</td>
        <td style="font-size:.82rem">${e.partner || '—'}</td>
        <td style="text-align:right;font-size:.82rem;font-weight:600">₹${fmtDec(e.amount)}</td>
        <td style="font-size:.78rem;color:#666">${e.reference || '—'}</td>
        <td style="font-size:.78rem;color:#666">${e.remarks || '—'}</td>
      </tr>`).join('') || '<tr><td colspan="6" style="color:#888">No entries yet.</td></tr>';
  } catch (e) {
    badge.textContent = '';
    tbody.innerHTML = `<tr><td colspan="6" style="color:#c00">${e.message}</td></tr>`;
  }
}

async function saveCapitalLogEntry() {
  const data = {
    date:      document.getElementById('cl-date').value.trim(),
    type:      document.getElementById('cl-type').value,
    partner:   document.getElementById('cl-partner').value.trim(),
    amount:    document.getElementById('cl-amount').value,
    reference: document.getElementById('cl-reference').value.trim(),
    remarks:   document.getElementById('cl-remarks').value.trim(),
  };
  if (!data.date || !data.amount)
    return showStatus('capital-log-status', 'error', 'Date and Amount are required.');
  const btn = event.target; btn.disabled = true; btn.textContent = 'Saving...';
  try {
    const r = await (await fetch('/capital-log', {method:'POST',
      headers:{'Content-Type':'application/json'}, body: JSON.stringify(data)})).json();
    if (r.ok) {
      showStatus('capital-log-status', 'success', `✅ Saved — new net capital: ₹${fmtDec(r.new_running_balance)}`);
      document.getElementById('cl-partner').value = '';
      document.getElementById('cl-amount').value = '';
      document.getElementById('cl-reference').value = '';
      document.getElementById('cl-remarks').value = '';
      loadCapitalLog();
      loadSolvencyCheck();
    } else {
      showStatus('capital-log-status', 'error', '❌ ' + r.error);
    }
  } catch (e) {
    showStatus('capital-log-status', 'error', '❌ ' + e.message);
  } finally {
    btn.disabled = false; btn.textContent = '💾 Add Capital Log Entry';
  }
}

// ── Bank File tab ─────────────────────────────────────────────────────────────
let bulkRequests = [];

function _fmtInr(v) {
  const n = Number(v);
  return n > 0 ? '₹' + n.toLocaleString('en-IN', {maximumFractionDigits:0}) : (v || '—');
}

async function loadBulkRequests() {
  const tbody = document.getElementById('bulk-tbody');
  const empty = document.getElementById('bulk-empty');
  const table = document.getElementById('bulk-table');
  empty.style.display = ''; empty.textContent = 'Loading...';
  try {
    // include=exported so already-exported requests stay visible and can be
    // re-exported (file rejected by bank, details corrected, etc.) — they
    // get an amber "Exported" badge and a double-disbursement warning at
    // export time instead of being hidden entirely.
    const items = await (await fetch('/requests/pending?include=exported')).json();
    if (items.error) throw new Error(items.error);
    bulkRequests = items;
    const pendingCount = items.filter(r => r.status === 'Pending').length;
    document.getElementById('bulk-count').textContent = items.length ? `(${pendingCount} pending)` : '';
    if (!items.length) {
      table.style.display = 'none';
      empty.textContent = 'No pending requests';
      updateBulkFooter();
      return;
    }
    tbody.innerHTML = '';
    items.forEach(r => {
      const tr = document.createElement('tr');
      tr.style.borderBottom = '1px solid #eef3f9';
      const exported = r.status === 'Exported';
      if (exported) tr.style.background = '#fffbf0';
      const badge = exported
        ? `<span style="background:#fff3cd;color:#8a6d1a;font-size:10px;font-weight:700;padding:1px 6px;border-radius:3px;margin-left:6px;">EXPORTED</span>`
        : '';
      tr.innerHTML = `
        <td style="padding:6px 8px;"><input type="checkbox" class="bulk-check" value="${r.request_id}" onchange="updateBulkFooter()"></td>
        <td style="padding:6px 8px;white-space:nowrap;color:#2a7ae2;">${r.request_id}${badge}</td>
        <td style="padding:6px 8px;white-space:nowrap;font-size:12px;color:#888;">${r.submitted_at}</td>
        <td style="padding:6px 8px;font-weight:600;">${r.customer}</td>
        <td style="padding:6px 8px;">${r.cluster} / ${r.branch}</td>
        <td style="padding:6px 8px;white-space:nowrap;">${_fmtInr(r.amount)}</td>
        <td style="padding:6px 8px;font-family:monospace;">${r.account_no || '—'}</td>
        <td style="padding:6px 8px;font-family:monospace;">${r.ifsc || '—'}</td>
        <td style="padding:6px 8px;">${r.bank || '—'}</td>
        <td style="padding:6px 8px;">${r.phone || '—'}</td>
        <td style="padding:6px 8px;">${r.so_name || '—'}</td>
        <td style="padding:6px 8px;">${r.gold_weight || '—'}</td>`;
      tbody.appendChild(tr);
    });
    table.style.display = '';
    empty.style.display = 'none';
    document.getElementById('bulk-select-all').checked = false;
    updateBulkFooter();
  } catch(e) {
    table.style.display = 'none';
    empty.textContent = 'Error: ' + e.message;
  }
}

function toggleBulkSelectAll(checked) {
  document.querySelectorAll('.bulk-check').forEach(c => c.checked = checked);
  updateBulkFooter();
}

function _selectedBulkIds() {
  return Array.from(document.querySelectorAll('.bulk-check:checked')).map(c => c.value);
}

function updateBulkFooter() {
  const ids = _selectedBulkIds();
  const footer = document.getElementById('bulk-footer');
  if (!ids.length) { footer.style.display = 'none'; return; }
  const total = bulkRequests.filter(r => ids.includes(r.request_id))
    .reduce((s, r) => s + (Number(String(r.amount).replace(/[^\\d.]/g,'')) || 0), 0);
  footer.style.display = '';
  footer.textContent = `${ids.length} selected · total ${_fmtInr(total)}`;
}

async function exportBulk() {
  const ids = _selectedBulkIds();
  if (!ids.length) { showStatus('bulk-status', 'error', '❌ Select at least one request'); return; }
  const bankSel = document.getElementById('bulk-bank');
  const bankLabel = bankSel.options[bankSel.selectedIndex].text;
  // A file built for the wrong bank's template will parse as garbage on
  // the actual portal (each bank's column order/meaning differs) — this
  // has caused real rejected uploads before. Force an explicit look at
  // which bank format is selected, every time, right before download.
  if (!confirm(`Export ${ids.length} request(s) as ${bankLabel} format?\n\nMake sure this matches the bank you will actually upload to — a file built for the wrong bank's template will be rejected or misread.`)) {
    return;
  }
  // Regeneration path: already-exported requests CAN be re-exported (bank
  // rejected the file, details changed, etc.) but only after an explicit
  // second confirmation naming the risk — if the earlier file was (or still
  // gets) uploaded too, the customer is paid TWICE.
  const reexports = bulkRequests.filter(r => ids.includes(r.request_id) && r.status === 'Exported');
  if (reexports.length) {
    const names = reexports.map(r => `${r.request_id} (${r.customer})`).join('\\n');
    if (!confirm(`⚠️ REGENERATING ${reexports.length} ALREADY-EXPORTED request(s):\n\n${names}\n\nBE CAREFUL OF DOUBLE DISBURSEMENT — make sure the earlier file was NOT uploaded to the bank (or was rejected). If both files get uploaded, these customers will be PAID TWICE.\n\nContinue with regeneration?`)) {
      return;
    }
  }
  const body = {
    request_ids: ids,
    bank: bankSel.value,
    debit_account: document.getElementById('bulk-debit').value.trim(),
    narration: document.getElementById('bulk-narration').value.trim(),
    value_date: document.getElementById('bulk-value-date').value.trim(),
    allow_reexport: reexports.length > 0,
  };
  showStatus('bulk-status', 'success', '⏳ Generating...');
  try {
    const r = await fetch('/requests/export-bulk', {method:'POST',
      headers:{'Content-Type':'application/json'}, body: JSON.stringify(body)});
    if (!r.ok) {
      const err = await r.json().catch(() => ({error: `HTTP ${r.status}`}));
      if (r.status === 409) {
        showStatus('bulk-status', 'error',
          `❌ Already exported/disbursed: ${(err.request_ids||[]).join(', ')} — list refreshed`);
        loadBulkRequests();
        return;
      }
      throw new Error(err.error || `HTTP ${r.status}`);
    }
    const blob = await r.blob();
    const disposition = r.headers.get('Content-Disposition') || '';
    const match = disposition.match(/filename="(.+)"/);
    const filename = match ? match[1] : 'bulk_payment.csv';
    const url = URL.createObjectURL(blob);
    const a = document.createElement('a');
    a.href = url; a.download = filename;
    document.body.appendChild(a); a.click(); a.remove();
    URL.revokeObjectURL(url);
    showStatus('bulk-status', 'success', `✅ Downloaded ${filename} — ${ids.length} request(s) marked Exported`);
    loadBulkRequests();
  } catch(e) {
    showStatus('bulk-status', 'error', '❌ ' + e.message);
  }
}

async function deleteBulkRequests() {
  const ids = _selectedBulkIds();
  if (!ids.length) { showStatus('bulk-delete-status', 'error', '❌ Select at least one request'); return; }
  const reason = prompt(`Delete ${ids.length} request(s) from the queue?\n\nReason (optional — saved to the sheet):`);
  if (reason === null) return;
  if (!confirm(`Confirm: remove ${ids.join(', ')} from the queue? The row stays in the sheet marked Deleted.`)) return;
  try {
    const r = await (await fetch('/requests/delete', {method:'POST',
      headers:{'Content-Type':'application/json'},
      body: JSON.stringify({request_ids: ids, reason: reason.trim()})})).json();
    if (!r.ok) throw new Error(r.error || 'Failed');
    showStatus('bulk-delete-status', 'success', `✅ Deleted ${r.deleted} request(s)`);
    loadBulkRequests();
  } catch(e) {
    showStatus('bulk-delete-status', 'error', '❌ ' + e.message);
  }
}

async function parseBulkMsg() {
  const msg = document.getElementById('bulk-paste').value;
  if (!msg.trim()) return;
  const r = await (await fetch('/requests/parse', {method:'POST',
    headers:{'Content-Type':'application/json'}, body: JSON.stringify({message: msg})})).json();
  const f = r.fields || {};
  document.getElementById('bp-customer').value = f.customer || '';
  document.getElementById('bp-amount').value = f.amount || '';
  document.getElementById('bp-account').value = f.account_no || '';
  document.getElementById('bp-ifsc').value = f.ifsc || '';
  document.getElementById('bp-bank').value = f.bank || '';
  document.getElementById('bp-phone').value = f.phone || '';
  document.getElementById('bp-cluster').value = f.cluster || '';
  document.getElementById('bp-branch').value = f.branch || '';
  document.getElementById('bp-so').value = f.so_name || '';
  document.getElementById('bp-gold').value = f.gold_weight || '';
  document.getElementById('bulk-parse-warnings').innerHTML =
    (r.warnings || []).map(w => `⚠ ${w}`).join('<br>');
  document.getElementById('bulk-parse-preview').style.display = '';
}

async function addParsedRequest() {
  const body = {
    customer: document.getElementById('bp-customer').value.trim(),
    amount: document.getElementById('bp-amount').value.trim(),
    account_no: document.getElementById('bp-account').value.trim(),
    ifsc: document.getElementById('bp-ifsc').value.trim().toUpperCase(),
    bank: document.getElementById('bp-bank').value.trim(),
    phone: document.getElementById('bp-phone').value.trim(),
    cluster: document.getElementById('bp-cluster').value.trim(),
    branch: document.getElementById('bp-branch').value.trim(),
    so_name: document.getElementById('bp-so').value.trim(),
    gold_weight: document.getElementById('bp-gold').value.trim(),
  };
  const r = await (await fetch('/requests/add', {method:'POST',
    headers:{'Content-Type':'application/json'}, body: JSON.stringify(body)})).json();
  if (r.ok) {
    showStatus('bulk-paste-status', 'success', `✅ Added to queue as ${r.request_id}`);
    document.getElementById('bulk-paste').value = '';
    document.getElementById('bulk-parse-preview').style.display = 'none';
    loadBulkRequests();
  } else {
    showStatus('bulk-paste-status', 'error', '❌ ' + (r.error || 'Failed'));
  }
}

window.onload = async () => {
  const today = new Date().toLocaleDateString('en-GB').replace(/\\//g,'-');
  document.getElementById('d-date').value = today;
  document.getElementById('r-date').value = today;
  // No reconciliation-date default any more: each uploaded statement carries
  // its own date, read from the file.
  document.getElementById('fd-date').value = today;
  document.getElementById('bulk-value-date').value = new Date().toLocaleDateString('en-GB');
  await loadCustomTypes();
  await loadBankAccounts();
  loadSummary();
  loadPendingRequests();
  try {
    const cfg = await (await fetch('/config')).json();
    document.getElementById('bulk-debit').value = cfg.bulk_debit_account || '';
    document.getElementById('bulk-narration').value = cfg.bulk_narration || '';
  } catch(e) {}
};

// ── Auto-refresh pending requests ─────────────────────────────────────────────
// New field-app submissions should appear without a manual page refresh.
// True server push isn't available on this serverless host, so poll every
// 45s while the page is visible, plus immediately on returning to the tab.
// The Bank File list shares the same data but is only auto-refreshed when
// that tab is active AND nothing is ticked — silently replacing rows under
// a half-built selection would be worse than a slightly stale list.
function _autoRefreshRequests() {
  if (document.visibilityState !== 'visible') return;
  loadPendingRequests();
  const bulkTab = document.getElementById('bulk');
  if (bulkTab && bulkTab.classList.contains('active') &&
      document.querySelectorAll('.bulk-check:checked').length === 0) {
    loadBulkRequests();
  }
}
setInterval(_autoRefreshRequests, 45000);
document.addEventListener('visibilitychange', () => {
  if (document.visibilityState === 'visible') _autoRefreshRequests();
});
</script>
</body></html>"""

@app.route('/')
def index():
    return render_template_string(HTML)

@app.route('/setup')
def setup():
    return render_template_string(SETUP_HTML)

@app.route('/extract/disbursement', methods=['POST'])
def api_extract_disbursement():
    return jsonify(extract_disbursement(request.json.get('message', '')))

@app.route('/extract/repayment', methods=['POST'])
def api_extract_repayment():
    return jsonify(extract_repayment(request.json.get('message', '')))

@app.route('/open-cases')
def api_open_cases():
    try:
        return jsonify(get_open_cases())
    except Exception as e:
        return jsonify({'error': str(e)}), 500

def _mark_request_disbursed(request_id, disb_id):
    _, ws = get_requests_sheet()
    all_vals = ws.get_all_values()
    for i, row in enumerate(all_vals[1:], start=2):
        if row and row[0] == request_id:
            ws.update_cell(i, 12, 'Disbursed')
            ws.update_cell(i, 13, disb_id)
            return

REQUEST_DISB_LOG_SHEET_NAME = "Request-Disb Log"
REQUEST_DISB_LOG_HEADERS = ["Request ID", "Disb ID", "Customer Name", "Date"]

def _log_request_disb_link(request_id, disb_id, customer, date_str):
    """Dedicated, append-only tracking tab mapping every field-request that
    became a real disbursement to its Disb ID. The Requests tab already
    carries this same link (its own Disb ID column, stamped by
    _mark_request_disbursed() above), but that tab accumulates other,
    unrelated columns over time -- this stays a flat, simple 4-column log
    purpose-built for that one lookup. Best-effort: never blocks the
    disbursement itself."""
    try:
        gc = get_gspread_client()
        sh = gc.open_by_key(SPREADSHEET_ID)
        try:
            ws = sh.worksheet(REQUEST_DISB_LOG_SHEET_NAME)
        except gspread.exceptions.WorksheetNotFound:
            ws = sh.add_worksheet(title=REQUEST_DISB_LOG_SHEET_NAME,
                                   rows=1000, cols=len(REQUEST_DISB_LOG_HEADERS))
            ws.append_row(REQUEST_DISB_LOG_HEADERS)
        ws.append_row([request_id, disb_id, customer, date_str])
    except Exception as e:
        notify_ops('request_disb_log', e)

def _lookup_request_bank_account(request_id):
    """The Request's 'Debit Account' column (an account NUMBER, stamped at
    /requests/export-bulk time) resolved to the matching Config bank_accounts
    entry's friendly NAME — the same identifier used everywhere else a
    disbursement/collection is tagged with a bank account. Read-only linear
    scan mirroring _mark_request_disbursed's own lookup; called from
    save_disbursement() to auto-fill the Accounts row's Bank Account when
    the form's manual field was left blank."""
    _, ws = get_requests_sheet()
    all_vals = ws.get_all_values()
    if not all_vals:
        return ''
    idx = {h: i for i, h in enumerate(all_vals[0])}
    da_col = idx.get('Debit Account', 15)
    debit_acct_num = ''
    for row in all_vals[1:]:
        if row and row[0] == request_id:
            debit_acct_num = row[da_col].strip() if len(row) > da_col else ''
            break
    if not debit_acct_num:
        return ''
    for acc in (load_config().get('bank_accounts') or []):
        if str(acc.get('account_number', '')).strip() == debit_acct_num:
            return acc.get('name', '')
    return ''

def _lookup_request_kyc_folder(request_id):
    """The Request's 'KYC Folder' column (a Drive folder URL, stamped by the
    field app's /api/upload-kyc-docs at scan-upload time) -- a straight
    passthrough of the stamped URL, unlike _lookup_request_bank_account()
    there's no Config cross-reference to resolve since there's no friendly
    name involved, just a link. Mirrors that function's scan/idx pattern."""
    _, ws = get_requests_sheet()
    all_vals = ws.get_all_values()
    if not all_vals:
        return ''
    idx = {h: i for i, h in enumerate(all_vals[0])}
    kf_col = idx.get('KYC Folder', 16)
    for row in all_vals[1:]:
        if row and row[0] == request_id:
            return row[kf_col].strip() if len(row) > kf_col else ''
    return ''

def _request_row_to_item(row, idx):
    def col(name, default_i):
        i = idx.get(name, default_i)
        return row[i] if len(row) > i else ''
    return {
        'request_id':  col('Request ID', 0),
        'submitted_at': col('Submitted At', 1),
        'customer':    col('Customer Name', 2),
        'cluster':     col('Cluster', 3),
        'branch':      col('Branch', 4),
        'amount':      col('Amount', 5),
        'account_no':  col('Account No', 6),
        'ifsc':        col('IFSC', 7),
        'phone':       col('Phone', 8),
        'so_name':     col('SO Name', 9),
        'gold_weight': col('Gold Weight', 10),
        'status':      col('Status', 11),
        'bank':        col('Bank', 14),
        'debit_account': col('Debit Account', 15),
        'kyc_folder':  col('KYC Folder', 16),
        'company':     col('Company', 17),
    }

@app.route('/requests/pending')
def api_requests_pending():
    try:
        # ?include=exported keeps just-exported requests visible (used by the
        # New Disbursement picker so click-to-fill still works after a bank
        # run); the Bank File tab calls this plain and sees Pending only.
        wanted = {'Pending'}
        if request.args.get('include') == 'exported':
            wanted.add('Exported')
        _, ws = get_requests_sheet()
        all_vals = ws.get_all_values()
        if len(all_vals) < 2:
            return jsonify([])
        idx = {h: i for i, h in enumerate(all_vals[0])}
        items = []
        for row in reversed(all_vals[1:]):
            if not row or not row[0]:
                continue
            status = row[idx.get('Status', 11)] if len(row) > 11 else ''
            if status not in wanted:
                continue
            items.append(_request_row_to_item(row, idx))
        return jsonify(items)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

def _build_bulk_csv(template, rows):
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow([h for h, _ in template['columns']])
    for r in rows:
        w.writerow([_resolve_bulk_field(f, r) for _, f in template['columns']])
    return buf.getvalue().encode('utf-8')

IDFC_TEMPLATE_PATH = os.path.join(ASSETS_DIR, 'templates', 'idfc_bulkpay_template.xlsx')

def _build_idfc_bulk_xlsx(template, rows):
    """Starts from IDFC's own downloaded template file byte-for-byte instead
    of building a fresh workbook from scratch — a from-scratch rebuild
    matched every column name, order, and even the cell text/General
    formats, and IDFC's portal still rejected it. Reusing the actual file
    (its exact styles, column widths, any workbook-level metadata a
    stricter validator might check) is the only way left to guarantee
    fidelity. IDFC's portal requires row 1 (header) and row 2 (their own
    instructions) to stay untouched exactly as downloaded — real data
    must start at row 3; a prior version of this function deleted row 2
    and started data there, which the portal silently rejected even
    though every column/value looked correct."""
    wb = openpyxl.load_workbook(IDFC_TEMPLATE_PATH)
    ws = wb.active

    ncols = len(template['columns'])
    # Verified against a real IDFC batch that the portal actually accepted:
    # text ('@') for Name/Account/IFSC/Transaction Type/Currency, General
    # for everything else. Hardcoded rather than sampled from a template
    # row, since nothing but the header+instructions rows should ever be
    # trusted as present in the bundled template file.
    text_cols = {1, 2, 3, 4, 8}
    formats = ['@' if (c in text_cols) else 'General' for c in range(1, ncols + 1)]

    # Strip everything from row 3 down (leftover data rows) — rows 1-2
    # (header + IDFC's own instructions) are never touched.
    if ws.max_row > 2:
        ws.delete_rows(3, ws.max_row - 2)

    for r_idx, r in enumerate(rows, start=3):
        for c_idx, (_, field) in enumerate(template['columns'], start=1):
            cell = ws.cell(r_idx, c_idx, _resolve_bulk_field(field, r))
            cell.number_format = formats[c_idx - 1]

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()

def _build_bulk_xlsx(template, rows):
    wb = openpyxl.Workbook()
    ws = wb.active
    headers = [h for h, _ in template['columns']]
    ws.append(headers)
    for r in rows:
        ws.append([_resolve_bulk_field(f, r) for _, f in template['columns']])
    # Some portals (IDFC especially) validate a cell's stored FORMAT, not
    # just its value — the real downloaded template has Beneficiary Name/
    # Account Number/IFSC/Transaction Type/Currency stored as explicit text
    # ('@'), and openpyxl's default 'General' format was silently causing
    # rejections even though the values themselves looked right.
    text_cols = template.get('text_columns')
    if text_cols:
        col_idxs = [i for i, h in enumerate(headers, start=1) if h in text_cols]
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for i in col_idxs:
                row[i - 1].number_format = '@'
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()

def _resolve_bulk_field(field, row):
    if isinstance(field, tuple) and field[0] == 'const':
        return field[1]
    return row.get(field, '')

def _clean_request_amount(raw):
    """'₹11,75,000' / 'Rs. 1175000.00' → numeric; portals reject ₹ and commas.
    Extracts the first number pattern rather than stripping globally, so the
    dot in a 'Rs.' prefix can't corrupt the value. Distinct from the recon
    _clean_amount(), which must preserve negative signs."""
    m = re.search(r'\d[\d,]*(?:\.\d{1,2})?', str(raw or ''))
    if not m:
        return 0
    val = float(m.group(0).replace(',', ''))
    return int(val) if val == int(val) else val

def _next_idfc_batch_filename(all_vals):
    """IDFC requires the filename itself under 14 characters, no spaces:
    Batch{N}-ddmmyy.xlsx (e.g. 'Batch1-110726.xlsx' = 19 chars total, but
    the STEM 'Batch1-110726' the portal actually validates is 13). N
    increments per day across every IDFC export today, scanned from the
    'Exported ...' stamps already written into Requests' Notes column
    (same scan-based sequencing pattern as Request ID generation — no
    separate counter to keep in sync)."""
    today = datetime.now(IST).strftime('%d%m%y')
    seq = 1
    pattern = re.compile(rf'Batch(\d+)-{today}\.xlsx')
    for row in all_vals[1:] if all_vals else []:
        if not row:
            continue
        notes = row[13] if len(row) > 13 else ''
        for m in pattern.finditer(notes or ''):
            n = int(m.group(1))
            if n >= seq:
                seq = n + 1
    return f'Batch{seq}-{today}.xlsx'

@app.route('/requests/export-bulk', methods=['POST'])
def api_requests_export_bulk():
    try:
        body = request.json or {}
        ids = body.get('request_ids') or []
        bank_key = body.get('bank', '')
        template = BANK_TEMPLATES.get(bank_key)
        if not template:
            return jsonify({'ok': False, 'error': f'Unknown bank format: {bank_key}'}), 400
        if not ids:
            return jsonify({'ok': False, 'error': 'No requests selected'}), 400

        cfg = load_config()
        # IDFC's own bulk-pay debit account is BridgeLine's fixed IDFC
        # account — always this value for that template regardless of
        # Config, since it's not really a "which account to debit" choice
        # like the other templates have.
        default_debit = IDFC_DEBIT_ACCOUNT if bank_key == 'idfc' else cfg.get('bulk_debit_account', '')
        debit_account = body.get('debit_account') or default_debit
        narration = body.get('narration') or cfg.get('bulk_narration', '')
        value_date = body.get('value_date') or datetime.now(IST).strftime('%d/%m/%Y')

        _, ws = get_requests_sheet()
        all_vals = ws.get_all_values()
        idx = {h: i for i, h in enumerate(all_vals[0])} if all_vals else {}
        by_id = {}
        status_i = idx.get('Status', 11)
        for rownum, row in enumerate(all_vals[1:], start=2):
            if row and row[0] in ids:
                # Duplicate Request IDs happen (double submissions leave an
                # 'ID COLLISION' pair, one Pending + one Deleted). Last-row-
                # wins used to let a Deleted duplicate shadow the live
                # Pending row and 409 the whole export — prefer the Pending
                # row whenever an ID appears more than once.
                prev = by_id.get(row[0])
                prev_status = prev[1][status_i] if prev and len(prev[1]) > status_i else ''
                if prev is None or prev_status != 'Pending':
                    by_id[row[0]] = (rownum, row)

        missing = [i for i in ids if i not in by_id]
        if missing:
            return jsonify({'ok': False, 'error': 'Requests not found',
                            'request_ids': missing}), 404
        # Double-export guard: anything not Pending has already been exported
        # or disbursed — refuse the whole batch so nothing is paid twice.
        # With allow_reexport (the widget sends it only after an explicit
        # double-disbursement warning the user confirmed), 'Exported'
        # requests may be regenerated — a rejected/incorrect file is a
        # legitimate reason to need the same request in a fresh file.
        # 'Disbursed' stays refused unconditionally: money already left the
        # bank, so regenerating that file IS the double payment.
        allow_reexport = bool(body.get('allow_reexport'))
        allowed_statuses = {'Pending', 'Exported'} if allow_reexport else {'Pending'}
        conflicts = [i for i in ids
                     if by_id[i][1][idx.get('Status', 11)] not in allowed_statuses]
        if conflicts:
            return jsonify({'ok': False, 'error': 'Not pending (already exported or disbursed?)',
                            'request_ids': conflicts}), 409

        export_rows = []
        for req_id in ids:
            item = _request_row_to_item(by_id[req_id][1], idx)
            amount = _clean_request_amount(item['amount'])
            row_extra = {
                'amount': amount,
                'mode': 'RTGS' if amount >= RTGS_MIN_AMOUNT else 'NEFT',
                'debit_account': debit_account,
                'narration': narration,
                'value_date': value_date,
                'bank_name': item['bank'],
            }
            if bank_key == 'idfc':
                # Beneficiary already holds an IDFC FIRST account = an
                # in-bank transfer: IFT, no IFSC needed. Anything else is
                # inter-bank: NEFT or RTGS by amount, IFSC required. IMPS
                # is NOT a valid value here — the template's own Transaction
                # Type instructions (cell D2) explicitly list only
                # IFT/NEFT/RTGS as accepted; picking IMPS would just trade
                # one rejection for another.
                if 'idfc' in (item.get('bank') or '').lower():
                    row_extra['txn_type'] = 'IFT'
                    row_extra['ifsc'] = ''
                else:
                    row_extra['txn_type'] = 'RTGS' if amount >= RTGS_MIN_AMOUNT else 'NEFT'
            export_rows.append({**item, **row_extra})

        # Last-line defense before a real bank file gets built: a 2-digit
        # account number once made it all the way to a real disbursement
        # with no check anywhere. Real Indian bank account numbers run
        # 9-18 digits — refuse the WHOLE batch (same philosophy as the
        # double-export guard above) rather than silently dropping one bad
        # row, since a partial file is its own source of confusion.
        bad_accounts = [
            (r['request_id'], r.get('account_no', '')) for r in export_rows
            if not (9 <= len(re.sub(r'\D', '', str(r.get('account_no') or ''))) <= 18)
        ]
        if bad_accounts:
            return jsonify({'ok': False,
                             'error': 'Invalid account number(s) — must be 9-18 digits',
                             'details': [f"{rid}: '{acc}'" for rid, acc in bad_accounts]}), 400

        if template['filetype'] == 'csv':
            file_bytes = _build_bulk_csv(template, export_rows)
            mimetype = 'text/csv'
        elif bank_key == 'idfc':
            file_bytes = _build_idfc_bulk_xlsx(template, export_rows)
            mimetype = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        else:
            file_bytes = _build_bulk_xlsx(template, export_rows)
            mimetype = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        filename = (_next_idfc_batch_filename(all_vals) if bank_key == 'idfc'
                    else template['filename'].format(date=datetime.now(IST).strftime('%d-%m-%Y')))

        # Mark AFTER the file is built (a marked-but-unbuilt file is the bad
        # direction), in ONE batch_update so a partial write can't split the
        # batch between Exported and Pending.
        # Filename included in the stamp so _next_idfc_batch_filename() can
        # scan prior batches' Notes to find the next number to use.
        ts = datetime.now(IST).strftime('%d-%m-%Y %H:%M')
        stamp = f"Exported {template['label']} ({filename}) {ts} IST"
        # Re-exports get a louder, distinct stamp — the Notes column is the
        # audit trail, and a regeneration must be visually unmistakable when
        # someone later asks why one request appears in two bank files.
        restamp = f"RE-EXPORTED {template['label']} ({filename}) {ts} IST — verify earlier file was not also uploaded (double disbursement risk)"
        notes_col = idx.get('Notes', 13)
        status_letter = chr(ord('A') + idx.get('Status', 11))
        notes_letter = chr(ord('A') + notes_col)
        # Debit Account is a structured column (not just the free-text Notes
        # stamp above) so save_disbursement() can look it up later and
        # auto-tag the resulting Accounts row with the bank account this
        # export actually used — previously this value was known here and
        # then lost forever once the request became a disbursement.
        debit_letter = chr(ord('A') + idx.get('Debit Account', 15))
        updates = []
        for req_id in ids:
            rownum, row = by_id[req_id]
            was_exported = (row[idx.get('Status', 11)] == 'Exported')
            this_stamp = restamp if was_exported else stamp
            old_note = row[notes_col] if len(row) > notes_col else ''
            note = f"{old_note} | {this_stamp}" if old_note.strip() else this_stamp
            updates.append({'range': f'{status_letter}{rownum}', 'values': [['Exported']]})
            updates.append({'range': f'{notes_letter}{rownum}', 'values': [[note]]})
            updates.append({'range': f'{debit_letter}{rownum}', 'values': [[debit_account]]})
        ws.batch_update(updates)

        return Response(file_bytes, mimetype=mimetype,
                        headers={'Content-Disposition': f'attachment; filename="{filename}"'})
    except Exception as e:
        notify_ops('requests_export_bulk', e)
        return jsonify({'ok': False, 'error': str(e)}), 500

REQUEST_LABELS = [
    ('customer',    ['customer name', 'name']),
    ('phone',       ['phone no', 'phone number', 'phone', 'mobile']),
    ('amount',      ['amount required', 'amount']),
    ('account_no',  ['account no', 'account number', 'a/c no', 'ac no']),
    ('ifsc',        ['ifsc code', 'ifsc']),
    ('bank',        ['bank name', 'bank']),
    ('cluster',     ['cluster']),
    ('branch',      ['branch']),
    ('so_name',     ['so name']),
    ('gold_weight', ['gold wt (gms)', 'gold wt', 'gold weight']),
    ('date',        ['date']),
]

def parse_request_message(text):
    fields, warnings = {}, []
    # Longest label first so "Customer Name" wins over "Name" and
    # "IFSC Code" over "IFSC".
    label_map = sorted(
        [(lbl, key) for key, labels in REQUEST_LABELS for lbl in labels],
        key=lambda x: -len(x[0]))
    for line in (text or '').splitlines():
        m = re.match(r'^\s*([^:\-]+?)\s*[:\-]\s*(.+?)\s*$', line)
        if not m:
            continue
        label = re.sub(r'\s+', ' ', m.group(1).strip().lower())
        value = m.group(2).strip()
        for lbl, key in label_map:
            if label == lbl and key not in fields:
                fields[key] = value
                break
    if 'amount' in fields:
        fields['amount'] = _clean_request_amount(fields['amount'])
        if not fields['amount']:
            warnings.append('Amount could not be read as a number')
    if 'ifsc' in fields:
        fields['ifsc'] = fields['ifsc'].upper().replace(' ', '')
        if not re.match(r'^[A-Z]{4}0[A-Z0-9]{6}$', fields['ifsc']):
            warnings.append(f"IFSC '{fields['ifsc']}' doesn't look valid (AAAA0XXXXXX)")
    if 'phone' in fields:
        fields['phone'] = re.sub(r'\D', '', fields['phone'])
        if len(fields['phone']) != 10:
            warnings.append(f"Phone '{fields['phone']}' is not 10 digits")
    if 'account_no' in fields:
        fields['account_no'] = re.sub(r'\s', '', fields['account_no'])
    for required in ('customer', 'amount', 'account_no', 'ifsc'):
        if not fields.get(required):
            warnings.append(f'Missing: {required}')
    territory_warning = _bib_territory_mismatch_warning(fields.get('cluster', ''), fields.get('branch', ''))
    if territory_warning:
        warnings.append(territory_warning)
    return fields, warnings

def generate_request_id(ws):
    # Ported verbatim from disbursement-form-deploy/api/submit-request.py so
    # widget-added and field-app requests share one REQ-DDMMYY-NNN sequence.
    # IST, not server time (Vercel's Python runtime is UTC) — otherwise a
    # request added after 5:30pm IST gets tomorrow's date-prefix.
    today = datetime.now(IST)
    prefix = f"REQ-{today.strftime('%d%m%y')}-"
    seq = 1
    for row in ws.get_all_values()[1:]:
        if row and row[0].startswith(prefix):
            try:
                n = int(row[0].split("-")[-1])
                if n >= seq:
                    seq = n + 1
            except ValueError:
                pass
    return f"{prefix}{seq:03d}"

@app.route('/requests/parse', methods=['POST'])
def api_requests_parse():
    fields, warnings = parse_request_message((request.json or {}).get('message', ''))
    return jsonify({'fields': fields, 'warnings': warnings})

@app.route('/requests/delete', methods=['POST'])
def api_requests_delete():
    try:
        body = request.json or {}
        ids = body.get('request_ids') or []
        reason = (body.get('reason') or '').strip()
        if not ids:
            return jsonify({'ok': False, 'error': 'No requests selected'}), 400
        _, ws = get_requests_sheet()
        all_vals = ws.get_all_values()
        idx = {h: i for i, h in enumerate(all_vals[0])} if all_vals else {}
        by_id = {}
        for rownum, row in enumerate(all_vals[1:], start=2):
            if row and row[0] in ids:
                by_id[row[0]] = (rownum, row)
        missing = [i for i in ids if i not in by_id]
        if missing:
            return jsonify({'ok': False, 'error': 'Requests not found',
                            'request_ids': missing}), 404
        # Soft delete: Status → Deleted, reason + timestamp appended to Notes.
        # The row stays in the sheet as the audit trail of what was removed
        # and why; every queue view filters on Status so it disappears
        # from pending/export/disb-picker immediately.
        stamp = f"Deleted{': ' + reason if reason else ''} {datetime.now(IST).strftime('%d-%m-%Y %H:%M')} IST"
        notes_col = idx.get('Notes', 13)
        status_letter = chr(ord('A') + idx.get('Status', 11))
        notes_letter = chr(ord('A') + notes_col)
        updates = []
        for req_id in ids:
            rownum, row = by_id[req_id]
            old_note = row[notes_col] if len(row) > notes_col else ''
            note = f"{old_note} | {stamp}" if old_note.strip() else stamp
            updates.append({'range': f'{status_letter}{rownum}', 'values': [['Deleted']]})
            updates.append({'range': f'{notes_letter}{rownum}', 'values': [[note]]})
        ws.batch_update(updates)
        return jsonify({'ok': True, 'deleted': len(ids)})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500

def _flag_id_collision(ws, req_id):
    """Best-effort detection (not prevention) of the ID-generation race: two
    concurrent submissions can compute the same REQ-DDMMYY-NNN before either
    has appended its row. Re-scanning right after our own append can catch
    it — not airtight (both racers may read a self-consistent but stale
    view), but turns a silent duplicate ID into a flagged one a human will
    notice instead of nothing. Mirrors disbursement-form-deploy/api/
    submit-request.py so both writers behave the same way. Never blocks the
    submission itself."""
    try:
        all_vals = ws.get_all_values()
        header = all_vals[0]
        notes_col = header.index('Notes') if 'Notes' in header else 13
        matches = [i for i, row in enumerate(all_vals[1:], start=2) if row and row[0] == req_id]
        if len(matches) <= 1:
            return
        marker = f"ID COLLISION: {len(matches)} requests share {req_id} - verify manually"
        for rownum in matches:
            row = all_vals[rownum - 1]
            old_note = row[notes_col] if len(row) > notes_col else ''
            if 'ID COLLISION' in old_note:
                continue
            note = f"{old_note} | {marker}" if old_note.strip() else marker
            ws.update_cell(rownum, notes_col + 1, note)
    except Exception:
        pass  # auxiliary safety net — a failure here must never surface as a submit failure

def _find_duplicate_request(all_vals, account_no, amount):
    """Same beneficiary account + same amount submitted today = duplicate,
    whatever its status — a request already exported or disbursed today is
    the worst kind of duplicate. Digits-only compare so '2,00,000' matches."""
    digits = lambda s: re.sub(r'\D', '', str(s or ''))
    today = datetime.now(IST).strftime('%d-%m-%Y')
    acct, amt = digits(account_no), digits(amount)
    if not acct or not amt:
        return None
    for row in all_vals[1:]:
        if len(row) < 12 or not row[0] or not row[1].startswith(today):
            continue
        if row[11] == 'Deleted':
            continue  # deleted requests don't block resubmission
        if digits(row[6]) == acct and digits(row[5]) == amt:
            return {'request_id': row[0], 'submitted_at': row[1], 'status': row[11]}
    return None

@app.route('/requests/add', methods=['POST'])
def api_requests_add():
    try:
        body = request.json or {}
        if not body.get('customer') or not body.get('account_no'):
            return jsonify({'ok': False, 'error': 'Customer and account number are required'}), 400
        _, ws = get_requests_sheet()
        all_vals = ws.get_all_values()
        dup = _find_duplicate_request(all_vals, body.get('account_no'), body.get('amount'))
        if dup:
            return jsonify({'ok': False, 'duplicate': True,
                            'error': f"Duplicate of {dup['request_id']} ({dup['status']}, {dup['submitted_at']})"}), 409
        req_id = generate_request_id(ws)
        ws.append_row([
            req_id,
            datetime.now(IST).strftime("%d-%m-%Y %H:%M IST"),
            body.get('customer', ''),
            body.get('cluster', ''),
            body.get('branch', ''),
            str(body.get('amount', '')),
            body.get('account_no', ''),
            body.get('ifsc', ''),
            body.get('phone', ''),
            body.get('so_name', ''),
            str(body.get('gold_weight', '')),
            'Pending',
            '',
            'Added via widget paste',
            body.get('bank', ''),
        ])
        _flag_id_collision(ws, req_id)
        return jsonify({'ok': True, 'request_id': req_id})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500

@app.route('/rebuild-ledger', methods=['POST'])
def api_rebuild_ledger():
    """Fired by the frontend as a non-blocking (not-awaited) call right
    after a save shows success, instead of running inline during the save
    itself — see rebuild_ledger_now()'s docstring for why. This is a
    genuinely independent request, so Vercel's normal guarantee (a function
    completes once it returns a response) applies; the response body isn't
    meant to be consulted by the caller."""
    try:
        rebuild_ledger_now()
    except Exception:
        pass
    return jsonify({'ok': True})

@app.route('/save/disbursement', methods=['POST'])
def api_save_disbursement():
    try:
        data = request.json or {}
        disb_id = save_disbursement(data)
        resp = {'ok': True, 'disb_id': disb_id}
        warning = _bib_territory_mismatch_warning(data.get('cluster', ''), data.get('branch', ''))
        if warning:
            resp['warning'] = warning
        return jsonify(resp)
    except Exception as e:
        notify_ops('save_disbursement', e)
        return jsonify({'ok': False, 'error': str(e)})

@app.route('/save/repayment', methods=['POST'])
def api_save_repayment():
    try:
        return jsonify({'ok': True, **save_repayment(request.json)})
    except Exception as e:
        notify_ops('save_repayment', e)
        return jsonify({'ok': False, 'error': str(e)})

@app.route('/case/recent')
def api_case_recent():
    try:
        days = int(request.args.get('days', 7))
        return jsonify({'ok': True, 'events': get_recent_activity(days)})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500

@app.route('/case/<disb_id>/detail')
def api_case_detail(disb_id):
    try:
        detail = _case_detail(disb_id)
        if not detail:
            return jsonify({'ok': False, 'error': f'{disb_id.upper()} not found'}), 404
        return jsonify({'ok': True, 'case': detail})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500

@app.route('/case/update-disbursement', methods=['POST'])
def api_case_update_disbursement():
    try:
        known = update_disbursement(request.json)
        return jsonify({'ok': True, 'case': _case_detail(request.json['disb_id'], known=known)})
    except Exception as e:
        notify_ops('update_disbursement', e)
        return jsonify({'ok': False, 'error': str(e)})

@app.route('/case/update-repayment', methods=['POST'])
def api_case_update_repayment():
    try:
        known = update_repayment(request.json)
        return jsonify({'ok': True, 'case': _case_detail(request.json['disb_id'], known=known)})
    except Exception as e:
        notify_ops('update_repayment', e)
        return jsonify({'ok': False, 'error': str(e)})

@app.route('/case/delete-repayment', methods=['POST'])
def api_case_delete_repayment():
    try:
        known = delete_repayment(request.json)
        return jsonify({'ok': True, 'case': _case_detail(request.json['disb_id'], known=known)})
    except Exception as e:
        notify_ops('delete_repayment', e)
        return jsonify({'ok': False, 'error': str(e)})

@app.route('/case/add-repayment', methods=['POST'])
def api_case_add_repayment():
    try:
        known = add_repayment(request.json)
        return jsonify({'ok': True, 'case': _case_detail(request.json['disb_id'], known=known)})
    except Exception as e:
        notify_ops('add_repayment', e)
        return jsonify({'ok': False, 'error': str(e)})

def _load_for_invoice(sh, disb_id):
    """Read only Accounts + M Coll (2 API calls) instead of the full
    load_data_from_sheet() which also reads archives, Contact, DashBoard."""
    acc_vals = sh.worksheet(SHEET_NAME).get_all_values()
    header_idx = next((i for i, r in enumerate(acc_vals) if r and 'Disbursement ID' in r), 1)
    raw_rows = [r for r in acc_vals[header_idx + 1:] if r and str(r[0]).strip().startswith('BLP-')]

    # Only scan archive tabs if the case isn't in the main Accounts sheet
    if not any(str(r[0]).strip().upper() == disb_id for r in raw_rows):
        for archive_name in get_archive_tab_names():
            try:
                archive_rows = [r for r in sh.worksheet(archive_name).get_all_values()
                                if r and str(r[0]).strip().startswith('BLP-')]
                raw_rows += archive_rows
                if any(str(r[0]).strip().upper() == disb_id for r in archive_rows):
                    break
            except gspread.exceptions.WorksheetNotFound:
                continue

    rows = []
    for raw in raw_rows:
        # Widened from 26 to 28 to also reach col AB (request_id, 0-indexed
        # 27) for the invoice header -- col Z (charge_plan) was the previous
        # ceiling here.
        row = list(raw[:28]) + [None] * max(0, 28 - len(raw))
        row = [None if c == '' else c for c in row]
        for i in _MIS_NUMERIC_COLS:
            row[i] = _clean_numeric_cell(row[i])
        if row[0]:
            rows.append(row)

    sheets = {}
    try:
        mcoll_vals = sh.worksheet('M Coll').get_all_values()
        mcoll_vals = [[_clean_numeric_cell(c) if i == 2 else c for i, c in enumerate(r)]
                      for r in mcoll_vals]
        sheets['M Coll'] = _SheetShim(mcoll_vals)
    except gspread.exceptions.WorksheetNotFound:
        pass

    mcoll = mis.load_mcoll(_WorkbookShim(sheets))
    return rows, mcoll

@app.route('/generate-invoice', methods=['POST'])
def api_generate_invoice():
    try:
        disb_id = (request.json or {}).get('disb_id', '').strip().upper()
        if not disb_id:
            return jsonify({'error': 'disb_id required'}), 400
        ensure_mis_assets_cached()
        sh = get_gspread_client().open_by_key(SPREADSHEET_ID)
        rows, mcoll = _load_for_invoice(sh, disb_id)
        _, _, all_cases_full = mis.parse_cases(rows, mcoll)
        case = next((c for c in all_cases_full if c['id'].upper() == disb_id), None)
        if not case:
            return jsonify({'error': f'{disb_id} not found'}), 404
        mcoll_entry = mcoll.get(disb_id)
        paid = case['status'] == 'Closed'
        pdf_bytes = mis.generate_invoice_ledger(case, mcoll_entry=mcoll_entry, paid_in_full=paid)
        safe = case['customer'].replace('/', '-')
        filename = f"{disb_id} {safe} Invoice and Ledger.pdf"
        return Response(pdf_bytes, mimetype='application/pdf',
            headers={'Content-Disposition': f'attachment; filename="{filename}"'})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/generate-mis', methods=['POST'])
def api_generate_mis():
    """Runs the daily MIS pipeline live from this Google Sheet (no manual
    Excel export / CLI step) and returns the same ZIP generate_mis.py's CLI
    path produces."""
    try:
        ensure_mis_assets_cached()

        sh = get_gspread_client().open_by_key(SPREADSHEET_ID)
        rows, db_raw, mcoll, (cluster_mgrs, branch_contacts, area_mgrs, territory_to_area), capital_log = load_data_from_sheet(sh)
        open_cases, all_cases, all_cases_full = mis.parse_cases(rows, mcoll)
        metrics = mis.compute_dashboard_metrics(all_cases_full, db_raw, capital_log)

        try:
            mis.write_claude_dashboard_to_sheet(sh, metrics)
        except Exception as e:
            print(f'WARNING: Could not write Claude_Dashboard: {e}')

        active_clusters = sorted(set(c['cluster'] for c in open_cases))
        zip_bytes = mis.build_zip(open_cases, all_cases, all_cases_full, metrics,
                                   cluster_mgrs, branch_contacts, active_clusters, mcoll_raw=mcoll,
                                   area_mgrs=area_mgrs, territory_to_area=territory_to_area)

        date_human = mis.TODAY.strftime('%d-%b-%Y')
        filename = f'{date_human} BridgeLine MIS Package.zip'
        return Response(
            zip_bytes,
            mimetype='application/zip',
            headers={'Content-Disposition': f'attachment; filename="{filename}"'},
        )
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500

@app.route('/summary')
def api_summary():
    try:
        return jsonify(get_today_summary())
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/config', methods=['GET'])
def api_get_config():
    return jsonify(load_config())

@app.route('/config', methods=['POST'])
def api_save_config():
    try:
        cfg = load_config_strict()
        cfg.update(request.json)
        save_config(cfg)
        return jsonify({'ok': True})
    except Exception as e:
        # A failed read here must abort the save outright (not fall back to
        # DEFAULT_CONFIG) — see load_config_strict()'s docstring. The user
        # sees a failed save and can retry; nothing in Config gets touched.
        return jsonify({'ok': False, 'error': f'Config save aborted (read failed, nothing was overwritten): {e}'})

@app.route('/reconcile/parse', methods=['POST'])
def api_reconcile_parse():
    """Parse one OR MANY bank statements in a single request.

    Uploading every account's statement together (13-08-2026, Prem's call) is
    what lets transfers between our own accounts be identified STRUCTURALLY —
    same amount, opposite direction, different account, same few days — instead
    of being inferred from narration text. See _pair_uploaded_statements().

    Still accepts a lone 'file' so an older cached page keeps working; that
    path is just the N=1 case of the same code.
    """
    files = [f for f in (request.files.getlist('files') or []) if f and f.filename]
    if not files:
        files = [f for f in (request.files.getlist('file') or []) if f and f.filename]
    if not files:
        return jsonify({'ok': False, 'error': 'No file uploaded'})

    import tempfile
    tmp_paths = []
    try:
        # One read of the books for the whole batch, not one per file.
        try:
            records = read_accounts_from_gsheet()
            sh      = get_gspread_client().open_by_key(SPREADSHEET_ID)
            mc_rows = sh.worksheet('M Coll').get_all_values()[1:]
        except Exception:
            records, mc_rows, sh = [], [], None
        bank_accounts = (load_config().get('bank_accounts') or [])

        typed_date  = request.form.get('date', '').strip()
        # Manual per-file account overrides, positional and optional — used
        # only where the statement doesn't name an account we recognise.
        forced_accts = request.form.getlist('accounts') or []
        legacy_acct  = request.form.get('account', '').strip()

        statements = []
        for idx, f in enumerate(files):
            suffix = '.' + f.filename.rsplit('.', 1)[-1]
            with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
                f.save(tmp.name)
                tmp_paths.append(tmp.name)
            try:
                result = parse_bank_statement(tmp_paths[-1], f.filename, bank_accounts)
            except Exception as e:
                # One unreadable file must not lose the others' work — report
                # it against its own row and carry on with the rest.
                statements.append({'filename': f.filename, 'error': str(e),
                                   'transactions': [], 'account': '', 'recon_date': ''})
                continue

            # The statement's own declared period wins over whatever was typed
            # — a hand-typed date is the one part of this flow with no source
            # of truth behind it (see _extract_statement_period()).
            stmt_to = result.get('statement_to')
            recon_date = stmt_to or typed_date
            forced = (forced_accts[idx].strip() if idx < len(forced_accts) else '')
            account = (forced or result.get('statement_account')
                       or (legacy_acct if len(files) == 1 else '') or '')
            statements.append({
                'filename': f.filename,
                'transactions': _match_transactions(result['transactions'], records, mc_rows),
                'opening_balance': result['opening_balance'],
                'closing_balance': result['closing_balance'],
                'account': account,
                'account_source': ('you selected' if forced
                                   else 'read from the statement' if result.get('statement_account')
                                   else 'not identified'),
                'account_number': result.get('statement_account_no'),
                'recon_date': recon_date,
                'statement_from': result.get('statement_from'),
                'statement_to': stmt_to,
                'date_corrected': bool(stmt_to and typed_date and stmt_to != typed_date),
                'typed_date': typed_date,
            })

        # Cross-account pairing runs BEFORE the review split, so a transfer
        # correctly identified here never reaches the review queue at all.
        cross_pairs = _pair_uploaded_statements([s for s in statements if not s.get('error')])

        def _needs_review(tx):
            if tx['match_basis'] in ('UTR', 'Manual', 'Cross-account'): return False
            if tx['type'] in ('FD Booking', 'Test Credit'):             return False
            if tx['credit'] > 0 and tx['credit'] < 100:                 return False
            return True

        # Every already-saved period, read once and reused by each statement's
        # completeness check — the whole point of the check is the WHOLE
        # month's picture, not just the file in hand.
        try:
            saved_periods = _load_recon_periods(get_recon_txns_sheet(sh)) if sh else []
        except Exception:
            saved_periods = []

        for st in statements:
            if st.get('error'):
                continue
            classified = st['transactions']
            st['confident'] = [tx for tx in classified if not _needs_review(tx)]
            st['review']    = [tx for tx in classified if _needs_review(tx)]

            try:
                period_month = datetime.strptime(st['recon_date'], '%d-%m-%Y').strftime('%b %Y')
            except Exception:
                period_month = ''
            st['period_month'] = period_month

            st['book_completeness'] = None
            if period_month:
                completeness_txns = list(classified)
                # Other accounts uploaded in this same batch count too — a
                # collection banked into IDFC is not "missing from the books"
                # just because we're looking at the HDFC statement.
                for other in statements:
                    if other is not st and not other.get('error'):
                        completeness_txns.extend(other['transactions'])
                for p in saved_periods:
                    if p['period_label'] == period_month:
                        completeness_txns.extend(p['txns'])
                st['book_completeness'] = _book_completeness(records, mc_rows,
                                                             completeness_txns, period_month)

            total_dr = sum(tx['debit'] for tx in classified)
            total_cr = sum(tx['credit'] for tx in classified)
            expected_closing = round(st['opening_balance'] + total_cr - total_dr, 2)
            variance = round(st['closing_balance'] - expected_closing, 2)
            st['statement_check'] = {
                'total_debit': round(total_dr, 2), 'total_credit': round(total_cr, 2),
                'expected_closing': expected_closing, 'variance': variance,
                'ok': abs(variance) <= 1.0,
            }
            st['account_tieout'] = (
                _account_tieout(records, mc_rows, st['account'], period_month, classified,
                                st['closing_balance'], recon_date=st['recon_date'])
                if (period_month and st['account']) else {'status': 'no_account_selected'})

        ok_stmts = [s for s in statements if not s.get('error')]
        # Days one account is silent while another moved money — the only
        # reliable way to tell "no activity" from "never uploaded".
        try:
            gap_periods = list(saved_periods) + [
                {'account': s['account'] or s['filename'], 'txns': s['transactions']}
                for s in ok_stmts]
            account_day_gaps = _account_day_gaps(gap_periods)
        except Exception:
            account_day_gaps = {}
        # Two accounts reconciled to DIFFERENT dates is the exact state in
        # which a transfer between them vanishes from the combined bank total
        # — it has left one account but not yet arrived in the other. Uploading
        # together is the fix, so say so when the dates don't line up.
        dates = {s['recon_date'] for s in ok_stmts if s['recon_date']}
        date_mismatch = sorted(dates) if len(dates) > 1 else None

        first = ok_stmts[0] if ok_stmts else {}
        return jsonify({
            'ok': True,
            'statements': statements,
            'cross_account_pairs': cross_pairs,
            'cross_account_total': round(sum(p['amount'] for p in cross_pairs), 2),
            'date_mismatch': date_mismatch,
            'account_day_gaps': account_day_gaps,
            # Flat mirror of the first statement, so an older cached page that
            # still expects the single-file shape keeps rendering.
            'opening_balance':  first.get('opening_balance', 0),
            'closing_balance':  first.get('closing_balance', 0),
            'transactions':     first.get('transactions', []),
            'confident':        first.get('confident', []),
            'review':           first.get('review', []),
            'confident_count':  len(first.get('confident', [])),
            'review_count':     len(first.get('review', [])),
            'book_completeness': first.get('book_completeness'),
            'statement_check':  first.get('statement_check'),
            'account_tieout':   first.get('account_tieout'),
            'recon_date':       first.get('recon_date', ''),
            'statement_from':   first.get('statement_from'),
            'statement_to':     first.get('statement_to'),
            'date_corrected':   first.get('date_corrected', False),
            'typed_date':       typed_date,
        })
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)})
    finally:
        for p in tmp_paths:
            try: os.unlink(p)
            except Exception: pass

@app.route('/contacts', methods=['GET'])
def api_contacts_get():
    try:
        return jsonify({'ok': True, 'contacts': read_contacts()})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e), 'contacts': []})

@app.route('/contacts', methods=['POST'])
def api_contacts_post():
    try:
        contacts = request.json.get('contacts', [])
        save_contacts(contacts)
        return jsonify({'ok': True})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)})

@app.route('/reconcile/records', methods=['GET'])
def api_reconcile_records():
    try:
        records = read_accounts_from_gsheet()
        out = [{'id': r.get('Disbursement ID',''), 'customer': r.get('Customer Name',''),
                'branch': r.get('Branch','')} for r in records if r.get('Disbursement ID','').strip()]
        return jsonify({'ok': True, 'records': out})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e), 'records': []})

@app.route('/capital-log', methods=['GET'])
def api_capital_log_get():
    try:
        sh = get_gspread_client().open_by_key(SPREADSHEET_ID)
        cl = read_capital_log(sh)
        if not cl or not cl.get('available'):
            return jsonify({'ok': False, 'available': False,
                             'error': (cl or {}).get('error', 'Capital Log tab not found')})
        entries = [{**e, 'date_str': e['date_str']} for e in cl['entries']]
        for e in entries:
            e.pop('date', None)  # datetime.date isn't JSON-serializable; date_str already has it
        return jsonify({'ok': True, 'available': True, 'entries': entries, 'net_capital': cl['net_capital']})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)})

@app.route('/capital-log', methods=['POST'])
def api_capital_log_post():
    try:
        data = request.json or {}
        missing = [k for k in ('date', 'type', 'amount') if not str(data.get(k, '')).strip()]
        if missing:
            return jsonify({'ok': False, 'error': f"Missing required field(s): {', '.join(missing)}"})
        if data['type'].strip().upper() not in mis.CAPITAL_LOG_KNOWN_TYPES:
            return jsonify({'ok': False, 'error': f"Unknown TYPE — must be one of {sorted(mis.CAPITAL_LOG_KNOWN_TYPES)}"})
        new_balance = append_capital_log_entry(data)
        return jsonify({'ok': True, 'new_running_balance': new_balance})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)})

@app.route('/solvency-check', methods=['GET'])
def api_solvency_check():
    try:
        return jsonify(_solvency_check())
    except Exception as e:
        return jsonify({'status': 'error', 'error': str(e)}), 500

@app.route('/charge-integrity')
def api_charge_integrity():
    try:
        return jsonify({'ok': True, **_charge_integrity_check()})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500

@app.route('/money-integrity', methods=['GET'])
def api_money_integrity():
    try:
        return jsonify(_money_integrity_check())
    except Exception as e:
        import traceback
        return jsonify({'all_clear': False, 'error': str(e),
                         'trace': traceback.format_exc()}), 500

# ── Reports ──────────────────────────────────────────────────────────────────
# Data-driven catalogue: the frontend renders whatever this returns, so adding
# a report later needs no frontend change at all.
REPORT_CATALOGUE = [
    {'id': 'pl', 'label': 'Profit & Loss Statement', 'group': 'Governance',
     'needs_period': True, 'needs_cluster': False,
     'desc': 'Cash-basis P&L: charges earned on collection, less expenses by category. '
             'GST and partner drawings shown below the line.'},
    {'id': 'integrity', 'label': 'Money Integrity & Audit Report', 'group': 'Governance',
     'needs_period': False, 'needs_cluster': False,
     'desc': 'Every distinct way a rupee could go missing, each proven. The report to hand '
             'partners and investors.'},
    {'id': 'position', 'label': 'Statement of Position', 'group': 'Governance',
     'needs_period': False, 'needs_cluster': False,
     'desc': 'Assets (bank + FD + receivables) against how they are funded. Deployable cash '
             'is the headline figure.'},
    {'id': 'gst', 'label': 'GST Register', 'group': 'Governance',
     'needs_period': False, 'needs_cluster': False,
     'desc': 'Month-wise from Apr-2026, collection basis. Mirrors the GST Register sheet exactly.'},
    {'id': 'consolidated_mis', 'label': 'Consolidated MIS', 'group': 'Portfolio',
     'needs_period': False, 'needs_cluster': False,
     'desc': 'Full portfolio dashboard: KPIs, open cases, cluster analytics, ROI, monthly and MTD/YTD.'},
    {'id': 'cluster_mis', 'label': 'Cluster MIS', 'group': 'Portfolio',
     'needs_period': False, 'needs_cluster': True,
     'desc': 'One cluster\'s own pack: KPI strip, MTD/YTD, branch leaderboard ranked by TAT, open cases.'},
    {'id': 'escalation', 'label': 'Collection Follow-Up / Escalation', 'group': 'Collections',
     'needs_period': False, 'needs_cluster': False,
     'desc': 'Calling list for every open case with the full escalation chain - field officer, '
             'territory, area and cluster manager, with phone numbers.'},
]

@app.route('/reports/catalogue', methods=['GET'])
def api_reports_catalogue():
    try:
        clusters = []
        try:
            records = read_accounts_from_gsheet()
            clusters = sorted({(r.get('Cluster', '') or '').strip() for r in records
                               if (r.get('Cluster', '') or '').strip()})
        except Exception:
            pass
        return jsonify({'ok': True, 'reports': REPORT_CATALOGUE, 'clusters': clusters})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500

@app.route('/reports/generate', methods=['POST'])
def api_reports_generate():
    """One route, dispatching on `report`. Deliberately READ-ONLY: unlike
    /generate-mis it never calls write_claude_dashboard_to_sheet(), so pulling
    a report can never mutate the spreadsheet."""
    try:
        data = request.json or {}
        report = (data.get('report') or '').strip()
        ensure_mis_assets_cached()
        today = date.today()

        def _d(key, default=None):
            v = (data.get(key) or '').strip()
            return _parse_flex_date(v) if v else default

        if report == 'pl':
            frm = _d('from', date(today.year if today.month >= 4 else today.year - 1, 4, 1))
            to = _d('to', today)
            fin = _period_financials(from_date=frm, to_date=to)
            label = f"{frm.strftime('%d-%b-%Y')} to {to.strftime('%d-%b-%Y')}"
            pdf_bytes = mis.generate_pl_statement(fin, label)
            fname = f"BridgeLine P&L {frm.strftime('%d%m%y')}-{to.strftime('%d%m%y')}.pdf"

        elif report == 'integrity':
            integrity = _money_integrity_check()
            summary = get_today_summary()
            pdf_bytes = mis.generate_money_integrity_report(integrity, summary=summary)
            fname = f"BridgeLine Money Integrity {today.strftime('%d%m%y')}.pdf"

        elif report == 'position':
            summary = get_today_summary()
            try:
                capital_log = read_capital_log(get_gspread_client().open_by_key(SPREADSHEET_ID))
            except Exception:
                capital_log = None
            pdf_bytes = mis.generate_position_statement(summary, capital_log=capital_log)
            fname = f"BridgeLine Position {today.strftime('%d%m%y')}.pdf"

        elif report == 'gst':
            sh = get_gspread_client().open_by_key(SPREADSHEET_ID)
            records = read_accounts_from_gsheet()
            periods = _load_recon_periods(get_recon_txns_sheet(sh))
            months, y, m = [], 2026, 4
            while (y, m) <= (today.year, today.month):
                frm = date(y, m, 1)
                to = date(y + (m == 12), (m % 12) + 1, 1) - timedelta(days=1)
                fin = _period_financials(from_date=frm, to_date=min(to, today),
                                          periods=periods, records=records, sh=sh)
                months.append({'label': frm.strftime('%b %Y'), 'charges': fin['charges_earned'],
                                'gst': fin['gst_collected'], 'cases': fin['cases_closed']})
                m += 1
                if m > 12:
                    m = 1; y += 1
            pdf_bytes = mis.generate_gst_register_pdf(months)
            fname = f"BridgeLine GST Register {today.strftime('%d%m%y')}.pdf"

        elif report in ('consolidated_mis', 'cluster_mis', 'escalation'):
            sh = get_gspread_client().open_by_key(SPREADSHEET_ID)
            rows, db_raw, mcoll, (cluster_mgrs, branch_contacts, area_mgrs,
                                   territory_to_area), capital_log = load_data_from_sheet(sh)
            open_cases, all_cases, all_cases_full = mis.parse_cases(rows, mcoll)
            metrics = mis.compute_dashboard_metrics(all_cases_full, db_raw, capital_log)
            if report == 'consolidated_mis':
                pdf_bytes = mis.generate_consolidated_mis(open_cases, all_cases, all_cases_full, metrics)
                fname = f"BridgeLine Consolidated MIS {today.strftime('%d%m%y')}.pdf"
            elif report == 'cluster_mis':
                cluster = (data.get('cluster') or '').strip()
                if not cluster:
                    return jsonify({'ok': False, 'error': 'Please choose a cluster.'}), 400
                pdf_bytes = mis.generate_cluster_mis(cluster, open_cases, all_cases, metrics)
                fname = f"BridgeLine {cluster} MIS {today.strftime('%d%m%y')}.pdf"
            else:
                pdf_bytes = mis.generate_calling_followup_pdf(
                    open_cases, cluster_mgrs, branch_contacts, area_mgrs, territory_to_area)
                fname = f"BridgeLine Collection Follow-Up {today.strftime('%d%m%y')}.pdf"
        else:
            return jsonify({'ok': False, 'error': f'Unknown report: {report}'}), 400

        return Response(pdf_bytes, mimetype='application/pdf',
                        headers={'Content-Disposition': f'attachment; filename="{fname}"'})
    except Exception as e:
        import traceback
        return jsonify({'ok': False, 'error': str(e), 'trace': traceback.format_exc()}), 500

@app.route('/period-financials', methods=['GET'])
def api_period_financials():
    try:
        frm = _parse_flex_date(request.args.get('from', '') or '')
        to = _parse_flex_date(request.args.get('to', '') or '')
        return jsonify({'ok': True, **_period_financials(from_date=frm, to_date=to)})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500

@app.route('/fd-balance', methods=['GET'])
def api_fd_balance_get():
    try:
        bal, d = get_latest_fd_balance()
        return jsonify({'ok': True, 'balance': bal, 'date': d.strftime('%d-%m-%Y') if d else None})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)})

@app.route('/fd-balance', methods=['POST'])
def api_fd_balance_post():
    try:
        data = request.json or {}
        if not str(data.get('date', '')).strip() or str(data.get('balance', '')).strip() == '':
            return jsonify({'ok': False, 'error': 'Date and balance are required.'})
        bal, d = save_fd_balance(data['date'], data['balance'])
        return jsonify({'ok': True, 'balance': bal, 'date': d.strftime('%d-%m-%Y')})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)})

@app.route('/reconcile/save', methods=['POST'])
def api_reconcile_save():
    """Persists the period to the 'Recon Txns' sheet tab and returns the FULL
    cumulative reconciliation workbook (all saved periods) as a direct
    download — see save_reconciliation()'s docstring. No file re-upload is
    needed to keep history; the Google Sheet is the system of record.
    Summary numbers (rows saved, closing balance) are echoed back in response
    headers so the frontend can show a status message without parsing a JSON
    body on a binary response.
    """
    try:
        data = request.json
        # Multi-account batch, or the single-statement shape wrapped as a
        # one-element batch so there is only ever one code path.
        batch = data.get('statements') or [{
            'date':         data.get('date'),
            'opening':      data.get('opening'),
            'closing':      data.get('closing'),
            'transactions': data.get('transactions'),
            'remarks':      data.get('remarks', ''),
            'remarks_map':  data.get('remarks_map', {}),
            'account':      data.get('account', ''),
        }]

        # Saved one at a time — the Sheet append is not transactional, so a
        # failure partway must report exactly which accounts made it in
        # rather than implying all or nothing. save_reconciliation() replaces
        # any existing (date, account) period, so re-running after a partial
        # failure is safe and won't duplicate what already saved.
        saved, result = [], None
        for st in batch:
            if not (st.get('account') or '').strip():
                raise ValueError('Every statement needs its bank account set before saving.')
            if not (st.get('date') or '').strip():
                raise ValueError(f"No reconciliation date for the {st['account']} statement.")
            try:
                result = save_reconciliation(
                    recon_date=st['date'],
                    opening_balance=float(st.get('opening') or 0),
                    closing_balance=float(st.get('closing') or 0),
                    transactions=st['transactions'],
                    remarks=st.get('remarks', ''),
                    remarks_map=st.get('remarks_map', {}),
                    account=st.get('account', ''),
                )
                saved.append({'account': st['account'], 'date': st['date'],
                              'rows': result['rows_saved'], 'closing': result['closing'],
                              'replaced': result.get('rows_superseded', 0)})
            except Exception as e:
                done = ', '.join(f"{s['account']} ({s['rows']} rows)" for s in saved) or 'none'
                raise RuntimeError(
                    f"Saved: {done}. Then FAILED on {st.get('account')}: {e}. "
                    f"Re-run the save — already-saved accounts are replaced, not duplicated.")

        # The workbook is cumulative across every period ever saved, so the
        # last call's bytes already contain all of this batch.
        return Response(
            result['file_bytes'],
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={
                'Content-Disposition': f'attachment; filename="{result["filename"]}"',
                'X-Rows-Saved': str(sum(s['rows'] for s in saved)),
                'X-Closing-Balance': str(result['closing']),
                'X-Accounts-Saved': json.dumps(saved),
            },
        )
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500

# Hosted on Vercel via api/index.py (WSGI). No local dev-server / ngrok /
# browser-open block needed here — that only applied to the Mac-only version.
